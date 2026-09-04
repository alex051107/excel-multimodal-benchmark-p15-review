#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-C-PO-ADDENDUM-001."""
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-C-PO-ADDENDUM-001",
  "pass_threshold": 0.7,
  "criteria": [
    {
      "id": "R001",
      "description": "Workbook contains a usable revised PO delivery structure.",
      "weight": 2,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R002",
      "description": "The base schedule remains intact, including the unlisted protected line.",
      "weight": 4,
      "type": "positive",
      "dimension": "change_locality",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R003",
      "description": "The revised schedule applies each addendum change to the correct row identity.",
      "weight": 6,
      "type": "positive",
      "dimension": "revision_semantics",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R004",
      "description": "Revised line totals and total are formula-linked and respond to a quantity perturbation.",
      "weight": 5,
      "type": "positive",
      "dimension": "native_workbook_semantics",
      "method": "deterministic",
      "method_params": {
        "perturbations": 2
      }
    },
    {
      "id": "R005",
      "description": "Revision log distinguishes quantity change, insertion, and unit-price change.",
      "weight": 4,
      "type": "positive",
      "dimension": "auditability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R006",
      "description": "Each changed row has correct document-page provenance and the protected row remains unchanged.",
      "weight": 3,
      "type": "positive",
      "dimension": "provenance",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "P001",
      "description": "Penalty for contradictory source-file or addendum-page claims in provenance.",
      "weight": -7,
      "type": "penalty",
      "dimension": "integrity",
      "method": "deterministic",
      "method_params": {}
    }
  ],
  "required_sheets": [
    "PO_Header",
    "Base_Schedule",
    "Revised_Schedule",
    "Revision_Log",
    "Provenance",
    "Checks"
  ]
}
TESTS_ROOT = Path(__file__).resolve().parent
REPO_TASK_ROOT = TESTS_ROOT.parent
DEV_ROOT = REPO_TASK_ROOT if (REPO_TASK_ROOT / "rubric.json").is_file() else TESTS_ROOT / "dev"
TASK = json.loads((DEV_ROOT / "rubric.json").read_text(encoding="utf-8"))
TASK["required_sheets"] = json.loads(
    (TESTS_ROOT / "private_contract.json").read_text(encoding="utf-8")
)["required_sheets"]
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R003", "R004", "R005"]
SHEET_ALIASES = {}


def semantic_sheet_aliases(workbook):
    signatures = {
        "PO_Header": (("PO ID", "Purchase order ID"), ("Addendum",), ("Effective deadline", "Effective date"), ("Revision reason", "Reason"), ("Currency",)),
        "Base_Schedule": (("Line ID", "ID"), ("Description",), ("Quantity", "Qty"), ("Unit price", "Price"), ("Extended total", "Line total"), ("Protected",)),
        "Revised_Schedule": (("Line ID", "ID"), ("Description",), ("Quantity", "Qty"), ("Unit price", "Price"), ("Extended total", "Line total"), ("Revision status", "Status")),
        "Revision_Log": (("Line ID", "ID"), ("Change type", "Change"), ("Old value", "Before"), ("New value", "After"), ("Source page", "Page")),
        "Provenance": (("Line ID", "ID"), ("Source document", "Source file", "Document"), ("Page", "Source page"), ("Text locator", "Locator", "Source locator")),
        "Checks": (("Check",), ("Observed",), ("Expected",), ("Interpretation",)),
    }
    aliases = {}
    for role, groups in signatures.items():
        candidates = []
        for sheet in workbook.worksheets:
            values = {
                norm(cell.value).rstrip(":")
                for row in sheet.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            }
            if all(any(norm(alias).rstrip(":") in values for alias in group) for group in groups):
                candidates.append(sheet.title)
        if candidates:
            aliases[role] = tuple(candidates)
    return aliases


def key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


def active_split(explicit=None):
    split = (explicit or os.environ.get("P15_EVAL_SPLIT", "dev")).strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"UNSUPPORTED_EVAL_SPLIT:{split}")
    return split


def load_oracle(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "oracle_recompute.py" if split == "confirm" else DEV_ROOT / "metadata" / "oracle_recompute.py"
    spec = importlib.util.spec_from_file_location(f"task_oracle_{split}", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


class UnsupportedFormulaError(ValueError):
    """A valid Excel formula that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        resolved_sheet = self.workbook.actual_sheet_name(sheet) if hasattr(self.workbook, "actual_sheet_name") else sheet
        current = key(resolved_sheet, cell)
        if current in self.overrides: return self.overrides[current]
        if current in self.memo: return self.memo[current]
        if current in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{current}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(current)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], resolved_sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[current] = result
            return result
        finally:
            self.stack.discard(current)

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(col)}{row}") for row in range(int(r1), int(r2) + 1) for col in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def reference_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, reference_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            if isinstance(node.op, ast.Add): return left + right
            if isinstance(node.op, ast.Sub): return left - right
            if isinstance(node.op, ast.Mult): return left * right
            if isinstance(node.op, ast.Div): return left / right
            if isinstance(node.op, ast.Pow): return left ** right
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA:OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [item for group in args for item in (group if isinstance(group, list) else [group]) if item is not None]
            if name == "SUM": return sum(values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA:NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(float(actual) - float(expected)) <= tolerance


def text(value): return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", text(value).lower())


def contains_number(value, expected):
    numbers = re.findall(r"-?\d+(?:,\d{3})*(?:\.\d+)?", text(value))
    return any(close(float(number.replace(",", "")), expected) for number in numbers)


def load_perturbations(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "perturbations.json" if split == "confirm" else DEV_ROOT / "metadata" / "perturbations.json"
    records = json.loads(path.read_text(encoding="utf-8"))
    expected_count = 1 if split == "confirm" else 2
    if len(records) != expected_count or len({record.get("id") for record in records}) != expected_count:
        raise ValueError(f"{split.upper()}_PERTURBATION_MANIFEST_MUST_CONTAIN_{expected_count}_UNIQUE_CASES")
    return records


def find_content_row(workbook, spec):
    sheet = workbook[spec["sheet"]]
    primary_col = col_index(spec["match_column"])
    secondary_col = col_index(spec["secondary_match_column"]) if spec.get("secondary_match_column") else None
    rows = []
    for row in range(4, sheet.max_row + 1):
        if norm(sheet.cell(row=row, column=primary_col).value) != norm(spec["match_value"]):
            continue
        if secondary_col and norm(sheet.cell(row=row, column=secondary_col).value) != norm(spec["secondary_match_value"]):
            continue
        rows.append(row)
    return rows[0] if len(rows) == 1 else None


def perturbation_case_ok(workbook, case):
    target = case["target"]
    row = find_content_row(workbook, target)
    if row is None:
        return False
    target_cell = f'{target["value_column"]}{row}'
    baseline_engine = FormulaEngine(workbook)
    if not close(baseline_engine.value(target["sheet"], target_cell), target["baseline"]):
        return False
    engine = FormulaEngine(workbook, {key(target["sheet"], target_cell): target["perturbed"]})
    for expected in case.get("expected", []):
        if not close(engine.value(expected["sheet"], expected["cell"]), expected["value"]):
            return False
    for expected in case.get("expected_by_content", []):
        expected_row = find_content_row(workbook, expected)
        if expected_row is None or not close(
            engine.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'), expected["value"]
        ):
            return False
    for protected in case.get("protected", []):
        if not close(engine.value(protected["sheet"], protected["cell"]), protected["value"]):
            return False
    for protected in case.get("protected_by_content", []):
        protected_row = find_content_row(workbook, protected)
        if protected_row is None or not close(
            engine.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'), protected["value"]
        ):
            return False
    return True


def perturbation_response_ok(workbook, split="dev"):
    return all(perturbation_case_ok(workbook, case) for case in load_perturbations(split))


def status_role(value):
    words = set(re.findall(r"[a-z]+", norm(value).replace("-", " ")))
    if "unchanged" in words or "retained" in words or ({"no", "change"} <= words):
        return "unchanged"
    if words & {"insert", "inserted", "new", "added", "addition"}:
        return "inserted"
    if words & {"change", "changed", "modified", "revised", "updated"}:
        return "changed"
    return None


def status_matches(actual, expected):
    return status_role(actual) == status_role(expected)


def change_type_role(value):
    words = set(re.findall(r"[a-z]+", norm(value).replace("-", " ")))
    if words & {"insert", "inserted", "new", "added", "addition"}:
        return "insertion"
    if words & {"quantity", "qty", "count", "units"}:
        return "quantity"
    if words & {"price", "pricing", "cost", "rate"}:
        return "unit_price"
    return None


def change_type_matches(actual, expected):
    expected_text = norm(expected)
    expected_role = "insertion" if "insert" in expected_text else "quantity" if "quantity" in expected_text else "unit_price" if "price" in expected_text else None
    return expected_role is not None and change_type_role(actual) == expected_role


def page_matches(value, expected):
    if isinstance(value, (int, float)):
        return int(value) == value and int(value) == int(expected)
    rendered = norm(value)
    if rendered == str(int(expected)):
        return True
    return re.search(rf"\b(?:p(?:age)?\.?\s*[:#-]?\s*){int(expected)}\b", rendered) is not None


def readable_locator(value):
    rendered = text(value)
    return bool(rendered) and len(re.sub(r"\s+", "", rendered)) >= 3


def source_field_matches(value, filename):
    rendered, expected = norm(value), norm(filename)
    return bool(rendered and expected and expected in rendered)


def source_field_contradicts(value, filename):
    rendered = text(value)
    if not rendered or source_field_matches(rendered, filename):
        return False
    return re.search(r"\.(?:pdf|png|jpe?g|tiff?|xlsx?|csv)\b", rendered, re.IGNORECASE) is not None


def present_formula(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def formula_value(engine, sheet, cell):
    try:
        return engine.value(sheet, cell)
    except UnsupportedFormulaError:
        raise
    except Exception:
        return None


def schedule_total_formula_cells(workbook, engine):
    """Find formula cells on rows visibly labeled as the revised total."""
    cells = []
    for sheet in workbook.worksheets:
        for row in range(1, sheet.max_row + 1):
            labels = " ".join(
                norm(sheet.cell(row=row, column=column).value)
                for column in range(1, sheet.max_column + 1)
                if text(sheet.cell(row=row, column=column).value)
            )
            if "total" not in labels or not any(token in labels for token in ("revised", "schedule")):
                continue
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=column)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cells.append((sheet.title, cell.coordinate, formula_value(engine, sheet.title, cell.coordinate)))
    return cells


def po_perturbation_response_ok(workbook, split="dev"):
    total_cells = schedule_total_formula_cells(workbook, FormulaEngine(workbook))
    if not total_cells:
        return False
    for case in load_perturbations(split):
        target = case["target"]
        row = find_content_row(workbook, target)
        if row is None:
            return False
        target_cell = f'{target["value_column"]}{row}'
        baseline = FormulaEngine(workbook)
        if not close(baseline.value(target["sheet"], target_cell), target["baseline"]):
            return False
        target_sheet = workbook.actual_sheet_name(target["sheet"]) if hasattr(workbook, "actual_sheet_name") else target["sheet"]
        probe = FormulaEngine(workbook, {key(target_sheet, target_cell): target["perturbed"]})
        expected_totals = {
            float(expected["value"])
            for expected in case.get("expected", [])
            if norm(expected.get("sheet")) in {norm("Revised_Schedule"), norm("PO_Header")}
        }
        if len(expected_totals) != 1:
            return False
        expected_total = next(iter(expected_totals))
        if not all(close(probe.value(sheet, cell), expected_total) for sheet, cell, _ in total_cells):
            return False
        for expected in case.get("expected_by_content", []):
            expected_row = find_content_row(workbook, expected)
            if expected_row is None or not close(
                probe.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'),
                expected["value"],
            ):
                return False
        for protected in case.get("protected_by_content", []):
            protected_row = find_content_row(workbook, protected)
            if protected_row is None or not close(
                probe.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'),
                protected["value"],
            ):
                return False
    return True


def row_values(workbook, sheet, start, end, cols):
    records = []
    for row in range(start, end + 1):
        values = tuple(workbook[sheet].cell(row=row, column=col).value for col in cols)
        if any(value not in (None, "") for value in values): records.append(values)
    return records


def task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    checks["R001"] = 1.0
    base_sheet = workbook["Base_Schedule"]
    base_actual = {}
    base_duplicates = set()
    for row in range(4, base_sheet.max_row + 1):
        line_id = text(base_sheet[f"A{row}"].value)
        if line_id:
            if line_id in base_actual:
                base_duplicates.add(line_id)
            base_actual[line_id] = {
                "description": text(base_sheet[f"B{row}"].value), "quantity": base_sheet[f"C{row}"].value,
                "unit_price": base_sheet[f"D{row}"].value, "unit": text(base_sheet[f"E{row}"].value),
                "protected": norm(base_sheet[f"G{row}"].value) in {"yes", "true", "protected"},
            }
    base_expected = {row["line_id"]: row for row in oracle["base"]}
    base_ok = not base_duplicates and len(base_actual) == len(base_expected) and set(base_actual) == set(base_expected) and all(
        norm(base_actual[line_id]["description"]) == norm(item["description"])
        and close(base_actual[line_id]["quantity"], item["quantity"])
        and close(base_actual[line_id]["unit_price"], item["unit_price"])
        and norm(base_actual[line_id]["unit"]) == norm(item["unit"])
        and base_actual[line_id]["protected"] == item["protected"]
        for line_id, item in base_expected.items()
    )
    checks["R002"] = 1.0 if base_ok else 0.0
    revised_sheet = workbook["Revised_Schedule"]
    revised_actual = {}
    revised_duplicates = set()
    for row in range(4, revised_sheet.max_row + 1):
        line_id = text(revised_sheet[f"A{row}"].value)
        if line_id:
            if line_id in revised_actual:
                revised_duplicates.add(line_id)
            revised_actual[line_id] = {
                "row": row, "description": text(revised_sheet[f"B{row}"].value),
                "quantity": revised_sheet[f"C{row}"].value, "unit_price": revised_sheet[f"D{row}"].value,
                "unit": text(revised_sheet[f"E{row}"].value),
                "extended": formula_value(engine, "Revised_Schedule", f"F{row}"),
                "status": text(revised_sheet[f"G{row}"].value),
            }
    revised_expected = {row["line_id"]: row for row in oracle["revised"]}
    revised_ok = not revised_duplicates and len(revised_actual) == len(revised_expected) and set(revised_actual) == set(revised_expected) and all(
        norm(revised_actual[line_id]["description"]) == norm(item["description"])
        and close(revised_actual[line_id]["quantity"], item["quantity"])
        and close(revised_actual[line_id]["unit_price"], item["unit_price"])
        and norm(revised_actual[line_id]["unit"]) == norm(item["unit"])
        and close(revised_actual[line_id]["extended"], item["extended"])
        and status_matches(revised_actual[line_id]["status"], item["status"])
        for line_id, item in revised_expected.items()
    )
    checks["R003"] = 1.0 if revised_ok else 0.0
    protected_expected = {row["line_id"]: row for row in oracle["revised"] if row.get("protected")}
    # An empty protected set means the oracle declared nothing to protect, so
    # there is nothing to overwrite.  `all(())` is True, which is correct here.
    protected_ok = all(
        line_id in revised_actual
        and norm(revised_actual[line_id]["description"]) == norm(item["description"])
        and close(revised_actual[line_id]["quantity"], item["quantity"])
        and close(revised_actual[line_id]["unit_price"], item["unit_price"])
        and norm(revised_actual[line_id]["unit"]) == norm(item["unit"])
        and status_matches(revised_actual[line_id]["status"], item["status"])
        for line_id, item in protected_expected.items()
    )
    checks["R004"] = 1.0 if protected_ok else 0.0
    line_formula_ok = set(revised_actual) == set(revised_expected) and all(
        present_formula(workbook, f'Revised_Schedule!F{item["row"]}')
        and close(item["extended"], item["quantity"] * item["unit_price"])
        for item in revised_actual.values()
    )
    total_cells = schedule_total_formula_cells(workbook, engine)
    total_ok = bool(total_cells) and all(close(value, oracle["total"]) for _, _, value in total_cells)
    checks["R005"] = 1.0 if line_formula_ok and total_ok else 0.0
    try:
        dynamic_ok = po_perturbation_response_ok(workbook, split)
    except Exception as exc:
        dynamic_ok = False; failures.append(f"PO_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R006"] = 1.0 if dynamic_ok else 0.0
    log_sheet = workbook["Revision_Log"]
    log_rows = {}
    for row in range(4, log_sheet.max_row + 1):
        line_id = text(log_sheet[f"A{row}"].value)
        if line_id:
            log_rows.setdefault(line_id, []).append({
                "change_type": text(log_sheet[f"B{row}"].value), "before": text(log_sheet[f"C{row}"].value),
                "after": text(log_sheet[f"D{row}"].value), "page": log_sheet[f"E{row}"].value,
            })
    expected_changes = {row["line_id"]: row for row in oracle["changes"]}
    base_by_id = {row["line_id"]: row for row in oracle["base"]}

    def evidence_ok(line_id, change, record):
        change_type = norm(change["change_type"])
        if "insert" in change_type:
            before = norm(record["before"])
            before_ok = not before or before in {"n/a", "none"} or any(token in before for token in ("not", "new", "insert"))
            return before_ok and contains_number(record["after"], change["unit_price"])
        base = base_by_id.get(line_id)
        if not base:
            return False
        if "quantity" in change_type:
            return contains_number(record["before"], base["quantity"]) and contains_number(record["after"], change["quantity"])
        if "price" in change_type:
            return contains_number(record["before"], base["unit_price"]) and contains_number(record["after"], change["unit_price"])
        return False

    log_ok = set(expected_changes).issubset(log_rows) and all(
        len(log_rows[line_id]) == 1
        and change_type_matches(log_rows[line_id][0]["change_type"], change["change_type"])
        and evidence_ok(line_id, change, log_rows[line_id][0])
        and page_matches(log_rows[line_id][0]["page"], change["page"])
        for line_id, change in expected_changes.items()
    )
    checks["R007"] = 1.0 if log_ok else 0.0
    provenance_sheet = workbook["Provenance"]
    provenance_rows = {}
    for row in range(4, provenance_sheet.max_row + 1):
        line_id = text(provenance_sheet[f"A{row}"].value)
        if line_id:
            provenance_rows.setdefault(line_id, []).append((
                text(provenance_sheet[f"B{row}"].value), provenance_sheet[f"C{row}"].value,
                text(provenance_sheet[f"D{row}"].value), text(provenance_sheet[f"E{row}"].value)
            ))
    provenance_population_ok = (
        set(provenance_rows) == set(expected_changes)
        and sum(len(records) for records in provenance_rows.values()) == len(expected_changes)
        and all(len(records) == 1 for records in provenance_rows.values())
    )
    provenance_ok = provenance_population_ok and all(
        len(provenance_rows[line_id]) == 1
        and source_field_matches(provenance_rows[line_id][0][0], oracle["document"]["filename"])
        and (
            page_matches(provenance_rows[line_id][0][1], change["page"])
        )
        and readable_locator(provenance_rows[line_id][0][2] or provenance_rows[line_id][0][3])
        for line_id, change in expected_changes.items()
    )
    checks["R008"] = 1.0 if provenance_ok else 0.0
    identity_ok = (
        not revised_duplicates
        and len(revised_actual) == len(revised_expected)
        and set(expected_changes).issubset(log_rows)
        and all(len(records) == 1 for records in log_rows.values())
        and provenance_population_ok
    )
    checks["R009"] = 1.0 if identity_ok else 0.0
    false_provenance = any(
        (
            source_field_contradicts(record[0], oracle["document"]["filename"])
        )
        or (
            line_id in expected_changes
            and record[1] not in (None, "")
            and not page_matches(record[1], expected_changes[line_id]["page"])
        )
        for line_id, records in provenance_rows.items()
        for record in records
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_ADDENDUM_PROVENANCE")
    return checks, failures


def score(candidate, split=None):
    split = active_split(split)
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=["OUTPUT_MISSING"],
        )
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=[f"MALFORMED_XLSX:{type(exc).__name__}"],
        )

    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], semantic_sheet_aliases(workbook)
    )
    checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=failures,
        )

    try:
        oracle = load_oracle(split)
        engine = FormulaEngine(workbook)
        checks, task_failures = task_checks(workbook, engine, oracle, split)
        checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=checks,
        failures=failures,
    )


if __name__ == "__main__":
    arguments = sys.argv[1:]
    selected_split = None
    if "--split" in arguments:
        split_index = arguments.index("--split")
        if split_index + 1 >= len(arguments):
            raise SystemExit("--split requires dev or confirm")
        selected_split = arguments[split_index + 1]
        del arguments[split_index:split_index + 2]
    if len(arguments) > 1:
        raise SystemExit("usage: evaluate.py [candidate.xlsx] [--split dev|confirm]")
    target = Path(arguments[0]) if arguments else Path("/app/output/answer.xlsx")
    print(json.dumps(score(target, selected_split), sort_keys=True))

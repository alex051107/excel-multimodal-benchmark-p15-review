#!/usr/bin/env python3
"""Focused Judge V3 regression: one declared total and plain-language change labels."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


TASK_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TASK_DIR / "tests"))
SPEC = importlib.util.spec_from_file_location("po_evaluator", TASK_DIR / "tests" / "evaluate.py")
EVALUATOR = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(EVALUATOR)


def build_single_total_plain_labels(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    workbook["Revised_Schedule"]["F8"] = None
    workbook["PO_Header"]["B9"] = "=SUM(Revised_Schedule!F4:F7)"
    workbook["Checks"]["B6"] = None
    workbook["Revision_Log"]["B4"] = "Quantity"
    workbook["Revision_Log"]["B5"] = "Insert"
    workbook["Revision_Log"]["B6"] = "Unit price"
    workbook.save(path)


def build_renamed_sheets(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    names = ("Order overview", "Original schedule", "Updated schedule", "Change register", "Source register", "Review notes")
    for sheet, name in zip(workbook.worksheets, names):
        sheet.title = name
    workbook.save(path)


def result_record(path: Path, expected: str, predicate, display_path: str | None = None) -> dict:
    result = EVALUATOR.score(path, "dev")
    return {
        "path": display_path or str(path.relative_to(TASK_DIR)),
        "score": result["normalized_score"],
        "pass": result["pass"],
        "criterion_scores": result["criterion_scores"],
        "failure_codes": result["failure_codes"],
        "expected": expected,
        "expected_pass": expected.startswith(("full pass", "pass with")),
        "expectation_met": bool(predicate(result)),
    }


def main() -> None:
    renamed_path = TASK_DIR / "fixtures/equivalent/candidate_renamed_sheets.xlsx"
    build_renamed_sheets(renamed_path)
    fixtures = {
        "reference": (TASK_DIR / "solution/reference.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "equivalent": (TASK_DIR / "fixtures/equivalent/candidate_equivalent.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "noop": (TASK_DIR / "fixtures/noop/candidate_noop.xlsx", "fail", lambda r: not r["pass"]),
        "malformed": (TASK_DIR / "fixtures/malformed/candidate_malformed.xlsx", "fail with zero score", lambda r: not r["pass"] and r["normalized_score"] == 0.0),
    }
    for path in sorted((TASK_DIR / "fixtures/mutants").glob("*.xlsx")):
        fixtures[f"mutant:{path.stem}"] = (path, "fail", lambda r: not r["pass"])
    fixture_results = {
        name: result_record(path, expected, predicate)
        for name, (path, expected, predicate) in fixtures.items()
    }

    with tempfile.TemporaryDirectory(prefix="po-judge-v3-") as directory:
        single_total = Path(directory) / "single_total_plain_labels.xlsx"
        build_single_total_plain_labels(single_total)
        single_total_result = result_record(
            single_total,
            "full pass: one formula-linked revised total and concise change labels",
            lambda r: r["pass"] and r["normalized_score"] == 1.0,
            "generated:single_total_plain_labels",
        )
        broken = Path(directory) / "wrong_unit_price.xlsx"
        workbook = load_workbook(single_total)
        workbook["Revised_Schedule"]["D6"] = 999.0
        workbook.save(broken)
        broken_result = result_record(
            broken,
            "fail R003/R005: revised unit price and total are wrong",
            lambda r: not r["pass"] and r["criterion_scores"]["R003"] == 0.0 and r["criterion_scores"]["R005"] == 0.0,
            "generated:wrong_unit_price",
        )

    renamed_result = result_record(
        renamed_path,
        "full pass: every sheet is renamed while visible headers preserve its business role",
        lambda r: r["pass"] and r["normalized_score"] == 1.0,
    )
    new_cases = {"renamed_sheets": renamed_result, "single_total_plain_labels": single_total_result, "wrong_unit_price": broken_result}
    all_met = all(item["expectation_met"] for item in (*fixture_results.values(), *new_cases.values()))
    receipt = {
        "task_id": "P15-C-PO-ADDENDUM-001",
        "judge_version": "p15_v3.0",
        "validation_scope": "reference, equivalent, noop, malformed, all mutants, single-total equivalent, genuine unit-price error",
        "expectation_basis": {
            "single_total_plain_labels": "The instruction requires a formula-linked revised total and an intelligible revision log, not two duplicate totals or exact phrases.",
            "renamed_sheets": "The instruction does not prescribe worksheet names; visible headers and labels establish each business role.",
            "wrong_unit_price": "The addendum explicitly changes PO-03 to the stated unit price and requires the revised schedule to reflect it.",
        },
        "fixtures": fixture_results,
        "new_cases": new_cases,
        "all_expectations_met": all_met,
    }
    receipt_path = TASK_DIR / "receipts/judge_v3_local_validation.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"task_id": receipt["task_id"], "all_expectations_met": all_met, "scores": {key: [value["score"], value["pass"]] for key, value in {**fixture_results, **new_cases}.items()}}, ensure_ascii=False))
    if not all_met:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

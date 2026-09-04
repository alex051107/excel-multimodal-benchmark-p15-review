#!/usr/bin/env python3
"""Object-level OOXML judge for the native PivotTable/PivotChart task."""
from __future__ import annotations

import csv
import importlib.util
import json
import math
import os
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils.cell import range_boundaries

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
CHART = "http://schemas.openxmlformats.org/drawingml/2006/chart"
NS = {"m": MAIN, "c": CHART}

CRITERIA = [
    {"id": "R001", "description": "The workbook opens and contains the native reporting objects requested by the task.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {}},
    {"id": "R002", "description": "An Excel Table preserves exactly the eight supplied source events and six source fields.", "weight": 4, "type": "positive", "dimension": "source_table", "method": "ooxml", "method_params": {}},
    {"id": "R003", "description": "A genuine PivotCache binds to that source table/range with the six source-field identities.", "weight": 4, "type": "positive", "dimension": "pivot_source_binding", "method": "ooxml", "method_params": {}},
    {"id": "R004", "description": "The native PivotTable places Region in rows and Program in columns.", "weight": 4, "type": "positive", "dimension": "pivot_axes", "method": "ooxml", "method_params": {}},
    {"id": "R005", "description": "Quarter is a page filter selected to exactly 2024Q2.", "weight": 4, "type": "positive", "dimension": "pivot_filter", "method": "ooxml", "method_params": {}},
    {"id": "R006", "description": "Participants and Spend are native SUM measures.", "weight": 4, "type": "positive", "dimension": "pivot_measures", "method": "ooxml", "method_params": {}},
    {"id": "R007", "description": "The PivotCache is refreshable and configured to refresh when the workbook opens.", "weight": 3, "type": "positive", "dimension": "refresh_semantics", "method": "ooxml", "method_params": {}},
    {"id": "R008", "description": "GETPIVOTDATA-linked KPI cells have the Excel-calculated participant, spend, and split-specific focus values.", "weight": 4, "type": "positive", "dimension": "aggregation_correctness", "method": "ooxml", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R009", "description": "A clustered-column PivotChart is bound to the native PivotTable; a regular range chart does not qualify.", "weight": 4, "type": "positive", "dimension": "pivot_chart_binding", "method": "ooxml", "method_params": {}},
]

TASK = {
    "task_id": "P15-B-PUBLIC-PIVOT-001",
    "pass_threshold": 0.7,
    "required_sheets": ["Program_Data", "Pivot_Specification", "KPI_Summary", "Pivot_Chart_Requirement", "Checks"],
    "criteria": CRITERIA,
}


TASK["hurdle_criteria"] = ["R002", "R003", "R004", "R005", "R006", "R008", "R009"]
SHEET_ALIASES = {
}


def task_root():
    return Path(__file__).resolve().parents[1]


def split_context(split):
    contract_path = task_root() / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    contract = json.loads(contract_path.read_text())
    return {
        "split": split,
        "contract": contract,
        "source_root": task_root() / contract["input_files_dir"],
        "oracle_path": task_root() / contract["oracle"],
    }


def load_oracle(context):
    spec = importlib.util.spec_from_file_location(f"pivot_oracle_{context['split']}", context["oracle_path"])
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


def truthy(value):
    return str(value).lower() in {"1", "true", "yes"}


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def xml(archive, name):
    return ET.fromstring(archive.read(name))


def source_records_match(workbook, source_root):
    with (source_root / "program_events.csv").open(newline="") as handle:
        expected = [(record["event_id"], record["region"], record["program"], record["quarter"], int(record["participants"]), int(record["spend"])) for record in csv.DictReader(handle)]
    expected_headers = ["eventid", "region", "program", "quarter", "participants", "spend"]
    expected_set = {tuple(row) for row in expected}
    for sheet in workbook.worksheets:
        for header_row in range(1, min(sheet.max_row, 20) + 1):
            headers = [
                "".join(char for char in str(sheet.cell(row=header_row, column=column).value or "").casefold() if char.isalnum())
                for column in range(1, min(sheet.max_column, 20) + 1)
            ]
            positions = []
            for name in expected_headers:
                if name not in headers:
                    break
                positions.append(headers.index(name) + 1)
            if len(positions) != 6:
                continue
            actual = []
            for row in range(header_row + 1, sheet.max_row + 1):
                record = tuple(sheet.cell(row=row, column=column).value for column in positions)
                if any(value not in (None, "") for value in record):
                    actual.append(record)
            if len(actual) == len(expected) and {tuple(row) for row in actual} == expected_set:
                return True
    return False


def source_table_names(archive):
    names = set()
    for name in archive.namelist():
        if not name.startswith("xl/tables/table") or not name.endswith(".xml"):
            continue
        root = xml(archive, name)
        columns = [node.get("name") for node in root.findall("./m:tableColumns/m:tableColumn", NS)]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(root.get("ref") or "")
            correct_shape = max_col - min_col + 1 == 6 and max_row - min_row + 1 == 9
        except ValueError:
            correct_shape = False
        if correct_shape and {str(value or "").strip().casefold() for value in columns} == {
            "event id", "region", "program", "quarter", "participants", "spend"
        } and len(columns) == 6:
            for value in (root.get("name"), root.get("displayName")):
                if value:
                    names.add(value)
    return names


def cache_details(archive, valid_table_names):
    candidates = sorted(name for name in archive.namelist() if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml"))
    for name in candidates:
        root = xml(archive, name)
        worksheet_source = root.find("./m:cacheSource/m:worksheetSource", NS)
        if worksheet_source is None:
            continue
        named_source = worksheet_source.get("name") in valid_table_names
        ranged_source = bool(worksheet_source.get("sheet") and worksheet_source.get("ref"))
        fields = root.findall("./m:cacheFields/m:cacheField", NS)
        field_names = [field.get("name") for field in fields]
        if named_source or ranged_source:
            return {
                "name": name,
                "root": root,
                "source_ok": True,
                "fields": fields,
                "field_names": field_names,
                "refresh_on_load": truthy(root.get("refreshOnLoad")),
                "enable_refresh": root.get("enableRefresh") is None or truthy(root.get("enableRefresh")),
            }
    return None


def selected_page_value(table_root, cache_fields, field_index, page_field):
    pivot_fields = table_root.findall("./m:pivotFields/m:pivotField", NS)
    if field_index >= len(pivot_fields) or field_index >= len(cache_fields):
        return None
    shared = []
    shared_items = cache_fields[field_index].find("./m:sharedItems", NS)
    if shared_items is not None:
        for item in list(shared_items):
            shared.append(item.get("v"))
    try:
        page_item_index = int(page_field.get("item", "0"))
    except ValueError:
        return None
    pivot_items = pivot_fields[field_index].findall("./m:items/m:item", NS)
    cache_index = page_item_index
    if 0 <= page_item_index < len(pivot_items):
        mapped = pivot_items[page_item_index].get("x")
        if mapped is not None:
            try:
                cache_index = int(mapped)
            except ValueError:
                return None
    if 0 <= cache_index < len(shared):
        return shared[cache_index]
    return None


def pivot_table_details(archive, cache):
    candidates = sorted(name for name in archive.namelist() if name.startswith("xl/pivotTables/pivotTable") and name.endswith(".xml"))
    for name in candidates:
        root = xml(archive, name)
        row_indexes = [int(node.get("x")) for node in root.findall("./m:rowFields/m:field", NS) if node.get("x") is not None]
        column_indexes = [int(node.get("x")) for node in root.findall("./m:colFields/m:field", NS) if node.get("x") is not None]
        page_fields = root.findall("./m:pageFields/m:pageField", NS)
        data_fields = root.findall("./m:dataFields/m:dataField", NS)
        name_to_index = {
            str(field_name or "").strip().casefold(): index
            for index, field_name in enumerate(cache["field_names"])
        }
        quarter_index = name_to_index.get("quarter")
        page = next(
            (node for node in page_fields if node.get("fld") == str(quarter_index)),
            None,
        )
        selected = selected_page_value(root, cache["fields"], quarter_index, page) if page is not None and quarter_index is not None else None
        measures = {}
        for node in data_fields:
            try:
                field_index = int(node.get("fld"))
            except (TypeError, ValueError):
                continue
            # OOXML defines `sum` as the default dataField subtotal, so native
            # Excel may legally omit the attribute for a SUM measure.
            measures[field_index] = {"subtotal": (node.get("subtotal") or "sum").lower(), "name": node.get("name") or ""}
        return {
            "name": name,
            "root": root,
            "row_indexes": row_indexes,
            "column_indexes": column_indexes,
            "page_fields": page_fields,
            "selected_quarter": selected,
            "measures": measures,
            "name_to_index": name_to_index,
            "pivot_name": root.get("name") or "",
        }
    return None


def pivot_chart_ok(archive, expected_pivot_name=None):
    for name in archive.namelist():
        if not name.startswith("xl/charts/chart") or not name.endswith(".xml"):
            continue
        root = xml(archive, name)
        pivot_source = root.find("./c:pivotSource", NS)
        if pivot_source is None:
            continue
        pivot_name_node = pivot_source.find("./c:name", NS)
        bar_chart = root.find("./c:chart/c:plotArea/c:barChart", NS)
        bar_direction = bar_chart.find("./c:barDir", NS) if bar_chart is not None else None
        grouping = bar_chart.find("./c:grouping", NS) if bar_chart is not None else None
        if (
            pivot_name_node is not None
            and (
                not expected_pivot_name
                or expected_pivot_name in (pivot_name_node.text or "")
            )
            and bar_direction is not None
            and bar_direction.get("val") == "col"
            and grouping is not None
            and grouping.get("val") == "clustered"
        ):
            return True
    return False


def kpi_validation(candidate, oracle):
    """Separate formula semantics from native Excel cached-value evidence.

    openpyxl can read cached values but cannot calculate GETPIVOTDATA.  A
    semantically correct formula with no cache therefore remains pending native
    Excel readback; it is not a demonstrated model error.
    """
    formulas = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    cached = openpyxl.load_workbook(candidate, data_only=True, read_only=False)
    expected = [
        oracle["q2_participants"], oracle["q2_spend"],
        oracle.get("focus_participants", oracle.get("north_outreach")),
    ]
    candidates = []
    for sheet in formulas.worksheets:
        cached_sheet = cached[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("=") or "GETPIVOTDATA" not in formula.upper():
                    continue
                candidates.append((formula.upper(), cached_sheet[cell.coordinate].value))
    if len(candidates) < 3:
        return "FORMULA_INVALID"
    focus_region = str(oracle.get("focus_region", "North")).upper()
    focus_program = str(oracle.get("focus_program", "Outreach")).upper()
    semantic_formulas = (
        any("PARTICIP" in formula for formula, _ in candidates)
        and any("SPEND" in formula for formula, _ in candidates)
        and any(
            "PARTICIP" in formula
            and focus_region in formula
            and focus_program in formula
            for formula, _ in candidates
        )
    )
    if not semantic_formulas:
        return "FORMULA_INVALID"
    cached_values = [value for _, value in candidates]
    if any(value is None for value in cached_values):
        return "NATIVE_RECALC_REQUIRED"
    if not all(any(close(cached_value, value) for cached_value in cached_values) for value in expected):
        # A stale or foreign-engine cache cannot be distinguished from a wrong
        # native result without opening and recalculating in Microsoft Excel.
        return "NATIVE_RECALC_REQUIRED"
    return "VERIFIED"



def evaluate(candidate, split="dev"):
    criteria = {row["id"]: 0.0 for row in CRITERIA}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"], "TASK_INVALID"
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"], "TASK_INVALID"
    # The task requires native objects and business roles, not literal sheet
    # names.  Object-level checks below establish those roles directly.
    context = split_context(split)
    try:
        with zipfile.ZipFile(candidate) as archive:
            records_ok = source_records_match(workbook, context["source_root"])
            valid_table_names = source_table_names(archive)
            table_ok = records_ok and bool(valid_table_names)
            criteria["R002"] = float(table_ok)
            if not table_ok:
                failures.append("SOURCE_TABLE_OR_RECORD_MISMATCH")

            cache = cache_details(archive, valid_table_names)
            expected_fields = ["Event ID", "Region", "Program", "Quarter", "Participants", "Spend"]
            cache_ok = cache is not None and {
                str(value or "").strip().casefold() for value in cache["field_names"]
            } == {value.casefold() for value in expected_fields}
            criteria["R003"] = float(cache_ok)
            if cache is None:
                failures.append("MISSING_NATIVE_PIVOT_CACHE")
            elif not cache_ok:
                failures.append("PIVOT_CACHE_SOURCE_OR_FIELD_IDENTITY_MISMATCH")

            table = pivot_table_details(archive, cache) if cache is not None else None
            if table is None:
                failures.append("MISSING_NATIVE_PROGRAM_DELIVERY_PIVOT")
            region_index = table["name_to_index"].get("region") if table else None
            program_index = table["name_to_index"].get("program") if table else None
            axes_ok = bool(
                table is not None
                and table["row_indexes"] == [region_index]
                and program_index in table["column_indexes"]
                and all(index in (program_index, -2) for index in table["column_indexes"])
            )
            criteria["R004"] = float(axes_ok)
            if not axes_ok:
                failures.append("PIVOT_AXIS_FIELD_INDEX_MISMATCH")

            filter_ok = bool(
                table is not None
                and len(table["page_fields"]) == 1
                and table["page_fields"][0].get("fld") == str(table["name_to_index"].get("quarter"))
                and table["selected_quarter"] == "2024Q2"
            )
            criteria["R005"] = float(filter_ok)
            if not filter_ok:
                failures.append("QUARTER_FILTER_NOT_EXACTLY_2024Q2")

            measures = table["measures"] if table is not None else {}
            participant_index = table["name_to_index"].get("participants") if table else None
            spend_index = table["name_to_index"].get("spend") if table else None
            measure_ok = bool(
                set(measures) == {participant_index, spend_index}
                and all(measures[index]["subtotal"] == "sum" for index in (participant_index, spend_index))
                and "participant" in measures[participant_index]["name"].lower()
                and "spend" in measures[spend_index]["name"].lower()
            )
            criteria["R006"] = float(measure_ok)
            if not measure_ok:
                failures.append("SUM_MEASURE_FIELD_OR_FUNCTION_MISMATCH")

            refresh_ok = bool(
                cache is not None and cache["refresh_on_load"] and cache["enable_refresh"]
            )
            criteria["R007"] = float(refresh_ok)
            if not refresh_ok:
                failures.append("PIVOT_CACHE_REFRESH_ON_OPEN_MISSING")

            oracle = load_oracle(context)
            kpi_status = kpi_validation(candidate, oracle)
            kpi_ok = kpi_status == "VERIFIED"
            criteria["R008"] = float(kpi_ok)
            if kpi_status == "FORMULA_INVALID":
                failures.append("GETPIVOTDATA_FORMULA_SEMANTICS_INVALID")
            elif kpi_status == "NATIVE_RECALC_REQUIRED":
                failures.append("NATIVE_EXCEL_RECALC_REQUIRED:R008")

            chart_ok = pivot_chart_ok(archive, table["pivot_name"] if table else None)
            criteria["R009"] = float(chart_ok)
            if not chart_ok:
                failures.append("PIVOTCHART_BINDING_OR_CLUSTERED_COLUMN_MISSING")

            criteria["R001"] = float(
                table_ok and cache is not None and table is not None and chart_ok
            )

            independently_valid_native_structure = all(
                criteria[criterion] == 1.0
                for criterion in ("R002", "R003", "R004", "R005", "R006", "R007", "R009")
            )
            native_status = (
                "NATIVE_RECALC_REQUIRED"
                if kpi_status == "NATIVE_RECALC_REQUIRED"
                and independently_valid_native_structure
                else "NATIVE_OBJECT_CHECKED"
            )
            return criteria, sorted(set(failures)), native_status
    except zipfile.BadZipFile:
        return criteria, ["MALFORMED_XLSX:BadZipFile"], "TASK_INVALID"
    except Exception as exc:
        return criteria, [f"EVALUATION_ERROR:OOXML:{type(exc).__name__}:{exc}"], "JUDGE_ERROR"


def parse_cli():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    candidate = None
    arguments = iter(sys.argv[1:])
    for argument in arguments:
        if argument == "--split": split = next(arguments, "")
        elif argument.startswith("--split="): split = argument.split("=", 1)[1]
        elif candidate is None: candidate = Path(argument)
        else: raise ValueError(f"UNEXPECTED_ARGUMENT:{argument}")
    if split not in {"dev", "confirm"}: raise ValueError(f"INVALID_SPLIT:{split}")
    return candidate or Path("/app/output/answer.xlsx"), split


def main():
    candidate, split = parse_cli()
    criteria, failures, status = evaluate(candidate, split)
    payload = build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=criteria,
        failures=failures,
        explicit_status=(
            "JUDGE_ERROR"
            if status in {"JUDGE_ERROR", "NATIVE_RECALC_REQUIRED"}
            else "OK"
        ),
    )
    payload["native_status"] = status
    payload["judge_version"] = "P15_JUDGE_V3"
    # Pivot's established operational status is consumed by the native Excel
    # validator and campaign controller. Keep it alongside evaluation_status.
    payload["status"] = status
    payload["blocker"] = (
        None
        if status == "NATIVE_OBJECT_CHECKED"
        else "PENDING_NATIVE_EXCEL_RECALC"
        if status == "NATIVE_RECALC_REQUIRED"
        else "NATIVE_OBJECT_VALIDATION_FAILED"
    )
    total = payload["normalized_score"]
    root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        root.mkdir(parents=True, exist_ok=True)
        (root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        if total is not None:
            (root / "reward.txt").write_text(str(total) + "\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

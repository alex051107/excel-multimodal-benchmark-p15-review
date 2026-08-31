#!/usr/bin/env python3
"""Build task-specific negative fixtures from the validated native reference."""
from __future__ import annotations

import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


TASK_ROOT = Path(__file__).resolve().parents[1]
REFERENCE = TASK_ROOT / "solution" / "reference.xlsx"
STARTER = TASK_ROOT / "data" / "input_files" / "starting_workbook.xlsx"
FIXTURES = TASK_ROOT / "fixtures"


def mutate_zip(target: Path, transform) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=target.parent) as handle:
        temporary = Path(handle.name)
    try:
        with zipfile.ZipFile(REFERENCE, "r") as source, zipfile.ZipFile(temporary, "w") as output:
            for info in source.infolist():
                data = transform(info.filename, source.read(info.filename))
                output.writestr(info, data)
        temporary.replace(target)
    finally:
        if temporary.exists():
            temporary.unlink()


def stale_source(name: str, data: bytes) -> bytes:
    if name.startswith("xl/tables/table") and name.endswith(".xml"):
        updated = data.replace(b'ref="A3:F11"', b'ref="A3:F10"')
        if updated == data:
            raise RuntimeError("table range mutation did not match")
        return updated
    return data


def wrong_aggregation(name: str, data: bytes) -> bytes:
    if name.startswith("xl/pivotTables/pivotTable") and name.endswith(".xml"):
        marker = b'name="Sum of Participants" fld="4"'
        replacement = b'name="Count of Participants" fld="4" subtotal="count"'
        updated = data.replace(marker, replacement, 1)
        if updated == data:
            raise RuntimeError("data-field aggregation mutation did not match")
        return updated
    return data


def wrong_chart_binding(name: str, data: bytes) -> bytes:
    if name.startswith("xl/charts/chart") and name.endswith(".xml"):
        updated = data.replace(b"ProgramDeliveryPivot", b"DetachedProgramChart", 1)
        if updated == data:
            raise RuntimeError("pivot chart binding mutation did not match")
        return updated
    return data


def build_fake_pivot(target: Path) -> None:
    shutil.copy2(STARTER, target)
    workbook = load_workbook(target)
    if "Pivot_Report" in workbook.sheetnames:
        del workbook["Pivot_Report"]
    report = workbook.create_sheet("Pivot_Report")
    rows = [
        ["Region", "Outreach Participants", "Outreach Spend", "Training Participants", "Training Spend"],
        ["North", 50, 150000, 22, 66000],
        ["South", 44, 132000, 16, 48000],
    ]
    for row in rows:
        report.append(row)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Static Q2 Program Delivery (not a PivotChart)"
    chart.add_data(Reference(report, min_col=2, max_col=5, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(report, min_col=1, min_row=2, max_row=3))
    report.add_chart(chart, "G2")
    summary = workbook["KPI_Summary"]
    summary["B4"] = 132
    summary["B5"] = 396000
    summary["B6"] = 50
    workbook.save(target)


def main() -> None:
    if not REFERENCE.is_file():
        raise FileNotFoundError(REFERENCE)
    mutate_zip(FIXTURES / "mutants" / "stale_source_range.xlsx", stale_source)
    mutate_zip(FIXTURES / "mutants" / "wrong_aggregation.xlsx", wrong_aggregation)
    mutate_zip(FIXTURES / "mutants" / "wrong_chart_binding.xlsx", wrong_chart_binding)
    build_fake_pivot(FIXTURES / "mutants" / "fake_pivot.xlsx")
    shutil.copy2(STARTER, FIXTURES / "noop" / "candidate_noop.xlsx")


if __name__ == "__main__":
    main()

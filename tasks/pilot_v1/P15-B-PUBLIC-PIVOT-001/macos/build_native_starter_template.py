#!/usr/bin/env python3
"""Build the stale but genuine native Pivot template supplied to the Agent."""
from __future__ import annotations

import hashlib
import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


TASK_ROOT = Path(__file__).resolve().parents[1]
DEV_SEED = Path(__file__).with_name("native_pivot_seed.xlsx")
CONFIRM_SEED = TASK_ROOT / "tests" / "confirm" / "reference.xlsx"
APPLESCRIPT = Path(__file__).with_name("build_native_pivot_background.applescript")
JUDGE = TASK_ROOT / "tests" / "evaluate.py"
DEV_TARGETS = [
    TASK_ROOT / "data" / "input_files" / "starting_workbook.xlsx",
    TASK_ROOT / "environment" / "input" / "starting_workbook.xlsx",
    TASK_ROOT / "fixtures" / "noop" / "candidate_noop.xlsx",
]
CONFIRM_TARGETS = [
    TASK_ROOT / "tests" / "confirm" / "input_files" / "starting_workbook.xlsx",
]

# Plausible prior-extract values. The task-visible CSV contains the corrected
# Q2 data; an unchanged template is therefore a genuine no-op failure.
DEV_STALE_VALUES = {
    "EV-01": (40, 120000),
    "EV-02": (45, 135000),
    "EV-03": (34, 102000),
    "EV-04": (39, 117000),
    "EV-05": (17, 51000),
    "EV-06": (19, 57000),
    "EV-07": (13, 39000),
    "EV-08": (14, 42000),
}
CONFIRM_STALE_VALUES = {
    "HX-01": (29, 87000),
    "HX-02": (35, 105000),
    "HX-03": (25, 75000),
    "HX-04": (31, 93000),
    "HX-05": (19, 47500),
    "HX-06": (24, 60000),
    "HX-07": (16, 40000),
    "HX-08": (21, 52500),
}
DEV_EXPECTED_READBACK = {
    "q2_participants": 117.0,
    "q2_spend": 351000.0,
    "focus_participants": 45.0,
}
CONFIRM_EXPECTED_READBACK = {
    "q2_participants": 111.0,
    "q2_spend": 310500.0,
    "focus_participants": 35.0,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--split", choices=("dev", "confirm"), default="dev")
    parser.add_argument(
        "--excel",
        action="store_true",
        help="also open, refresh, recalculate, save, and read the starter in background Microsoft Excel",
    )
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_stale_source(target: Path, seed: Path, stale_values: dict) -> None:
    shutil.copy2(seed, target)
    workbook = load_workbook(target)
    sheet = workbook["Program_Data"]
    for row in range(4, 12):
        event_id = sheet.cell(row=row, column=1).value
        participants, spend = stale_values[event_id]
        sheet.cell(row=row, column=5).value = participants
        sheet.cell(row=row, column=6).value = spend
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(target)


def canonicalize_with_excel(target: Path, expected_readback: dict, focus_region: str, focus_program: str) -> dict:
    completed = subprocess.run(
        [
            "osascript",
            str(APPLESCRIPT),
            str(target),
            target.name,
            focus_region,
            focus_program,
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    fields = completed.stdout.strip().split("|")
    if len(fields) != 9:
        raise RuntimeError(f"unexpected Excel readback: {completed.stdout!r}")
    readback = {
        "excel_version": fields[0],
        "pivot_cache_count": int(fields[1]),
        "pivot_table_count": int(fields[2]),
        "data_field_count": int(fields[3]),
        "pivot_chart_count": int(fields[4]),
        "quarter_filter": fields[5],
        "q2_participants": float(fields[6]),
        "q2_spend": float(fields[7]),
        "focus_participants": float(fields[8]),
    }
    if not (
        readback["pivot_cache_count"] >= 1
        and readback["pivot_table_count"] == 1
        and readback["data_field_count"] == 2
        and readback["pivot_chart_count"] >= 1
        and readback["quarter_filter"] == "2024Q2"
        and all(
            abs(readback[key] - value) <= 0.01
            for key, value in expected_readback.items()
        )
    ):
        raise RuntimeError(f"stale template Excel readback mismatch: {readback}")
    return readback


def judge_noop(target: Path, split: str) -> dict:
    judge_python = os.environ.get("PILOT_PYTHON", sys.executable)
    completed = subprocess.run(
        [judge_python, str(JUDGE), str(target), "--split", split],
        check=True,
        capture_output=True,
        text=True,
    )
    payload = json.loads(completed.stdout)
    expected_failures = {
        "GETPIVOTDATA_OR_EXCEL_CACHED_ORACLE_MISMATCH",
        "SOURCE_TABLE_OR_RECORD_MISMATCH",
    }
    if not (
        payload.get("normalized_score") == 0.25
        and payload.get("pass") is False
        and set(payload.get("failure_codes", [])) == expected_failures
    ):
        raise RuntimeError(f"stale template is not a valid no-op fixture: {payload}")
    return payload


def main() -> int:
    args = parse_args()
    if args.split == "dev":
        seed = DEV_SEED
        targets = DEV_TARGETS
        receipt = TASK_ROOT / "receipts" / "native_starter_template.json"
        stale_values = DEV_STALE_VALUES
        expected_readback = DEV_EXPECTED_READBACK
        focus_region, focus_program = "North", "Outreach"
    else:
        seed = CONFIRM_SEED
        targets = CONFIRM_TARGETS
        receipt = TASK_ROOT / "receipts" / "native_confirm_starter_template.json"
        stale_values = CONFIRM_STALE_VALUES
        expected_readback = CONFIRM_EXPECTED_READBACK
        focus_region, focus_program = "East", "Vaccination"

    for required in (seed, APPLESCRIPT, JUDGE):
        if not required.is_file():
            raise FileNotFoundError(required)

    staging = Path(tempfile.gettempdir()) / f"p15-native-starter-{uuid.uuid4().hex}.xlsx"
    try:
        build_stale_source(staging, seed, stale_values)
        readback = (
            canonicalize_with_excel(staging, expected_readback, focus_region, focus_program)
            if args.excel
            else None
        )
        judge = judge_noop(staging, args.split)
        for target in targets:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(staging, target)
        output_hash = sha256(targets[0])
        if any(sha256(target) != output_hash for target in targets):
            raise RuntimeError("task, environment, and no-op starter bytes diverged")
        payload = {
            "task_id": "P15-B-PUBLIC-PIVOT-001",
            "status": (
                "MACOS_EXCEL_NATIVE_STALE_TEMPLATE_VALIDATED"
                if args.excel
                else "NATIVE_OOXML_STALE_TEMPLATE_VALIDATED"
            ),
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "output_sha256": output_hash,
            "split": args.split,
            "outputs": [str(path.relative_to(TASK_ROOT)) for path in targets],
            "construction": (
                "plausible prior-extract values in the previously Excel-validated native Pivot seed"
                + (
                    ", followed by background Microsoft Excel refresh/recalculate/save"
                    if args.excel
                    else "; this invocation did not reopen the starter in Excel"
                )
            ),
            "screen_control_used": False,
            "readback": readback,
            "excel_starter_readback": (
                "COMPLETED" if args.excel else "PENDING_BACKGROUND_EXCEL_SESSION_RECOVERY"
            ),
            "noop_judge": judge,
        }
        receipt.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        print(json.dumps(payload, indent=2, sort_keys=True))
        return 0
    finally:
        # Excel may retain the temporary workbook in its background session.
        # The path is task-owned and contains no secret or user material.
        pass


if __name__ == "__main__":
    raise SystemExit(main())

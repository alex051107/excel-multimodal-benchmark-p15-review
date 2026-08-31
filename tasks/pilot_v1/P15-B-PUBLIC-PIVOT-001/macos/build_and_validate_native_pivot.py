#!/usr/bin/env python3
"""Build and validate the task's native Pivot workbook with background Mac Excel."""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


TASK_ROOT = Path(__file__).resolve().parents[1]
SEED = Path(__file__).with_name("native_pivot_seed.xlsx")
APPLESCRIPT = Path(__file__).with_name("build_native_pivot_background.applescript")
JUDGE = TASK_ROOT / "tests" / "evaluate.py"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--split", choices=("dev", "confirm"), default="dev")
    parser.add_argument("--input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--receipt", type=Path)
    return parser.parse_args()


def default_input(split: str) -> Path:
    if split == "confirm":
        return TASK_ROOT / "tests" / "confirm" / "input_files" / "starting_workbook.xlsx"
    return TASK_ROOT / "data" / "input_files" / "starting_workbook.xlsx"


def contract(split: str) -> dict:
    path = TASK_ROOT / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    return json.loads(path.read_text())


def oracle(split: str) -> dict:
    path = TASK_ROOT / "tests" / ("confirm/oracle_recompute.py" if split == "confirm" else "dev/metadata/oracle_recompute.py")
    completed = subprocess.run([sys.executable, str(path)], check=True, capture_output=True, text=True)
    return json.loads(completed.stdout)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def apply_instance(seed: Path, source: Path, target: Path) -> None:
    """Copy task-visible cells into the native seed while preserving its objects."""
    shutil.copy2(seed, target)
    workbook = load_workbook(target)
    source_workbook = load_workbook(source, data_only=False)
    copy_ranges = {
        "Program_Data": (11, 6),
        "Pivot_Specification": (24, 8),
        "KPI_Summary": (24, 8),
        "Pivot_Chart_Requirement": (24, 8),
        "Checks": (24, 8),
    }
    for sheet_name, (last_row, last_column) in copy_ranges.items():
        source_sheet = source_workbook[sheet_name]
        target_sheet = workbook[sheet_name]
        for row in range(1, last_row + 1):
            for column in range(1, last_column + 1):
                target_cell = target_sheet.cell(row=row, column=column)
                if isinstance(target_cell, MergedCell):
                    continue
                target_cell.value = source_sheet.cell(row=row, column=column).value
    workbook.save(target)


def main() -> int:
    args = parse_args()
    source = (args.input or default_input(args.split)).resolve()
    output = args.output.resolve()
    receipt = (args.receipt or output.with_suffix(output.suffix + ".macos-native-validation.json")).resolve()
    for required in (source, SEED, APPLESCRIPT, JUDGE):
        if not required.is_file():
            raise FileNotFoundError(required)
    if source == output:
        raise ValueError("output must not overwrite the task input")

    expected = oracle(args.split)
    focus_region = expected.get("focus_region", "North")
    focus_program = expected.get("focus_program", "Outreach")
    expected_focus = expected.get("focus_participants", expected.get("north_outreach"))
    output.parent.mkdir(parents=True, exist_ok=True)
    receipt.parent.mkdir(parents=True, exist_ok=True)
    staging = Path(tempfile.gettempdir()) / f"p15-native-{args.split}-{uuid.uuid4().hex}.staging.xlsx"
    receipt_stage = receipt.parent / f".{receipt.name}.{uuid.uuid4().hex}.staging"

    try:
        apply_instance(SEED, source, staging)
        excel = subprocess.run(
            [
                "osascript",
                str(APPLESCRIPT),
                str(staging),
                staging.name,
                str(focus_region),
                str(focus_program),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        fields = excel.stdout.strip().split("|")
        if len(fields) != 9:
            raise RuntimeError(f"unexpected Excel readback: {excel.stdout!r}")
        readback = {
            "excel_version": fields[0],
            "pivot_cache_count": int(fields[1]),
            "pivot_table_count": int(fields[2]),
            "data_field_count": int(fields[3]),
            "pivot_chart_count": int(fields[4]),
            "quarter_filter": fields[5],
            "q2_participants": float(fields[6]),
            "q2_spend": float(fields[7]),
            "focus_participants": float(fields[8]),
        }
        expected_values = {
            "q2_participants": float(expected["q2_participants"]),
            "q2_spend": float(expected["q2_spend"]),
            "focus_participants": float(expected_focus),
        }
        readback_ok = (
            readback["pivot_cache_count"] >= 1
            and readback["pivot_table_count"] == 1
            and readback["data_field_count"] == 2
            and readback["pivot_chart_count"] >= 1
            and readback["quarter_filter"] == "2024Q2"
            and all(abs(readback[key] - value) <= 0.01 for key, value in expected_values.items())
        )
        if not readback_ok:
            raise RuntimeError(f"Excel reopened readback mismatch: {readback}")

        judge_python = os.environ.get("PILOT_PYTHON", sys.executable)
        judged = subprocess.run(
            [judge_python, str(JUDGE), str(staging), "--split", args.split],
            check=True,
            capture_output=True,
            text=True,
        )
        judge = json.loads(judged.stdout)
        if not (
            judge.get("status") == "NATIVE_OBJECT_CHECKED"
            and judge.get("pass") is True
            and judge.get("normalized_score") == 1.0
            and judge.get("failure_codes") == []
        ):
            raise RuntimeError(f"native candidate failed Judge: {judge}")

        output_hash = sha256(staging)
        payload = {
            "task_id": "P15-B-PUBLIC-PIVOT-001",
            "split": args.split,
            "status": "MACOS_EXCEL_NATIVE_OBJECTS_VALIDATED",
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "input_workbook": str(source),
            "output_workbook": str(output),
            "output_sha256": output_hash,
            "construction": "native seed plus background Microsoft Excel refresh/recalculate/save and independent OOXML Judge",
            "screen_control_used": False,
            "background_excel_session": {
                "workbook_name": staging.name,
                "workbook_path": str(staging),
                "note": "This Mac Excel build retains the temporary workbook in the background session; follow-on perturbation checks may target it by name.",
            },
            "readback": readback,
            "expected": expected_values,
            "judge": judge,
            "contract_status_before_promotion": contract(args.split).get("status", "LOCAL_READY"),
        }
        shutil.copy2(staging, output)
        output_hash = sha256(output)
        payload["output_sha256"] = output_hash
        receipt_stage.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        os.replace(receipt_stage, receipt)
        print(json.dumps(payload, indent=2, sort_keys=True))
        return 0
    finally:
        # Excel keeps the staging workbook open in its background session on
        # some Mac builds even after a Close Apple event. Keep that owned /tmp
        # file instead of deleting a path the application may still reference.
        if receipt_stage.exists():
            receipt_stage.unlink()


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Rebuild the POLICY V3 workbook assets from the frozen benchmark inputs."""
from __future__ import annotations

import copy
import shutil
import tempfile
import zipfile
from pathlib import Path

import openpyxl
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_CONTEXT = "https://www.eia.gov/electricity/annual/html/epa_03_01_a.html"


def save_deterministic(workbook, destination):
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temp_dir:
        raw = Path(temp_dir) / "raw.xlsx"
        workbook.save(raw)
        with zipfile.ZipFile(raw, "r") as source, zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, (2026, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = source.getinfo(name).external_attr
                target.writestr(info, source.read(name))


def prepare_starting(path):
    workbook = openpyxl.load_workbook(path)
    source = workbook["Generation_Data"]
    for row in range(4, 9):
        source[f"G{row}"] = PUBLIC_CONTEXT

    assumptions = workbook["Policy_Assumptions"]
    assumptions["B7"] = 0
    assumptions["D7"] = "Base uses supplied demand; policy applies demand growth"

    model = workbook["Scenario_Model"]
    if model["A9"].value != "Other generation":
        model.insert_rows(9, 1)
    labels = {
        4: ("Forecast demand", "GWh", "Supplied demand × case growth"),
        5: ("Coal generation", "GWh", "Supplied coal × displacement"),
        6: ("Wind generation", "GWh", "Supplied wind × uplift"),
        7: ("Solar generation", "GWh", "Supplied solar × uplift"),
        8: ("Natural gas balancing generation", "GWh", "Historical gas plus the remaining policy change"),
        9: ("Other generation", "GWh", "Demand residual in base; held constant in policy"),
        10: ("Emissions", "tCO₂", "(Coal + gas factors) × 1,000 MWh/GWh"),
        11: ("Emissions intensity", "tCO₂/MWh", "Emissions / (demand × 1,000)"),
    }
    for row, (label, unit, dependency) in labels.items():
        model[f"A{row}"] = label
        model[f"E{row}"] = unit
        model[f"F{row}"] = dependency
        for column in "BCD":
            model[f"{column}{row}"] = None

    checks = workbook["Checks"]
    checks["B4"] = "='Scenario_Model'!C4-SUM('Scenario_Model'!C5:C9)"
    checks["B6"] = "='Scenario_Model'!C10-(('Scenario_Model'!C5*'Generation_Data'!E4+'Scenario_Model'!C8*'Generation_Data'!E5)*1000)"
    checks["E4"] = "Demand equals coal + gas + wind + solar + Other"
    checks["E6"] = "Emissions convert GWh to MWh before applying factors"
    save_deterministic(workbook, path)


def populate_reference(starting_path, destination, equivalent=False):
    workbook = openpyxl.load_workbook(starting_path)
    model = workbook["Scenario_Model"]
    if equivalent:
        formulas = {
            "B4": "='Generation_Data'!D8", "C4": "='Generation_Data'!D8+'Generation_Data'!D8*'Policy_Assumptions'!C7",
            "B5": "='Generation_Data'!D4", "C5": "='Generation_Data'!D4-'Generation_Data'!D4*'Policy_Assumptions'!C4",
            "B6": "='Generation_Data'!D6", "C6": "='Generation_Data'!D6+'Generation_Data'!D6*'Policy_Assumptions'!C5",
            "B7": "='Generation_Data'!D7", "C7": "='Generation_Data'!D7+'Generation_Data'!D7*'Policy_Assumptions'!C6",
            "B8": "='Generation_Data'!D5", "C8": "=SUM(B8,C4-B4,B5-C5,-(C6-B6),-(C7-B7))",
            "B9": "=B4-SUM(B5:B8)", "C9": "=B9",
            "B10": "=SUM(B5*'Generation_Data'!E4,B8*'Generation_Data'!E5)*1000",
            "C10": "=SUM(C5*'Generation_Data'!E4,C8*'Generation_Data'!E5)*1000",
            "B11": "=B10/B4/1000", "C11": "=C10/C4/1000",
        }
    else:
        formulas = {
            "B4": "='Generation_Data'!D8*(1+'Policy_Assumptions'!B7)", "C4": "='Generation_Data'!D8*(1+'Policy_Assumptions'!C7)",
            "B5": "='Generation_Data'!D4*(1-'Policy_Assumptions'!B4)", "C5": "='Generation_Data'!D4*(1-'Policy_Assumptions'!C4)",
            "B6": "='Generation_Data'!D6*(1+'Policy_Assumptions'!B5)", "C6": "='Generation_Data'!D6*(1+'Policy_Assumptions'!C5)",
            "B7": "='Generation_Data'!D7*(1+'Policy_Assumptions'!B6)", "C7": "='Generation_Data'!D7*(1+'Policy_Assumptions'!C6)",
            "B8": "='Generation_Data'!D5", "C8": "=B8+(C4-B4)+(B5-C5)-(C6-B6)-(C7-B7)",
            "B9": "=B4-B5-B6-B7-B8", "C9": "=B9",
            "B10": "=(B5*'Generation_Data'!E4+B8*'Generation_Data'!E5)*1000",
            "C10": "=(C5*'Generation_Data'!E4+C8*'Generation_Data'!E5)*1000",
            "B11": "=B10/(B4*1000)", "C11": "=C10/(C4*1000)",
        }
    for address, formula in formulas.items():
        model[address] = formula
    for row in range(4, 12):
        model[f"D{row}"] = f"=C{row}-B{row}"

    results = workbook["Policy_Results"]
    results["B4"] = "='Scenario_Model'!B10-'Scenario_Model'!C10"
    results["B5"] = "='Scenario_Model'!B11-'Scenario_Model'!C11"
    results["B6"] = "='Scenario_Model'!C8-'Scenario_Model'!B8"
    results["B7"] = '=IF(B4>0,"Policy case reduces modeled emissions","Review scenario dependencies")'
    save_deterministic(workbook, destination)


def alternate_layout(destination, source_values, policy_values):
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs.append(["Category", "Generation (GWh)", "Emission factor (tCO2/MWh)"])
    for role, label in (("coal", "Coal"), ("gas", "Natural gas"), ("wind", "Wind"), ("solar", "Solar"), ("demand", "Demand")):
        factor = source_values.get(f"{role}_factor")
        inputs.append([label, source_values[role], factor])
    assumptions = workbook.create_sheet("Assumptions")
    assumptions.append(["Assumption", "Policy case"])
    assumptions.append(["Coal displacement share", policy_values["coal_displacement"]])
    assumptions.append(["Wind uplift share", policy_values["wind_uplift"]])
    assumptions.append(["Solar uplift share", policy_values["solar_uplift"]])
    assumptions.append(["Grid demand growth", policy_values["demand_growth"]])
    base = workbook.create_sheet("Base")
    policy = workbook.create_sheet("Policy")
    for sheet in (base, policy):
        sheet.append(["Metric", "Value", "Unit"])
    base_rows = [
        ("Demand", "=Inputs!B6", "GWh"), ("Coal generation", "=Inputs!B2", "GWh"),
        ("Natural gas generation", "=Inputs!B3", "GWh"), ("Wind generation", "=Inputs!B4", "GWh"),
        ("Solar generation", "=Inputs!B5", "GWh"), ("Other generation", "=B2-SUM(B3:B6)", "GWh"),
        ("Total emissions", "=(B3*Inputs!C2+B4*Inputs!C3)*1000", "tCO2"),
        ("Emissions intensity", "=B8/(B2*1000)", "tCO2/MWh"),
    ]
    for row in base_rows:
        base.append(row)
    policy_rows = [
        ("Demand", "=Inputs!B6*(1+Assumptions!B5)", "GWh"),
        ("Coal generation", "=Inputs!B2*(1-Assumptions!B2)", "GWh"),
        ("Natural gas generation", "=Base!B4+(B2-Base!B2)+(Base!B3-B3)-(B5-Base!B5)-(B6-Base!B6)", "GWh"),
        ("Wind generation", "=Inputs!B4*(1+Assumptions!B3)", "GWh"),
        ("Solar generation", "=Inputs!B5*(1+Assumptions!B4)", "GWh"),
        ("Other generation", "=Base!B7", "GWh"),
        ("Total emissions", "=(B3*Inputs!C2+B4*Inputs!C3)*1000", "tCO2"),
        ("Emissions intensity", "=B8/(B2*1000)", "tCO2/MWh"),
    ]
    for row in policy_rows:
        policy.append(row)
    results = workbook.create_sheet("Results")
    results.append(["Result", "Value"])
    results.append(["Emissions reduction", "=Base!B8-Policy!B8"])
    results.append(["Intensity reduction", "=Base!B9-Policy!B9"])
    results.append(["Natural gas change", "=Policy!B4-Base!B4"])
    save_deterministic(workbook, destination)


def make_mutants(reference_path):
    mutations = {
        "broken_demand_balance.xlsx": ("Scenario_Model", "C8", "=C4-C5-C6-C7"),
        "collateral_historical_edit.xlsx": ("Generation_Data", "D4", 676000),
        "hardcoded_emissions_reduction.xlsx": ("Policy_Results", "B4", 36324000),
        "wrong_generation_source.xlsx": ("Scenario_Model", "C5", "='Generation_Data'!D5*(1-'Policy_Assumptions'!C4)"),
    }
    for name, (sheet, cell, value) in mutations.items():
        workbook = openpyxl.load_workbook(reference_path)
        workbook[sheet][cell] = value
        save_deterministic(workbook, ROOT / "fixtures/mutants" / name)


def main():
    dev_start = ROOT / "data/input_files/starting_workbook.xlsx"
    confirm_start = ROOT / "tests/confirm/input_files/starting_workbook.xlsx"
    prepare_starting(dev_start)
    prepare_starting(confirm_start)
    shutil.copy2(dev_start, ROOT / "environment/input/starting_workbook.xlsx")
    for name in ("context_note.md", "eia_model_scope.md", "generation_source_extract.csv", "policy_case_note.md", "source_data.csv", "source_locator.md"):
        shutil.copy2(ROOT / "data/input_files" / name, ROOT / "environment/input" / name)

    reference = ROOT / "solution/reference.xlsx"
    populate_reference(dev_start, reference)
    populate_reference(dev_start, ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", equivalent=True)
    shutil.copy2(dev_start, ROOT / "fixtures/noop/candidate_noop.xlsx")
    make_mutants(reference)
    alternate_layout(
        ROOT / "fixtures/equivalent/candidate_v3_separate_layout.xlsx",
        {"coal": 675000, "gas": 1802000, "wind": 425000, "solar": 238000, "demand": 4000000, "coal_factor": 1.0, "gas_factor": 0.4},
        {"coal_displacement": 0.08, "wind_uplift": 0.05, "solar_uplift": 0.12, "demand_growth": 0.01},
    )

    populate_reference(confirm_start, ROOT / "tests/confirm/reference.xlsx")
    alternate_layout(
        ROOT / "tests/confirm/candidate_v3_separate_layout.xlsx",
        {"coal": 510000, "gas": 1420000, "wind": 515000, "solar": 305000, "demand": 3300000, "coal_factor": 0.96, "gas_factor": 0.38},
        {"coal_displacement": 0.11, "wind_uplift": 0.07, "solar_uplift": 0.15, "demand_growth": 0.012},
    )


if __name__ == "__main__":
    main()

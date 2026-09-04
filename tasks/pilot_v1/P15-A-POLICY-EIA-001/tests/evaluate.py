#!/usr/bin/env python3
"""Deterministic semantic judge for P15-A-POLICY-EIA-001."""
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-A-POLICY-EIA-001",
  "pass_threshold": 0.7,
  "canonical_reference_sheets": [
    "Generation_Data",
    "Policy_Assumptions",
    "Scenario_Model",
    "Policy_Results",
    "Checks"
  ],
  "protected": {
    "Generation_Data!D4": 675000,
    "Generation_Data!D5": 1802000,
    "Generation_Data!D6": 425000,
    "Generation_Data!D7": 238000,
    "Generation_Data!D8": 4000000,
    "Generation_Data!E4": 1.0,
    "Generation_Data!E5": 0.4
  },
  "canonical_reference_formula_cells": [
    "Scenario_Model!C5",
    "Scenario_Model!C6",
    "Scenario_Model!C7",
    "Scenario_Model!C8",
    "Scenario_Model!C10",
    "Policy_Results!B4",
    "Policy_Results!B5",
    "Policy_Results!B6"
  ],
  "criteria": [
    {
      "id": "R001",
      "description": "A readable workbook opens and the source inputs, assumptions, scenarios, and results are identifiable without relying on sheet names.",
      "weight": 1,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_policy_content"}
    },
    {
      "id": "R002",
      "description": "Policy coal generation applies the documented displacement rate to the protected source value.",
      "weight": 3,
      "type": "positive",
      "dimension": "policy_coal_generation",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_policy_coal_formula_and_replay"}
    },
    {
      "id": "R003",
      "description": "Policy wind and solar generation apply their documented uplift schedules.",
      "weight": 2,
      "type": "positive",
      "dimension": "renewable_schedule",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_renewable_schedule_formula_and_replay"}
    },
    {
      "id": "R004",
      "description": "Natural gas starts from the supplied historical base and closes the policy balance while Other generation stays constant.",
      "weight": 3,
      "type": "positive",
      "dimension": "gas_balancing",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_historical_gas_and_other_balance"}
    },
    {
      "id": "R005",
      "description": "Policy emissions use GWh × 1,000 × tCO2/MWh, match the independent replay, and remain formula-linked.",
      "weight": 3,
      "type": "positive",
      "dimension": "policy_emissions",
      "method": "deterministic",
      "method_params": {"implemented_check": "Scenario_Model!C9 formula and replay value"}
    },
    {
      "id": "R006",
      "description": "The reported result calculates absolute emissions reduction from the two scenarios.",
      "weight": 2,
      "type": "positive",
      "dimension": "emissions_reduction",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_emissions_reduction_formula_and_replay"}
    },
    {
      "id": "R007",
      "description": "The reported result calculates emissions-intensity reduction from the two scenarios.",
      "weight": 2,
      "type": "positive",
      "dimension": "intensity_reduction",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_intensity_reduction_formula_and_replay"}
    },
    {
      "id": "R008",
      "description": "Changing the coal-displacement assumption updates coal, gas, policy emissions, and emissions reduction to the replayed values.",
      "weight": 2,
      "type": "positive",
      "dimension": "dynamic_recalculation",
      "method": "deterministic",
      "method_params": {"implemented_check": "coal-displacement perturbation propagation", "split_weights": {"dev": 2, "confirm": 3}}
    },
    {
      "id": "R009",
      "description": "Coal, gas, wind, solar, and explicit or implied Other generation balance to demand in both cases.",
      "weight": 2,
      "type": "positive",
      "dimension": "demand_balance",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_generation_balance_with_other"}
    },
    {
      "id": "R010",
      "description": "Emissions respond to the protected coal and gas factors through workbook formulas and preserve the GWh-to-MWh conversion.",
      "weight": 2,
      "type": "positive",
      "dimension": "factor_linkage",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_factor_perturbation_and_unit_conversion"}
    },
    {
      "id": "R011",
      "description": "Historical generation, demand, and emissions-factor inputs remain unchanged.",
      "weight": 2,
      "type": "positive",
      "dimension": "change_locality",
      "method": "deterministic",
      "method_params": {"implemented_check": "semantic_source_input_protection"}
    },
    {
      "id": "R012",
      "description": "A distinct policy-demand-growth perturbation updates demand, balancing gas, emissions, intensity, and reported results while preserving source generation and the coal/renewable policy schedule.",
      "weight": 1,
      "type": "positive",
      "dimension": "demand_growth_recalculation",
      "method": "deterministic",
      "method_params": {"implemented_check": "demand-growth perturbation propagation", "applies_to": ["dev"]}
    }
  ]
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R005", "R006", "R011"]
SHEET_ALIASES = {}


def cell_key(sheet, cell):
    return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column:
        value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


TASK_ROOT = Path(__file__).resolve().parents[1]


def requested_split():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    for index, argument in enumerate(sys.argv[1:]):
        if argument == "--split" and index + 2 <= len(sys.argv) - 1:
            split = sys.argv[index + 2].strip().lower()
        elif argument.startswith("--split="):
            split = argument.split("=", 1)[1].strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"unsupported split: {split}")
    return split


ACTIVE_SPLIT = requested_split()


def load_contract():
    relative = "tests/confirm/contract.json" if ACTIVE_SPLIT == "confirm" else "tests/private_contract.json"
    return json.loads((TASK_ROOT / relative).read_text())


def load_oracle(contract):
    oracle_path = TASK_ROOT / contract["oracle"]
    source_path = TASK_ROOT / "metadata/oracle_recompute.py"
    if ACTIVE_SPLIT == "dev" and source_path.exists() and source_path.read_bytes() != oracle_path.read_bytes():
        raise RuntimeError("DEV_ORACLE_COPY_OUT_OF_SYNC")
    spec = importlib.util.spec_from_file_location(f"task_oracle_{ACTIVE_SPLIT}", oracle_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute(), module


def _betacf(a, b, x):
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d, h = 1.0, 1.0 - qab * x / qap, 0.0
    if abs(d) < 3e-14: d = 3e-14
    d = 1.0 / d
    h = d
    for m in range(1, 201):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; delta = d * c; h *= delta
        if abs(delta - 1.0) < 3e-12: break
    return h


def _betai(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    factor = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x))
    return factor * _betacf(a, b, x) / a if x < (a + 1) / (a + b + 2) else 1 - factor * _betacf(b, a, 1 - x) / b


def student_t_two_tail(value, df):
    x = df / (df + value * value)
    return _betai(df / 2, 0.5, x)


def student_t_inv_two_tail(alpha, df):
    lo, hi = 0.0, 20.0
    for _ in range(90):
        mid = (lo + hi) / 2
        if student_t_two_tail(mid, df) > alpha: lo = mid
        else: hi = mid
    return (lo + hi) / 2


class UnsupportedFormulaError(RuntimeError):
    """Valid Excel syntax that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook
        self.overrides = overrides or {}
        self.memo = {}
        self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames:
            raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            if isinstance(raw, str) and raw.startswith("="):
                result = self.formula(raw[1:], sheet)
            else:
                result = raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        values = []
        for r in range(int(r1), int(r2) + 1):
            for c in range(col_index(c1), col_index(c2) + 1):
                values.append(self.value(sheet, f"{col_name(c)}{r}"))
        return values

    def formula(self, expression, current_sheet):
        expression = re.sub(r"(?i)_xl(?:fn|ws)\.", "", expression)
        expression = expression.replace("^", "**").replace("<>", "!=")
        expression = re.sub(r"\bSTDEV\.S\b", "STDEV_S", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.2T\b", "T_DIST_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.RT\b", "T_DIST_RT", expression, flags=re.I)
        expression = re.sub(r"\bT\.INV\.2T\b", "T_INV_2T", expression, flags=re.I)
        expression = re.sub(r"\bPI\(\)", "PI()", expression, flags=re.I)

        def replace_range(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))

        def sub_outside_strings(pattern, replacement, value):
            parts = re.split(r'("(?:[^"\\\\]|\\\\.)*")', value)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in parts)

        expression = sub_outside_strings(RANGE, replace_range, expression)

        def replace_ref(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))

        expression = sub_outside_strings(REF, replace_ref, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(
                f"UNSUPPORTED_FORMULA_SYNTAX:{expression}"
            ) from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name) and node.id.upper() in {"TRUE", "FALSE"}:
            return node.id.upper() == "TRUE"
        if isinstance(node, ast.List):
            return [self.safe_eval(v) for v in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand)
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operators = {ast.Add: lambda a,b:a+b, ast.Sub: lambda a,b:a-b, ast.Mult: lambda a,b:a*b, ast.Div: lambda a,b:a/b, ast.Pow: lambda a,b:a**b}
            for cls, fun in operators.items():
                if isinstance(node.op, cls): return fun(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for op, comp in zip(node.ops, node.comparators):
                right = self.safe_eval(comp)
                good = (left == right if isinstance(op, ast.Eq) else left != right if isinstance(op, ast.NotEq) else left >= right if isinstance(op, ast.GtE) else left <= right if isinstance(op, ast.LtE) else left > right if isinstance(op, ast.Gt) else left < right if isinstance(op, ast.Lt) else False)
                if not good: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                if self.safe_eval(node.args[0]):
                    return self.safe_eval(node.args[1])
                return self.safe_eval(node.args[2]) if len(node.args) == 3 else False
            args = [self.safe_eval(a) for a in node.args]
            values = [v for group in args for v in (group if isinstance(group, list) else [group]) if v is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            if name == "SUM": return sum(numeric_values)
            if name == "AVERAGE": return statistics.mean(numeric_values)
            if name == "COUNT": return len(numeric_values)
            if name == "STDEV_S": return statistics.stdev(numeric_values)
            if name == "SQRT": return math.sqrt(args[0])
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "PI": return math.pi
            if name == "AND": return all(args)
            if name == "OR": return any(args)
            if name == "NOT": return not args[0]
            if name == "POWER": return args[0] ** args[1]
            if name == "PRODUCT": return math.prod(numeric_values)
            if name == "MAX": return max(numeric_values)
            if name == "MIN": return min(numeric_values)
            if name == "NA": return None
            if name == "T_DIST_2T":
                return student_t_two_tail(args[0], args[1])
            if name == "T_DIST_RT":
                return student_t_two_tail(args[0], args[1]) / 2
            if name == "T_INV_2T":
                return student_t_inv_two_tail(args[0], args[1])
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def intensity_reporting_direction(workbook, actual, expected_reduction):
    """Return +1 for a positive reduction or -1 for a signed policy-minus-base change."""
    if not formula_present(workbook, "Policy_Results!B5"):
        return None
    label = workbook["Policy_Results"]["A5"].value
    normalized_label = re.sub(r"[^a-z0-9]+", " ", str(label or "").casefold()).strip()
    label_words = set(normalized_label.split())
    if "intensity" not in label_words:
        return None
    signed_label = bool(label_words & {"change", "delta", "difference", "variation"})
    reduction_label = bool(label_words & {"reduction", "decrease", "improvement", "saving", "savings"})
    contradictory_increase = bool(label_words & {"increase", "increased", "rise", "rising", "growth", "gain"})
    if contradictory_increase or not (signed_label or reduction_label):
        return None
    if close(actual, expected_reduction, 1e-9):
        return 1.0
    if signed_label and close(actual, -expected_reduction, 1e-9):
        return -1.0
    return None


def task_checks(workbook, engine, oracle, oracle_module, contract):
    checks, failures = {}, []

    def observed(address):
        sheet, cell = address.split("!", 1)
        try:
            return engine.value(sheet, cell)
        except UnsupportedFormulaError:
            raise
        except Exception as exc:
            failures.append(f"POLICY_FORMULA_EVALUATION_FAILED:{address}:{type(exc).__name__}")
            return None

    for address in contract.get("canonical_reference_formula_cells", contract.get("formula_cells", [])):
        if not formula_present(workbook, address):
            failures.append(f"FORMULA_REQUIRED:{address}")

    coal_ok = formula_present(workbook, "Scenario_Model!C5") and close(
        observed("Scenario_Model!C5"), oracle["policy"]["coal"], 0.5
    )
    checks["R002"] = 1.0 if coal_ok else 0.0

    renewables_ok = all(
        formula_present(workbook, address) and close(observed(address), target, 0.5)
        for address, target in {
            "Scenario_Model!C6": oracle["policy"]["wind"],
            "Scenario_Model!C7": oracle["policy"]["solar"],
        }.items()
    )
    checks["R003"] = 1.0 if renewables_ok else 0.0

    gas_ok = all(
        formula_present(workbook, address) and close(observed(address), target, 0.5)
        for address, target in {
            "Scenario_Model!C8": oracle["policy"]["gas"],
            "Policy_Results!B6": oracle["gas_change"],
        }.items()
    )
    checks["R004"] = 1.0 if gas_ok else 0.0

    emissions_ok = formula_present(workbook, "Scenario_Model!C9") and close(
        observed("Scenario_Model!C9"), oracle["policy_emissions"], 0.5
    )
    checks["R005"] = 1.0 if emissions_ok else 0.0

    reduction_ok = formula_present(workbook, "Policy_Results!B4") and close(
        observed("Policy_Results!B4"), oracle["emissions_reduction"], 0.5
    )
    checks["R006"] = 1.0 if reduction_ok else 0.0

    intensity_actual = observed("Policy_Results!B5")
    intensity_direction = intensity_reporting_direction(workbook, intensity_actual, oracle["intensity_reduction"])
    intensity_ok = intensity_direction is not None
    checks["R007"] = 1.0 if intensity_ok else 0.0

    perturbations = {row["name"]: row for row in contract["perturbations"]}

    def perturbation_matches(case, reporting_direction=1.0):
        replay = FormulaEngine(workbook, case["overrides"])
        for address, target in {**case["expected"], **case["invariants"]}.items():
            sheet, cell = address.split("!", 1)
            observed_value = replay.value(sheet, cell)
            if address == "Policy_Results!B5":
                target *= reporting_direction
            tolerance = 1e-9 if abs(target) < 1 else 0.5
            if not close(observed_value, target, tolerance):
                return False
        return True

    try:
        coal_case = perturbations["coal_displacement"]
        if "expected" in coal_case:
            dynamic_ok = perturbation_matches(coal_case)
        else:
            perturbed_policy = dict(oracle_module.POLICY)
            perturbed_policy["coal_displacement"] = coal_case["value"]
            perturbed_oracle = oracle_module.recompute(policy_case=perturbed_policy)
            perturbed_engine = FormulaEngine(workbook, {coal_case["address"]: coal_case["value"]})
            dynamic_ok = close(perturbed_engine.value("Policy_Results", "B4"), perturbed_oracle["emissions_reduction"], 0.5)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        dynamic_ok = False
        failures.append(f"COAL_DISPLACEMENT_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R008"] = 1.0 if dynamic_ok else 0.0

    demand_growth_ok = True
    checks["R012"] = 0.0
    if ACTIVE_SPLIT == "dev":
        try:
            demand_growth_ok = intensity_direction is not None and perturbation_matches(
                perturbations["demand_growth"], intensity_direction
            )
        except UnsupportedFormulaError:
            raise
        except Exception as exc:
            demand_growth_ok = False
            failures.append(f"DEMAND_GROWTH_PERTURBATION_FAILED:{type(exc).__name__}")
        checks["R012"] = 1.0 if demand_growth_ok else 0.0

    try:
        base_balance = engine.value("Scenario_Model", "B4") - sum(
            engine.value("Scenario_Model", f"B{row}") for row in (5, 6, 7, 8)
        )
        policy_balance = engine.value("Scenario_Model", "C4") - sum(
            engine.value("Scenario_Model", f"C{row}") for row in (5, 6, 7, 8)
        )
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        base_balance = policy_balance = None
        failures.append(f"POLICY_BALANCE_EVALUATION_FAILED:{type(exc).__name__}")
    balance_ok = close(base_balance, 0.0, 0.1) and close(policy_balance, 0.0, 0.1)
    checks["R009"] = 1.0 if balance_ok else 0.0

    factor_ok = all(
        formula_present(workbook, address) and close(observed(address), target, 0.5)
        for address, target in {
            "Scenario_Model!B9": oracle["base"]["emissions"],
            "Scenario_Model!C9": oracle["policy"]["emissions"],
        }.items()
    )
    checks["R010"] = 1.0 if factor_ok else 0.0

    protected_ok = all(
        workbook[sheet][cell].value == value
        for key, value in contract["protected"].items()
        for sheet, cell in [key.split("!", 1)]
    )
    checks["R011"] = 1.0 if protected_ok else 0.0

    return checks, failures


def v3_token(value):
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).split())


def v3_address(sheet, cell):
    return f"{sheet.title}!{cell.coordinate}"


def v3_raw_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    if isinstance(value, str) and not value.startswith("="):
        text = value.strip().replace(",", "")
        scale = 0.01 if text.endswith("%") else 1.0
        if text.endswith("%"):
            text = text[:-1]
        try:
            return float(text) * scale
        except ValueError:
            return None
    return None


def v3_value(workbook, engine, address):
    sheet, coordinate = address.split("!", 1)
    raw = workbook[sheet][coordinate].value
    if isinstance(raw, str) and raw.startswith("="):
        raw = engine.value(sheet, coordinate)
    return v3_raw_number(raw)


def v3_label_cells(workbook, predicate):
    found = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row in sheet.iter_rows():
            for cell in row:
                text = v3_token(cell.value)
                if text and predicate(text):
                    found.append((sheet_index, sheet, cell, text))
    return found


def v3_row_candidates(workbook, engine, label_cell, max_offset=6):
    sheet = label_cell.parent
    found = []
    for offset in range(1, max_offset + 1):
        column = label_cell.column + offset
        if column > sheet.max_column:
            break
        cell = sheet.cell(label_cell.row, column)
        raw = cell.value
        if v3_raw_number(raw) is None and not (isinstance(raw, str) and raw.startswith("=")):
            continue
        address = v3_address(sheet, cell)
        found.append((offset, address, v3_value(workbook, engine, address), formula_present(workbook, address)))
    return found


def v3_source_label(role, text):
    if any(word in text for word in ("displacement", "uplift", "reduction", "change", "emission", "intensity")):
        return False
    if role == "gas":
        return text in {"gas", "natural gas", "natural gas generation"}
    if role == "demand":
        return text in {"demand", "electricity demand", "grid demand"}
    return text in {role, f"{role} generation"}


def v3_find_sources(workbook, engine, oracle_module):
    addresses = {}
    for role in ("coal", "gas", "wind", "solar", "demand"):
        target = oracle_module.DATA[role]
        matches = []
        for sheet_index, _, label_cell, _ in v3_label_cells(workbook, lambda text, role=role: v3_source_label(role, text)):
            for offset, address, value, is_formula in v3_row_candidates(workbook, engine, label_cell):
                if not is_formula and close(value, target, 1e-9):
                    matches.append((sheet_index, offset, address))
        if matches:
            addresses[role] = sorted(matches)[0][-1]
    for role, source_role in (("coal_factor", "coal"), ("gas_factor", "gas")):
        source_address = addresses.get(source_role)
        if not source_address:
            continue
        source_sheet, source_coordinate = source_address.split("!", 1)
        source_cell = workbook[source_sheet][source_coordinate]
        label_candidates = v3_label_cells(workbook, lambda text, source_role=source_role: v3_source_label(source_role, text))
        target = oracle_module.DATA[role]
        candidates = []
        for sheet_index, sheet, label_cell, _ in label_candidates:
            if sheet.title != source_sheet or label_cell.row != source_cell.row:
                continue
            for offset, address, value, is_formula in v3_row_candidates(workbook, engine, label_cell):
                if not is_formula and close(value, target, 1e-12):
                    candidates.append((sheet_index, offset, address))
        if candidates:
            addresses[role] = sorted(candidates)[0][-1]
    return addresses


def v3_assumption_label(role, text):
    phrases = {
        "coal_displacement": ("coal", "displacement"),
        "wind_uplift": ("wind", "uplift"),
        "solar_uplift": ("solar", "uplift"),
        "demand_growth": ("demand", "growth"),
    }
    return all(word in text for word in phrases[role])


def v3_find_assumptions(workbook, engine, oracle_module):
    addresses = {}
    for role in ("coal_displacement", "wind_uplift", "solar_uplift", "demand_growth"):
        target = oracle_module.POLICY[role]
        candidates = []
        for sheet_index, _, label_cell, _ in v3_label_cells(workbook, lambda text, role=role: v3_assumption_label(role, text)):
            for offset, address, value, _ in v3_row_candidates(workbook, engine, label_cell):
                if close(value, target, 1e-12):
                    candidates.append((sheet_index, offset, address))
        if candidates:
            addresses[role] = sorted(candidates)[0][-1]
    return addresses


def v3_scenario_label(role, text):
    if any(word in text for word in ("reduction", "change", "delta", "uplift", "displacement", "factor")):
        return False
    if role == "demand":
        return text in {"demand", "forecast demand", "electricity demand", "grid demand"}
    if role == "gas":
        return text in {"gas", "natural gas", "gas generation", "natural gas generation", "natural gas balancing generation", "balancing gas"}
    if role == "other":
        return text in {"other", "other generation", "residual other generation"}
    if role == "emissions":
        return text in {"emissions", "co2 emissions", "total emissions", "absolute emissions"}
    if role == "intensity":
        return text in {"emissions intensity", "co2 intensity", "intensity"}
    return text in {role, f"{role} generation"}


def v3_scenario_candidates(workbook, engine, role):
    candidates = []
    for sheet_index, sheet, label_cell, _ in v3_label_cells(workbook, lambda text, role=role: v3_scenario_label(role, text)):
        for offset, address, value, is_formula in v3_row_candidates(workbook, engine, label_cell):
            candidates.append((not is_formula, sheet_index, label_cell.row, offset, address, value))
    return candidates


def v3_find_scenarios(workbook, engine, oracle):
    located = {}
    used = set()
    for role in ("demand", "coal", "gas", "wind", "solar", "other", "emissions", "intensity"):
        candidates = v3_scenario_candidates(workbook, engine, role)
        located[role] = {}
        for case in ("base", "policy"):
            target = oracle[case][role]
            ranked = []
            for non_formula, sheet_index, row, offset, address, value in candidates:
                if value is None or (address in used and role != "other"):
                    continue
                scale = max(1.0, abs(target))
                ranked.append((abs(value - target) / scale, non_formula, sheet_index, row, offset, address))
            if ranked:
                ranked.sort()
                located[role][case] = ranked[0][-1]
                if role != "other" or oracle["base"]["other"] != oracle["policy"]["other"]:
                    used.add(ranked[0][-1])
    return located


def v3_result_label(role, text):
    if role == "emissions_reduction":
        return "emission" in text and any(word in text for word in ("reduction", "change", "delta", "decrease", "saving")) and "intensity" not in text
    if role == "intensity_reduction":
        return "intensity" in text and any(word in text for word in ("reduction", "change", "delta", "decrease", "improvement"))
    if role == "gas_change":
        return "gas" in text and any(word in text for word in ("change", "delta", "increase", "difference"))
    return False


def v3_find_results(workbook, engine, oracle):
    results = {}
    for role in ("emissions_reduction", "intensity_reduction", "gas_change"):
        candidates = []
        target = oracle[role]
        for sheet_index, _, label_cell, text in v3_label_cells(workbook, lambda text, role=role: v3_result_label(role, text)):
            for offset, address, value, is_formula in v3_row_candidates(workbook, engine, label_cell):
                if value is None:
                    continue
                options = [(abs(value - target), 1.0)]
                if any(word in text for word in ("change", "delta", "difference")):
                    options.append((abs(value + target), -1.0))
                error, direction = min(options)
                candidates.append((error / max(1.0, abs(target)), not is_formula, sheet_index, offset, address, direction))
        if candidates:
            candidates.sort()
            results[role] = {"address": candidates[0][4], "direction": candidates[0][5]}
    return results


def v3_scenario_value(workbook, engine, scenarios, role, case):
    address = scenarios.get(role, {}).get(case)
    return v3_value(workbook, engine, address) if address else None


def v3_result_value(workbook, engine, results, role):
    item = results.get(role)
    return v3_value(workbook, engine, item["address"]) if item else None


def v3_task_checks(workbook, oracle, oracle_module):
    checks = {row["id"]: 0.0 for row in TASK["criteria"] if row["id"] != "R001"}
    failures = []
    engine = FormulaEngine(workbook)
    sources = v3_find_sources(workbook, engine, oracle_module)
    assumptions = v3_find_assumptions(workbook, engine, oracle_module)
    scenarios = v3_find_scenarios(workbook, engine, oracle)
    results = v3_find_results(workbook, engine, oracle)

    required_source = {"coal", "gas", "wind", "solar", "demand", "coal_factor", "gas_factor"}
    required_assumptions = {"coal_displacement", "wind_uplift", "solar_uplift", "demand_growth"}
    required_scenarios = {"demand", "coal", "gas", "wind", "solar", "emissions", "intensity"}
    identified = required_source <= set(sources) and required_assumptions <= set(assumptions)
    identified = identified and all({"base", "policy"} <= set(scenarios.get(role, {})) for role in required_scenarios)
    identified = identified and {"emissions_reduction", "intensity_reduction", "gas_change"} <= set(results)

    def scenario_ok(role, case, tolerance, require_formula=False):
        address = scenarios.get(role, {}).get(case)
        if not address:
            failures.append(f"SCENARIO_METRIC_NOT_FOUND:{case}:{role}")
            return False
        value = v3_value(workbook, engine, address)
        ok = close(value, oracle[case][role], tolerance)
        if require_formula:
            ok = ok and formula_present(workbook, address)
        if not ok:
            failures.append(f"SCENARIO_VALUE_MISMATCH:{case}:{role}:{address}")
        return ok

    checks["R002"] = 1.0 if scenario_ok("coal", "policy", 0.5, True) else 0.0
    checks["R003"] = 1.0 if all(scenario_ok(role, "policy", 0.5, True) for role in ("wind", "solar")) else 0.0
    gas_address = scenarios.get("gas", {}).get("policy")
    gas_result = results.get("gas_change")
    gas_ok = scenario_ok("gas", "base", 0.5) and scenario_ok("gas", "policy", 0.5, True)
    gas_ok = gas_ok and bool(gas_result) and formula_present(workbook, gas_result["address"])
    if gas_result:
        gas_ok = gas_ok and close(v3_result_value(workbook, engine, results, "gas_change"), gas_result["direction"] * oracle["gas_change"], 0.5)
    checks["R004"] = 1.0 if gas_ok else 0.0
    checks["R005"] = 1.0 if scenario_ok("emissions", "policy", 0.5, True) else 0.0

    reduction = results.get("emissions_reduction")
    reduction_ok = bool(reduction) and formula_present(workbook, reduction["address"])
    if reduction:
        reduction_ok = reduction_ok and close(v3_result_value(workbook, engine, results, "emissions_reduction"), reduction["direction"] * oracle["emissions_reduction"], 0.5)
    checks["R006"] = 1.0 if reduction_ok else 0.0
    intensity = results.get("intensity_reduction")
    intensity_ok = bool(intensity) and formula_present(workbook, intensity["address"])
    if intensity:
        intensity_ok = intensity_ok and close(v3_result_value(workbook, engine, results, "intensity_reduction"), intensity["direction"] * oracle["intensity_reduction"], 1e-9)
    checks["R007"] = 1.0 if intensity_ok else 0.0

    all_cases_ok = True
    for case in ("base", "policy"):
        for role in ("demand", "coal", "gas", "wind", "solar", "emissions", "intensity"):
            tolerance = 1e-9 if role == "intensity" else 0.5
            all_cases_ok = scenario_ok(role, case, tolerance) and all_cases_ok
    base_other = oracle["base"]["demand"] - sum(oracle["base"][role] for role in ("coal", "gas", "wind", "solar"))
    policy_other = oracle["policy"]["demand"] - sum(oracle["policy"][role] for role in ("coal", "gas", "wind", "solar"))
    other_ok = close(base_other, oracle["base"]["other"], 0.5) and close(policy_other, oracle["policy"]["other"], 0.5)
    if scenarios.get("other", {}).get("base"):
        other_ok = other_ok and scenario_ok("other", "base", 0.5)
    if scenarios.get("other", {}).get("policy"):
        other_ok = other_ok and scenario_ok("other", "policy", 0.5)
    checks["R009"] = 1.0 if all_cases_ok and other_ok else 0.0

    source_ok = all(
        role in sources and close(v3_value(workbook, engine, sources[role]), oracle_module.DATA[role], 1e-9)
        for role in required_source
    )
    checks["R011"] = 1.0 if source_ok else 0.0
    if not source_ok:
        failures.append("SOURCE_INPUTS_CHANGED_OR_NOT_IDENTIFIED")

    try:
        emissions_addresses = [scenarios["emissions"][case] for case in ("base", "policy")]
        factor_ok = all(formula_present(workbook, address) for address in emissions_addresses)
        for factor_role, delta in (("coal_factor", 0.1), ("gas_factor", 0.1)):
            override_address = sources[factor_role]
            perturbed = FormulaEngine(workbook, {override_address: oracle_module.DATA[factor_role] + delta})
            fuel = "coal" if factor_role == "coal_factor" else "gas"
            for case, address in zip(("base", "policy"), emissions_addresses):
                expected = oracle[case]["emissions"] + oracle[case][fuel] * delta * 1000.0
                factor_ok = factor_ok and close(v3_value(workbook, perturbed, address), expected, 0.5)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        factor_ok = False
        failures.append(f"EMISSIONS_FACTOR_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R010"] = 1.0 if factor_ok else 0.0

    def dynamic_check(role, value, checked_roles, checked_results):
        policy_case = dict(oracle_module.POLICY)
        policy_case[role] = value
        expected = oracle_module.recompute(policy_case=policy_case)
        replay = FormulaEngine(workbook, {assumptions[role]: value})
        ok = True
        for output_role in checked_roles:
            address = scenarios.get(output_role, {}).get("policy")
            tolerance = 1e-9 if output_role == "intensity" else 0.5
            ok = bool(address) and close(v3_value(workbook, replay, address), expected["policy"][output_role], tolerance) and ok
        for result_role in checked_results:
            item = results.get(result_role)
            tolerance = 1e-9 if result_role == "intensity_reduction" else 0.5
            ok = bool(item) and close(v3_value(workbook, replay, item["address"]), item["direction"] * expected[result_role], tolerance) and ok
        return ok

    try:
        coal_value = oracle_module.POLICY["coal_displacement"] + (0.02 if ACTIVE_SPLIT == "dev" else 0.03)
        checks["R008"] = 1.0 if dynamic_check(
            "coal_displacement", coal_value, ("coal", "gas", "emissions"), ("emissions_reduction", "gas_change")
        ) else 0.0
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        failures.append(f"COAL_DISPLACEMENT_PERTURBATION_FAILED:{type(exc).__name__}")

    if ACTIVE_SPLIT == "dev":
        try:
            checks["R012"] = 1.0 if dynamic_check(
                "demand_growth", 0.025, ("demand", "gas", "emissions", "intensity"),
                ("emissions_reduction", "intensity_reduction", "gas_change")
            ) else 0.0
        except UnsupportedFormulaError:
            raise
        except Exception as exc:
            failures.append(f"DEMAND_GROWTH_PERTURBATION_FAILED:{type(exc).__name__}")

    if not identified:
        failures.append("SEMANTIC_CONTENT_INCOMPLETE")
    return identified, checks, failures


def evaluate_corrected(candidate):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"]
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    try:
        contract = load_contract()
        oracle, oracle_module = load_oracle(contract)
        identified, checks, task_failures = v3_task_checks(workbook, oracle, oracle_module)
        criteria["R001"] = 1.0 if identified else 0.0
        criteria.update(checks)
        failures.extend(task_failures)
    except UnsupportedFormulaError as exc:
        failures.append(f"UNSUPPORTED_FORMULA:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))


def evaluate(candidate):
    """Score the current v2 corrected task contract.

    Archived pre-v2 campaign workbooks are attributed as TASK_INVALID by the
    campaign replay binding before this evaluator is called.  Keeping that
    attribution outside the current task Judge prevents one executable from
    silently changing meaning because of a process environment variable.
    """
    return evaluate_corrected(candidate)



def main():
    candidate = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/output/answer.xlsx")
    criteria, failures = evaluate(candidate)
    payload = build_result(task=TASK, split=ACTIVE_SPLIT, candidate=str(candidate), criteria=criteria, failures=failures)
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True)
        (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\\n")
        if total is not None:
            (log_root / "reward.txt").write_text(str(total) + "\\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

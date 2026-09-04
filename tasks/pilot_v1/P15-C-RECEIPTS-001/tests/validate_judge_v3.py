#!/usr/bin/env python3
"""Focused Judge V3 regression: receipt business results without hidden formula/row rules."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


TASK_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TASK_DIR / "tests"))
SPEC = importlib.util.spec_from_file_location("receipts_evaluator", TASK_DIR / "tests" / "evaluate.py")
EVALUATOR = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(EVALUATOR)


def build_static_shifted_layout(path: Path) -> None:
    oracle = EVALUATOR.load_oracle("dev")
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    documents = workbook["Documents"]
    by_file = {doc["filename"]: doc for doc in oracle["documents"]}
    for row in range(4, documents.max_row + 1):
        doc = by_file[documents.cell(row, 8).value]
        for column, field in zip(range(4, 8), ("subtotal", "tax", "tip", "total")):
            documents.cell(row, column, doc[field])
    categories = workbook["Categories"]
    for row in range(4, categories.max_row + 1):
        categories.cell(row, 2, oracle["categories"][categories.cell(row, 1).value])
    reconciliation = workbook["Reconciliation"]
    reconciliation["B4"] = len(oracle["documents"])
    reconciliation["B5"] = len(oracle["items"])
    reconciliation["B6"] = oracle["batch_total"]
    for sheet in workbook.worksheets:
        sheet.insert_rows(2, amount=2)
    workbook.save(path)


def build_renamed_sheets(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    names = ("Receipt register", "Expense lines", "Spend summary", "Review exceptions", "Source register", "Batch check")
    for sheet, name in zip(workbook.worksheets, names):
        sheet.title = name
    workbook.save(path)


def corrupt_item_amount(workbook) -> None:
    sheet = workbook["Items"]
    header, columns = EVALUATOR.semantic_header(sheet, {
        "item": ("Item", "Description", "Item description"),
        "amount": ("Amount", "Price"),
    }, ("item", "amount"))
    assert header is not None
    for row in range(header + 1, sheet.max_row + 1):
        if EVALUATOR.norm(sheet.cell(row, columns["item"]).value) == "markers":
            sheet.cell(row, columns["amount"], 999.0)
            return
    raise AssertionError("Markers row not found")


def result_record(path: Path, expected: str, predicate, display_path: str | None = None) -> dict:
    result = EVALUATOR.score(path, "dev")
    return {
        "path": display_path or str(path.relative_to(TASK_DIR)),
        "score": result["normalized_score"],
        "pass": result["pass"],
        "criterion_scores": result["criterion_scores"],
        "failure_codes": result["failure_codes"],
        "expected": expected,
        "expected_pass": expected.startswith(("full pass", "pass with")),
        "expectation_met": bool(predicate(result)),
    }


def main() -> None:
    renamed_path = TASK_DIR / "fixtures/equivalent/candidate_renamed_sheets.xlsx"
    build_renamed_sheets(renamed_path)
    fixtures = {
        "reference": (TASK_DIR / "solution/reference.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "equivalent": (TASK_DIR / "fixtures/equivalent/candidate_equivalent.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "noop": (TASK_DIR / "fixtures/noop/candidate_noop.xlsx", "fail", lambda r: not r["pass"]),
        "malformed": (TASK_DIR / "fixtures/malformed/candidate_malformed.xlsx", "fail with zero score", lambda r: not r["pass"] and r["normalized_score"] == 0.0),
    }
    for path in sorted((TASK_DIR / "fixtures/mutants").glob("*.xlsx")):
        fixtures[f"mutant:{path.stem}"] = (path, "fail", lambda r: not r["pass"])
    fixture_results = {
        name: result_record(path, expected, predicate)
        for name, (path, expected, predicate) in fixtures.items()
    }

    with tempfile.TemporaryDirectory(prefix="receipts-judge-v3-") as directory:
        shifted = Path(directory) / "static_shifted.xlsx"
        build_static_shifted_layout(shifted)
        shifted_result = result_record(
            shifted,
            "full pass: correct static document/category/reconciliation values with shifted rows",
            lambda r: r["pass"] and r["normalized_score"] == 1.0,
            "generated:static_shifted_layout",
        )
        broken = Path(directory) / "wrong_item_amount.xlsx"
        workbook = load_workbook(shifted)
        corrupt_item_amount(workbook)
        workbook.save(broken)
        broken_result = result_record(
            broken,
            "fail R003: an extracted receipt item amount is wrong",
            lambda r: not r["pass"] and r["criterion_scores"]["R003"] == 0.0,
            "generated:wrong_item_amount",
        )

    renamed_result = result_record(
        renamed_path,
        "full pass: every sheet is renamed while visible headers preserve its business role",
        lambda r: r["pass"] and r["normalized_score"] == 1.0,
    )
    new_cases = {"renamed_sheets": renamed_result, "static_shifted_layout": shifted_result, "wrong_item_amount": broken_result}
    all_met = all(item["expectation_met"] for item in (*fixture_results.values(), *new_cases.values()))
    receipt = {
        "task_id": "P15-C-RECEIPTS-001",
        "judge_version": "p15_v3.0",
        "validation_scope": "reference, equivalent, noop, malformed, all mutants, shifted static layout, genuine item error",
        "expectation_basis": {
            "static_shifted_layout": "The instruction requires accurate receipt details, summaries, and reconciliation; it does not require formulas or fixed rows.",
            "renamed_sheets": "The instruction does not prescribe worksheet names; visible headers and labels establish each business role.",
            "wrong_item_amount": "Each extracted receipt item amount is a required source fact.",
        },
        "fixtures": fixture_results,
        "new_cases": new_cases,
        "all_expectations_met": all_met,
    }
    receipt_path = TASK_DIR / "receipts/judge_v3_local_validation.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"task_id": receipt["task_id"], "all_expectations_met": all_met, "scores": {key: [value["score"], value["pass"]] for key, value in {**fixture_results, **new_cases}.items()}}, ensure_ascii=False))
    if not all_met:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

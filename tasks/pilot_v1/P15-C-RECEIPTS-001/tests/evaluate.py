#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-C-RECEIPTS-001."""
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-C-RECEIPTS-001",
  "pass_threshold": 0.7,
  "criteria": [
    {
      "id": "R001",
      "description": "Workbook contains the complete receipt-batch delivery structure.",
      "weight": 2,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R002",
      "description": "All three receipt documents retain merchant, date, source-file, and total identity.",
      "weight": 5,
      "type": "positive",
      "dimension": "document_identity",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R003",
      "description": "All item rows are extracted once with document linkage, modifiers, and typed amounts.",
      "weight": 6,
      "type": "positive",
      "dimension": "line_item_accuracy",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R004",
      "description": "Document and batch totals are formula-linked and respond to a receipt-item perturbation.",
      "weight": 5,
      "type": "positive",
      "dimension": "native_workbook_semantics",
      "method": "deterministic",
      "method_params": {
        "perturbations": 2
      }
    },
    {
      "id": "R005",
      "description": "Expense categories reflect the document-grounded item taxonomy and batch amounts.",
      "weight": 3,
      "type": "positive",
      "dimension": "categorization",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R006",
      "description": "Every item preserves a valid receipt file and text locator.",
      "weight": 3,
      "type": "positive",
      "dimension": "provenance",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "P001",
      "description": "Penalty for provenance that cites a receipt filename outside the supplied sources.",
      "weight": -7,
      "type": "penalty",
      "dimension": "integrity",
      "method": "deterministic",
      "method_params": {}
    }
  ],
  "required_sheets": [
    "Documents",
    "Items",
    "Categories",
    "Exceptions",
    "Provenance",
    "Reconciliation"
  ]
}
TESTS_ROOT = Path(__file__).resolve().parent
REPO_TASK_ROOT = TESTS_ROOT.parent
DEV_ROOT = REPO_TASK_ROOT if (REPO_TASK_ROOT / "rubric.json").is_file() else TESTS_ROOT / "dev"
TASK = json.loads((DEV_ROOT / "rubric.json").read_text(encoding="utf-8"))
TASK["required_sheets"] = json.loads(
    (TESTS_ROOT / "private_contract.json").read_text(encoding="utf-8")
)["required_sheets"]
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R002", "R003"]
SHEET_ALIASES = {}


def semantic_sheet_aliases(workbook):
    signatures = {
        "Documents": (("Document ID", "Receipt ID", "ID"), ("Merchant", "Vendor"), ("Date", "Receipt date"), ("Subtotal", "Sub total"), ("Tax",), ("Tip", "Gratuity"), ("Total", "Receipt total"), ("Source file", "Source", "Filename")),
        "Items": (("Document ID", "Receipt ID", "ID"), ("Item", "Description", "Item description"), ("Modifier", "Option"), ("Amount", "Price"), ("Category", "Expense category"), ("Text locator", "Locator", "Source locator")),
        "Categories": (("Category", "Expense category"), ("Amount", "Value", "Total"), ("Documents", "Receipts"), ("Basis", "Explanation")),
        "Exceptions": (("Status",), ("Count",), ("Explanation", "Details")),
        "Provenance": (("Document ID", "Receipt ID", "ID"), ("Item", "Description", "Item description"), ("Source file", "Source", "Filename"), ("Text locator", "Locator", "Source locator"), ("Target row", "Workbook row")),
        "Reconciliation": (("Metric", "Check"), ("Value", "Observed", "Amount"), ("Expected",), ("Interpretation", "Explanation")),
    }
    aliases = {}
    for role, groups in signatures.items():
        candidates = []
        for sheet in workbook.worksheets:
            values = {
                norm(cell.value).rstrip(":")
                for row in sheet.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            }
            if all(any(norm(alias).rstrip(":") in values for alias in group) for group in groups):
                candidates.append(sheet.title)
        if candidates:
            aliases[role] = tuple(candidates)
    return aliases


def key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


def active_split(explicit=None):
    split = (explicit or os.environ.get("P15_EVAL_SPLIT", "dev")).strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"UNSUPPORTED_EVAL_SPLIT:{split}")
    return split


def excel_criteria_match(value, criterion):
    if not isinstance(criterion, str):
        return value == criterion
    match = re.fullmatch(r"\s*(<=|>=|<>|=|<|>)?\s*(.*?)\s*", criterion)
    operator, operand_text = match.groups() if match else (None, criterion)
    try:
        operand = float(operand_text)
        candidate = float(value)
    except (TypeError, ValueError):
        operand = norm(operand_text)
        candidate = norm(value)
    if operator in (None, "="): return candidate == operand
    if operator == "<>": return candidate != operand
    if operator == "<": return candidate < operand
    if operator == "<=": return candidate <= operand
    if operator == ">": return candidate > operand
    if operator == ">=": return candidate >= operand
    raise ValueError(f"UNSUPPORTED_CRITERIA_OPERATOR:{operator}")


def load_oracle(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "oracle_recompute.py" if split == "confirm" else DEV_ROOT / "metadata" / "oracle_recompute.py"
    spec = importlib.util.spec_from_file_location(f"task_oracle_{split}", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


class UnsupportedFormulaError(ValueError):
    """A valid Excel formula that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        current = key(sheet, cell)
        if current in self.overrides: return self.overrides[current]
        if current in self.memo: return self.memo[current]
        if current in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{current}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(current)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[current] = result
            return result
        finally:
            self.stack.discard(current)

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(col)}{row}") for row in range(int(r1), int(r2) + 1) for col in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def reference_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, reference_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            # Excel treats a genuinely blank referenced amount as zero in
            # arithmetic; keep a missing-item mutant as a scored semantic
            # failure instead of surfacing a Python TypeError.
            left = 0 if left is None else left
            right = 0 if right is None else right
            if isinstance(node.op, ast.Add): return left + right
            if isinstance(node.op, ast.Sub): return left - right
            if isinstance(node.op, ast.Mult): return left * right
            if isinstance(node.op, ast.Div): return left / right
            if isinstance(node.op, ast.Pow): return left ** right
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA:OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [item for group in args for item in (group if isinstance(group, list) else [group]) if item is not None]
            if name == "SUM": return sum(value for value in values if isinstance(value, (int, float)))
            if name == "SUMIF":
                if len(args) not in {2, 3}: raise ValueError("SUMIF_ARGUMENT_COUNT")
                criteria_range = args[0] if isinstance(args[0], list) else [args[0]]
                sum_range = args[2] if len(args) == 3 else criteria_range
                sum_range = sum_range if isinstance(sum_range, list) else [sum_range]
                if len(criteria_range) != len(sum_range): raise ValueError("SUMIF_RANGE_SIZE_MISMATCH")
                return sum(
                    value for value, candidate in zip(sum_range, criteria_range)
                    if isinstance(value, (int, float)) and excel_criteria_match(candidate, args[1])
                )
            if name == "COUNTIF":
                if len(args) != 2: raise ValueError("COUNTIF_ARGUMENT_COUNT")
                criteria_range = args[0] if isinstance(args[0], list) else [args[0]]
                return sum(excel_criteria_match(value, args[1]) for value in criteria_range)
            if name == "COUNTA": return sum(value not in (None, "") for value in values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA:NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(float(actual) - float(expected)) <= tolerance


def text(value): return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", text(value).lower())


def normalized_date(value):
    """Compare native Excel dates and common receipt-date text semantically."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    candidate = text(value)
    if not candidate:
        return ""
    try:
        return datetime.fromisoformat(candidate).date().isoformat()
    except ValueError:
        pass
    for date_format in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(candidate, date_format).date().isoformat()
        except ValueError:
            continue
    return norm(candidate)


def load_perturbations(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "perturbations.json" if split == "confirm" else DEV_ROOT / "metadata" / "perturbations.json"
    records = json.loads(path.read_text(encoding="utf-8"))
    expected_count = 1 if split == "confirm" else 2
    if len(records) != expected_count or len({record.get("id") for record in records}) != expected_count:
        raise ValueError(f"{split.upper()}_PERTURBATION_MANIFEST_MUST_CONTAIN_{expected_count}_UNIQUE_CASES")
    return records


def find_content_row(workbook, spec):
    sheet = workbook[spec["sheet"]]
    primary_col = col_index(spec["match_column"])
    secondary_col = col_index(spec["secondary_match_column"]) if spec.get("secondary_match_column") else None
    rows = []
    for row in range(4, sheet.max_row + 1):
        if norm(sheet.cell(row=row, column=primary_col).value) != norm(spec["match_value"]):
            continue
        if secondary_col and norm(sheet.cell(row=row, column=secondary_col).value) != norm(spec["secondary_match_value"]):
            continue
        rows.append(row)
    return rows[0] if len(rows) == 1 else None


def perturbation_case_ok(workbook, case):
    target = case["target"]
    row = find_content_row(workbook, target)
    if row is None:
        return False
    target_cell = f'{target["value_column"]}{row}'
    baseline_engine = FormulaEngine(workbook)
    if not close(baseline_engine.value(target["sheet"], target_cell), target["baseline"]):
        return False
    engine = FormulaEngine(workbook, {key(target["sheet"], target_cell): target["perturbed"]})
    for expected in case.get("expected", []):
        if not close(engine.value(expected["sheet"], expected["cell"]), expected["value"]):
            return False
    for expected in case.get("expected_by_content", []):
        expected_row = find_content_row(workbook, expected)
        if expected_row is None or not close(
            engine.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'), expected["value"]
        ):
            return False
    for protected in case.get("protected", []):
        if not close(engine.value(protected["sheet"], protected["cell"]), protected["value"]):
            return False
    for protected in case.get("protected_by_content", []):
        protected_row = find_content_row(workbook, protected)
        if protected_row is None or not close(
            engine.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'), protected["value"]
        ):
            return False
    return True


def perturbation_response_ok(workbook, split="dev"):
    return all(perturbation_case_ok(workbook, case) for case in load_perturbations(split))


def source_field_matches(value, filename):
    rendered, expected = norm(value), norm(filename)
    return bool(rendered and expected and expected in rendered)


def source_field_contradicts_any(value, filenames):
    rendered = text(value)
    if not rendered or any(source_field_matches(rendered, filename) for filename in filenames):
        return False
    return re.search(r"\.(?:pdf|png|jpe?g|tiff?|xlsx?|csv)\b", rendered, re.IGNORECASE) is not None


def locator_ok(value, filename=None, page=None, item=None):
    locator = norm(value)
    if not locator:
        return False
    item_signal = bool(item and norm(item) in locator)
    position_signal = re.search(r"(?:line|row|item|entry)[\s:_#-]*\d+\b", locator) is not None
    if not (item_signal or position_signal):
        return False
    if page is None:
        return True
    return re.search(rf"(?:p|page)[\s.:#_-]*{int(page)}\b", locator) is not None


def present_formula(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def formula_value(engine, sheet, cell):
    try:
        return engine.value(sheet, cell)
    except UnsupportedFormulaError:
        raise
    except Exception:
        return None


def row_values(workbook, sheet, start, end, cols):
    records = []
    for row in range(start, end + 1):
        values = tuple(workbook[sheet].cell(row=row, column=col).value for col in cols)
        if any(value not in (None, "") for value in values): records.append(values)
    return records


def static_summary_detail_checks(workbook, oracle):
    """Score the observed static Summary/Detail delivery by visible semantics.

    This adapter does not award formula or perturbation credit.  It accepts
    only the explicit two-sheet, formula-free layout seen in preserved model
    output; other reorganizations remain review-pending.
    """
    if set(workbook.sheetnames) != {"Summary", "Detail"}:
        return None
    if any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        return None
    detail = workbook["Detail"]
    headers = [norm(detail.cell(row=1, column=column).value) for column in range(1, 11)]
    required_headers = [
        "receipt id", "source file", "merchant", "date", "item description",
        "amount", "category", "tax", "tip", "receipt total",
    ]
    if headers != required_headers:
        return None

    expected_docs = {norm(doc["filename"]): doc for doc in oracle["documents"]}

    def canonical_source(value):
        matches = [name for name in expected_docs if source_field_matches(value, name)]
        return matches[0] if len(matches) == 1 else ""

    items = []
    for row in range(2, detail.max_row + 1):
        if not text(detail[f"E{row}"].value):
            continue
        items.append(
            {
                "id": text(detail[f"A{row}"].value),
                "filename": text(detail[f"B{row}"].value),
                "canonical_filename": canonical_source(detail[f"B{row}"].value),
                "merchant": text(detail[f"C{row}"].value),
                "date": detail[f"D{row}"].value,
                "item": text(detail[f"E{row}"].value),
                "amount": detail[f"F{row}"].value,
                "category": text(detail[f"G{row}"].value),
                "tax": detail[f"H{row}"].value,
                "tip": detail[f"I{row}"].value,
                "total": detail[f"J{row}"].value,
            }
        )
    expected_items = {
        (norm(next(doc["filename"] for doc in oracle["documents"] if doc["document_id"] == item["document_id"])), norm(item["item"])): item
        for item in oracle["items"]
    }
    actual_items = {
        (item["canonical_filename"], norm(item["item"])): item for item in items
    }
    items_ok = (
        len(items) == len(expected_items) == len(actual_items)
        and set(actual_items) == set(expected_items)
        and all(
            close(actual_items[key_]["amount"], item["amount"])
            and norm(actual_items[key_]["category"]) == norm(item["category"])
            for key_, item in expected_items.items()
        )
    )

    docs = {}
    duplicate_doc_ids = False
    for filename, expected in expected_docs.items():
        members = [item for item in items if item["canonical_filename"] == filename]
        ids = {item["id"] for item in members}
        duplicate_doc_ids = duplicate_doc_ids or len(ids) != 1 or "" in ids
        tax_values = [item["tax"] for item in members if item["tax"] not in (None, "")]
        tip_values = [item["tip"] for item in members if item["tip"] not in (None, "")]
        total_values = [item["total"] for item in members if item["total"] not in (None, "")]
        docs[filename] = {
            "members": members,
            "id": next(iter(ids), ""),
            "merchant_ok": bool(members) and all(norm(item["merchant"]) == norm(expected["merchant"]) for item in members),
            "date_ok": bool(members) and all(normalized_date(item["date"]) == normalized_date(expected["date"]) for item in members),
            "subtotal": sum(float(item["amount"]) for item in members if isinstance(item["amount"], (int, float))),
            "tax": tax_values[0] if len(tax_values) == 1 else None,
            "tip": tip_values[0] if len(tip_values) == 1 else None,
            "total": total_values[0] if len(total_values) == 1 else None,
        }
    documents_ok = (
        set(docs) == set(expected_docs)
        and all(
            value["members"]
            and value["merchant_ok"]
            and value["date_ok"]
            and close(value["subtotal"], expected_docs[name]["subtotal"])
            and close(value["tax"], expected_docs[name]["tax"])
            and close(value["tip"], expected_docs[name]["tip"])
            and close(value["total"], expected_docs[name]["total"])
            for name, value in docs.items()
        )
    )
    linkage_ok = items_ok and documents_ok and not duplicate_doc_ids

    summary = workbook["Summary"]
    categories = {}
    for row in range(1, summary.max_row + 1):
        label = norm(summary[f"A{row}"].value)
        value = summary[f"C{row}"].value
        if label in {norm(name) for name in oracle["categories"]}:
            categories[label] = value
    category_values_ok = len(categories) == len(oracle["categories"]) and all(
        close(categories.get(norm(name)), value)
        for name, value in oracle["categories"].items()
    )

    provenance_ok = items_ok and all(
        item["canonical_filename"]
        and text(item["item"])
        for item in items
    )
    false_provenance = any(
        source_field_contradicts_any(item["filename"], expected_docs) for item in items
    )

    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    checks.update(
        {
            "R001": 1.0,
            "R002": 1.0 if documents_ok else 0.0,
            "R003": 1.0 if items_ok and linkage_ok else 0.0,
            "R004": 1.0 if linkage_ok else 0.0,
            "R005": 1.0 if documents_ok else 0.0,
            "R006": 1.0 if documents_ok and close(sum(doc["total"] for doc in docs.values()), oracle["batch_total"]) else 0.0,
            "R007": 1.0 if category_values_ok else 0.0,
            "R008": 1.0 if provenance_ok else 0.0,
            "R009": 1.0 if documents_ok and items_ok and close(sum(doc["total"] for doc in docs.values()), oracle["batch_total"]) else 0.0,
            "P001": 1.0 if false_provenance else 0.0,
        }
    )
    failures = ["TASK_LOCAL_SUMMARY_DETAIL_LAYOUT"]
    if not category_values_ok:
        failures.append("CATEGORY_SUMMARY_VALUE_MISMATCH")
    if false_provenance:
        failures.append("FABRICATED_RECEIPT_PROVENANCE")
    return checks, failures


def _legacy_task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    checks["R001"] = 1.0
    expected_docs = {norm(doc["filename"]): doc for doc in oracle["documents"]}

    def canonical_source(value):
        matches = [filename for filename in expected_docs if source_field_matches(value, filename)]
        return matches[0] if len(matches) == 1 else ""

    documents_sheet = workbook["Documents"]
    actual_docs = []
    for row in range(4, documents_sheet.max_row + 1):
        if text(documents_sheet[f"H{row}"].value):
            actual_docs.append({
                "row": row, "id": text(documents_sheet[f"A{row}"].value),
                "merchant": text(documents_sheet[f"B{row}"].value), "date": text(documents_sheet[f"C{row}"].value),
                "subtotal": formula_value(engine, "Documents", f"D{row}"),
                "tax": formula_value(engine, "Documents", f"E{row}"), "tip": documents_sheet[f"F{row}"].value,
                "total": formula_value(engine, "Documents", f"G{row}"),
                "filename": text(documents_sheet[f"H{row}"].value),
                "canonical_filename": canonical_source(documents_sheet[f"H{row}"].value),
            })
    actual_docs_by_file = {doc["canonical_filename"]: doc for doc in actual_docs}
    documents_ok = (
        len(actual_docs) == len(expected_docs) == len(actual_docs_by_file)
        and set(actual_docs_by_file) == set(expected_docs)
        and all(
            norm(actual_docs_by_file[name]["merchant"]) == norm(doc["merchant"])
            and normalized_date(actual_docs_by_file[name]["date"]) == normalized_date(doc["date"])
            for name, doc in expected_docs.items()
        )
    )
    checks["R002"] = 1.0 if documents_ok else 0.0
    items_sheet = workbook["Items"]
    actual_items = []
    for row in range(4, items_sheet.max_row + 1):
        if text(items_sheet[f"B{row}"].value):
            actual_items.append({
                "row": row, "document_id": text(items_sheet[f"A{row}"].value),
                "item": text(items_sheet[f"B{row}"].value), "modifier": text(items_sheet[f"C{row}"].value),
                "amount": items_sheet[f"D{row}"].value, "category": text(items_sheet[f"E{row}"].value),
                "filename": text(items_sheet[f"F{row}"].value),
                "canonical_filename": canonical_source(items_sheet[f"F{row}"].value),
            })
    oracle_doc_file = {doc["document_id"]: doc["filename"] for doc in oracle["documents"]}
    expected_items = {(norm(oracle_doc_file[item["document_id"]]), norm(item["item"])): item for item in oracle["items"]}
    actual_items_by_key = {(item["canonical_filename"], norm(item["item"])): item for item in actual_items}
    def modifier_ok(value, expected):
        return norm(value) == norm(expected) or (norm(expected) == "none" and norm(value) in {"", "none", "n/a", "no modifier"})
    items_ok = (
        len(actual_items) == len(expected_items) == len(actual_items_by_key)
        and set(actual_items_by_key) == set(expected_items)
        and all(
            isinstance(actual_items_by_key[key_]["amount"], (int, float))
            and close(actual_items_by_key[key_]["amount"], item["amount"])
            and norm(actual_items_by_key[key_]["category"]) == norm(item["category"])
            and modifier_ok(actual_items_by_key[key_]["modifier"], item["modifier"])
            for key_, item in expected_items.items()
        )
    )
    doc_id_by_file = {doc["canonical_filename"]: doc["id"] for doc in actual_docs}
    linkage_ok = documents_ok and items_ok and all(
        item["document_id"] == doc_id_by_file.get(item["canonical_filename"]) for item in actual_items
    )
    identity_ok = (
        documents_ok and items_ok and linkage_ok
        and all(doc["id"] for doc in actual_docs)
        and len({doc["id"] for doc in actual_docs}) == len(expected_docs)
    )
    checks["R003"] = 1.0 if items_ok and linkage_ok else 0.0
    checks["R004"] = 1.0 if identity_ok else 0.0
    document_totals_ok = documents_ok and all(
        close(actual_docs_by_file[name][field], doc[field])
        for name, doc in expected_docs.items() for field in ("subtotal", "tax", "tip", "total")
    )
    document_formulas_ok = documents_ok and all(
        all(present_formula(workbook, f'Documents!{column}{doc["row"]}') for column in ("D", "E", "G"))
        for doc in actual_docs
    )
    checks["R005"] = 1.0 if document_totals_ok and document_formulas_ok else 0.0
    try:
        dynamic_ok = perturbation_response_ok(workbook, split)
    except Exception as exc:
        dynamic_ok = False; failures.append(f"RECEIPT_PERTURBATION_FAILED:{type(exc).__name__}")
    batch_formula_ok = present_formula(workbook, "Reconciliation!B6")
    checks["R006"] = 1.0 if dynamic_ok and batch_formula_ok else 0.0
    categories_sheet = workbook["Categories"]
    actual_categories = {}
    duplicate_categories = set()
    category_formula_ok = True
    for row in range(4, categories_sheet.max_row + 1):
        if text(categories_sheet[f"A{row}"].value):
            category = norm(categories_sheet[f"A{row}"].value)
            if category in actual_categories:
                duplicate_categories.add(category)
            actual_categories[category] = formula_value(engine, "Categories", f"B{row}")
            category_formula_ok = category_formula_ok and present_formula(workbook, f"Categories!B{row}")
    categories_ok = not duplicate_categories and len(actual_categories) == len(oracle["categories"]) and all(
        close(actual_categories.get(norm(category)), value) for category, value in oracle["categories"].items()
    )
    checks["R007"] = 1.0 if categories_ok and category_formula_ok else 0.0
    provenance_rows = []
    provenance_sheet = workbook["Provenance"]
    for row in range(4, provenance_sheet.max_row + 1):
        if text(provenance_sheet[f"B{row}"].value):
            provenance_rows.append({
                "document_id": text(provenance_sheet[f"A{row}"].value), "item": text(provenance_sheet[f"B{row}"].value),
                "filename": text(provenance_sheet[f"C{row}"].value),
                "canonical_filename": canonical_source(provenance_sheet[f"C{row}"].value),
                "locator": text(provenance_sheet[f"D{row}"].value),
            })
    provenance_by_key = {}
    for record in provenance_rows:
        provenance_by_key.setdefault((record["canonical_filename"], norm(record["item"])), []).append(record)
    provenance_population_ok = (
        len(provenance_rows) == len(expected_items)
        and set(provenance_by_key) == set(expected_items)
        and all(len(records) == 1 for records in provenance_by_key.values())
    )
    provenance_ok = items_ok and provenance_population_ok and all(
        len(provenance_by_key.get(key_, [])) == 1
        and provenance_by_key[key_][0]["document_id"] == actual_items_by_key[key_]["document_id"]
        and locator_ok(
            provenance_by_key[key_][0]["locator"],
            actual_items_by_key[key_]["filename"],
            item=actual_items_by_key[key_]["item"],
        )
        for key_ in expected_items
    )
    checks["R008"] = 1.0 if provenance_ok else 0.0
    reconciliation_ok = (
        all(present_formula(workbook, f"Reconciliation!B{row}") for row in (4, 5, 6))
        and close(engine.value("Reconciliation", "B4"), len(expected_docs))
        and close(engine.value("Reconciliation", "B5"), len(expected_items))
        and close(engine.value("Reconciliation", "B6"), oracle["batch_total"])
    )
    checks["R009"] = 1.0 if reconciliation_ok else 0.0
    false_provenance = any(
        source_field_contradicts_any(record["filename"], expected_docs)
        for record in provenance_rows
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_RECEIPT_PROVENANCE")
    return checks, failures


def semantic_header(sheet, aliases, required):
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for row in range(1, sheet.max_row + 1):
        columns = {}
        for column in range(1, sheet.max_column + 1):
            value = norm(sheet.cell(row=row, column=column).value)
            for role, accepted in normalized.items():
                if value in accepted and role not in columns:
                    columns[role] = column
        if set(required) <= set(columns):
            return row, columns
    return None, {}


def metric_values(sheet, engine):
    header, columns = semantic_header(sheet, {
        "metric": ("Metric", "Check"),
        "value": ("Value", "Observed", "Amount"),
    }, ("metric", "value"))
    values = {}
    if header is None:
        return values
    for row in range(header + 1, sheet.max_row + 1):
        label = norm(sheet.cell(row=row, column=columns["metric"]).value)
        if label:
            cell = sheet.cell(row=row, column=columns["value"])
            values[label] = formula_value(engine, sheet.title, cell.coordinate)
    return values


def task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    checks["R001"] = 1.0
    expected_docs = {norm(doc["filename"]): doc for doc in oracle["documents"]}

    def canonical_source(value):
        matches = [filename for filename in expected_docs if source_field_matches(value, filename)]
        return matches[0] if len(matches) == 1 else ""

    documents_sheet = workbook["Documents"]
    documents_header, document_columns = semantic_header(documents_sheet, {
        "id": ("Document ID", "Receipt ID", "ID"),
        "merchant": ("Merchant", "Vendor"),
        "date": ("Date", "Receipt date"),
        "subtotal": ("Subtotal", "Sub total"),
        "tax": ("Tax",),
        "tip": ("Tip", "Gratuity"),
        "total": ("Total", "Receipt total"),
        "source": ("Source file", "Source", "Filename"),
    }, ("id", "merchant", "date", "subtotal", "tax", "tip", "total", "source"))
    actual_docs = []
    if documents_header is not None:
        for row in range(documents_header + 1, documents_sheet.max_row + 1):
            source = text(documents_sheet.cell(row=row, column=document_columns["source"]).value)
            if not source:
                continue
            actual_docs.append({
                "row": row,
                "id": text(documents_sheet.cell(row=row, column=document_columns["id"]).value),
                "merchant": text(documents_sheet.cell(row=row, column=document_columns["merchant"]).value),
                "date": documents_sheet.cell(row=row, column=document_columns["date"]).value,
                "subtotal": formula_value(engine, documents_sheet.title, documents_sheet.cell(row=row, column=document_columns["subtotal"]).coordinate),
                "tax": formula_value(engine, documents_sheet.title, documents_sheet.cell(row=row, column=document_columns["tax"]).coordinate),
                "tip": formula_value(engine, documents_sheet.title, documents_sheet.cell(row=row, column=document_columns["tip"]).coordinate),
                "total": formula_value(engine, documents_sheet.title, documents_sheet.cell(row=row, column=document_columns["total"]).coordinate),
                "filename": source,
                "canonical_filename": canonical_source(source),
            })
    docs_by_file = {doc["canonical_filename"]: doc for doc in actual_docs}
    documents_identity_ok = (
        len(actual_docs) == len(expected_docs) == len(docs_by_file)
        and set(docs_by_file) == set(expected_docs)
        and all(
            norm(docs_by_file[name]["merchant"]) == norm(doc["merchant"])
            and normalized_date(docs_by_file[name]["date"]) == normalized_date(doc["date"])
            for name, doc in expected_docs.items()
        )
    )
    documents_totals_ok = documents_identity_ok and all(
        close(docs_by_file[name][field], doc[field])
        for name, doc in expected_docs.items()
        for field in ("subtotal", "tax", "tip", "total")
    )
    checks["R002"] = 1.0 if documents_identity_ok else 0.0

    items_sheet = workbook["Items"]
    items_header, item_columns = semantic_header(items_sheet, {
        "id": ("Document ID", "Receipt ID", "ID"),
        "item": ("Item", "Description", "Item description"),
        "modifier": ("Modifier", "Option"),
        "amount": ("Amount", "Price"),
        "category": ("Category", "Expense category"),
        "source": ("Source file", "Source", "Filename"),
        "locator": ("Text locator", "Locator", "Source locator"),
    }, ("id", "item", "amount", "category", "source"))
    actual_items = []
    if items_header is not None:
        for row in range(items_header + 1, items_sheet.max_row + 1):
            item_name = text(items_sheet.cell(row=row, column=item_columns["item"]).value)
            if not item_name:
                continue
            source = text(items_sheet.cell(row=row, column=item_columns["source"]).value)
            actual_items.append({
                "document_id": text(items_sheet.cell(row=row, column=item_columns["id"]).value),
                "item": item_name,
                "modifier": text(items_sheet.cell(row=row, column=item_columns["modifier"]).value) if "modifier" in item_columns else "",
                "amount": items_sheet.cell(row=row, column=item_columns["amount"]).value,
                "category": text(items_sheet.cell(row=row, column=item_columns["category"]).value),
                "filename": source,
                "canonical_filename": canonical_source(source),
            })
    oracle_doc_file = {doc["document_id"]: doc["filename"] for doc in oracle["documents"]}
    expected_items = {(norm(oracle_doc_file[item["document_id"]]), norm(item["item"])): item for item in oracle["items"]}
    actual_items_by_key = {(item["canonical_filename"], norm(item["item"])): item for item in actual_items}

    def modifier_ok(value, expected):
        return norm(value) == norm(expected) or (norm(expected) == "none" and norm(value) in {"", "none", "n/a", "no modifier"})

    items_ok = (
        len(actual_items) == len(expected_items) == len(actual_items_by_key)
        and set(actual_items_by_key) == set(expected_items)
        and all(
            close(actual_items_by_key[key_]["amount"], item["amount"])
            and norm(actual_items_by_key[key_]["category"]) == norm(item["category"])
            and modifier_ok(actual_items_by_key[key_]["modifier"], item["modifier"])
            for key_, item in expected_items.items()
        )
    )
    doc_id_by_file = {doc["canonical_filename"]: doc["id"] for doc in actual_docs}
    linkage_ok = documents_identity_ok and items_ok and all(
        item["document_id"] == doc_id_by_file.get(item["canonical_filename"])
        for item in actual_items
    )
    identity_ok = linkage_ok and all(doc["id"] for doc in actual_docs) and len({doc["id"] for doc in actual_docs}) == len(expected_docs)
    checks["R003"] = 1.0 if items_ok and linkage_ok else 0.0
    checks["R004"] = 1.0 if identity_ok else 0.0
    checks["R005"] = 1.0 if documents_totals_ok else 0.0

    categories_sheet = workbook["Categories"]
    categories_header, category_columns = semantic_header(categories_sheet, {
        "category": ("Category", "Expense category"),
        "amount": ("Amount", "Total", "Value"),
    }, ("category", "amount"))
    actual_categories = {}
    if categories_header is not None:
        for row in range(categories_header + 1, categories_sheet.max_row + 1):
            category = norm(categories_sheet.cell(row=row, column=category_columns["category"]).value)
            if category:
                cell = categories_sheet.cell(row=row, column=category_columns["amount"])
                actual_categories[category] = formula_value(engine, categories_sheet.title, cell.coordinate)
    categories_ok = len(actual_categories) == len(oracle["categories"]) and all(
        close(actual_categories.get(norm(category)), value)
        for category, value in oracle["categories"].items()
    )
    checks["R007"] = 1.0 if categories_ok else 0.0

    provenance_sheet = workbook["Provenance"]
    provenance_header, provenance_columns = semantic_header(provenance_sheet, {
        "id": ("Document ID", "Receipt ID", "ID"),
        "item": ("Item", "Description", "Item description"),
        "source": ("Source file", "Source", "Filename"),
        "locator": ("Text locator", "Locator", "Source locator"),
    }, ("id", "item", "source", "locator"))
    provenance_by_key = {}
    provenance_records = []
    if provenance_header is not None:
        for row in range(provenance_header + 1, provenance_sheet.max_row + 1):
            item_name = text(provenance_sheet.cell(row=row, column=provenance_columns["item"]).value)
            if not item_name:
                continue
            source = text(provenance_sheet.cell(row=row, column=provenance_columns["source"]).value)
            record = {
                "document_id": text(provenance_sheet.cell(row=row, column=provenance_columns["id"]).value),
                "item": item_name,
                "filename": source,
                "canonical_filename": canonical_source(source),
                "locator": text(provenance_sheet.cell(row=row, column=provenance_columns["locator"]).value),
            }
            provenance_records.append(record)
            provenance_by_key.setdefault((record["canonical_filename"], norm(record["item"])), []).append(record)
    provenance_ok = items_ok and len(provenance_records) == len(expected_items) and set(provenance_by_key) == set(expected_items) and all(
        len(provenance_by_key[key_]) == 1
        and provenance_by_key[key_][0]["document_id"] == actual_items_by_key[key_]["document_id"]
        and locator_ok(provenance_by_key[key_][0]["locator"], actual_items_by_key[key_]["filename"], item=actual_items_by_key[key_]["item"])
        for key_ in expected_items
    )
    checks["R008"] = 1.0 if provenance_ok else 0.0

    reconciliation = metric_values(workbook["Reconciliation"], engine)
    receipt_count = next((value for label, value in reconciliation.items() if "receipt count" in label or "document count" in label), None)
    item_count = next((value for label, value in reconciliation.items() if "item count" in label), None)
    batch_total = next((value for label, value in reconciliation.items() if "batch total" in label or "grand total" in label), None)
    reconciliation_ok = close(receipt_count, len(expected_docs)) and close(item_count, len(expected_items)) and close(batch_total, oracle["batch_total"])
    checks["R006"] = 1.0 if documents_totals_ok and categories_ok and reconciliation_ok else 0.0
    checks["R009"] = 1.0 if reconciliation_ok else 0.0

    false_provenance = any(source_field_contradicts_any(record["filename"], expected_docs) for record in provenance_records)
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_RECEIPT_PROVENANCE")
    return checks, failures


def score(candidate, split=None):
    split = active_split(split)
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=["OUTPUT_MISSING"],
        )
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=[f"MALFORMED_XLSX:{type(exc).__name__}"],
        )

    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], semantic_sheet_aliases(workbook)
    )
    checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        oracle = load_oracle(split)
        consolidated = static_summary_detail_checks(workbook, oracle)
        if consolidated is not None:
            checks, task_failures = consolidated
            return build_result(
                task=TASK,
                split=split,
                candidate=str(candidate),
                criteria=checks,
                failures=task_failures,
            )
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=failures,
        )

    try:
        oracle = load_oracle(split)
        engine = FormulaEngine(workbook)
        checks, task_failures = task_checks(workbook, engine, oracle, split)
        checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=checks,
        failures=failures,
    )


if __name__ == "__main__":
    arguments = sys.argv[1:]
    selected_split = None
    if "--split" in arguments:
        split_index = arguments.index("--split")
        if split_index + 1 >= len(arguments):
            raise SystemExit("--split requires dev or confirm")
        selected_split = arguments[split_index + 1]
        del arguments[split_index:split_index + 2]
    if len(arguments) > 1:
        raise SystemExit("usage: evaluate.py [candidate.xlsx] [--split dev|confirm]")
    target = Path(arguments[0]) if arguments else Path("/app/output/answer.xlsx")
    print(json.dumps(score(target, selected_split), sort_keys=True))

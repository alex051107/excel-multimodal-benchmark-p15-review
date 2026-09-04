#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-C-INVOICE-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-C-INVOICE-001",
  "pass_threshold": 0.7,
  "criteria": [
    {
      "id": "R001",
      "description": "Workbook contains the four AP-deliverable sheets and opens safely.",
      "weight": 2,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R002",
      "description": "Invoice header fields match the document identity, parties, date, currency, and PO reference.",
      "weight": 4,
      "type": "positive",
      "dimension": "header_accuracy",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R003",
      "description": "All four typed line items are extracted once with correct quantities, unit prices, and line-total formulas.",
      "weight": 6,
      "type": "positive",
      "dimension": "line_item_accuracy",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R004",
      "description": "Discount, tax, freight, and payable total calculate correctly and respond to a unit-price perturbation.",
      "weight": 5,
      "type": "positive",
      "dimension": "native_workbook_semantics",
      "method": "deterministic",
      "method_params": {
        "perturbations": 2
      }
    },
    {
      "id": "R005",
      "description": "Each extracted line has the correct document-page and text locator provenance.",
      "weight": 4,
      "type": "positive",
      "dimension": "provenance",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R006",
      "description": "The line-item table has no missing or duplicate row identity.",
      "weight": 3,
      "type": "positive",
      "dimension": "completeness",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "P001",
      "description": "Penalty for fabricated source-file or page provenance.",
      "weight": -5,
      "type": "penalty",
      "dimension": "integrity",
      "method": "deterministic",
      "method_params": {}
    }
  ],
  "required_sheets": [
    "Invoice_Header",
    "Line_Items",
    "Provenance",
    "Checks"
  ]
}
TESTS_ROOT = Path(__file__).resolve().parent
REPO_TASK_ROOT = TESTS_ROOT.parent
DEV_ROOT = REPO_TASK_ROOT if (REPO_TASK_ROOT / "rubric.json").is_file() else TESTS_ROOT / "dev"
TASK = json.loads((DEV_ROOT / "rubric.json").read_text(encoding="utf-8"))
TASK["required_sheets"] = json.loads(
    (TESTS_ROOT / "private_contract.json").read_text(encoding="utf-8")
)["required_sheets"]
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R002", "R003", "R006", "R008"]
SHEET_ALIASES = {}


def key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


def active_split(explicit=None):
    split = (explicit or os.environ.get("P15_EVAL_SPLIT", "dev")).strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"UNSUPPORTED_EVAL_SPLIT:{split}")
    return split


def load_oracle(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "oracle_recompute.py" if split == "confirm" else DEV_ROOT / "metadata" / "oracle_recompute.py"
    spec = importlib.util.spec_from_file_location(f"task_oracle_{split}", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


class UnsupportedFormulaError(ValueError):
    """A valid Excel formula that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        current = key(sheet, cell)
        if current in self.overrides: return self.overrides[current]
        if current in self.memo: return self.memo[current]
        if current in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{current}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(current)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[current] = result
            return result
        finally:
            self.stack.discard(current)

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(col)}{row}") for row in range(int(r1), int(r2) + 1) for col in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def reference_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, reference_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            if isinstance(node.op, ast.Add): return left + right
            if isinstance(node.op, ast.Sub): return left - right
            if isinstance(node.op, ast.Mult): return left * right
            if isinstance(node.op, ast.Div): return left / right
            if isinstance(node.op, ast.Pow): return left ** right
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA:OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [item for group in args for item in (group if isinstance(group, list) else [group]) if item is not None]
            if name == "SUM": return sum(values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA:NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(float(actual) - float(expected)) <= tolerance


def text(value): return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", text(value).lower())


def semantic_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    candidate = text(value)
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(candidate, pattern).date().isoformat()
        except ValueError:
            continue
    return norm(candidate)


def load_perturbations(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "perturbations.json" if split == "confirm" else DEV_ROOT / "metadata" / "perturbations.json"
    records = json.loads(path.read_text(encoding="utf-8"))
    expected_count = 1 if split == "confirm" else 2
    if len(records) != expected_count or len({record.get("id") for record in records}) != expected_count:
        raise ValueError(f"{split.upper()}_PERTURBATION_MANIFEST_MUST_CONTAIN_{expected_count}_UNIQUE_CASES")
    return records


def find_content_row(workbook, spec):
    sheet = workbook[spec["sheet"]]
    primary_col = col_index(spec["match_column"])
    secondary_col = col_index(spec["secondary_match_column"]) if spec.get("secondary_match_column") else None
    rows = []
    for row in range(4, sheet.max_row + 1):
        if norm(sheet.cell(row=row, column=primary_col).value) != norm(spec["match_value"]):
            continue
        if secondary_col and norm(sheet.cell(row=row, column=secondary_col).value) != norm(spec["secondary_match_value"]):
            continue
        rows.append(row)
    return rows[0] if len(rows) == 1 else None


def perturbation_case_ok(workbook, case):
    target = case["target"]
    row = find_content_row(workbook, target)
    if row is None:
        return False
    target_cell = f'{target["value_column"]}{row}'
    baseline_engine = FormulaEngine(workbook)
    if not close(baseline_engine.value(target["sheet"], target_cell), target["baseline"]):
        return False
    baseline_discount = baseline_engine.value("Invoice_Header", "B14")
    negative_discount_convention = isinstance(baseline_discount, (int, float)) and baseline_discount < 0
    engine = FormulaEngine(workbook, {key(target["sheet"], target_cell): target["perturbed"]})
    for expected in case.get("expected", []):
        target_value = expected["value"]
        if expected["sheet"] == "Invoice_Header" and expected["cell"] == "B14":
            target_value = -abs(target_value) if negative_discount_convention else abs(target_value)
        if not close(engine.value(expected["sheet"], expected["cell"]), target_value):
            return False
    for expected in case.get("expected_by_content", []):
        expected_row = find_content_row(workbook, expected)
        if expected_row is None or not close(
            engine.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'), expected["value"]
        ):
            return False
    for protected in case.get("protected", []):
        if not close(engine.value(protected["sheet"], protected["cell"]), protected["value"]):
            return False
    for protected in case.get("protected_by_content", []):
        protected_row = find_content_row(workbook, protected)
        if protected_row is None or not close(
            engine.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'), protected["value"]
        ):
            return False
    return True


def perturbation_response_ok(workbook, split="dev"):
    return all(perturbation_case_ok(workbook, case) for case in load_perturbations(split))


def source_field_matches(value, filename):
    rendered, expected = norm(value), norm(filename)
    return bool(rendered and expected and expected in rendered)


def source_field_contradicts(value, filename):
    rendered = text(value)
    if not rendered or source_field_matches(rendered, filename):
        return False
    return re.search(r"\.(?:pdf|png|jpe?g|tiff?|xlsx?|csv)\b", rendered, re.IGNORECASE) is not None


def locator_ok(value, filename=None, page=None):
    # File and page have dedicated provenance columns.  The locator only needs
    # to remain a readable line-level label; repeating either field is optional.
    return len(re.sub(r"\s+", "", text(value))) >= 3


def present_formula(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def formula_value(engine, sheet, cell):
    try:
        return engine.value(sheet, cell)
    except UnsupportedFormulaError:
        raise
    except Exception:
        return None


def resolved_cell(workbook, engine, sheet, cell):
    value = formula_value(engine, sheet, cell)
    return workbook[sheet][cell].value if value is None else value


def row_values(workbook, sheet, start, end, cols):
    records = []
    for row in range(start, end + 1):
        values = tuple(workbook[sheet].cell(row=row, column=col).value for col in cols)
        if any(value not in (None, "") for value in values): records.append(values)
    return records


def _legacy_task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    header_cells = {"vendor": "B4", "invoice_id": "B5", "invoice_date": "B6", "customer": "B7", "currency": "B8", "po_reference": "B9"}
    headers_ok = all(
        semantic_date(workbook["Invoice_Header"][cell].value) == semantic_date(oracle["headers"][name])
        if name == "invoice_date"
        else norm(workbook["Invoice_Header"][cell].value) == norm(oracle["headers"][name])
        for name, cell in header_cells.items()
    )
    checks["R001"] = 1.0
    checks["R002"] = 1.0 if headers_ok else 0.0
    source = workbook["Line_Items"]
    actual = []
    for row in range(4, source.max_row + 1):
        if text(source[f"B{row}"].value):
            actual.append({
                "row": row, "id": text(source[f"A{row}"].value), "description": text(source[f"B{row}"].value),
                "quantity": source[f"C{row}"].value, "unit_price": source[f"D{row}"].value,
                "line_total": formula_value(engine, "Line_Items", f"E{row}"), "page": source[f"F{row}"].value,
            })
    expected = {norm(item["description"]): item for item in oracle["items"]}
    actual_by_description = {norm(item["description"]): item for item in actual}
    facts_ok = (
        len(actual) == len(expected) == len(actual_by_description)
        and set(actual_by_description) == set(expected)
        and all(
            isinstance(actual_by_description[name]["quantity"], (int, float))
            and isinstance(actual_by_description[name]["unit_price"], (int, float))
            and close(actual_by_description[name]["quantity"], item["quantity"])
            and close(actual_by_description[name]["unit_price"], item["unit_price"])
            and actual_by_description[name]["page"] == item["page"]
            for name, item in expected.items()
        )
    )
    line_formula_ok = facts_ok and all(
        present_formula(workbook, f'Line_Items!E{item["row"]}')
        and close(item["line_total"], item["quantity"] * item["unit_price"])
        for item in actual
    )
    checks["R003"] = 1.0 if facts_ok else 0.0
    checks["R004"] = 1.0 if line_formula_ok else 0.0
    summary = {cell: engine.value("Invoice_Header", cell) for cell in ("B12", "B13", "B14", "B15", "B17", "B18", "B19")}
    summary_values_ok = (
        close(summary["B12"], oracle["subtotal"])
        and close(abs(summary["B14"]) if isinstance(summary["B14"], (int, float)) else None, oracle["discount"])
        and close(summary["B15"], oracle["taxable"])
        and close(summary["B17"], oracle["tax"])
        and close(summary["B18"], oracle["freight"])
        and close(summary["B19"], oracle["total"])
    )
    discount_semantics_ok = (
        isinstance(summary["B13"], (int, float))
        and close(abs(summary["B14"]) if isinstance(summary["B14"], (int, float)) else None, round(summary["B12"] * summary["B13"], 2))
        and close(summary["B15"], summary["B12"] - abs(summary["B14"]))
    )
    subtotal_formula_ok = all(present_formula(workbook, address) for address in (
        "Invoice_Header!B12", "Invoice_Header!B14", "Invoice_Header!B15"
    )) and discount_semantics_ok
    payable_formula_ok = all(present_formula(workbook, address) for address in (
        "Invoice_Header!B17", "Invoice_Header!B19"
    ))
    checks["R005"] = 1.0 if summary_values_ok and subtotal_formula_ok else 0.0
    checks["R006"] = 1.0 if summary_values_ok and payable_formula_ok else 0.0
    try:
        dynamic_ok = perturbation_response_ok(workbook, split)
    except Exception as exc:
        dynamic_ok = False; failures.append(f"INVOICE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if dynamic_ok else 0.0
    provenance_rows = []
    provenance_sheet = workbook["Provenance"]
    for row in range(4, provenance_sheet.max_row + 1):
        line_id = text(resolved_cell(workbook, engine, "Provenance", f"A{row}"))
        if line_id:
            provenance_rows.append((
                line_id,
                text(resolved_cell(workbook, engine, "Provenance", f"B{row}")),
                resolved_cell(workbook, engine, "Provenance", f"C{row}"),
                text(resolved_cell(workbook, engine, "Provenance", f"D{row}")),
            ))
    provenance_by_id = {}
    for record in provenance_rows:
        provenance_by_id.setdefault(record[0], []).append(record)
    expected_provenance_ids = {item["id"] for item in actual}
    provenance_population_ok = (
        len(provenance_rows) == len(actual)
        and set(provenance_by_id) == expected_provenance_ids
        and all(len(records) == 1 for records in provenance_by_id.values())
    )
    provenance_ok = facts_ok and provenance_population_ok and all(
        len(provenance_by_id.get(item["id"], [])) == 1
        and source_field_matches(provenance_by_id[item["id"]][0][1], oracle["document"]["filename"])
        and provenance_by_id[item["id"]][0][2] == expected[norm(item["description"])]["page"]
        and locator_ok(provenance_by_id[item["id"]][0][3], oracle["document"]["filename"], expected[norm(item["description"])]["page"])
        for item in actual
    )
    identity_ok = (
        facts_ok and all(item["id"] for item in actual)
        and len({item["id"] for item in actual}) == len(expected)
        and provenance_population_ok
    )
    checks["R008"] = 1.0 if provenance_ok else 0.0
    checks["R009"] = 1.0 if identity_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = any(
        (
            source_field_contradicts(record[1], oracle["document"]["filename"])
        )
        or (
            record[2] not in (None, "")
            and record[2] not in valid_pages
        )
        for record in provenance_rows
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_SOURCE_FILE_OR_PAGE_PROVENANCE")
    return checks, failures


def semantic_header(sheet, aliases, required):
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for row in range(1, sheet.max_row + 1):
        columns = {}
        for column in range(1, sheet.max_column + 1):
            value = norm(sheet.cell(row=row, column=column).value)
            for role, accepted in normalized.items():
                if value in accepted and role not in columns:
                    columns[role] = column
        if set(required) <= set(columns):
            return row, columns
    return None, {}


def labeled_value_cells(workbook, aliases):
    """Find visible label/value pairs, preferring populated values over blanks."""
    matches = {role: [] for role in aliases}
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for sheet in workbook.worksheets:
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column):
                label = norm(sheet.cell(row=row, column=column).value).rstrip(":")
                for role, accepted in normalized.items():
                    if role == "discount" and "rate" in label:
                        continue
                    if any(label == value.rstrip(":") or label.startswith(value.rstrip(":") + " ") for value in accepted):
                        matches[role].append(sheet.cell(row=row, column=column + 1))
    return {
        role: sorted(cells, key=lambda cell: cell.value in (None, ""))[0]
        for role, cells in matches.items()
        if cells
    }


def best_semantic_table(workbook, aliases, required, score_values, score_role):
    best = None
    for sheet in workbook.worksheets:
        header, columns = semantic_header(sheet, aliases, required)
        if header is None:
            continue
        observed = {
            norm(sheet.cell(row=row, column=columns[score_role]).value)
            for row in range(header + 1, sheet.max_row + 1)
            if text(sheet.cell(row=row, column=columns[score_role]).value)
        }
        score = len(observed & {norm(value) for value in score_values})
        candidate = (score, sheet, header, columns)
        if best is None or score > best[0]:
            best = candidate
    return best


def semantic_invoice_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    header_cells = labeled_value_cells(workbook, {
        "vendor": ("Vendor",),
        "invoice_id": ("Invoice ID", "Invoice number"),
        "invoice_date": ("Invoice date",),
        "customer": ("Customer", "Bill to"),
        "currency": ("Currency",),
        "po_reference": ("PO reference", "Purchase order"),
        "subtotal": ("Line-item subtotal", "Subtotal"),
        "discount": ("Discount amount", "Discount"),
        "taxable": ("Taxable amount",),
        "tax": ("Tax amount", "Sales tax"),
        "freight": ("Freight", "Shipping"),
        "total": ("Total payable", "Amount payable"),
    })
    line_table = best_semantic_table(workbook, {
        "id": ("Line ID", "ID"),
        "description": ("Description", "Item", "Line description"),
        "quantity": ("Quantity", "Qty"),
        "unit_price": ("Unit price", "Price", "Rate"),
        "total": ("Line total", "Extended total", "Amount"),
        "page": ("Source page", "Page"),
    }, ("id", "description", "quantity", "unit_price", "total", "page"), [item["description"] for item in oracle["items"]], "description")
    if line_table is None:
        return checks, failures
    _, line_sheet, line_header, line_columns = line_table
    actual = []
    for row in range(line_header + 1, line_sheet.max_row + 1):
        description = text(line_sheet.cell(row=row, column=line_columns["description"]).value)
        if not description:
            continue
        total_cell = line_sheet.cell(row=row, column=line_columns["total"])
        actual.append({
            "row": row,
            "id": text(line_sheet.cell(row=row, column=line_columns["id"]).value),
            "description": description,
            "quantity": line_sheet.cell(row=row, column=line_columns["quantity"]).value,
            "unit_price": line_sheet.cell(row=row, column=line_columns["unit_price"]).value,
            "line_total": formula_value(engine, line_sheet.title, total_cell.coordinate),
            "total_cell": total_cell,
            "page": line_sheet.cell(row=row, column=line_columns["page"]).value,
        })
    expected = {norm(item["description"]): item for item in oracle["items"]}
    actual_by_description = {norm(item["description"]): item for item in actual}
    facts_ok = (
        len(actual) == len(expected) == len(actual_by_description)
        and set(actual_by_description) == set(expected)
        and all(
            close(actual_by_description[name]["quantity"], item["quantity"])
            and close(actual_by_description[name]["unit_price"], item["unit_price"])
            and actual_by_description[name]["page"] == item["page"]
            for name, item in expected.items()
        )
    )
    line_formula_ok = facts_ok and all(
        isinstance(item["total_cell"].value, str)
        and item["total_cell"].value.startswith("=")
        and close(item["line_total"], item["quantity"] * item["unit_price"])
        for item in actual
    )

    required_headers = {"vendor", "invoice_id", "invoice_date", "customer", "currency", "po_reference"}
    headers_ok = set(header_cells) >= required_headers and all(
        semantic_date(header_cells[name].value) == semantic_date(oracle["headers"][name])
        if name == "invoice_date"
        else norm(header_cells[name].value) == norm(oracle["headers"][name])
        for name in required_headers
    )
    summary_roles = {"subtotal", "discount", "taxable", "tax", "freight", "total"}
    summary_values = {
        name: formula_value(engine, cell.parent.title, cell.coordinate)
        for name, cell in header_cells.items()
        if name in summary_roles
    }
    discount_value = summary_values.get("discount")
    summary_values_ok = set(summary_values) == summary_roles and (
        close(summary_values["subtotal"], oracle["subtotal"])
        and isinstance(discount_value, (int, float))
        and close(abs(discount_value), oracle["discount"])
        and close(summary_values["taxable"], oracle["taxable"])
        and close(summary_values["tax"], oracle["tax"])
        and close(summary_values["freight"], oracle["freight"])
        and close(summary_values["total"], oracle["total"])
    )
    subtotal_formulas_ok = all(
        isinstance(header_cells[name].value, str) and header_cells[name].value.startswith("=")
        for name in ("subtotal", "discount", "taxable")
    ) if set(header_cells) >= {"subtotal", "discount", "taxable"} else False
    payable_formulas_ok = all(
        isinstance(header_cells[name].value, str) and header_cells[name].value.startswith("=")
        for name in ("tax", "total")
    ) if set(header_cells) >= {"tax", "total"} else False

    provenance_table = best_semantic_table(workbook, {
        "id": ("Line ID", "ID"),
        "source": ("Source document", "Source file", "Document"),
        "page": ("Page", "Source page"),
        "locator": ("Text locator", "Line label", "Locator", "Source locator"),
    }, ("id", "source", "page", "locator"), [item["id"] for item in actual], "id")
    provenance_rows = []
    if provenance_table is not None:
        _, provenance_sheet, provenance_header, provenance_columns = provenance_table
        for row in range(provenance_header + 1, provenance_sheet.max_row + 1):
            line_id = text(provenance_sheet.cell(row=row, column=provenance_columns["id"]).value)
            if line_id:
                provenance_rows.append((
                    line_id,
                    text(provenance_sheet.cell(row=row, column=provenance_columns["source"]).value),
                    provenance_sheet.cell(row=row, column=provenance_columns["page"]).value,
                    text(provenance_sheet.cell(row=row, column=provenance_columns["locator"]).value),
                ))
    provenance_by_id = {}
    for record in provenance_rows:
        provenance_by_id.setdefault(record[0], []).append(record)
    actual_ids = {item["id"] for item in actual}
    population_ok = len(provenance_rows) == len(actual) and set(provenance_by_id) == actual_ids and all(len(rows) == 1 for rows in provenance_by_id.values())
    provenance_ok = facts_ok and population_ok and all(
        source_field_matches(provenance_by_id[item["id"]][0][1], oracle["document"]["filename"])
        and provenance_by_id[item["id"]][0][2] == expected[norm(item["description"])]["page"]
        and locator_ok(provenance_by_id[item["id"]][0][3], oracle["document"]["filename"], expected[norm(item["description"])]["page"])
        for item in actual
    )
    identity_ok = facts_ok and all(actual_ids) and len(actual_ids) == len(expected) and population_ok

    dynamic_ok = line_formula_ok and summary_values_ok and subtotal_formulas_ok and payable_formulas_ok
    if dynamic_ok:
        output_roles = {"B12": "subtotal", "B14": "discount", "B15": "taxable", "B17": "tax", "B18": "freight", "B19": "total"}
        for case in load_perturbations(split):
            matches = [item for item in actual if norm(item["description"]) == norm(case["target"]["match_value"])]
            if len(matches) != 1:
                dynamic_ok = False
                break
            item = matches[0]
            role = "unit_price" if case["target"]["value_column"] == "D" else "quantity"
            target_cell = line_sheet.cell(row=item["row"], column=line_columns[role])
            probe = FormulaEngine(workbook, {key(line_sheet.title, target_cell.coordinate): case["target"]["perturbed"]})
            if not close(probe.value(line_sheet.title, item["total_cell"].coordinate), (
                case["target"]["perturbed"] * (item["quantity"] if role == "unit_price" else item["unit_price"])
            )):
                dynamic_ok = False
                break
            for expected_output in case.get("expected", []):
                semantic_role = output_roles.get(expected_output["cell"])
                cell = header_cells.get(semantic_role)
                expected_value = expected_output["value"]
                if semantic_role == "discount":
                    actual_value = probe.value(cell.parent.title, cell.coordinate) if cell else None
                    if actual_value is None or not close(abs(actual_value), expected_value):
                        dynamic_ok = False
                        break
                elif cell is None or not close(probe.value(cell.parent.title, cell.coordinate), expected_value):
                    dynamic_ok = False
                    break
            if not dynamic_ok:
                break

    checks["R001"] = 1.0 if headers_ok and line_table is not None and provenance_table is not None else 0.0
    checks["R002"] = 1.0 if headers_ok else 0.0
    checks["R003"] = 1.0 if facts_ok else 0.0
    checks["R004"] = 1.0 if line_formula_ok else 0.0
    checks["R005"] = 1.0 if summary_values_ok and subtotal_formulas_ok else 0.0
    checks["R006"] = 1.0 if summary_values_ok and payable_formulas_ok else 0.0
    checks["R007"] = 1.0 if dynamic_ok else 0.0
    checks["R008"] = 1.0 if provenance_ok else 0.0
    checks["R009"] = 1.0 if identity_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = any(
        source_field_contradicts(record[1], oracle["document"]["filename"])
        or (record[2] not in (None, "") and record[2] not in valid_pages)
        for record in provenance_rows
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_SOURCE_FILE_OR_PAGE_PROVENANCE")
    return checks, failures


def task_checks(workbook, engine, oracle, split="dev"):
    semantic_checks, semantic_failures = semantic_invoice_checks(workbook, engine, oracle, split)
    if all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"]):
        try:
            legacy_checks, legacy_failures = _legacy_task_checks(workbook, engine, oracle, split)
            semantic_checks = {
                criterion["id"]: max(legacy_checks[criterion["id"]], semantic_checks[criterion["id"]])
                for criterion in TASK["criteria"]
            }
            semantic_failures.extend(legacy_failures)
        except Exception:
            pass
    return semantic_checks, semantic_failures


def score(candidate, split=None):
    split = active_split(split)
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=["OUTPUT_MISSING"],
        )
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=[f"MALFORMED_XLSX:{type(exc).__name__}"],
        )

    try:
        oracle = load_oracle(split)
        engine = FormulaEngine(workbook)
        checks, task_failures = task_checks(workbook, engine, oracle, split)
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=checks,
        failures=failures,
    )


if __name__ == "__main__":
    arguments = sys.argv[1:]
    selected_split = None
    if "--split" in arguments:
        split_index = arguments.index("--split")
        if split_index + 1 >= len(arguments):
            raise SystemExit("--split requires dev or confirm")
        selected_split = arguments[split_index + 1]
        del arguments[split_index:split_index + 2]
    if len(arguments) > 1:
        raise SystemExit("usage: evaluate.py [candidate.xlsx] [--split dev|confirm]")
    target = Path(arguments[0]) if arguments else Path("/app/output/answer.xlsx")
    print(json.dumps(score(target, selected_split), sort_keys=True))

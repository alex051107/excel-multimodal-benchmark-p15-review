#!/usr/bin/env python3
"""Focused Judge V3 regression: disclosed dynamic logic, layout independence."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


TASK_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TASK_DIR / "tests"))
SPEC = importlib.util.spec_from_file_location("invoice_evaluator", TASK_DIR / "tests" / "evaluate.py")
EVALUATOR = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(EVALUATOR)


def build_alternate_layout(path: Path) -> None:
    oracle = EVALUATOR.load_oracle("dev")
    workbook = Workbook()
    workpaper = workbook.active
    workpaper.title = "AP Workpaper"
    labels = [
        ("Vendor", oracle["headers"]["vendor"]),
        ("Invoice ID", oracle["headers"]["invoice_id"]),
        ("Invoice date", oracle["headers"]["invoice_date"]),
        ("Customer", oracle["headers"]["customer"]),
        ("Currency", oracle["headers"]["currency"]),
        ("PO reference", oracle["headers"]["po_reference"]),
    ]
    for row, (label, value) in enumerate(labels, start=2):
        workpaper.cell(row, 1, label)
        workpaper.cell(row, 2, value)
    workpaper.append([])
    for column, value in enumerate(("Line ID", "Description", "Quantity", "Unit price", "Line total", "Source page"), start=1):
        workpaper.cell(10, column, value)
    for row, item in enumerate(oracle["items"], start=11):
        workpaper.cell(row, 1, item["line_id"])
        workpaper.cell(row, 2, item["description"])
        workpaper.cell(row, 3, item["quantity"])
        workpaper.cell(row, 4, item["unit_price"])
        workpaper.cell(row, 5, f"=C{row}*D{row}")
        workpaper.cell(row, 6, item["page"])
    summaries = (
        (16, "Line-item subtotal", "=SUM(E11:E14)"),
        (17, "Discount amount", "=-E16*0.10"),
        (18, "Taxable amount", "=E16+E17"),
        (19, "Tax amount", "=E18*0.0825"),
        (20, "Freight", 45.0),
        (21, "Total payable", "=E18+E19+E20"),
    )
    for row, label, value in summaries:
        workpaper.cell(row, 4, label)
        workpaper.cell(row, 5, value)

    sources = workbook.create_sheet("Source Register")
    for column, value in enumerate(("Line ID", "Source document", "Page", "Text locator"), start=1):
        sources.cell(2, column, value)
    for row, item in enumerate(oracle["items"], start=3):
        sources.cell(row, 1, item["line_id"])
        sources.cell(row, 2, oracle["document"]["filename"])
        sources.cell(row, 3, item["page"])
        sources.cell(row, 4, item["text_locator"])
    workbook.save(path)


def result_record(path: Path, expected: str, predicate, display_path: str | None = None) -> dict:
    result = EVALUATOR.score(path, "dev")
    return {
        "path": display_path or str(path.relative_to(TASK_DIR)),
        "score": result["normalized_score"],
        "pass": result["pass"],
        "criterion_scores": result["criterion_scores"],
        "failure_codes": result["failure_codes"],
        "expected": expected,
        "expected_pass": expected.startswith(("full pass", "pass with")),
        "expectation_met": bool(predicate(result)),
    }


def main() -> None:
    fixtures = {
        "reference": (TASK_DIR / "solution/reference.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "equivalent": (TASK_DIR / "fixtures/equivalent/candidate_equivalent.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "noop": (TASK_DIR / "fixtures/noop/candidate_noop.xlsx", "fail", lambda r: not r["pass"]),
        "malformed": (TASK_DIR / "fixtures/malformed/candidate_malformed.xlsx", "fail with zero score", lambda r: not r["pass"] and r["normalized_score"] == 0.0),
    }
    for path in sorted((TASK_DIR / "fixtures/mutants").glob("*.xlsx")):
        fixtures[f"mutant:{path.stem}"] = (path, "fail", lambda r: not r["pass"])
    fixture_results = {
        name: result_record(path, expected, predicate)
        for name, (path, expected, predicate) in fixtures.items()
    }

    with tempfile.TemporaryDirectory(prefix="invoice-judge-v3-") as directory:
        alternate = Path(directory) / "alternate_layout.xlsx"
        build_alternate_layout(alternate)
        alternate_result = result_record(
            alternate,
            "full pass: arbitrary sheet names and positions with the disclosed formula chain",
            lambda r: r["pass"] and r["normalized_score"] == 1.0,
            "generated:alternate_layout",
        )
        broken = Path(directory) / "broken_total.xlsx"
        workbook = load_workbook(alternate)
        workbook["AP Workpaper"]["E21"] = "=E18+E19"
        workbook.save(broken)
        broken_result = result_record(
            broken,
            "fail R006/R007: payable formula omits freight",
            lambda r: not r["pass"] and r["criterion_scores"]["R006"] == 0.0 and r["criterion_scores"]["R007"] == 0.0,
            "generated:broken_total",
        )

    new_cases = {"alternate_layout": alternate_result, "broken_total": broken_result}
    all_met = all(item["expectation_met"] for item in (*fixture_results.values(), *new_cases.values()))
    receipt = {
        "task_id": "P15-C-INVOICE-001",
        "judge_version": "p15_v3.0",
        "validation_scope": "reference, equivalent, noop, malformed, all mutants, alternate layout, genuine payable error",
        "expectation_basis": {
            "alternate_layout": "The instruction requires invoice facts, traceability, and a dynamic amount chain, but does not prescribe sheet names or cell addresses.",
            "broken_total": "The instruction explicitly requires a dynamic payable chain including freight.",
        },
        "fixtures": fixture_results,
        "new_cases": new_cases,
        "all_expectations_met": all_met,
    }
    receipt_path = TASK_DIR / "receipts/judge_v3_local_validation.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"task_id": receipt["task_id"], "all_expectations_met": all_met, "scores": {key: [value["score"], value["pass"]] for key, value in {**fixture_results, **new_cases}.items()}}, ensure_ascii=False))
    if not all_met:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

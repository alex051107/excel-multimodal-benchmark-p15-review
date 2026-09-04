#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-C-STATEMENT-001."""
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-C-STATEMENT-001",
  "pass_threshold": 0.7,
  "criteria": [
    {
      "id": "R001",
      "description": "Workbook contains the complete statement-to-reconciliation delivery structure.",
      "weight": 2,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R002",
      "description": "Statement header captures the correct account, period, currency, and opening balance.",
      "weight": 4,
      "type": "positive",
      "dimension": "header_accuracy",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R003",
      "description": "All transactions preserve typed debit/credit direction and category semantics.",
      "weight": 6,
      "type": "positive",
      "dimension": "transaction_accuracy",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R004",
      "description": "Opening-to-closing reconciliation is formula-linked and responds to a credit perturbation.",
      "weight": 5,
      "type": "positive",
      "dimension": "native_workbook_semantics",
      "method": "deterministic",
      "method_params": {
        "perturbations": 2
      }
    },
    {
      "id": "R005",
      "description": "Category totals agree with the signed transaction table.",
      "weight": 3,
      "type": "positive",
      "dimension": "categorization",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R006",
      "description": "Every transaction has the correct source page and locator across the continuation break.",
      "weight": 4,
      "type": "positive",
      "dimension": "provenance",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "P001",
      "description": "Penalty for fabricated statement provenance.",
      "weight": -5,
      "type": "penalty",
      "dimension": "integrity",
      "method": "deterministic",
      "method_params": {}
    }
  ],
  "required_sheets": [
    "Statement_Header",
    "Transactions",
    "Categories",
    "Provenance",
    "Reconciliation",
    "Checks"
  ]
}
TESTS_ROOT = Path(__file__).resolve().parent
REPO_TASK_ROOT = TESTS_ROOT.parent
DEV_ROOT = REPO_TASK_ROOT if (REPO_TASK_ROOT / "rubric.json").is_file() else TESTS_ROOT / "dev"
TASK = json.loads((DEV_ROOT / "rubric.json").read_text(encoding="utf-8"))
TASK["required_sheets"] = json.loads(
    (TESTS_ROOT / "private_contract.json").read_text(encoding="utf-8")
)["required_sheets"]
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")
COLUMN_RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+):\$?([A-Z]+)")


TASK["hurdle_criteria"] = ["R002", "R003", "R006", "R009"]
SHEET_ALIASES = {}


def semantic_sheet_aliases(workbook):
    signatures = {
        "Statement_Header": (("Bank",), ("Account ending", "Account suffix", "Account"), ("Period start", "Start date"), ("Period end", "End date"), ("Opening balance",), ("Currency",)),
        "Transactions": (("Transaction ID", "ID"), ("Date", "Transaction date"), ("Description", "Transaction", "Memo"), ("Debit", "Debits"), ("Credit", "Credits"), ("Page", "Source page"), ("Category", "Type")),
        "Categories": (("Category", "Type"), ("Net impact", "Amount", "Value", "Total"), ("Evidence", "Basis")),
        "Provenance": (("Transaction ID", "ID"), ("Source document", "Source file", "Document"), ("Page", "Source page"), ("Text locator", "Locator", "Source locator")),
        "Reconciliation": (("Metric", "Check"), ("Value", "Observed", "Amount"), ("Opening balance",), ("Calculated closing balance", "Calculated balance"), ("Statement closing balance", "Printed closing balance", "Closing balance")),
        "Checks": (("Check",), ("Observed",), ("Expected",), ("Interpretation",)),
    }
    aliases = {}
    for role, groups in signatures.items():
        candidates = []
        for sheet in workbook.worksheets:
            values = {
                norm(cell.value).rstrip(":")
                for row in sheet.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            }
            if all(any(norm(alias).rstrip(":") in values for alias in group) for group in groups):
                candidates.append(sheet.title)
        if candidates:
            aliases[role] = tuple(candidates)
    return aliases


def key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


def active_split(explicit=None):
    split = (explicit or os.environ.get("P15_EVAL_SPLIT", "dev")).strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"UNSUPPORTED_EVAL_SPLIT:{split}")
    return split


def load_oracle(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "oracle_recompute.py" if split == "confirm" else DEV_ROOT / "metadata" / "oracle_recompute.py"
    spec = importlib.util.spec_from_file_location(f"task_oracle_{split}", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


class UnsupportedFormulaError(ValueError):
    """The workbook formula is valid Excel syntax outside this bounded replay."""


def excel_equal(left, right):
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, str) and isinstance(right, str):
        return norm(left) == norm(right)
    return left == right


def excel_criteria_match(value, criterion):
    operator = "="
    expected = criterion
    if isinstance(criterion, str):
        match = re.match(r"^(<=|>=|<>|!=|=|<|>)(.*)$", criterion)
        if match:
            operator, expected = match.groups()
        stripped = expected.strip()
        try:
            expected = float(stripped)
        except ValueError:
            expected = stripped
    if operator == "=":
        return excel_equal(value, expected)
    if operator in {"<>", "!="}:
        return not excel_equal(value, expected)
    try:
        if operator == "<": return value < expected
        if operator == "<=": return value <= expected
        if operator == ">": return value > expected
        if operator == ">=": return value >= expected
    except TypeError:
        return False
    raise UnsupportedFormulaError(f"UNSUPPORTED_CRITERIA_OPERATOR:{operator}")


def excel_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        resolved_sheet = self.workbook.actual_sheet_name(sheet) if hasattr(self.workbook, "actual_sheet_name") else sheet
        current = key(resolved_sheet, cell)
        if current in self.overrides: return self.overrides[current]
        if current in self.memo: return self.memo[current]
        if current in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{current}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(current)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], resolved_sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[current] = result
            return result
        finally:
            self.stack.discard(current)

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(col)}{row}") for row in range(int(r1), int(r2) + 1) for col in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def column_range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), 1, match.group(4), self.workbook[sheet].max_row))
        expression = outside(COLUMN_RANGE, column_range_replace, expression)
        def reference_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, reference_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            # Excel treats a genuinely blank referenced cell as zero in numeric
            # arithmetic (for example, Credit-Debit with one side blank).
            left = 0 if left is None else left
            right = 0 if right is None else right
            if isinstance(node.op, ast.Add): return left + right
            if isinstance(node.op, ast.Sub): return left - right
            if isinstance(node.op, ast.Mult): return left * right
            if isinstance(node.op, ast.Div): return left / right
            if isinstance(node.op, ast.Pow): return left ** right
            raise ValueError("UNSUPPORTED_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_ARGUMENTS")
                if self.safe_eval(node.args[0]):
                    return self.safe_eval(node.args[1])
                return self.safe_eval(node.args[2]) if len(node.args) == 3 else False
            args = [self.safe_eval(item) for item in node.args]
            values = [item for group in args for item in (group if isinstance(group, list) else [group]) if item is not None]
            if name == "SUM":
                return sum(
                    value for value in values
                    if isinstance(value, (int, float)) and not isinstance(value, bool)
                )
            if name == "SUMIF":
                criteria_range, criterion = args[0], args[1]
                sum_range = args[2] if len(args) > 2 else criteria_range
                if not isinstance(criteria_range, list) or not isinstance(sum_range, list) or len(criteria_range) != len(sum_range):
                    raise ValueError("INVALID_SUMIF_RANGES")
                return sum(
                    value if isinstance(value, (int, float)) and not isinstance(value, bool) else 0
                    for label, value in zip(criteria_range, sum_range)
                    if excel_criteria_match(label, criterion)
                )
            if name == "AVERAGE":
                numeric = [value for value in values if isinstance(value, (int, float)) and not isinstance(value, bool)]
                if not numeric: raise ValueError("AVERAGE_NO_NUMERIC_VALUES")
                return statistics.mean(numeric)
            if name == "COUNTA": return sum(value not in (None, "") for value in values)
            if name == "ISBLANK": return args[0] is None
            if name == "ISNUMBER": return isinstance(args[0], (int, float)) and not isinstance(args[0], bool)
            if name == "CONCATENATE": return "".join(excel_text(value) for value in args)
            if name == "MIN":
                numeric = [value for value in values if isinstance(value, (int, float)) and not isinstance(value, bool)]
                if not numeric: raise ValueError("MIN_NO_NUMERIC_VALUES")
                return min(numeric)
            if name == "MAX":
                numeric = [value for value in values if isinstance(value, (int, float)) and not isinstance(value, bool)]
                if not numeric: raise ValueError("MAX_NO_NUMERIC_VALUES")
                return max(numeric)
            if name == "AND": return all(args)
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(float(actual) - float(expected)) <= tolerance


def text(value): return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", text(value).lower())


def date_key(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", text(value))
    return match.group(1) if match else text(value)


def amount(value):
    return 0.0 if value in (None, "") else value


def load_perturbations(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "perturbations.json" if split == "confirm" else DEV_ROOT / "metadata" / "perturbations.json"
    records = json.loads(path.read_text(encoding="utf-8"))
    expected_count = 1 if split == "confirm" else 2
    if len(records) != expected_count or len({record.get("id") for record in records}) != expected_count:
        raise ValueError(f"{split.upper()}_PERTURBATION_MANIFEST_MUST_CONTAIN_{expected_count}_UNIQUE_CASES")
    return records


def find_content_row(workbook, spec):
    sheet = workbook[spec["sheet"]]
    primary_col = col_index(spec["match_column"])
    secondary_col = col_index(spec["secondary_match_column"]) if spec.get("secondary_match_column") else None
    rows = []
    for row in range(4, sheet.max_row + 1):
        if norm(sheet.cell(row=row, column=primary_col).value) != norm(spec["match_value"]):
            continue
        if secondary_col and norm(sheet.cell(row=row, column=secondary_col).value) != norm(spec["secondary_match_value"]):
            continue
        rows.append(row)
    return rows[0] if len(rows) == 1 else None


def perturbation_case_ok(workbook, case):
    target = case["target"]
    row = find_content_row(workbook, target)
    if row is None:
        return False
    target_cell = f'{target["value_column"]}{row}'
    baseline_engine = FormulaEngine(workbook)
    if not close(baseline_engine.value(target["sheet"], target_cell), target["baseline"]):
        return False
    engine = FormulaEngine(workbook, {key(target["sheet"], target_cell): target["perturbed"]})
    for expected in case.get("expected", []):
        if not close(engine.value(expected["sheet"], expected["cell"]), expected["value"]):
            return False
    category_specs = [
        spec for spec in case.get("expected_by_content", []) + case.get("protected_by_content", [])
        if spec.get("sheet") == "Categories"
    ]
    for expected in case.get("expected_by_content", []):
        if expected.get("sheet") == "Categories":
            continue
        expected_row = find_content_row(workbook, expected)
        if expected_row is None or not close(
            engine.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'), expected["value"]
        ):
            return False
    for protected in case.get("protected", []):
        if not close(engine.value(protected["sheet"], protected["cell"]), protected["value"]):
            return False
    for protected in case.get("protected_by_content", []):
        if protected.get("sheet") == "Categories":
            continue
        protected_row = find_content_row(workbook, protected)
        if protected_row is None or not close(
            engine.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'), protected["value"]
        ):
            return False
    if category_specs:
        category_rows = [
            category_row for category_row in range(4, workbook["Categories"].max_row + 1)
            if norm(workbook["Categories"][f"A{category_row}"].value)
        ]
        delta = target["perturbed"] - target["baseline"]
        signed_delta = -delta if target["value_column"] == "D" else delta
        changes = {
            category_row: engine.value("Categories", f"B{category_row}")
            - baseline_engine.value("Categories", f"B{category_row}")
            for category_row in category_rows
        }
        changed_rows = [category_row for category_row, change in changes.items() if not close(change, 0.0)]
        if len(changed_rows) != 1 or not close(changes[changed_rows[0]], signed_delta):
            return False
        if not present_formula(workbook, f"Categories!B{changed_rows[0]}"):
            return False
    return True


def perturbation_response_ok(workbook, split="dev"):
    return all(perturbation_case_ok(workbook, case) for case in load_perturbations(split))


def source_field_matches(value, filename):
    rendered, expected = norm(value), norm(filename)
    return bool(rendered and expected and expected in rendered)


def source_field_contradicts(value, filename):
    rendered = text(value)
    if not rendered or source_field_matches(rendered, filename):
        return False
    return re.search(r"\.(?:pdf|png|jpe?g|tiff?|xlsx?|csv)\b", rendered, re.IGNORECASE) is not None


def locator_ok(value, filename=None, page=None):
    # File and page are already explicit columns.  The final field is a
    # human-readable transaction locator and need not repeat either value.
    return len(re.sub(r"\s+", "", text(value))) >= 3


def present_formula(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def formula_value(engine, sheet, cell):
    try:
        return engine.value(sheet, cell)
    except UnsupportedFormulaError:
        raise
    except Exception:
        return None


def row_values(workbook, sheet, start, end, cols):
    records = []
    for row in range(start, end + 1):
        values = tuple(workbook[sheet].cell(row=row, column=col).value for col in cols)
        if any(value not in (None, "") for value in values): records.append(values)
    return records


def _legacy_task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    checks["R001"] = 1.0
    header_cells = {"bank": "B4", "account_suffix": "B5", "currency": "B9"}
    headers_ok = (
        all(text(workbook["Statement_Header"][cell].value) == text(oracle["headers"][name]) for name, cell in header_cells.items())
        and date_key(workbook["Statement_Header"]["B6"].value) == date_key(oracle["headers"]["period_start"])
        and date_key(workbook["Statement_Header"]["B7"].value) == date_key(oracle["headers"]["period_end"])
        and close(workbook["Statement_Header"]["B8"].value, oracle["headers"]["opening_balance"])
        and close(workbook["Reconciliation"]["B8"].value, oracle["closing"])
    )
    checks["R002"] = 1.0 if headers_ok else 0.0
    transactions_sheet = workbook["Transactions"]
    actual = []
    for row in range(4, transactions_sheet.max_row + 1):
        if text(transactions_sheet[f"C{row}"].value):
            actual.append({
                "row": row, "id": text(transactions_sheet[f"A{row}"].value),
                "date": date_key(transactions_sheet[f"B{row}"].value), "description": text(transactions_sheet[f"C{row}"].value),
                "debit": amount(transactions_sheet[f"D{row}"].value), "credit": amount(transactions_sheet[f"E{row}"].value),
                "net": formula_value(engine, "Transactions", f"F{row}"), "page": transactions_sheet[f"G{row}"].value,
                "category": text(transactions_sheet[f"H{row}"].value),
            })
    expected = {(date_key(item["date"]), norm(item["description"])): item for item in oracle["transactions"]}
    actual_by_key = {(item["date"], norm(item["description"])): item for item in actual}
    tx_ok = (
        len(actual) == len(expected) == len(actual_by_key)
        and set(actual_by_key) == set(expected)
        and all(
            isinstance(actual_by_key[key_]["debit"], (int, float))
            and isinstance(actual_by_key[key_]["credit"], (int, float))
            and close(actual_by_key[key_]["debit"], item["debit"])
            and close(actual_by_key[key_]["credit"], item["credit"])
            and actual_by_key[key_]["page"] == item["page"]
            for key_, item in expected.items()
        )
    )
    checks["R003"] = 1.0 if tx_ok else 0.0
    identity_ok = tx_ok and all(item["id"] for item in actual) and len({item["id"] for item in actual}) == len(expected)
    net_formula_ok = tx_ok and all(
        present_formula(workbook, f'Transactions!F{item["row"]}')
        and close(item["net"], item["credit"] - item["debit"])
        for item in actual
    )
    checks["R005"] = 1.0 if net_formula_ok else 0.0
    recon_ok = (
        close(engine.value("Reconciliation", "B4"), oracle["headers"]["opening_balance"])
        and close(engine.value("Reconciliation", "B5"), oracle["credits"])
        and close(engine.value("Reconciliation", "B6"), oracle["debits"])
        and close(engine.value("Reconciliation", "B7"), oracle["closing"])
        and close(workbook["Reconciliation"]["B8"].value, oracle["closing"])
    )
    recon_formula_ok = all(present_formula(workbook, address) for address in (
        "Reconciliation!B4", "Reconciliation!B5", "Reconciliation!B6", "Reconciliation!B7"
    ))
    checks["R006"] = 1.0 if recon_ok and recon_formula_ok else 0.0
    try:
        dynamic_ok = perturbation_response_ok(workbook, split)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        dynamic_ok = False; failures.append(f"STATEMENT_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if dynamic_ok else 0.0
    def category_signal_negated(value, signals):
        label = norm(value).replace("-", " ")
        for signal in signals:
            if re.search(rf"\b(?:not|non|without|exclude|excluded|excluding)\s+(?:a\s+|an\s+)?{re.escape(signal)}\b", label):
                return True
        return False

    def category_label_unambiguous(value):
        label = norm(value)
        revenue_signal = any(token in label for token in ("receipt", "revenue", "customer payment", "income", "deposit"))
        expense_signal = any(token in label for token in ("equipment", "supply", "expense", "travel", "lodging", "fee", "charge"))
        return not (revenue_signal and expense_signal)

    def category_semantics_ok(description, value):
        description, label = norm(description), norm(value)
        if not label or not category_label_unambiguous(label):
            return False
        if "receipt" in description or "ach credit" in description:
            tokens = ("receipt", "revenue", "customer payment", "income", "deposit")
            return not category_signal_negated(label, tokens) and any(token in label for token in tokens)
        if any(token in description for token in ("equipment", "instrument", "supply")):
            tokens = ("equipment", "supply", "operating expense", "business expense")
            return not category_signal_negated(label, tokens) and any(token in label for token in tokens)
        if any(token in description for token in ("travel", "lodging")):
            tokens = ("travel", "lodging", "operating expense", "business expense")
            return not category_signal_negated(label, tokens) and any(token in label for token in tokens)
        if "fee" in description:
            tokens = ("fee", "bank charge", "operating expense", "business expense")
            return not category_signal_negated(label, tokens) and any(token in label for token in tokens)
        return False
    transaction_categories_ok = tx_ok and all(
        category_semantics_ok(item["description"], actual_by_key[key_]["category"])
        for key_, item in expected.items()
    )
    categories_sheet = workbook["Categories"]
    category_rows = []
    for row in range(4, categories_sheet.max_row + 1):
        if text(categories_sheet[f"A{row}"].value):
            category_rows.append((
                norm(categories_sheet[f"A{row}"].value), formula_value(engine, "Categories", f"B{row}"),
                present_formula(workbook, f"Categories!B{row}")
            ))
    def transaction_kind(description):
        description = norm(description)
        if "receipt" in description or "ach credit" in description: return "receipts"
        if any(token in description for token in ("equipment", "instrument", "supply")): return "equipment"
        if any(token in description for token in ("travel", "lodging")): return "travel"
        if "fee" in description: return "fees"
        return ""
    def summary_label_ok(label, kinds):
        if not kinds or "" in kinds or not category_label_unambiguous(label):
            return False
        if "receipts" in kinds:
            tokens = ("receipt", "revenue", "customer payment", "income", "deposit")
            return kinds == {"receipts"} and not category_signal_negated(label, tokens) and any(token in label for token in tokens)
        generic_tokens = ("operating expense", "business expense", "general expense")
        if any(token in label for token in generic_tokens):
            return not category_signal_negated(label, generic_tokens)
        tokens = {
            "equipment": ("equipment", "supply", "instrument"),
            "travel": ("travel", "lodging"),
            "fees": ("fee", "charge"),
        }
        return all(
            not category_signal_negated(label, tokens[kind])
            and any(token in label for token in tokens[kind])
            for kind in kinds
        )
    category_rows_unique = len({label for label, _, _ in category_rows}) == len(category_rows)
    assignments = {row: [] for row in range(4, categories_sheet.max_row + 1) if text(categories_sheet[f"A{row}"].value)}
    summary_probe_ok = category_rows_unique and all(record[2] for record in category_rows)
    baseline_category_values = {
        row: engine.value("Categories", f"B{row}") for row in assignments
    }
    if summary_probe_ok:
        for item in actual:
            input_column = "E" if item["credit"] else "D"
            input_value = item["credit"] if item["credit"] else item["debit"]
            signed_delta = 1.0 if input_column == "E" else -1.0
            probe = FormulaEngine(workbook, {key("Transactions", f'{input_column}{item["row"]}'): input_value + 1.0})
            changes = {
                category_row: probe.value("Categories", f"B{category_row}") - baseline_value
                for category_row, baseline_value in baseline_category_values.items()
            }
            changed_rows = [category_row for category_row, change in changes.items() if not close(change, 0.0)]
            if len(changed_rows) != 1 or not close(changes[changed_rows[0]], signed_delta):
                summary_probe_ok = False
                break
            assignments[changed_rows[0]].append(item)
    summary_categories_ok = summary_probe_ok and all(
        assigned
        and close(
            baseline_category_values[category_row],
            sum(item["credit"] - item["debit"] for item in assigned),
        )
        and summary_label_ok(
            norm(categories_sheet[f"A{category_row}"].value),
            {transaction_kind(item["description"]) for item in assigned},
        )
        for category_row, assigned in assignments.items()
    )
    categories_ok = transaction_categories_ok and summary_categories_ok
    checks["R008"] = 1.0 if categories_ok else 0.0
    provenance_sheet = workbook["Provenance"]
    provenance_rows = {}
    for row in range(4, provenance_sheet.max_row + 1):
        transaction_id = text(provenance_sheet[f"A{row}"].value)
        if transaction_id:
            provenance_rows.setdefault(transaction_id, []).append((
                text(provenance_sheet[f"B{row}"].value), provenance_sheet[f"C{row}"].value,
                text(provenance_sheet[f"D{row}"].value)
            ))
    actual_ids = {item["id"] for item in actual}
    provenance_population_ok = (
        identity_ok
        and set(provenance_rows) == actual_ids
        and sum(len(records) for records in provenance_rows.values()) == len(actual)
    )
    checks["R004"] = 1.0 if provenance_population_ok else 0.0
    provenance_ok = provenance_population_ok and all(
        len(provenance_rows.get(item["id"], [])) == 1
        and source_field_matches(provenance_rows[item["id"]][0][0], oracle["document"]["filename"])
        and provenance_rows[item["id"]][0][1] == expected[(item["date"], norm(item["description"]))]["page"]
        and locator_ok(
            provenance_rows[item["id"]][0][2], oracle["document"]["filename"],
            expected[(item["date"], norm(item["description"]))]["page"]
        )
        for item in actual
    )
    checks["R009"] = 1.0 if provenance_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = bool(set(provenance_rows) - actual_ids) or any(
        (
            source_field_contradicts(record[0], oracle["document"]["filename"])
        )
        or (
            record[1] not in (None, "")
            and record[1] not in valid_pages
        )
        for records in provenance_rows.values()
        for record in records
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_STATEMENT_PROVENANCE")
    return checks, failures


def semantic_header(sheet, aliases, required):
    """Locate a table by its visible headers, independent of row and column."""
    normalized = {
        role: {norm(alias) for alias in values}
        for role, values in aliases.items()
    }
    for row in range(1, sheet.max_row + 1):
        columns = {}
        for column in range(1, sheet.max_column + 1):
            value = norm(sheet.cell(row=row, column=column).value)
            for role, accepted in normalized.items():
                if value in accepted and role not in columns:
                    columns[role] = column
        if set(required) <= set(columns):
            return row, columns
    return None, {}


def labeled_value_cells(sheet, aliases):
    result = {}
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column):
            label = norm(sheet.cell(row=row, column=column).value)
            for role, accepted in normalized.items():
                if label in accepted and role not in result:
                    result[role] = sheet.cell(row=row, column=column + 1)
    return result


def task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []

    header_cells = labeled_value_cells(workbook["Statement_Header"], {
        "bank": ("Bank",),
        "account_suffix": ("Account ending", "Account suffix", "Account"),
        "period_start": ("Period start", "Start date"),
        "period_end": ("Period end", "End date"),
        "opening_balance": ("Opening balance",),
        "currency": ("Currency",),
    })
    recon_cells = labeled_value_cells(workbook["Reconciliation"], {
        "opening": ("Opening balance",),
        "credits": ("Credits", "Total credits"),
        "debits": ("Debits", "Total debits"),
        "calculated": ("Calculated closing balance", "Calculated balance"),
        "statement": ("Statement closing balance", "Printed closing balance", "Closing balance"),
    })
    checks["R001"] = 1.0
    headers_ok = set(header_cells) == {"bank", "account_suffix", "period_start", "period_end", "opening_balance", "currency"} and (
        text(header_cells["bank"].value) == text(oracle["headers"]["bank"])
        and text(header_cells["account_suffix"].value) == text(oracle["headers"]["account_suffix"])
        and date_key(header_cells["period_start"].value) == date_key(oracle["headers"]["period_start"])
        and date_key(header_cells["period_end"].value) == date_key(oracle["headers"]["period_end"])
        and close(header_cells["opening_balance"].value, oracle["headers"]["opening_balance"])
        and norm(header_cells["currency"].value) == norm(oracle["headers"]["currency"])
        and "statement" in recon_cells
        and close(recon_cells["statement"].value, oracle["closing"])
    )
    checks["R002"] = 1.0 if headers_ok else 0.0

    transaction_sheet = workbook["Transactions"]
    tx_header, tx_columns = semantic_header(transaction_sheet, {
        "id": ("Transaction ID", "ID"),
        "date": ("Date", "Transaction date"),
        "description": ("Description", "Transaction", "Memo"),
        "debit": ("Debit", "Debits"),
        "credit": ("Credit", "Credits"),
        "net": ("Net impact", "Net amount", "Net"),
        "page": ("Page", "Source page"),
        "category": ("Category", "Type"),
    }, ("id", "date", "description", "debit", "credit", "page", "category"))
    actual = []
    if tx_header is not None:
        for row in range(tx_header + 1, transaction_sheet.max_row + 1):
            description = text(transaction_sheet.cell(row=row, column=tx_columns["description"]).value)
            if not description:
                continue
            net_cell = transaction_sheet.cell(row=row, column=tx_columns.get("net", tx_columns["credit"]))
            actual.append({
                "row": row,
                "id": text(transaction_sheet.cell(row=row, column=tx_columns["id"]).value),
                "date": date_key(transaction_sheet.cell(row=row, column=tx_columns["date"]).value),
                "description": description,
                "debit": amount(transaction_sheet.cell(row=row, column=tx_columns["debit"]).value),
                "credit": amount(transaction_sheet.cell(row=row, column=tx_columns["credit"]).value),
                "net": formula_value(engine, transaction_sheet.title, net_cell.coordinate) if "net" in tx_columns else None,
                "net_cell": net_cell,
                "page": transaction_sheet.cell(row=row, column=tx_columns["page"]).value,
                "category": text(transaction_sheet.cell(row=row, column=tx_columns["category"]).value),
            })
    expected = {(date_key(item["date"]), norm(item["description"])): item for item in oracle["transactions"]}
    actual_by_key = {(item["date"], norm(item["description"])): item for item in actual}
    tx_ok = (
        len(actual) == len(expected) == len(actual_by_key)
        and set(actual_by_key) == set(expected)
        and all(
            close(actual_by_key[key_]["debit"], item["debit"])
            and close(actual_by_key[key_]["credit"], item["credit"])
            and actual_by_key[key_]["page"] == item["page"]
            for key_, item in expected.items()
        )
    )
    checks["R003"] = 1.0 if tx_ok else 0.0
    identity_ok = tx_ok and all(item["id"] for item in actual) and len({item["id"] for item in actual}) == len(expected)

    net_values_ok = tx_ok and all(
        item["net"] is None or close(item["net"], item["credit"] - item["debit"])
        for item in actual
    )
    checks["R005"] = 1.0 if net_values_ok else 0.0

    recon_formula_ok = set(recon_cells) >= {"opening", "credits", "debits", "calculated", "statement"} and all(
        isinstance(recon_cells[name].value, str) and recon_cells[name].value.startswith("=")
        for name in ("opening", "credits", "debits", "calculated")
    )
    recon_values = {
        name: formula_value(engine, cell.parent.title, cell.coordinate)
        for name, cell in recon_cells.items()
    }
    recon_ok = recon_formula_ok and (
        close(recon_values.get("opening"), oracle["headers"]["opening_balance"])
        and close(recon_values.get("credits"), oracle["credits"])
        and close(recon_values.get("debits"), oracle["debits"])
        and close(recon_values.get("calculated"), oracle["closing"])
        and close(recon_values.get("statement"), oracle["closing"])
    )
    checks["R006"] = 1.0 if recon_ok else 0.0

    dynamic_ok = recon_ok
    if dynamic_ok:
        for case in load_perturbations(split):
            target = case["target"]
            matches = [item for item in actual if norm(item["description"]) == norm(target["match_value"])]
            if len(matches) != 1:
                dynamic_ok = False
                break
            item = matches[0]
            role = "credit" if target["value_column"] == "E" else "debit"
            target_cell = transaction_sheet.cell(row=item["row"], column=tx_columns[role])
            probe = FormulaEngine(workbook, {key(transaction_sheet.title, target_cell.coordinate): target["perturbed"]})
            canonical_to_role = {"B4": "opening", "B5": "credits", "B6": "debits", "B7": "calculated", "B8": "statement"}
            for expected_output in case.get("expected", []):
                role_name = canonical_to_role.get(expected_output["cell"])
                if role_name and not close(
                    probe.value(recon_cells[role_name].parent.title, recon_cells[role_name].coordinate),
                    expected_output["value"],
                ):
                    dynamic_ok = False
                    break
            for protected_output in case.get("protected", []):
                role_name = canonical_to_role.get(protected_output["cell"])
                if role_name and not close(
                    probe.value(recon_cells[role_name].parent.title, recon_cells[role_name].coordinate),
                    protected_output["value"],
                ):
                    dynamic_ok = False
                    break
            if not dynamic_ok:
                break
    checks["R007"] = 1.0 if dynamic_ok else 0.0

    def category_semantics_ok(description, label):
        description, label = norm(description), norm(label)
        if "receipt" in description or "ach credit" in description:
            return any(token in label for token in ("receipt", "revenue", "customer payment", "income", "deposit"))
        if any(token in description for token in ("equipment", "instrument", "supply")):
            return any(token in label for token in ("equipment", "supply", "expense"))
        if any(token in description for token in ("travel", "lodging")):
            return any(token in label for token in ("travel", "lodging", "expense"))
        if "fee" in description:
            return any(token in label for token in ("fee", "charge", "expense"))
        return False

    category_sheet = workbook["Categories"]
    category_header, category_columns = semantic_header(category_sheet, {
        "category": ("Category", "Type"),
        "value": ("Net impact", "Amount", "Value", "Total"),
    }, ("category", "value"))
    summary_values = {}
    if category_header is not None:
        for row in range(category_header + 1, category_sheet.max_row + 1):
            label = text(category_sheet.cell(row=row, column=category_columns["category"]).value)
            if label:
                cell = category_sheet.cell(row=row, column=category_columns["value"])
                summary_values[norm(label)] = formula_value(engine, category_sheet.title, cell.coordinate)
    transaction_categories_ok = tx_ok and all(category_semantics_ok(item["description"], item["category"]) for item in actual)
    def transaction_kind(description):
        description = norm(description)
        if "receipt" in description or "ach credit" in description:
            return "receipts"
        if any(token in description for token in ("equipment", "instrument", "supply")):
            return "equipment"
        if any(token in description for token in ("travel", "lodging")):
            return "travel"
        if "fee" in description:
            return "fees"
        return ""

    def summary_accepts(label, kind):
        label = norm(label)
        if kind == "receipts":
            return any(token in label for token in ("receipt", "revenue", "customer payment", "income", "deposit"))
        if kind == "equipment":
            return any(token in label for token in ("equipment", "supply", "instrument", "operating expense", "business expense", "general expense"))
        if kind == "travel":
            return any(token in label for token in ("travel", "lodging", "operating expense", "business expense", "general expense"))
        if kind == "fees":
            return any(token in label for token in ("fee", "charge", "operating expense", "business expense", "general expense"))
        return False

    expected_summary = {label: 0.0 for label in summary_values}
    unambiguous_assignment = bool(summary_values)
    for item in actual:
        matches = [label for label in summary_values if summary_accepts(label, transaction_kind(item["description"]))]
        if len(matches) != 1:
            unambiguous_assignment = False
            break
        expected_summary[matches[0]] += item["credit"] - item["debit"]
    summary_ok = unambiguous_assignment and all(
        close(summary_values[label], value) for label, value in expected_summary.items()
    )
    checks["R008"] = 1.0 if transaction_categories_ok and summary_ok else 0.0

    provenance_sheet = workbook["Provenance"]
    provenance_header, provenance_columns = semantic_header(provenance_sheet, {
        "id": ("Transaction ID", "ID"),
        "source": ("Source document", "Source file", "Document"),
        "page": ("Page", "Source page"),
        "locator": ("Text locator", "Locator", "Source locator"),
    }, ("id", "source", "page", "locator"))
    provenance_rows = {}
    if provenance_header is not None:
        for row in range(provenance_header + 1, provenance_sheet.max_row + 1):
            transaction_id = text(provenance_sheet.cell(row=row, column=provenance_columns["id"]).value)
            if transaction_id:
                provenance_rows.setdefault(transaction_id, []).append((
                    text(provenance_sheet.cell(row=row, column=provenance_columns["source"]).value),
                    provenance_sheet.cell(row=row, column=provenance_columns["page"]).value,
                    text(provenance_sheet.cell(row=row, column=provenance_columns["locator"]).value),
                ))
    actual_ids = {item["id"] for item in actual}
    provenance_population_ok = identity_ok and set(provenance_rows) == actual_ids and sum(map(len, provenance_rows.values())) == len(actual)
    checks["R004"] = 1.0 if provenance_population_ok else 0.0
    provenance_ok = provenance_population_ok and all(
        len(provenance_rows[item["id"]]) == 1
        and source_field_matches(provenance_rows[item["id"]][0][0], oracle["document"]["filename"])
        and provenance_rows[item["id"]][0][1] == expected[(item["date"], norm(item["description"]))]["page"]
        and locator_ok(provenance_rows[item["id"]][0][2], oracle["document"]["filename"], expected[(item["date"], norm(item["description"]))]["page"])
        for item in actual
    )
    checks["R009"] = 1.0 if provenance_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = bool(set(provenance_rows) - actual_ids) or any(
        source_field_contradicts(record[0], oracle["document"]["filename"])
        or (record[1] not in (None, "") and record[1] not in valid_pages)
        for records in provenance_rows.values() for record in records
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_STATEMENT_PROVENANCE")
    return checks, failures


def score(candidate, split=None):
    split = active_split(split)
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=["OUTPUT_MISSING"],
        )
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=[f"MALFORMED_XLSX:{type(exc).__name__}"],
        )

    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], semantic_sheet_aliases(workbook)
    )
    checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=failures,
        )

    try:
        oracle = load_oracle(split)
        engine = FormulaEngine(workbook)
        checks, task_failures = task_checks(workbook, engine, oracle, split)
        checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=checks,
        failures=failures,
    )


if __name__ == "__main__":
    arguments = sys.argv[1:]
    selected_split = None
    if "--split" in arguments:
        split_index = arguments.index("--split")
        if split_index + 1 >= len(arguments):
            raise SystemExit("--split requires dev or confirm")
        selected_split = arguments[split_index + 1]
        del arguments[split_index:split_index + 2]
    if len(arguments) > 1:
        raise SystemExit("usage: evaluate.py [candidate.xlsx] [--split dev|confirm]")
    target = Path(arguments[0]) if arguments else Path("/app/output/answer.xlsx")
    print(json.dumps(score(target, selected_split), sort_keys=True))

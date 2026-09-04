#!/usr/bin/env python3
"""Focused Judge V3 regression: dynamic reconciliation independent of fixed rows/cells."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


TASK_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TASK_DIR / "tests"))
SPEC = importlib.util.spec_from_file_location("statement_evaluator", TASK_DIR / "tests" / "evaluate.py")
EVALUATOR = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(EVALUATOR)


def build_shifted_dynamic_layout(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    for sheet in workbook.worksheets:
        sheet.insert_rows(2, amount=2)
    for row in range(6, 11):
        workbook["Transactions"].cell(row, 6, f"=E{row}-D{row}")
    workbook["Categories"]["B6"] = "=Transactions!F6+Transactions!F8"
    workbook["Categories"]["B7"] = "=Transactions!F7"
    workbook["Categories"]["B8"] = "=Transactions!F9+Transactions!F10"
    workbook["Reconciliation"]["B6"] = "=Statement_Header!B10"
    workbook["Reconciliation"]["B7"] = "=Transactions!E6+Transactions!E7+Transactions!E8+Transactions!E9+Transactions!E10"
    workbook["Reconciliation"]["B8"] = "=Transactions!D6+Transactions!D7+Transactions!D8+Transactions!D9+Transactions!D10"
    workbook["Reconciliation"]["B9"] = "=B6+B7-B8"
    workbook["Reconciliation"]["B10"] = 8533.0
    workbook.save(path)


def build_renamed_sheets(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    names = ("Account overview", "Ledger lines", "Cash categories", "Source register", "Balance bridge", "Review notes")
    for sheet, name in zip(workbook.worksheets, names):
        sheet.title = name
    workbook.save(path)


def result_record(path: Path, expected: str, predicate, display_path: str | None = None) -> dict:
    result = EVALUATOR.score(path, "dev")
    return {
        "path": display_path or str(path.relative_to(TASK_DIR)),
        "score": result["normalized_score"],
        "pass": result["pass"],
        "criterion_scores": result["criterion_scores"],
        "failure_codes": result["failure_codes"],
        "expected": expected,
        "expected_pass": expected.startswith(("full pass", "pass with")),
        "expectation_met": bool(predicate(result)),
    }


def main() -> None:
    renamed_path = TASK_DIR / "fixtures/equivalent/candidate_renamed_sheets.xlsx"
    build_renamed_sheets(renamed_path)
    fixtures = {
        "reference": (TASK_DIR / "solution/reference.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "equivalent": (TASK_DIR / "fixtures/equivalent/candidate_equivalent.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "noop": (TASK_DIR / "fixtures/noop/candidate_noop.xlsx", "fail", lambda r: not r["pass"]),
        "malformed": (TASK_DIR / "fixtures/malformed/candidate_malformed.xlsx", "fail with zero score", lambda r: not r["pass"] and r["normalized_score"] == 0.0),
    }
    for path in sorted((TASK_DIR / "fixtures/mutants").glob("*.xlsx")):
        if path.stem == "wrong_debit_credit_sign":
            fixtures[f"mutant:{path.stem}"] = (
                path,
                "pass with score below 1 and R005 failure because the auxiliary net-impact display is not a hurdle",
                lambda r: r["normalized_score"] < 1.0 and r["criterion_scores"]["R005"] == 0.0,
            )
        else:
            fixtures[f"mutant:{path.stem}"] = (path, "fail", lambda r: not r["pass"])
    fixture_results = {
        name: result_record(path, expected, predicate)
        for name, (path, expected, predicate) in fixtures.items()
    }

    with tempfile.TemporaryDirectory(prefix="statement-judge-v3-") as directory:
        shifted = Path(directory) / "shifted_dynamic.xlsx"
        build_shifted_dynamic_layout(shifted)
        shifted_result = result_record(
            shifted,
            "full pass: reconciliation and category summary remain valid after all tables move two rows",
            lambda r: r["pass"] and r["normalized_score"] == 1.0,
            "generated:shifted_dynamic_layout",
        )
        broken = Path(directory) / "wrong_opening_link.xlsx"
        workbook = load_workbook(shifted)
        workbook["Reconciliation"]["B6"] = "=Statement_Header!B10+1"
        workbook.save(broken)
        broken_result = result_record(
            broken,
            "fail R006/R007: reconciliation links to the wrong opening amount",
            lambda r: not r["pass"] and r["criterion_scores"]["R006"] == 0.0 and r["criterion_scores"]["R007"] == 0.0,
            "generated:wrong_opening_link",
        )

    renamed_result = result_record(
        renamed_path,
        "full pass: every sheet is renamed while visible headers preserve its business role",
        lambda r: r["pass"] and r["normalized_score"] == 1.0,
    )
    new_cases = {"renamed_sheets": renamed_result, "shifted_dynamic_layout": shifted_result, "wrong_opening_link": broken_result}
    all_met = all(item["expectation_met"] for item in (*fixture_results.values(), *new_cases.values()))
    receipt = {
        "task_id": "P15-C-STATEMENT-001",
        "judge_version": "p15_v3.0",
        "validation_scope": "reference, equivalent, noop, malformed, all mutants, shifted dynamic layout, genuine reconciliation error",
        "expectation_basis": {
            "shifted_dynamic_layout": "The instruction explicitly requires a formula-linked, editable reconciliation, but does not prescribe table rows or cell addresses.",
            "renamed_sheets": "The instruction does not prescribe worksheet names; visible headers and labels establish each business role.",
            "wrong_opening_link": "Opening balance is part of the disclosed reconciliation identity and must link to the correct source amount.",
        },
        "fixtures": fixture_results,
        "new_cases": new_cases,
        "all_expectations_met": all_met,
    }
    receipt_path = TASK_DIR / "receipts/judge_v3_local_validation.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"task_id": receipt["task_id"], "all_expectations_met": all_met, "scores": {key: [value["score"], value["pass"]] for key, value in {**fixture_results, **new_cases}.items()}}, ensure_ascii=False))
    if not all_met:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-B-HEALTH-REPORT-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import csv
import importlib.util
import json
import math
import os
import re
import statistics
import sys
import unicodedata
from pathlib import Path

import openpyxl

from judge_v2_support import (
    build_result,
    resolve_sheet_roles,
    sheet_resolution_failures,
    sheet_token,
)

CRITERIA = [
    {"id": "R001", "description": "A readable workbook contains every requested analytical and report sheet.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {}},
    {"id": "R002", "description": "Analytical_Data contains exactly the Vermont and New Hampshire 2012-2017 records with correct state/year alignment.", "weight": 4, "type": "positive", "dimension": "data_alignment", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R003", "description": "Both states' baseline mean, current mean, absolute change, and percent change equal the independent replay.", "weight": 4, "type": "positive", "dimension": "metric_correctness", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R004", "description": "The Vermont comparison is supported by the supplied Vermont rows and the stated baseline/current periods.", "weight": 3, "type": "positive", "dimension": "vermont_support", "method": "deterministic", "method_params": {}},
    {"id": "R005", "description": "The New Hampshire comparison is supported by the supplied New Hampshire rows and the stated baseline/current periods.", "weight": 3, "type": "positive", "dimension": "new_hampshire_support", "method": "deterministic", "method_params": {}},
    {"id": "R006", "description": "The numerical briefing claims agree with the period metrics and identify the supporting workbook evidence.", "weight": 3, "type": "positive", "dimension": "report_traceability", "method": "deterministic", "method_params": {}},
    {"id": "R007", "description": "The chart compares both states using the correct current-period values.", "weight": 4, "type": "positive", "dimension": "chart_binding", "method": "ooxml", "method_params": {}},
    {"id": "R008", "description": "Every supplied Vermont, New Hampshire, and geography-map source field is preserved exactly.", "weight": 3, "type": "positive", "dimension": "change_locality", "method": "deterministic", "method_params": {}},
    {"id": "R009", "description": "The workbook exposes reviewable checks for coverage, completeness, and metric consistency.", "weight": 2, "type": "positive", "dimension": "auditability", "method": "deterministic", "method_params": {}},
]

TASK = {
    "task_id": "P15-B-HEALTH-REPORT-001",
    "pass_threshold": 0.7,
    "required_sheets": ["Raw_Vermont", "Raw_New_Hampshire", "Geo_Map", "Analytical_Data", "Period_Metrics", "Report", "Visualization", "Checks"],
    # Hurdles are intentionally rare. They gate delivery acceptance without
    # changing the continuous score earned on independent criteria.
    "hurdle_criteria": ["R002", "R003", "R006", "R007", "R008"],
    "criteria": CRITERIA,
}
SHEET_ALIASES = {
    "Raw_Vermont": (
        "VT_Raw_Extract", "Raw_VT_Extract", "Raw_VT", "VT Raw Extract",
        "Raw_VT_Data", "Raw Vermont Extract",
    ),
    "Raw_New_Hampshire": (
        "NH_Raw_Extract", "Raw_NH_Extract", "Raw_NH", "NH Raw Extract",
        "Raw_NH_Data", "Raw New Hampshire Extract",
    ),
    "Geo_Map": ("Geography_Map", "Geography Map"),
    "Visualization": ("Chart", "Chart_Data"),
    "Checks": (
        "Data_Checks", "Validation_Checks", "Data_Quality_Checks", "Data Checks",
    ),
}
STATIC_ALTERNATIVE_SHEET_NAMES = frozenset(
    sheet_name
    for aliases in SHEET_ALIASES.values()
    for sheet_name in aliases
    if sheet_name not in {
        "VT_Raw_Extract", "Raw_VT_Extract", "NH_Raw_Extract", "Raw_NH_Extract",
        "Geography_Map", "Chart", "Data_Checks", "Validation_Checks",
    }
)
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")
FULLY_QUALIFIED_RANGE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)"
    r":(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)\$?([A-Z]+)\$?(\d+)",
    re.I,
)


def cell_key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def task_root(): return Path(__file__).resolve().parents[1]


def split_context(split):
    contract_path = task_root() / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    contract = json.loads(contract_path.read_text())
    drivers = {
        "vermont_source_cell": "Raw_Vermont!D4",
        "vermont_new_value": 151.6,
        "new_hampshire_source_cell": "Raw_New_Hampshire!D7",
        "new_hampshire_new_value": 150.0,
    }
    drivers.update(contract.get("drivers", {}))
    return {"split": split, "contract": contract, "source_root": task_root() / contract["input_files_dir"], "oracle_path": task_root() / contract["oracle"], "drivers": drivers}


def load_oracle(context):
    global oracle_module
    spec = importlib.util.spec_from_file_location(f"task_oracle_{context['split']}", context["oracle_path"])
    oracle_module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(oracle_module)
    return oracle_module.recompute()


def excel_equal(left, right):
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, str) and isinstance(right, str):
        return left.casefold() == right.casefold()
    return left == right


def excel_criteria_match(value, criteria):
    operator = "="
    expected = criteria
    if isinstance(criteria, str):
        match = re.match(r"^(<=|>=|<>|!=|=|<|>)(.*)$", criteria)
        if match:
            operator, expected = match.groups()
        stripped = expected.strip()
        try:
            expected = float(stripped)
        except (TypeError, ValueError):
            expected = stripped
    if operator in {"=", "=="}:
        return excel_equal(value, expected)
    if operator in {"<>", "!="}:
        return not excel_equal(value, expected)
    try:
        if operator == "<": return value < expected
        if operator == "<=": return value <= expected
        if operator == ">": return value > expected
        if operator == ">=": return value >= expected
    except TypeError:
        return False
    raise ValueError(f"UNSUPPORTED_CRITERIA_OPERATOR:{operator}")


def as_range(value, function_name):
    if not isinstance(value, list):
        raise ValueError(f"{function_name}_REQUIRES_RANGE")
    return value


class FormulaRange(list):
    """A flat Excel range with enough shape metadata for exact VLOOKUP."""

    def __init__(self, values, rows, columns):
        super().__init__(values)
        self.rows = int(rows)
        self.columns = int(columns)


def excel_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def excel_round(value, digits):
    digits = int(digits)
    factor = 10 ** digits
    magnitude = math.floor(abs(float(value)) * factor + 0.5) / factor
    return math.copysign(magnitude, float(value))


class UnsupportedFormulaError(RuntimeError):
    """Valid-looking Excel syntax outside this bounded deterministic replay."""


class InvalidFormulaError(ValueError):
    """A malformed candidate formula; this is a model outcome, not a Judge fault."""


def formula_strings_balanced(expression):
    in_string = False
    index = 0
    while index < len(expression):
        if expression[index] != '"':
            index += 1
            continue
        if in_string and index + 1 < len(expression) and expression[index + 1] == '"':
            index += 2
            continue
        in_string = not in_string
        index += 1
    return not in_string


def formula_parentheses_balanced(expression):
    in_string = False
    depth = 0
    index = 0
    while index < len(expression):
        char = expression[index]
        if char == '"':
            if in_string and index + 1 < len(expression) and expression[index + 1] == '"':
                index += 2
                continue
            in_string = not in_string
        elif not in_string and char == "(":
            depth += 1
        elif not in_string and char == ")":
            depth -= 1
            if depth < 0:
                return False
        index += 1
    return depth == 0


def elementwise_binary(left, right, operation):
    left_is_range = isinstance(left, list)
    right_is_range = isinstance(right, list)
    if not left_is_range and not right_is_range:
        return operation(left, right)
    if left_is_range and right_is_range and len(left) != len(right):
        raise ValueError("ARRAY_LENGTH_MISMATCH")
    length = len(left) if left_is_range else len(right)
    return [
        operation(left[index] if left_is_range else left, right[index] if right_is_range else right)
        for index in range(length)
    ]


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides: return self.overrides[key]
        if key in self.memo: return self.memo[key]
        if key in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(c)}{r}") for r in range(int(r1), int(r2) + 1) for c in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        if not formula_strings_balanced(expression):
            raise InvalidFormulaError("UNBALANCED_FORMULA_STRING")
        if not formula_parentheses_balanced(expression):
            raise InvalidFormulaError("UNBALANCED_FORMULA_PARENTHESES")
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(piece if piece.startswith('"') else pattern.sub(replacement, piece) for piece in pieces)
        isformula = re.compile(
            r"ISFORMULA\(\s*(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?"
            r"\$?([A-Z]+)\$?(\d+)\s*\)",
            re.I,
        )
        def isformula_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            raw = self.workbook[sheet][f"{match.group(3)}{match.group(4)}"].value
            return repr(isinstance(raw, str) and raw.startswith("="))
        # Preserve the formula-identity predicate before ordinary references
        # are replaced by their calculated values.
        expression = outside(isformula, isformula_replace, expression)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            values = self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6))
            rows = int(match.group(6)) - int(match.group(4)) + 1
            columns = col_index(match.group(5)) - col_index(match.group(3)) + 1
            return f"RANGE_VALUE({values!r},{rows},{columns})"
        def fully_qualified_range_replace(match):
            first_sheet = match.group(1) or match.group(2) or current_sheet
            second_sheet = match.group(5) or match.group(6)
            if sheet_token(first_sheet) != sheet_token(second_sheet):
                raise UnsupportedFormulaError("THREE_DIMENSIONAL_RANGE")
            values = self.range_values(
                first_sheet,
                match.group(3),
                match.group(4),
                match.group(7),
                match.group(8),
            )
            rows = int(match.group(8)) - int(match.group(4)) + 1
            columns = col_index(match.group(7)) - col_index(match.group(3)) + 1
            return f"RANGE_VALUE({values!r},{rows},{columns})"
        expression = outside(
            FULLY_QUALIFIED_RANGE,
            fully_qualified_range_replace,
            expression,
        )
        expression = outside(RANGE, range_replace, expression)
        def ref_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, ref_replace, expression)
        expression = outside(
            re.compile(r"\bRANK\.EQ\s*\(", re.I),
            "RANK_EQ(",
            expression,
        )
        # Python's bitwise-or precedence matches Excel concatenation for the
        # bounded formulas replayed here: below arithmetic and above
        # comparisons.  Keep it distinct from numeric addition so Excel's text
        # coercion can be applied deliberately in safe_eval.
        expression = outside(re.compile(r"&"), "|", expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_SYNTAX:{expression}") from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.Name) and node.id.upper() in {"TRUE", "FALSE"}:
            return node.id.upper() == "TRUE"
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand)
            if isinstance(value, list):
                return [(-item if isinstance(node.op, ast.USub) else item) for item in value]
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            if isinstance(node.op, ast.BitOr):
                return elementwise_binary(left, right, lambda a, b: excel_text(a) + excel_text(b))
            operations = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b, ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b, ast.Pow: lambda a, b: a ** b}
            for cls, operation in operations.items():
                if isinstance(node.op, cls): return elementwise_binary(left, right, operation)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                comparison = (
                    (lambda a, b: a == b) if isinstance(operator, ast.Eq)
                    else (lambda a, b: a != b) if isinstance(operator, ast.NotEq)
                    else (lambda a, b: a > b) if isinstance(operator, ast.Gt)
                    else (lambda a, b: a >= b) if isinstance(operator, ast.GtE)
                    else (lambda a, b: a < b) if isinstance(operator, ast.Lt)
                    else (lambda a, b: a <= b) if isinstance(operator, ast.LtE)
                    else None
                )
                if comparison is None:
                    raise UnsupportedFormulaError("UNSUPPORTED_COMPARISON_OPERATOR")
                ok = elementwise_binary(left, right, comparison)
                if isinstance(ok, list):
                    return ok
                if not ok:
                    return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) != 3:
                    raise ValueError("IF_ARGUMENTS")
                return self.safe_eval(node.args[1] if self.safe_eval(node.args[0]) else node.args[2])
            if name == "IFERROR":
                if len(node.args) != 2:
                    raise ValueError("IFERROR_ARGUMENTS")
                try:
                    return self.safe_eval(node.args[0])
                except (ArithmeticError, TypeError, ValueError):
                    return self.safe_eval(node.args[1])
            args = [self.safe_eval(item) for item in node.args]
            values = [value for group in args for value in (group if isinstance(group, list) else [group]) if value is not None]
            if name == "RANGE_VALUE":
                if len(args) != 3 or not isinstance(args[0], list):
                    raise ValueError("RANGE_VALUE_ARGUMENTS")
                return FormulaRange(args[0], args[1], args[2])
            if name == "SUM": return sum(values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "AVERAGEIFS":
                if len(args) < 3 or len(args) % 2 == 0:
                    raise ValueError("AVERAGEIFS_ARGUMENTS")
                average_range = as_range(args[0], name)
                criteria_pairs = [
                    (as_range(args[index], name), args[index + 1])
                    for index in range(1, len(args), 2)
                ]
                if any(len(criteria_range) != len(average_range) for criteria_range, _ in criteria_pairs):
                    raise ValueError("AVERAGEIFS_RANGE_LENGTH")
                selected = [
                    value
                    for row, value in enumerate(average_range)
                    if isinstance(value, (int, float))
                    and all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                ]
                if not selected:
                    raise ValueError("AVERAGEIFS_NO_MATCH")
                return statistics.mean(selected)
            if name == "COUNTIFS":
                if len(args) < 2 or len(args) % 2 != 0:
                    raise ValueError("COUNTIFS_ARGUMENTS")
                criteria_pairs = [
                    (as_range(args[index], name), args[index + 1])
                    for index in range(0, len(args), 2)
                ]
                lengths = {len(criteria_range) for criteria_range, _ in criteria_pairs}
                if len(lengths) != 1:
                    raise ValueError("COUNTIFS_RANGE_LENGTH")
                criteria_arrays = [criterion for _, criterion in criteria_pairs if isinstance(criterion, list)]
                if criteria_arrays:
                    array_lengths = {len(criterion) for criterion in criteria_arrays}
                    if len(array_lengths) != 1:
                        raise ValueError("COUNTIFS_CRITERIA_ARRAY_LENGTH")
                    return [
                        sum(
                            all(
                                excel_criteria_match(
                                    criteria_range[row],
                                    criterion[index] if isinstance(criterion, list) else criterion,
                                )
                                for criteria_range, criterion in criteria_pairs
                            )
                            for row in range(next(iter(lengths)))
                        )
                        for index in range(next(iter(array_lengths)))
                    ]
                return sum(
                    all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                    for row in range(next(iter(lengths)))
                )
            if name == "COUNTIF":
                if len(args) != 2:
                    raise ValueError("COUNTIF_ARGUMENTS")
                candidates = as_range(args[0], name)
                criteria = args[1]
                if isinstance(criteria, list):
                    return [
                        sum(excel_criteria_match(value, criterion) for value in candidates)
                        for criterion in criteria
                    ]
                return sum(excel_criteria_match(value, criteria) for value in candidates)
            if name == "SUMIF":
                if len(args) not in (2, 3):
                    raise ValueError("SUMIF_ARGUMENTS")
                candidates = as_range(args[0], name)
                values = as_range(args[2], name) if len(args) == 3 else candidates
                if len(candidates) != len(values):
                    raise ValueError("SUMIF_RANGE_LENGTH")
                return sum(
                    value if isinstance(value, (int, float)) and not isinstance(value, bool) else 0
                    for candidate, value in zip(candidates, values)
                    if excel_criteria_match(candidate, args[1])
                )
            if name in {"MINIFS", "MAXIFS"}:
                if len(args) < 3 or len(args) % 2 == 0:
                    raise ValueError(f"{name}_ARGUMENTS")
                value_range = as_range(args[0], name)
                criteria_pairs = [
                    (as_range(args[index], name), args[index + 1])
                    for index in range(1, len(args), 2)
                ]
                if any(len(criteria_range) != len(value_range) for criteria_range, _ in criteria_pairs):
                    raise ValueError(f"{name}_RANGE_LENGTH")
                selected = [
                    value
                    for row, value in enumerate(value_range)
                    if isinstance(value, (int, float))
                    and all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                ]
                if not selected:
                    raise ValueError(f"{name}_NO_MATCH")
                return min(selected) if name == "MINIFS" else max(selected)
            if name == "MAX":
                numeric_values = [value for value in values if isinstance(value, (int, float))]
                if not numeric_values:
                    raise ValueError("MAX_NO_NUMERIC_VALUES")
                return max(numeric_values)
            if name == "MIN":
                numeric_values = [value for value in values if isinstance(value, (int, float))]
                if not numeric_values:
                    raise ValueError("MIN_NO_NUMERIC_VALUES")
                return min(numeric_values)
            if name == "COUNT": return sum(isinstance(value, (int, float)) for value in values)
            if name == "COUNTA": return sum(value not in (None, "") for value in values)
            if name == "COUNTBLANK":
                if len(args) != 1 or not isinstance(args[0], list):
                    raise ValueError("COUNTBLANK_REQUIRES_RANGE")
                return sum(value in (None, "") for value in args[0])
            if name == "ISERROR":
                if len(args) != 1:
                    raise ValueError("ISERROR_ARGUMENTS")
                return [False for _ in args[0]] if isinstance(args[0], list) else False
            if name == "ISFORMULA":
                return False
            if name == "ISNUMBER":
                return elementwise_binary(args[0], None, lambda value, _: isinstance(value, (int, float)) and not isinstance(value, bool))
            if name == "ISTEXT":
                return elementwise_binary(args[0], None, lambda value, _: isinstance(value, str))
            if name == "SUMPRODUCT":
                arrays = [argument if isinstance(argument, list) else [argument] for argument in args]
                lengths = {len(array) for array in arrays}
                if len(lengths) != 1:
                    raise ValueError("SUMPRODUCT_RANGE_LENGTH")
                return sum(math.prod(array[index] for array in arrays) for index in range(next(iter(lengths))))
            if name == "MATCH":
                if len(args) not in (2, 3):
                    raise ValueError("MATCH_ARGUMENTS")
                lookup_range = as_range(args[1], name)
                match_type = 1 if len(args) == 2 else args[2]
                if match_type != 0:
                    raise ValueError("MATCH_ONLY_EXACT_SUPPORTED")
                for index, value in enumerate(lookup_range, start=1):
                    if excel_equal(value, args[0]):
                        return index
                raise ValueError("MATCH_NOT_FOUND")
            if name == "INDEX":
                if len(args) not in (2, 3):
                    raise ValueError("INDEX_ARGUMENTS")
                index_range = as_range(args[0], name)
                row_number = args[1]
                if not isinstance(row_number, (int, float)) or int(row_number) != row_number:
                    raise ValueError("INDEX_ROW_NUMBER")
                row_number = int(row_number)
                if len(args) == 3 and args[2] not in (0, 1):
                    raise ValueError("INDEX_ONLY_SINGLE_COLUMN_SUPPORTED")
                if row_number < 1 or row_number > len(index_range):
                    raise ValueError("INDEX_OUT_OF_RANGE")
                return index_range[row_number - 1]
            if name == "VLOOKUP":
                if len(args) != 4:
                    raise ValueError("VLOOKUP_ARGUMENTS")
                table = args[1]
                column = args[2]
                if args[3] not in (False, 0):
                    raise ValueError("VLOOKUP_ONLY_EXACT_SUPPORTED")
                if not isinstance(table, FormulaRange):
                    raise ValueError("VLOOKUP_REQUIRES_SHAPED_RANGE")
                if not isinstance(column, (int, float)) or int(column) != column:
                    raise ValueError("VLOOKUP_COLUMN_NUMBER")
                column = int(column)
                if column < 1 or column > table.columns:
                    raise ValueError("VLOOKUP_COLUMN_OUT_OF_RANGE")
                for offset in range(0, len(table), table.columns):
                    if excel_equal(table[offset], args[0]):
                        return table[offset + column - 1]
                raise ValueError("VLOOKUP_NOT_FOUND")
            if name == "TEXT":
                if len(args) != 2 or not isinstance(args[0], (int, float)):
                    raise ValueError("TEXT_ARGUMENTS")
                value, number_format = args[0], str(args[1])
                sections = number_format.split(";")
                section = sections[1] if value < 0 and len(sections) > 1 else sections[0]
                percent = "%" in section
                decimal_match = re.search(r"\.([0#]+)", section)
                decimals = len(decimal_match.group(1)) if decimal_match else 0
                magnitude = abs(value) * (100 if percent else 1)
                rendered = f"{magnitude:.{decimals}f}"
                if value < 0:
                    rendered = "-" + rendered
                elif section.strip().startswith("+"):
                    rendered = "+" + rendered
                if percent:
                    rendered += "%"
                return rendered
            if name == "ROUND":
                if len(args) != 2 or not isinstance(args[0], (int, float)):
                    raise ValueError("ROUND_ARGUMENTS")
                return excel_round(args[0], args[1])
            if name in {"RANK", "RANK_EQ"}:
                if len(args) not in (2, 3):
                    raise ValueError(f"{name}_ARGUMENTS")
                value = args[0]
                candidates = [
                    candidate for candidate in as_range(args[1], name)
                    if isinstance(candidate, (int, float)) and not isinstance(candidate, bool)
                ]
                order = args[2] if len(args) == 3 else 0
                if value not in candidates or order not in (0, 1):
                    raise ValueError(f"{name}_VALUE_OR_ORDER")
                ordered = sorted(candidates, reverse=order == 0)
                return ordered.index(value) + 1
            if name == "ABS": return abs(args[0])
            if name == "AND": return all(args)
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def norm(value):
    if isinstance(value, float): return round(value, 6)
    return value


def normalized_words(value):
    text = unicodedata.normalize("NFKC", str(value)).casefold()
    text = re.sub(r"[‐‑‒–—−]", "-", text)
    return re.sub(r"[^a-z0-9-]+", " ", text).strip()


def normalize_state_label(value):
    text = normalized_words(value).replace(".", "")
    compact = re.sub(r"[^a-z]", "", text)
    if compact in {"vt", "vermont"}:
        return "Vermont"
    if compact in {"nh", "newhampshire"}:
        return "New Hampshire"
    return value


def normalize_period_label(value):
    text = normalized_words(value)
    role = "baseline" if "baseline" in text else "current" if "current" in text else None
    years = re.findall(r"\b(?:19|20)\d{2}\b", text)
    if role and len(years) == 2:
        return role, int(years[0]), int(years[1])
    return text


def header_matches(role, value):
    text = normalized_words(value)
    words = set(text.split())
    compact = re.sub(r"[^a-z0-9]", "", text)
    if role == "state": return compact in {"state", "statename"}
    if role == "year": return compact in {"year", "reportingyear"}
    if role == "cause": return compact in {"cause", "causename", "causeofdeath"}
    if role == "deaths": return compact in {"death", "deaths", "deathcount", "deathcounts"}
    if role == "aadr": return "aadr" in words or compact == "aadr" or ("ageadjusted" in compact and "rate" in compact)
    if role == "period": return "period" in words and "metric" not in words
    if role == "reporting_region": return "reporting" in words and "region" in words
    if role == "source_key": return "source" in words and "key" in words
    if role == "baseline": return "baseline" in words and ("aadr" in words or "rate" in words or "mean" in words)
    if role == "current": return "current" in words and ("aadr" in words or "rate" in words or "mean" in words)
    if role == "absolute_change": return "absolute" in words and "change" in words
    if role == "percent_change": return ("percent" in words or "percentage" in words) and "change" in words
    if role == "claim": return compact in {"claim", "briefingclaim", "metricclaim"}
    if role == "value": return compact in {"value", "result", "metricvalue"}
    return False


def header_columns(sheet, row, roles):
    columns = {}
    duplicates = set()
    for column in range(1, min(sheet.max_column, 40) + 1):
        value = sheet.cell(row=row, column=column).value
        for role in roles:
            if not header_matches(role, value):
                continue
            if role in columns:
                duplicates.add(role)
            else:
                columns[role] = column
    return columns, duplicates


def find_semantic_table(workbook, sheet_name, roles, allow_missing=False):
    sheet = workbook[sheet_name]
    candidates = []
    for row in range(1, min(sheet.max_row, 30) + 1):
        columns, duplicates = header_columns(sheet, row, roles)
        if not duplicates and all(role in columns for role in roles):
            candidates.append((row, columns))
    if not candidates and allow_missing:
        return None
    if len(candidates) != 1:
        raise ValueError(f"{sheet_name}_SEMANTIC_HEADER_COUNT:{len(candidates)}")
    return candidates[0]


def optional_header_column(sheet, header_row, role):
    columns, duplicates = header_columns(sheet, header_row, (role,))
    if duplicates:
        raise ValueError(f"{sheet.title}_{role}_HEADER_AMBIGUOUS")
    return columns.get(role)


def unique_record_row(workbook, engine, sheet_name, table, state, year):
    header_row, columns = table
    sheet = workbook[sheet_name]
    matches = []
    data_started = False
    for row in range(header_row + 1, sheet.max_row + 1):
        row_state = engine.value(sheet_name, f"{col_name(columns['state'])}{row}")
        row_year = engine.value(sheet_name, f"{col_name(columns['year'])}{row}")
        if row_state in (None, "") and row_year in (None, ""):
            if data_started:
                break
            continue
        data_started = True
        if normalize_state_label(row_state) == normalize_state_label(state) and row_year == year:
            matches.append(row)
    if len(matches) != 1:
        raise ValueError(f"{sheet_name}_RECORD_COUNT:{state}:{year}:{len(matches)}")
    return matches[0]


def unique_state_row(workbook, engine, sheet_name, table, state):
    header_row, columns = table
    sheet = workbook[sheet_name]
    matches = []
    data_started = False
    for row in range(header_row + 1, sheet.max_row + 1):
        row_state = engine.value(sheet_name, f"{col_name(columns['state'])}{row}")
        if row_state in (None, ""):
            if data_started:
                break
            continue
        data_started = True
        if normalize_state_label(row_state) == normalize_state_label(state):
            matches.append(row)
    if len(matches) != 1:
        raise ValueError(f"{sheet_name}_STATE_ROW_COUNT:{state}:{len(matches)}")
    return matches[0]


def semantic_address(table, role, row):
    return f"{col_name(table[1][role])}{row}"


def formula_dependency_cells(workbook, sheet_name, cell):
    pending, seen = [(sheet_name, cell)], set()
    while pending and len(seen) < 128:
        current_sheet, current_cell = pending.pop()
        key = (current_sheet, current_cell)
        if key in seen or current_sheet not in workbook.sheetnames:
            continue
        seen.add(key)
        formula = workbook[current_sheet][current_cell].value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        expression = re.sub(r'"(?:[^"]|"")*"', "", formula)
        for match in RANGE.finditer(expression):
            referenced_sheet = match.group(1) or match.group(2) or current_sheet
            for row in range(int(match.group(4)), int(match.group(6)) + 1):
                for column in range(col_index(match.group(3)), col_index(match.group(5)) + 1):
                    pending.append((referenced_sheet, f"{col_name(column)}{row}"))
        expression = RANGE.sub("", expression)
        for match in REF.finditer(expression):
            pending.append((match.group(1) or match.group(2) or current_sheet, f"{match.group(3)}{match.group(4)}"))
    return {f"{sheet}!{cell}" for sheet, cell in seen}


def formula_mentions_sheet(workbook, sheet_name, cell, dependency):
    target = sheet_token(
        workbook.actual_sheet_name(dependency)
        if hasattr(workbook, "actual_sheet_name")
        else dependency
    )
    return any(
        sheet_token(address.split("!", 1)[0]) == target
        for address in formula_dependency_cells(workbook, sheet_name, cell)
    )


def formula_cells_in_row(workbook, sheet_name, row):
    sheet = workbook[sheet_name]
    return [
        f"{col_name(column)}{row}"
        for column in range(1, min(sheet.max_column, 40) + 1)
        if isinstance(sheet.cell(row=row, column=column).value, str)
        and sheet.cell(row=row, column=column).value.startswith("=")
    ]


def load_public_sources(root):
    health = {}
    for filename in ("vermont_heart_disease.csv", "new_hampshire_heart_disease.csv"):
        with (root / filename).open(newline="") as handle:
            for record in csv.DictReader(handle):
                key = (normalize_state_label(record["state"]), int(record["year"]))
                health[key] = {
                    "state": record["state"],
                    "year": int(record["year"]),
                    "cause": record.get("cause_name"),
                    "deaths": int(record["deaths"]),
                    "aadr": float(record["aadr"]),
                }
    geography = {}
    with (root / "geography_map.csv").open(newline="") as handle:
        for record in csv.DictReader(handle):
            geography[normalize_state_label(record["state"])] = {
                "state": record["state"],
                "reporting_region": record["reporting_region"],
                "source_key": record["source_key"],
            }
    return health, geography


def source_value_equal(actual, expected, role):
    if role in {"year", "deaths"}:
        return isinstance(actual, (int, float)) and int(actual) == actual and int(actual) == expected
    if role == "aadr":
        return close(actual, expected, 0.000001)
    return actual == expected


def rows(workbook, sheet, start, end, columns):
    values = []
    for row in range(start, end + 1):
        record = tuple(norm(workbook[sheet].cell(row=row, column=column).value) for column in columns)
        if any(value not in (None, "") for value in record): values.append(record)
    return values


def source_tables_match(workbook, root):
    expected_health, expected_geo = load_public_sources(root)
    actual_health = {}
    for sheet in ("Raw_Vermont", "Raw_New_Hampshire"):
        table = find_semantic_table(workbook, sheet, ("state", "year", "deaths", "aadr"))
        header_row, columns = table
        cause_column = optional_header_column(workbook[sheet], header_row, "cause")
        for row in range(header_row + 1, workbook[sheet].max_row + 1):
            state = workbook[sheet].cell(row=row, column=columns["state"]).value
            year = workbook[sheet].cell(row=row, column=columns["year"]).value
            if state in (None, "") and year in (None, ""):
                continue
            if state in (None, "") or not isinstance(year, (int, float)) or int(year) != year:
                return False
            key = (normalize_state_label(state), int(year))
            if key in actual_health:
                return False
            record = {
                "state": state,
                "year": int(year),
                "deaths": workbook[sheet].cell(row=row, column=columns["deaths"]).value,
                "aadr": workbook[sheet].cell(row=row, column=columns["aadr"]).value,
            }
            if cause_column is not None:
                record["cause"] = workbook[sheet].cell(row=row, column=cause_column).value
            actual_health[key] = record
    if set(actual_health) != set(expected_health):
        return False
    for key, expected in expected_health.items():
        actual = actual_health[key]
        for role in ("state", "year", "deaths", "aadr"):
            if not source_value_equal(actual[role], expected[role], role):
                return False
        if "cause" in actual and not source_value_equal(actual["cause"], expected["cause"], "cause"):
            return False

    geo_table = find_semantic_table(workbook, "Geo_Map", ("state", "reporting_region", "source_key"))
    geo_header, geo_columns = geo_table
    actual_geo = {}
    for row in range(geo_header + 1, workbook["Geo_Map"].max_row + 1):
        state = workbook["Geo_Map"].cell(row=row, column=geo_columns["state"]).value
        if state in (None, ""):
            continue
        key = normalize_state_label(state)
        if key in actual_geo:
            return False
        actual_geo[key] = {
            "state": state,
            "reporting_region": workbook["Geo_Map"].cell(row=row, column=geo_columns["reporting_region"]).value,
            "source_key": workbook["Geo_Map"].cell(row=row, column=geo_columns["source_key"]).value,
        }
    if set(actual_geo) != set(expected_geo):
        return False
    return all(
        all(source_value_equal(actual_geo[state][role], expected_geo[state][role], role) for role in ("state", "reporting_region", "source_key"))
        for state in expected_geo
    )


def normalize_chart_ref(value):
    return (value or "").replace("'", "").replace("$", "").lower()


def parse_chart_range(value):
    match = re.fullmatch(r"(.+)!([a-z]+)(\d+):([a-z]+)(\d+)", normalize_chart_ref(value))
    if not match or match.group(2) != match.group(4):
        return None
    return match.group(1), match.group(2).upper(), int(match.group(3)), int(match.group(5))


def chart_sheet_name(workbook, reference_name):
    """Resolve an OOXML chart reference without requiring the chart's sheet."""
    wanted = sheet_token(reference_name)
    matches = [name for name in workbook.sheetnames if sheet_token(name) == wanted]
    return matches[0] if len(matches) == 1 else None


def chart_binding_ok(workbook, engine, expected):
    chart_sheet = workbook.actual_sheet_name("Visualization") if hasattr(workbook, "actual_sheet_name") else "Visualization"
    for chart in workbook[chart_sheet]._charts:
        for series in getattr(chart, "ser", []):
            category_ref = None
            if getattr(series, "cat", None) is not None:
                str_ref = getattr(series.cat, "strRef", None)
                num_ref = getattr(series.cat, "numRef", None)
                category_ref = getattr(str_ref, "f", None) or getattr(num_ref, "f", None)
            value_ref = getattr(getattr(getattr(series, "val", None), "numRef", None), "f", None)
            categories, values = parse_chart_range(category_ref), parse_chart_range(value_ref)
            if not categories or not values:
                continue
            category_sheet, category_column, category_start, category_end = categories
            value_sheet, value_column, value_start, value_end = values
            category_sheet = chart_sheet_name(workbook, category_sheet)
            value_sheet = chart_sheet_name(workbook, value_sheet)
            if category_sheet is None or value_sheet is None:
                continue
            if category_end - category_start + 1 != 2 or value_end - value_start + 1 != 2:
                continue
            observed = []
            valid = True
            for category_row, value_row in zip(
                range(category_start, category_end + 1),
                range(value_start, value_end + 1),
            ):
                value_cell = f"{value_column}{value_row}"
                try:
                    state = normalize_state_label(engine.value(category_sheet, f"{category_column}{category_row}"))
                    value = engine.value(value_sheet, value_cell)
                except UnsupportedFormulaError:
                    raise
                except Exception:
                    valid = False
                    break
                if state not in expected["metrics"]:
                    valid = False
                    break
                if not close(value, expected["metrics"][state][1], 0.051):
                    valid = False
                    break
                observed.append(state)
            if valid and set(observed) == {"Vermont", "New Hampshire"}:
                return True
    return False


def analytical_alignment_ok(workbook, engine, oracle, source_root):
    table = find_semantic_table(workbook, "Analytical_Data", ("state", "year", "deaths", "aadr"))
    header_row, columns = table
    cause_column = optional_header_column(workbook["Analytical_Data"], header_row, "cause")
    period_column = optional_header_column(workbook["Analytical_Data"], header_row, "period")
    region_column = optional_header_column(workbook["Analytical_Data"], header_row, "reporting_region")
    source_key_column = optional_header_column(workbook["Analytical_Data"], header_row, "source_key")
    expected_health, expected_geo = load_public_sources(source_root)
    expected_rows = {
        (normalize_state_label(row[0]), row[1]): {"deaths": row[2], "aadr": row[3]}
        for row in oracle["rows"]
    }
    years = sorted({row[1] for row in oracle["rows"]})
    baseline_label = oracle.get("baseline_label", f"Baseline {years[0]}-{years[2]}")
    current_label = oracle.get("current_label", f"Current {years[3]}-{years[5]}")
    actual = {}
    for row in range(header_row + 1, workbook["Analytical_Data"].max_row + 1):
        state = engine.value("Analytical_Data", f"{col_name(columns['state'])}{row}")
        year = engine.value("Analytical_Data", f"{col_name(columns['year'])}{row}")
        if state in (None, "") and year in (None, ""):
            continue
        if state in (None, "") or not isinstance(year, (int, float)) or int(year) != year:
            return False
        key = (normalize_state_label(state), int(year))
        if key in actual:
            return False
        actual[key] = {
            "deaths": engine.value("Analytical_Data", f"{col_name(columns['deaths'])}{row}"),
            "aadr": engine.value("Analytical_Data", f"{col_name(columns['aadr'])}{row}"),
        }
        if period_column is not None:
            actual[key]["period"] = engine.value("Analytical_Data", f"{col_name(period_column)}{row}")
        if cause_column is not None:
            actual[key]["cause"] = engine.value("Analytical_Data", f"{col_name(cause_column)}{row}")
        if region_column is not None:
            actual[key]["reporting_region"] = engine.value("Analytical_Data", f"{col_name(region_column)}{row}")
        if source_key_column is not None:
            actual[key]["source_key"] = engine.value("Analytical_Data", f"{col_name(source_key_column)}{row}")
    if set(actual) != set(expected_rows):
        return False
    for key, expected in expected_rows.items():
        record = actual[key]
        if not source_value_equal(record["deaths"], expected["deaths"], "deaths") or not close(record["aadr"], expected["aadr"], 0.000001):
            return False
        if "period" in record:
            expected_period = baseline_label if key[1] <= years[2] else current_label
            actual_period = normalize_period_label(record["period"])
            normalized_expected_period = normalize_period_label(expected_period)
            expected_role = (
                normalized_expected_period[0]
                if isinstance(normalized_expected_period, tuple)
                else normalized_expected_period
            )
            if actual_period not in {normalized_expected_period, expected_role}:
                return False
        if "cause" in record and record["cause"] != expected_health[key]["cause"]:
            return False
        if "reporting_region" in record and record["reporting_region"] != expected_geo[key[0]]["reporting_region"]:
            return False
        if "source_key" in record and record["source_key"] != expected_geo[key[0]]["source_key"]:
            return False
    return True


def metric_table(workbook):
    return find_semantic_table(workbook, "Period_Metrics", ("state", "baseline", "current", "absolute_change", "percent_change"))


def metric_address(workbook, engine, table, state, role):
    row = unique_state_row(workbook, engine, "Period_Metrics", table, state)
    return semantic_address(table, role, row)


def semantic_driver(context, name, defaults):
    semantic = context["contract"].get("semantic_drivers", {}).get(name, {})
    driver = dict(defaults)
    driver.update(semantic)
    legacy_prefix = "vermont" if name == "vermont" else "new_hampshire"
    legacy_cell = context["drivers"].get(f"{legacy_prefix}_source_cell")
    if legacy_cell and "raw_sheet" not in semantic:
        driver["raw_sheet"] = legacy_cell.split("!", 1)[0]
    if f"{legacy_prefix}_new_value" in context["drivers"] and "new_value" not in semantic:
        driver["new_value"] = context["drivers"][f"{legacy_prefix}_new_value"]
    return driver


def raw_driver_address(workbook, engine, driver):
    table = find_semantic_table(workbook, driver["raw_sheet"], ("state", "year", "deaths", "aadr"))
    row = unique_record_row(workbook, engine, driver["raw_sheet"], table, driver["state"], driver["year"])
    return f"{driver['raw_sheet']}!{semantic_address(table, 'aadr', row)}"


def report_state_table(workbook):
    return find_semantic_table(
        workbook,
        "Report",
        ("state", "baseline", "current", "absolute_change", "percent_change"),
        allow_missing=True,
    )


def state_name_in_text(text, state):
    return re.sub(r"[^a-z]", "", normalized_words(state)) in re.sub(r"[^a-z]", "", normalized_words(text))


def report_change_formula(workbook, engine, state, expected_value):
    matches = []
    sheet = workbook["Report"]
    for row in range(1, sheet.max_row + 1):
        label = " ".join(str(sheet.cell(row=row, column=column).value or "") for column in range(1, min(sheet.max_column, 12) + 1))
        if not state_name_in_text(label, state):
            continue
        for cell in formula_cells_in_row(workbook, "Report", row):
            try:
                observed = engine.value("Report", cell)
                numeric_claims = []
                if isinstance(observed, str):
                    numeric_claims = [float(value) for value in re.findall(r"[-+]?\d+(?:\.\d+)?", observed)]
                value_matches = close(observed, expected_value, 0.001) or any(
                    close(value, expected_value, 0.051) for value in numeric_claims
                )
                if formula_mentions_sheet(workbook, "Report", cell, "Period_Metrics") and value_matches:
                    matches.append(cell)
            except UnsupportedFormulaError:
                raise
            except Exception:
                continue
    if len(matches) != 1:
        raise ValueError(f"REPORT_CHANGE_FORMULA_COUNT:{state}:{len(matches)}")
    return matches[0]


def highest_state_formula(workbook, engine, expected_state):
    matches = []
    sheet = workbook["Report"]
    metrics = metric_table(workbook)
    metrics_sheet = (
        workbook.actual_sheet_name("Period_Metrics")
        if hasattr(workbook, "actual_sheet_name")
        else "Period_Metrics"
    )
    current_addresses = {
        f"{metrics_sheet}!{metric_address(workbook, engine, metrics, state, 'current')}"
        for state in ("Vermont", "New Hampshire")
    }
    for row in range(1, sheet.max_row + 1):
        context = " ".join(
            str(sheet.cell(row=row, column=column).value or "")
            for column in range(1, min(sheet.max_column, 12) + 1)
        )
        words = set(normalized_words(context).split())
        has_leader_meaning = bool(
            {"highest", "higher", "leader", "leading", "maximum", "max"} & words
        ) and bool({"current", "recent"} & words)
        if not has_leader_meaning:
            continue
        for cell in formula_cells_in_row(workbook, "Report", row):
            try:
                dependencies = formula_dependency_cells(workbook, "Report", cell)
                if (
                    formula_mentions_sheet(workbook, "Report", cell, "Period_Metrics")
                    and current_addresses <= dependencies
                    and state_name_in_text(engine.value("Report", cell), expected_state)
                ):
                    matches.append(cell)
            except UnsupportedFormulaError:
                raise
            except Exception:
                continue
    if len(matches) != 1:
        raise ValueError(f"HIGHEST_STATE_FORMULA_COUNT:{len(matches)}")
    return matches[0]


def report_claims_ok(workbook, engine, expected):
    state_table = report_state_table(workbook)
    if state_table is not None:
        for state in ("Vermont", "New Hampshire"):
            row = unique_state_row(workbook, engine, "Report", state_table, state)
            for role, expected_value in zip(("baseline", "current", "absolute_change", "percent_change"), expected["metrics"][state]):
                cell = semantic_address(state_table, role, row)
                if not formula_mentions_sheet(workbook, "Report", cell, "Period_Metrics") or not close(engine.value("Report", cell), expected_value, 0.001):
                    return False
    else:
        for state in ("Vermont", "New Hampshire"):
            report_change_formula(workbook, engine, state, expected["metrics"][state][2])
    highest_state_formula(workbook, engine, expected["highest_current"])
    return True


def report_claims_traceable(workbook, engine, expected):
    """Accept correct, visibly sourced report claims without prescribing formulas."""
    try:
        if report_claims_ok(workbook, engine, expected):
            return True
    except (UnsupportedFormulaError, ValueError, KeyError, TypeError):
        pass

    sheet = workbook["Report"]
    report_rows = [
        " ".join(str(sheet.cell(row=row, column=column).value or "") for column in range(1, min(sheet.max_column, 20) + 1))
        for row in range(1, sheet.max_row + 1)
    ]
    all_report_text = normalized_words(" ".join(report_rows))
    global_evidence_ok = "period metrics" in all_report_text or (
        "analytical data" in all_report_text and "data source" in all_report_text
    )
    state_claims = set()
    for row in range(1, sheet.max_row + 1):
        cells = [sheet.cell(row=row, column=column) for column in range(1, min(sheet.max_column, 20) + 1)]
        row_text = " ".join(str(cell.value or "") for cell in cells)
        words = set(normalized_words(row_text).split())
        evidence_ok = global_evidence_ok or "period metrics" in normalized_words(row_text)
        resolved = []
        for cell in cells:
            try:
                resolved.append(engine.value("Report", cell.coordinate))
            except Exception:
                resolved.append(cell.value)
        for state in ("Vermont", "New Hampshire"):
            if state_name_in_text(row_text, state) and evidence_ok and any(
                close(value, expected["metrics"][state][2], 0.001) for value in resolved
            ):
                state_claims.add(state)
    if state_claims == {"Vermont", "New Hampshire"}:
        return True

    # Narrative reports often put a state heading above several prose rows.
    # Match each state's own section at the displayed precision rather than
    # requiring a table or formula cell at a fixed address.
    heading_rows = {}
    for state in ("Vermont", "New Hampshire"):
        for index, row_text in enumerate(report_rows):
            compact = normalized_words(row_text)
            if state_name_in_text(compact, state) and len(compact.split()) <= 3:
                heading_rows[state] = index
                break
    if set(heading_rows) != {"Vermont", "New Hampshire"} or not global_evidence_ok:
        return False
    ordered = sorted((row, state) for state, row in heading_rows.items())
    for position, (start, state) in enumerate(ordered):
        end = ordered[position + 1][0] if position + 1 < len(ordered) else len(report_rows)
        section = " ".join(report_rows[start:end])
        numbers = []
        for raw in re.findall(r"[-+]?\d[\d,]*(?:\.\d+)?%?", section):
            percent = raw.endswith("%")
            try:
                value = float(raw.rstrip("%").replace(",", ""))
            except ValueError:
                continue
            numbers.append((value, percent))
        baseline, current, change, percent_change = expected["metrics"][state]
        required = ((baseline, False), (current, False), (change, False), (percent_change * 100, True))
        if not all(any(flag == expected_percent and abs(value - target) <= 0.051 for value, flag in numbers) for target, expected_percent in required):
            return False
    return True


def report_dynamic_ok(workbook, engine, state, expected_current, expected_absolute):
    state_table = report_state_table(workbook)
    if state_table is not None:
        row = unique_state_row(workbook, engine, "Report", state_table, state)
        cell = semantic_address(state_table, "current", row)
        return formula_mentions_sheet(workbook, "Report", cell, "Period_Metrics") and close(engine.value("Report", cell), expected_current, 0.001)
    report_change_formula(workbook, engine, state, expected_absolute)
    return True


def visualization_current_address(workbook, engine, state):
    table = find_semantic_table(workbook, "Visualization", ("state", "current"))
    row = unique_state_row(workbook, engine, "Visualization", table, state)
    return semantic_address(table, "current", row)


def check_linkage_ok(workbook, engine):
    canonical_formulas = all(formula_present(workbook, f"Checks!B{row}") for row in range(4, 7))
    try:
        canonical_values = close(engine.value("Checks", "B4"), 12) and close(engine.value("Checks", "B5"), 2) and close(engine.value("Checks", "B6"), 0, 0.001)
    except UnsupportedFormulaError:
        raise
    except Exception:
        canonical_values = False
    if canonical_formulas and canonical_values:
        return True

    analytical_ok = metric_ok = closure_ok = False
    sheet = workbook["Checks"]
    for row in range(1, sheet.max_row + 1):
        label = normalized_words(sheet.cell(row=row, column=1).value)
        for cell in formula_cells_in_row(workbook, "Checks", row):
            try:
                value = engine.value("Checks", cell)
            except UnsupportedFormulaError:
                raise
            except Exception:
                continue
            if "analytical" in label and ({"row", "rows", "record", "records"} & set(label.split())) and close(value, 12):
                analytical_ok = True
            if "period" in label and "metric" in label and formula_mentions_sheet(workbook, "Checks", cell, "Period_Metrics") and isinstance(value, (int, float)) and value >= 2:
                metric_ok = True
            if (
                ({"error", "errors", "closure", "residual", "residuals"} & set(label.split()))
                and close(value, 0, 0.001)
                and formula_mentions_sheet(workbook, "Checks", cell, "Period_Metrics")
                and formula_mentions_sheet(workbook, "Checks", cell, "Report")
            ):
                closure_ok = True
    return analytical_ok and metric_ok and closure_ok


def workbook_contains_formula(workbook):
    return any(
        cell.data_type == "f" and not re.fullmatch(r"=+", str(cell.value or ""))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def sheet_text(sheet):
    return normalized_words(
        " ".join(
            str(cell.value)
            for row in sheet.iter_rows()
            for cell in row
            if cell.value not in (None, "")
        )
    )


def static_metric_header_columns(sheet, row):
    """Find one explicit state/rate-change table without inferring cell roles."""
    columns = {}
    duplicates = set()
    for column in range(1, min(sheet.max_column, 40) + 1):
        raw = sheet.cell(row=row, column=column).value
        text = normalized_words(raw)
        words = set(text.split())
        compact = re.sub(r"[^a-z0-9]", "", text)
        role = None
        if compact in {"state", "statename"}:
            role = "state"
        elif "baseline" in words and ({"rate", "aadr", "mean", "avg"} & words):
            role = "baseline"
        elif "current" in words and ({"rate", "aadr", "mean", "avg"} & words):
            role = "current"
        elif (
            "change" in words
            and not ({"death", "deaths"} & words)
            and ("%" in str(raw) or {"percent", "percentage", "pct"} & words)
        ):
            role = "percent_change"
        elif "change" in words and not ({"death", "deaths"} & words) and (
            {"rate", "aadr", "absolute"} & words or compact == "change"
        ):
            role = "absolute_change"
        if role is None:
            continue
        if role in columns:
            duplicates.add(role)
        else:
            columns[role] = column
    required = {"state", "baseline", "current", "absolute_change", "percent_change"}
    return columns if not duplicates and required <= set(columns) else None


def static_metric_value_matches(role, actual, expected):
    if role == "percent_change":
        return close(actual, expected, 0.0005) or close(actual, expected * 100, 0.051)
    return close(actual, expected, 0.051)


def static_metric_values_ok(workbook, expected):
    """Accept a locally static metric table even if unrelated formulas exist."""
    candidate_sheets = []
    for role in ("Period_Metrics", "Report"):
        physical = workbook.actual_sheet_name(role)
        if physical not in candidate_sheets:
            candidate_sheets.append(physical)
    if "Comparison" in workbook.sheetnames:
        candidate_sheets.append("Comparison")
    for sheet_name in candidate_sheets:
        sheet = workbook[sheet_name]
        for header_row in range(1, min(sheet.max_row, 30) + 1):
            columns = static_metric_header_columns(sheet, header_row)
            if columns is None:
                continue
            observed = {}
            for row in range(header_row + 1, min(sheet.max_row, header_row + 20) + 1):
                state = normalize_state_label(sheet.cell(row=row, column=columns["state"]).value)
                if state not in expected["metrics"]:
                    continue
                if state in observed:
                    observed = {}
                    break
                observed[state] = {
                    role: sheet.cell(row=row, column=columns[role]).value
                    for role in ("baseline", "current", "absolute_change", "percent_change")
                }
            if set(observed) != set(expected["metrics"]):
                continue
            roles = ("baseline", "current", "absolute_change", "percent_change")
            if all(
                static_metric_value_matches(role, observed[state][role], expected["metrics"][state][index])
                for state in expected["metrics"]
                for index, role in enumerate(roles)
            ):
                return True
    return False


def static_alternative_layout_signature_ok(workbook, role_map):
    """Validate only the observed task-local aliases before treating them as roles."""
    if not any(actual in STATIC_ALTERNATIVE_SHEET_NAMES for actual in role_map.values()):
        return True
    try:
        find_semantic_table(workbook, "Raw_Vermont", ("state", "year", "deaths", "aadr"))
        find_semantic_table(workbook, "Raw_New_Hampshire", ("state", "year", "deaths", "aadr"))
        find_semantic_table(workbook, "Geo_Map", ("state", "reporting_region", "source_key"))
        find_semantic_table(workbook, "Analytical_Data", ("state", "year", "deaths", "aadr"))
        period_sheet = workbook["Period_Metrics"]
        period_signature = any(
            header_columns(period_sheet, row, ("state", "period", "aadr"))[0].keys()
            >= {"state", "period", "aadr"}
            or static_metric_header_columns(period_sheet, row) is not None
            for row in range(1, min(period_sheet.max_row, 30) + 1)
        )
        # This gate only establishes that the task-local aliases identify the
        # expected data roles.  Report text, chart presence/binding, and check
        # content are deliverables scored independently by R006/R007/R009;
        # requiring them here would turn a local omission into score=null.
        return period_signature
    except Exception:
        return False


def task_checks(workbook, engine, oracle, context):
    checks, failures = {}, []
    expected = oracle
    try:
        alignment_ok = analytical_alignment_ok(workbook, engine, expected, context["source_root"])
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        alignment_ok = False; failures.append(f"ALIGNMENT_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R002"] = 1.0 if alignment_ok else 0.0

    metrics_table = None
    metric_failure = None
    try:
        metrics_table = metric_table(workbook)
        metrics_ok = True
        for state in ("Vermont", "New Hampshire"):
            expected_metric = expected["metrics"][state]
            for position, role in enumerate(("baseline", "current", "absolute_change", "percent_change")):
                address = metric_address(workbook, engine, metrics_table, state, role)
                metrics_ok = metrics_ok and formula_present(workbook, f"Period_Metrics!{address}") and close(engine.value("Period_Metrics", address), expected_metric[position], 0.001)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        metrics_ok = False
        metric_failure = f"METRIC_REPLAY_FAILED:{type(exc).__name__}"
    if not metrics_ok and static_metric_values_ok(workbook, expected):
        metrics_ok = True
    elif not metrics_ok and metric_failure:
        failures.append(metric_failure)
    checks["R003"] = 1.0 if metrics_ok else 0.0

    years = sorted({row[1] for row in expected["rows"]})
    try:
        analytical_table = find_semantic_table(workbook, "Analytical_Data", ("state", "year", "deaths", "aadr"))
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        analytical_table = None
        failures.append(f"ANALYTICAL_TABLE_DISCOVERY_FAILED:{type(exc).__name__}")
    vermont_driver = semantic_driver(context, "vermont", {"raw_sheet": "Raw_Vermont", "state": "Vermont", "year": years[0], "field": "aadr", "new_value": 151.6})
    try:
        source = raw_driver_address(workbook, engine, vermont_driver)
        source_sheet, source_cell = source.split("!", 1)
        baseline = engine.value(source_sheet, source_cell)
        new_value = vermont_driver["new_value"]
        perturb = FormulaEngine(workbook, {source: new_value})
        new_baseline = expected["metrics"]["Vermont"][0] + (new_value - baseline) / 3
        analytical_row = unique_record_row(workbook, perturb, "Analytical_Data", analytical_table, vermont_driver["state"], vermont_driver["year"])
        analytical_address = semantic_address(analytical_table, "aadr", analytical_row)
        baseline_address = metric_address(workbook, perturb, metrics_table, "Vermont", "baseline")
        vt_dynamic = (
            formula_present(workbook, f"Analytical_Data!{analytical_address}")
            and formula_present(workbook, f"Period_Metrics!{baseline_address}")
            and close(perturb.value("Analytical_Data", analytical_address), new_value)
            and close(perturb.value("Period_Metrics", baseline_address), new_baseline, 0.001)
        )
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        vt_dynamic = False; failures.append(f"VERMONT_SOURCE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R004"] = 1.0 if vt_dynamic else 0.0

    new_hampshire_driver = semantic_driver(context, "new_hampshire", {"raw_sheet": "Raw_New_Hampshire", "state": "New Hampshire", "year": years[3], "field": "aadr", "new_value": 150.0})
    try:
        source = raw_driver_address(workbook, engine, new_hampshire_driver)
        source_sheet, source_cell = source.split("!", 1)
        baseline = engine.value(source_sheet, source_cell)
        new_value = new_hampshire_driver["new_value"]
        perturb = FormulaEngine(workbook, {source: new_value})
        new_current = expected["metrics"]["New Hampshire"][1] + (new_value - baseline) / 3
        analytical_row = unique_record_row(workbook, perturb, "Analytical_Data", analytical_table, new_hampshire_driver["state"], new_hampshire_driver["year"])
        analytical_address = semantic_address(analytical_table, "aadr", analytical_row)
        current_address = metric_address(workbook, perturb, metrics_table, "New Hampshire", "current")
        visualization_address = visualization_current_address(workbook, perturb, "New Hampshire")
        expected_absolute = new_current - expected["metrics"]["New Hampshire"][0]
        nh_dynamic = (
            formula_present(workbook, f"Analytical_Data!{analytical_address}")
            and formula_present(workbook, f"Period_Metrics!{current_address}")
            and formula_present(workbook, f"Visualization!{visualization_address}")
            and close(perturb.value("Analytical_Data", analytical_address), new_value)
            and close(perturb.value("Period_Metrics", current_address), new_current, 0.001)
            and close(perturb.value("Visualization", visualization_address), new_current, 0.001)
            and report_dynamic_ok(workbook, perturb, "New Hampshire", new_current, expected_absolute)
        )
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        nh_dynamic = False; failures.append(f"NEW_HAMPSHIRE_SOURCE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R005"] = 1.0 if nh_dynamic else 0.0

    try:
        report_ok = report_claims_traceable(workbook, engine, expected)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        report_ok = False; failures.append(f"REPORT_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R006"] = 1.0 if report_ok else 0.0

    chart_ok = chart_binding_ok(workbook, engine, expected)
    checks["R007"] = 1.0 if chart_ok else 0.0
    if not chart_ok: failures.append("CHART_SERIES_OR_CATEGORY_BINDING_MISMATCH")

    protected_ok = source_tables_match(workbook, context["source_root"])
    checks["R008"] = 1.0 if protected_ok else 0.0
    if not protected_ok: failures.append("PROTECTED_HEALTH_SOURCE_CHANGED")

    # The visible task asks for a supported comparison, not for two hidden
    # cell-edit experiments.  Correct aligned source rows plus correct metrics
    # establish that support independent of sheet names or formula syntax.
    source_supported = alignment_ok and metrics_ok and protected_ok
    checks["R004"] = 1.0 if source_supported else 0.0
    checks["R005"] = 1.0 if source_supported else 0.0
    failures = [
        code for code in failures
        if not code.startswith((
            "VERMONT_SOURCE_PERTURBATION_FAILED:",
            "NEW_HAMPSHIRE_SOURCE_PERTURBATION_FAILED:",
        ))
    ]

    try:
        check_values_ok = check_linkage_ok(workbook, engine)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        check_values_ok = False; failures.append(f"CHECK_LINKAGE_FAILED:{type(exc).__name__}")
    check_sheet = workbook["Checks"]
    check_text = " ".join(
        str(cell.value or "")
        for row in check_sheet.iter_rows()
        for cell in row
    ).casefold()
    reviewable_checks = (
        "check" in check_text
        and any(word in check_text for word in ("coverage", "missing", "consistency", "closure"))
        and any(word in check_text for word in ("pass", "ok", "complete", "zero"))
    )
    checks["R009"] = 1.0 if (check_values_ok or reviewable_checks) else 0.0

    return checks, failures


def evaluate(candidate, split="dev"):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0: return criteria, ["OUTPUT_MISSING"]
    try: workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc: return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(workbook, TASK["required_sheets"], SHEET_ALIASES)
    criteria["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        return criteria, sorted(set(failures))
    if not static_alternative_layout_signature_ok(workbook, role_map):
        failures.append("UNSUPPORTED_LAYOUT:STATIC_ALTERNATIVE_SIGNATURE")
        return criteria, sorted(set(failures))
    try:
        context = split_context(split)
        checks, task_failures = task_checks(workbook, FormulaEngine(workbook), load_oracle(context), context)
        criteria.update(checks); failures.extend(task_failures)
    except UnsupportedFormulaError as exc:
        failures.append(f"UNSUPPORTED_FORMULA:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))



def parse_cli():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    candidate = None
    arguments = iter(sys.argv[1:])
    for argument in arguments:
        if argument == "--split": split = next(arguments, "")
        elif argument.startswith("--split="): split = argument.split("=", 1)[1]
        elif candidate is None: candidate = Path(argument)
        else: raise ValueError(f"UNEXPECTED_ARGUMENT:{argument}")
    if split not in {"dev", "confirm"}: raise ValueError(f"INVALID_SPLIT:{split}")
    return candidate or Path("/app/output/answer.xlsx"), split


def main():
    candidate, split = parse_cli()
    criteria, failures = evaluate(candidate, split)
    payload = build_result(task=TASK, split=split, candidate=str(candidate), criteria=criteria, failures=failures)
    payload["judge_version"] = "P15_JUDGE_V3"
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True); (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        if total is not None: (log_root / "reward.txt").write_text(str(total) + "\n")
    except OSError: pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__": main()

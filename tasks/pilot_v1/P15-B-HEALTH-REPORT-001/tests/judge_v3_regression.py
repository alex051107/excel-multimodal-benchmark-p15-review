#!/usr/bin/env python3
"""Focused Judge V3 regression: equivalent layout plus real business errors."""
from __future__ import annotations

import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

import evaluate as judge

ROOT = Path(__file__).resolve().parents[1]


def run_case(path: Path, split: str = "dev") -> dict:
    criteria, failures = judge.evaluate(path, split)
    payload = judge.build_result(
        task=judge.TASK, split=split, candidate=str(path), criteria=criteria, failures=failures
    )
    return {
        "score": payload["normalized_score"], "pass": payload["pass"],
        "criteria": payload["criterion_scores"], "failure_codes": payload["failure_codes"],
    }


def alternative_layout(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Checks"].title = "Validation Checks"
    workbook.save(path)


def chart_directly_from_metrics(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    series = workbook["Visualization"]._charts[0].ser[0]
    series.cat.strRef.f = "'Period_Metrics'!$A$4:$A$5"
    series.val.numRef.f = "'Period_Metrics'!$C$4:$C$5"
    workbook.save(path)


def static_metrics_with_unrelated_formula(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    expected = judge.load_oracle(judge.split_context("dev"))
    for row, state in ((4, "Vermont"), (5, "New Hampshire")):
        for column, value in zip(("B", "C", "D", "E"), expected["metrics"][state]):
            workbook["Period_Metrics"][f"{column}{row}"] = value
    # Formula remains in another requested area.  It must not disqualify the
    # now-static but correct Period_Metrics table.
    assert workbook["Checks"]["B4"].data_type == "f"
    workbook.save(path)


def main() -> None:
    cases = {
        "reference": (ROOT / "solution/reference.xlsx", "dev", True),
        "confirm_reference": (ROOT / "tests/confirm/reference.xlsx", "confirm", True),
        "equivalent": (ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", "dev", True),
        "noop": (ROOT / "fixtures/noop/candidate_noop.xlsx", "dev", False),
        "malformed": (ROOT / "fixtures/malformed/candidate_malformed.xlsx", "dev", False),
    }
    cases.update({
        f"mutant:{path.stem}": (path, "dev", False)
        for path in sorted((ROOT / "fixtures/mutants").glob("*.xlsx"))
    })
    results = {}
    with tempfile.TemporaryDirectory() as directory:
        alternative = Path(directory) / "validation_checks_with_spaces.xlsx"
        direct_chart = Path(directory) / "chart_direct_period_metrics.xlsx"
        local_static = Path(directory) / "static_metrics_unrelated_formula.xlsx"
        alternative_layout(alternative)
        chart_directly_from_metrics(direct_chart)
        static_metrics_with_unrelated_formula(local_static)
        cases["alternative_layout:validation_checks_with_spaces"] = (alternative, "dev", True)
        cases["equivalent:chart_direct_period_metrics"] = (direct_chart, "dev", True)
        cases["equivalent:static_metrics_with_unrelated_formula"] = (local_static, "dev", True)
        for name, (path, split, expected_pass) in cases.items():
            result = run_case(path, split)
            assert result["pass"] is expected_pass, (name, result)
            results[name] = {**result, "expected_pass": expected_pass, "meets_expectation": True}
    receipt = {
        "task_id": judge.TASK["task_id"], "judge_version": "P15_JUDGE_V3",
        "generated_at": datetime.now(timezone.utc).isoformat(), "cases": results,
        "all_expectations_met": True,
    }
    (ROOT / "receipts/judge_v3_local_validation.json").write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n"
    )
    print(json.dumps(receipt, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Deterministic semantic judge for P15-A-STAT-EXPERIMENT-001."""
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-A-STAT-EXPERIMENT-001",
  "pass_threshold": 0.7,
  "canonical_reference_sheets": [
    "Data",
    "Plan",
    "Analysis",
    "Results",
    "Visualization",
    "Checks"
  ],
  "canonical_reference_examples": {
    "Data!B4": 0.7,
    "Data!C13": 3.4
  },
  "canonical_reference_formula_cells": [
    "Data!D4",
    "Analysis!B5",
    "Analysis!B10",
    "Results!B8"
  ],
  "criteria": [
    {"id":"R001","description":"The workbook opens and the paired data and requested analysis are identifiable without relying on sheet names.","weight":1,"type":"positive","dimension":"file_usability","method":"deterministic","method_params":{"implemented_check":"semantic_analysis_content"}},
    {"id":"R002","description":"Every paired-difference row uses Group 2 minus Group 1 and is numerically correct.","weight":2,"type":"positive","dimension":"paired_rows","method":"deterministic","method_params":{"implemented_check":"paired_differences","oracle":"split_contract"}},
    {"id":"R003","description":"Paired sample size and mean difference agree with the independent analysis.","weight":2,"type":"positive","dimension":"paired_location","method":"deterministic","method_params":{"implemented_check":"paired_n_mean","oracle":"split_contract"}},
    {"id":"R004","description":"Paired SD, SE, t statistic, and degrees of freedom agree with the independent analysis.","weight":3,"type":"positive","dimension":"paired_dispersion","method":"deterministic","method_params":{"implemented_check":"paired_sd_se_t_df","oracle":"split_contract"}},
    {"id":"R005","description":"The two-sided paired-test p-value uses the declared method and agrees with the independent analysis.","weight":3,"type":"positive","dimension":"two_sided_p_value","method":"deterministic","method_params":{"implemented_check":"two_sided_p_value","oracle":"split_contract"}},
    {"id":"R006","description":"Both 95% confidence-interval bounds agree with the independent paired analysis.","weight":2,"type":"positive","dimension":"confidence_interval","method":"deterministic","method_params":{"implemented_check":"confidence_interval","oracle":"split_contract"}},
    {"id":"R007","description":"The planned paired sample size equals the declared normal-approximation result.","weight":2,"type":"positive","dimension":"power_planning","method":"deterministic","method_params":{"implemented_check":"planned_sample_size","oracle":"split_contract"}},
    {"id":"R008","description":"A chart is bound to all ten paired-difference rows, directly or through formula-linked helper cells.","weight":2,"type":"positive","dimension":"chart_binding","method":"deterministic","method_params":{"implemented_check":"paired_chart_binding"}},
    {"id":"R009","description":"The paired-difference calculation and its chart-linked value update when a source observation changes.","weight":2,"type":"positive","dimension":"native_recalculation","method":"deterministic","method_params":{"implemented_check":"paired_chart_recalculation","split_weights":{"dev":2,"confirm":3}}},
    {"id":"R010","description":"The decision summary applies the declared two-sided 0.05 rule to the calculated p-value.","weight":2,"type":"positive","dimension":"decision_consistency","method":"deterministic","method_params":{"implemented_check":"decision_summary"}},
    {"id":"R011","description":"All ten source pairs and their subject order match the supplied observations.","weight":2,"type":"positive","dimension":"source_protection","method":"deterministic","method_params":{"implemented_check":"source_pairing_protection"}}
  ]
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R005", "R010", "R011"]
SHEET_ALIASES = {}


def cell_key(sheet, cell):
    return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column:
        value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


TASK_ROOT = Path(__file__).resolve().parents[1]


def requested_split():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    for index, argument in enumerate(sys.argv[1:]):
        if argument == "--split" and index + 2 <= len(sys.argv) - 1:
            split = sys.argv[index + 2].strip().lower()
        elif argument.startswith("--split="):
            split = argument.split("=", 1)[1].strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"unsupported split: {split}")
    return split


ACTIVE_SPLIT = requested_split()


def load_contract():
    relative = "tests/confirm/contract.json" if ACTIVE_SPLIT == "confirm" else "tests/private_contract.json"
    return json.loads((TASK_ROOT / relative).read_text())


def load_oracle(contract):
    oracle_path = TASK_ROOT / contract["oracle"]
    source_path = TASK_ROOT / "metadata/oracle_recompute.py"
    if ACTIVE_SPLIT == "dev" and source_path.exists() and source_path.read_bytes() != oracle_path.read_bytes():
        raise RuntimeError("DEV_ORACLE_COPY_OUT_OF_SYNC")
    spec = importlib.util.spec_from_file_location(f"task_oracle_{ACTIVE_SPLIT}", oracle_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute(), module


def _betacf(a, b, x):
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d, h = 1.0, 1.0 - qab * x / qap, 0.0
    if abs(d) < 3e-14: d = 3e-14
    d = 1.0 / d
    h = d
    for m in range(1, 201):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; delta = d * c; h *= delta
        if abs(delta - 1.0) < 3e-12: break
    return h


def _betai(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    factor = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x))
    return factor * _betacf(a, b, x) / a if x < (a + 1) / (a + b + 2) else 1 - factor * _betacf(b, a, 1 - x) / b


def student_t_two_tail(value, df):
    x = df / (df + value * value)
    return _betai(df / 2, 0.5, x)


def student_t_inv_two_tail(alpha, df):
    lo, hi = 0.0, 20.0
    for _ in range(90):
        mid = (lo + hi) / 2
        if student_t_two_tail(mid, df) > alpha: lo = mid
        else: hi = mid
    return (lo + hi) / 2


class UnsupportedFormulaError(RuntimeError):
    """Valid-looking Excel syntax outside this bounded deterministic replay."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook
        self.overrides = overrides or {}
        self.memo = {}
        self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames:
            raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            if isinstance(raw, str) and raw.startswith("="):
                result = self.formula(raw[1:], sheet)
            else:
                result = raw
            self.memo[key] = result
            return result
        finally:
            self.stack.discard(key)

    def range_values(self, sheet, c1, r1, c2, r2):
        values = []
        for r in range(int(r1), int(r2) + 1):
            for c in range(col_index(c1), col_index(c2) + 1):
                values.append(self.value(sheet, f"{col_name(c)}{r}"))
        return values

    def formula(self, expression, current_sheet):
        # Modern Excel may persist otherwise ordinary worksheet functions with
        # a compatibility prefix.  The prefix does not change their semantics.
        expression = re.sub(r"(?i)_xl(?:fn|ws)\.", "", expression)
        expression = expression.replace("^", "**").replace("<>", "!=")
        expression = re.sub(r"\bSTDEV\.S\b", "STDEV_S", expression, flags=re.I)
        # Excel retains STDEV as a backward-compatible alias for STDEV.S.
        # Treat both spellings identically so a semantically equivalent
        # workbook is not rejected by the deterministic formula engine.
        expression = re.sub(r"\bSTDEV\b", "STDEV_S", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.2T\b", "T_DIST_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.RT\b", "T_DIST_RT", expression, flags=re.I)
        expression = re.sub(r"\bT\.INV\.2T\b", "T_INV_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.TEST\b", "T_TEST", expression, flags=re.I)
        expression = re.sub(r"\bTTEST\b", "T_TEST", expression, flags=re.I)
        expression = re.sub(r"\bCEILING\.MATH\b", "CEILING_MATH", expression, flags=re.I)
        expression = re.sub(r"\bCEILING\.PRECISE\b", "CEILING_PRECISE", expression, flags=re.I)
        expression = re.sub(r"\bPI\(\)", "PI()", expression, flags=re.I)

        def replace_range(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))

        def sub_outside_strings(pattern, replacement, value):
            parts = re.split(r'("(?:[^"\\\\]|\\\\.)*")', value)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in parts)

        expression = sub_outside_strings(RANGE, replace_range, expression)

        def replace_ref(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))

        expression = sub_outside_strings(REF, replace_ref, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_SYNTAX:{expression}") from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name) and node.id.upper() in {"TRUE", "FALSE"}:
            return node.id.upper() == "TRUE"
        if isinstance(node, ast.List):
            return [self.safe_eval(v) for v in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand)
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operators = {ast.Add: lambda a,b:a+b, ast.Sub: lambda a,b:a-b, ast.Mult: lambda a,b:a*b, ast.Div: lambda a,b:a/b, ast.Pow: lambda a,b:a**b}
            for cls, fun in operators.items():
                if isinstance(node.op, cls): return fun(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for op, comp in zip(node.ops, node.comparators):
                right = self.safe_eval(comp)
                good = (left == right if isinstance(op, ast.Eq) else left != right if isinstance(op, ast.NotEq) else left >= right if isinstance(op, ast.GtE) else left <= right if isinstance(op, ast.LtE) else left > right if isinstance(op, ast.Gt) else left < right if isinstance(op, ast.Lt) else False)
                if not good: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                if self.safe_eval(node.args[0]):
                    return self.safe_eval(node.args[1])
                return self.safe_eval(node.args[2]) if len(node.args) == 3 else False
            args = [self.safe_eval(a) for a in node.args]
            values = [v for group in args for v in (group if isinstance(group, list) else [group]) if v is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            if name == "SUM": return sum(numeric_values)
            if name == "AVERAGE": return statistics.mean(numeric_values)
            if name == "COUNT": return len(numeric_values)
            if name in {"STDEV_S", "STDEV"}: return statistics.stdev(numeric_values)
            if name == "SQRT": return math.sqrt(args[0])
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "ROUNDUP":
                digits = int(args[1])
                factor = 10 ** digits
                return math.ceil(args[0] * factor) / factor if args[0] >= 0 else math.floor(args[0] * factor) / factor
            if name in {"CEILING", "CEILING_MATH", "CEILING_PRECISE"}:
                significance = abs(args[1]) if len(args) > 1 else 1
                if significance == 0:
                    return 0
                return math.ceil(args[0] / significance) * significance
            if name == "PI": return math.pi
            if name == "AND": return all(args)
            if name == "OR": return any(args)
            if name == "NOT": return not args[0]
            if name == "POWER": return args[0] ** args[1]
            if name == "PRODUCT": return math.prod(numeric_values)
            if name == "SUMPRODUCT":
                arrays = [argument if isinstance(argument, list) else [argument] for argument in args]
                lengths = {len(array) for array in arrays}
                if len(lengths) != 1:
                    raise ValueError("SUMPRODUCT_RANGE_LENGTH")
                return sum(math.prod(array[index] for array in arrays) for index in range(next(iter(lengths))))
            if name == "MAX": return max(numeric_values)
            if name == "MIN": return min(numeric_values)
            if name == "NA": return None
            if name == "T_DIST_2T":
                return student_t_two_tail(args[0], args[1])
            if name == "T_DIST_RT":
                return student_t_two_tail(args[0], args[1]) / 2
            if name == "T_INV_2T":
                return student_t_inv_two_tail(args[0], args[1])
            if name == "T_TEST":
                if len(args) != 4 or args[2] != 2 or args[3] != 1:
                    raise ValueError("T_TEST_ONLY_PAIRED_TWO_SIDED_SUPPORTED")
                if not isinstance(args[0], list) or not isinstance(args[1], list):
                    raise ValueError("T_TEST_REQUIRES_TWO_RANGES")
                pairs = [
                    (left, right)
                    for left, right in zip(args[0], args[1])
                    if isinstance(left, (int, float)) and isinstance(right, (int, float))
                ]
                if len(pairs) < 2 or len(pairs) != len(args[0]) or len(pairs) != len(args[1]):
                    raise ValueError("T_TEST_INVALID_PAIRED_RANGES")
                differences = [left - right for left, right in pairs]
                standard_deviation = statistics.stdev(differences)
                if standard_deviation == 0:
                    raise ValueError("T_TEST_ZERO_VARIANCE")
                statistic = statistics.mean(differences) / (standard_deviation / math.sqrt(len(differences)))
                return student_t_two_tail(abs(statistic), len(differences) - 1)
            if name == "NORM_S_INV":
                return statistics.NormalDist().inv_cdf(args[0])
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def normalized_formula(value):
    return re.sub(r"\s+", "", str(value or "")).replace("$", "").replace("'", "").upper()


def uses_two_sided_t_method(value):
    formula = normalized_formula(value).replace("_XLFN.", "").replace("_XLWS.", "")
    if re.search(r"(?:T\.TEST|TTEST)\([^)]*,2,1\)", formula):
        return True
    if "T.DIST.2T(" in formula:
        return True
    if "T.DIST.RT(" not in formula:
        return False
    return (
        "2*T.DIST.RT(" in formula
        or "*2" in formula
        or "/0.5" in formula
        or formula.count("T.DIST.RT(") >= 2
    )


def decision_matches(value, reject_null):
    text = " ".join(str(value or "").strip().casefold().split())
    non_rejection = (
        "do not reject h0",
        "do not reject the null",
        "fail to reject h0",
        "fail to reject the null",
        "not statistically significant",
        "no statistically significant",
        "non-significant",
        "nonsignificant",
        "insufficient evidence",
    )
    rejection = (
        "reject h0",
        "reject the null",
        "statistically significant",
        "significant paired effect",
        "evidence of a paired effect",
    )
    if reject_null:
        return not any(phrase in text for phrase in non_rejection) and any(phrase in text for phrase in rejection)
    return any(phrase in text for phrase in non_rejection)


def normalized_label(value):
    text = str(value or "").replace("−", "-").replace("–", "-").replace("—", "-")
    return " ".join(text.strip().casefold().split())


def formula_dependency_closure(workbook, address, limit=128):
    """Return formula cells reachable from ``address`` through ordinary A1 refs."""
    pending, seen, formulas = [address], set(), {}
    while pending and len(seen) < limit:
        current = pending.pop()
        if current in seen or "!" not in current:
            continue
        seen.add(current)
        sheet_name, cell = current.split("!", 1)
        if sheet_name not in workbook.sheetnames:
            continue
        raw = workbook[sheet_name][cell].value
        formulas[current] = raw if isinstance(raw, str) and raw.startswith("=") else None
        if not isinstance(raw, str) or not raw.startswith("="):
            continue
        expression = re.sub(r'"(?:[^"]|"")*"', "", raw)
        for match in RANGE.finditer(expression):
            dependency_sheet = match.group(1) or match.group(2) or sheet_name
            for row in range(int(match.group(4)), int(match.group(6)) + 1):
                for column in range(col_index(match.group(3)), col_index(match.group(5)) + 1):
                    pending.append(f"{dependency_sheet}!{col_name(column)}{row}")
        expression = RANGE.sub("", expression)
        for match in REF.finditer(expression):
            dependency_sheet = match.group(1) or match.group(2) or sheet_name
            pending.append(f"{dependency_sheet}!{match.group(3)}{match.group(4)}")
    return formulas


def formula_chain_text(workbook, address):
    return " ".join(normalized_formula(value) for value in formula_dependency_closure(workbook, address).values())


def find_labeled_value(workbook, sheet_name, aliases, prefer_first=False):
    worksheet = workbook[sheet_name]
    all_matches = []
    for alias in aliases:
        expected = normalized_label(alias)
        matches = []
        for row in worksheet.iter_rows():
            for cell in row:
                if normalized_label(cell.value) == expected and cell.column < worksheet.max_column:
                    matches.append(f"{sheet_name}!{worksheet.cell(cell.row, cell.column + 1).coordinate}")
        if matches:
            if len(matches) != 1:
                return None
            if prefer_first:
                return matches[0]
            all_matches.extend(matches)
    return all_matches[0] if len(all_matches) == 1 else None


def find_header_columns(workbook):
    candidates = []
    for row in range(1, min(workbook["Data"].max_row, 30) + 1):
        columns = {"header_row": row}
        duplicates = set()
        for cell in workbook["Data"][row]:
            label = normalized_label(cell.value)
            role = None
            if label in {"subject", "batch", "participant", "pair"}:
                role = "subject"
            elif "paired difference" in label or ("difference" in label and "group 2" in label and "group 1" in label):
                role = "raw_difference"
            elif "group 1" in label:
                role = "group_1"
            elif "group 2" in label:
                role = "group_2"
            elif ("analysis" in label or "included" in label) and ("difference" in label or "helper" in label):
                role = "analysis_difference"
            elif label in {"include", "included", "include?"}:
                role = "include"
            if role:
                if role in columns:
                    duplicates.add(role)
                columns[role] = col_name(cell.column)
        if not duplicates and {"subject", "raw_difference"} <= set(columns):
            candidates.append(columns)
    if len(candidates) != 1:
        raise ValueError(f"DATA_HEADER_COUNT:{len(candidates)}")
    return candidates[0]


def paired_data_rows(workbook, data_columns, expected_count):
    sheet = workbook["Data"]
    rows = []
    for row in range(data_columns["header_row"] + 1, sheet.max_row + 1):
        subject = sheet[f"{data_columns['subject']}{row}"].value
        difference = sheet[f"{data_columns['raw_difference']}{row}"].value
        # Locate observations independently of whether the candidate used a
        # formula. Formula-linkage is scored by R002 below; hard-coded or
        # missing calculations are model outcomes, not evaluator failures.
        if subject not in (None, "") and difference not in (None, ""):
            rows.append(row)
    return rows


def paired_n_address(workbook):
    direct = find_labeled_value(workbook, "Analysis", ("Paired sample size", "Complete pairs (n)"))
    if direct:
        return direct
    sheet = workbook["Analysis"]
    labels = [cell for row in sheet.iter_rows() for cell in row if normalized_label(cell.value) in {"n", "paired n"}]
    matches = []
    for label in labels:
        for header_row in range(max(1, label.row - 4), label.row):
            for header in sheet[header_row]:
                if "paired difference" in normalized_label(header.value):
                    matches.append(f"Analysis!{sheet.cell(label.row, header.column).coordinate}")
    return matches[0] if len(set(matches)) == 1 else None


def metric_addresses(workbook):
    return {
        "n": paired_n_address(workbook),
        "mean": find_labeled_value(workbook, "Analysis", ("Mean paired difference",)),
        "sd": find_labeled_value(workbook, "Analysis", ("SD of paired differences",)),
        "se": find_labeled_value(workbook, "Analysis", ("Standard error",)),
        "t": find_labeled_value(workbook, "Analysis", ("t statistic",)),
        "df": find_labeled_value(workbook, "Analysis", ("Degrees of freedom",)),
        "p": find_labeled_value(workbook, "Analysis", ("Two-sided p-value",)),
        "ci_lower": find_labeled_value(workbook, "Analysis", ("95% CI lower",)),
        "ci_upper": find_labeled_value(workbook, "Analysis", ("95% CI upper",)),
        "planned_n": find_labeled_value(workbook, "Analysis", ("Planned n at d=0.8", "Required complete paired subjects", "Planned n (rounded up)")),
        # The workbook may separately report test alpha and planning alpha.
        # Prefer the unqualified test label, but still reject a repeated exact
        # label inside either section.
        "alpha": find_labeled_value(workbook, "Analysis", ("Alpha", "Two-sided alpha"), prefer_first=True),
    }


def normalized_chart_ref(value):
    return str(value or "").replace("$", "").replace("'", "").replace(" ", "").upper()


def parse_chart_range(value):
    match = re.fullmatch(r"(.+)!([A-Z]+)(\d+):([A-Z]+)(\d+)", normalized_chart_ref(value))
    if not match or match.group(2) != match.group(4):
        return None
    return match.group(1), match.group(2), int(match.group(3)), int(match.group(5))


def chart_binding_details(workbook, engine, oracle, data_columns, data_rows):
    charts = getattr(workbook["Visualization"], "_charts", [])
    subject_column = data_columns.get("subject")
    difference_column = data_columns.get("raw_difference")
    for chart in charts:
        for series in chart.series:
            category_formula = getattr(getattr(getattr(series, "cat", None), "strRef", None), "f", "") or getattr(getattr(getattr(series, "cat", None), "numRef", None), "f", "")
            value_formula = getattr(getattr(getattr(series, "val", None), "numRef", None), "f", "")
            category, values = parse_chart_range(category_formula), parse_chart_range(value_formula)
            if not category or not values:
                continue
            category_sheet, category_column, category_start, category_end = category
            value_sheet, value_column, value_start, value_end = values
            if category_start != value_start or category_end != value_end or category_end - category_start + 1 != len(data_rows):
                continue
            if (category_sheet, value_sheet) == ("DATA", "DATA") and category_column == subject_column and value_column == difference_column and list(range(category_start, category_end + 1)) == data_rows:
                return {"sheet": "Data", "column": difference_column, "mode": "direct", "row_map": {row: row for row in data_rows}}
            if category_sheet != "VISUALIZATION" or value_sheet != "VISUALIZATION":
                continue
            helper_rows = list(range(value_start, value_end + 1))
            helper_ok = True
            for index, (data_row, helper_row) in enumerate(zip(data_rows, helper_rows)):
                value_address = f"Visualization!{value_column}{helper_row}"
                try:
                    category_value = engine.value("Visualization", f"{category_column}{helper_row}")
                    observed = engine.value("Visualization", f"{value_column}{helper_row}")
                except UnsupportedFormulaError:
                    raise
                except Exception:
                    helper_ok = False
                    break
                dependency = f"Data!{difference_column}{data_row}"
                dependencies = set(formula_dependency_closure(workbook, value_address))
                raw_pair_dependencies = {
                    f"Data!{data_columns.get('group_1', 'B')}{data_row}",
                    f"Data!{data_columns.get('group_2', 'C')}{data_row}",
                }
                linked = dependency in dependencies or raw_pair_dependencies <= dependencies
                if not formula_present(workbook, value_address) or not linked:
                    helper_ok = False
                    break
                expected = oracle.get("full_differences", oracle["differences"])[index]
                expected_subject = engine.value("Data", f"{subject_column}{data_row}")
                if category_value != expected_subject or not close(observed, expected, 0.0001):
                    helper_ok = False
                    break
            if helper_ok:
                return {"sheet": "Visualization", "column": value_column, "mode": "helper", "row_map": dict(zip(data_rows, helper_rows))}
    return None


def find_decision_formulas(workbook):
    matches = []
    for row in workbook["Results"].iter_rows():
        for cell in row:
            value = cell.value
            formula = normalized_formula(value)
            text = str(value or "").casefold()
            if isinstance(value, str) and value.startswith("=") and "IF(" in formula and "reject" in text:
                matches.append((f"Results!{cell.coordinate}", value))
    return matches


def decision_formula_ok(workbook, p_address, alpha_address):
    if not p_address:
        return False
    alpha_tokens = {"0.05"}
    if alpha_address:
        alpha_tokens.add(normalized_formula(alpha_address))
    for address, raw in find_decision_formulas(workbook):
        formula = normalized_formula(raw)
        dependencies = set(formula_dependency_closure(workbook, address))
        correct_rule = p_address in dependencies and any("<" + token in formula or token + ">" in formula for token in alpha_tokens)
        text = str(raw).casefold().replace("₀", "0")
        if correct_rule and "reject" in text and (
            "do not reject" in text or "fail to reject" in text
        ):
            return True
    return False


def _legacy_reference_layout_checks(workbook, engine, oracle, contract, oracle_module):
    """Unused V2 reference-layout helper retained only for fixture archaeology."""
    checks, failures = {}, []
    data_columns = find_header_columns(workbook)
    data_rows = paired_data_rows(workbook, data_columns, len(oracle.get("full_differences", oracle["differences"])))
    metrics = metric_addresses(workbook)

    paired_rows_ok = bool(data_columns.get("raw_difference")) and (
        len(data_rows) == len(oracle.get("full_differences", oracle["differences"]))
        and bool(data_rows)
        and data_rows == list(range(data_rows[0], data_rows[-1] + 1))
    )
    paired_columns = [data_columns["raw_difference"]] if paired_rows_ok else []
    if data_columns.get("analysis_difference"):
        paired_columns.append(data_columns["analysis_difference"])
    if not paired_rows_ok:
        failures.append("PAIRED_DIFFERENCE_COLUMN_NOT_FOUND")
    for index, row in enumerate(data_rows):
        for column in paired_columns:
            address = f"Data!{column}{row}"
            if not formula_present(workbook, address):
                paired_rows_ok = False
                failures.append(f"FORMULA_REQUIRED:{address}")
            try:
                observed = engine.value("Data", f"{column}{row}")
            except UnsupportedFormulaError:
                raise
            except Exception as exc:
                observed = None
                failures.append(f"PAIRED_ROW_EVALUATION_FAILED:{address}:{type(exc).__name__}")
            if not close(observed, oracle["differences"][index], 0.0001):
                paired_rows_ok = False
                failures.append(f"PAIRED_ROW_MISMATCH:{address}")
    checks["R002"] = 1.0 if paired_rows_ok else 0.0

    def validate_group(criterion_id, expected):
        ok = True
        for key, target in expected.items():
            address = metrics.get(key)
            if not address:
                ok = False
                failures.append(f"METRIC_NOT_FOUND:{key}")
                continue
            if not formula_present(workbook, address):
                ok = False
                failures.append(f"FORMULA_REQUIRED:{address}")
            sheet, cell = address.split("!", 1)
            try:
                observed = engine.value(sheet, cell)
            except UnsupportedFormulaError:
                raise
            except Exception as exc:
                observed = None
                failures.append(f"INFERENTIAL_EVALUATION_FAILED:{address}:{type(exc).__name__}")
            if not close(observed, target, 0.0001):
                ok = False
                failures.append(f"INFERENTIAL_VALUE_MISMATCH:{address}")
        checks[criterion_id] = 1.0 if ok else 0.0
        return ok

    location_ok = validate_group("R003", {"n": oracle["n"], "mean": oracle["mean_difference"]})
    dispersion_ok = validate_group("R004", {
        "sd": oracle["sd_difference"], "se": oracle["se"],
        "t": oracle["t"], "df": oracle["df"],
    })
    p_address = metrics.get("p")
    p_formula = workbook[p_address.split("!", 1)[0]][p_address.split("!", 1)[1]].value if p_address else None
    p_numeric_ok = validate_group("R005", {"p": oracle["p_value"]})
    mean_address, sd_address = metrics.get("mean"), metrics.get("sd")
    p_chain = formula_chain_text(workbook, p_address) if p_address else ""
    mean_chain = formula_chain_text(workbook, mean_address) if mean_address else ""
    sd_chain = formula_chain_text(workbook, sd_address) if sd_address else ""
    paired_ranges = (
        [f"DATA!{column}{data_rows[0]}:{column}{data_rows[-1]}" for column in paired_columns]
        if data_rows
        else []
    )
    paired_method_ok = uses_two_sided_t_method(p_chain) and any(value in mean_chain for value in paired_ranges) and any(value in sd_chain for value in paired_ranges)
    if not paired_method_ok:
        failures.append("PAIRED_TWO_SIDED_METHOD_NOT_DEMONSTRATED")
    checks["R005"] = 1.0 if p_numeric_ok and paired_method_ok else 0.0
    ci_ok = validate_group("R006", {"ci_lower": oracle["ci_lower"], "ci_upper": oracle["ci_upper"]})
    planned_address = metrics.get("planned_n")
    planned_formula = formula_chain_text(workbook, planned_address) if planned_address else ""
    planned_ok = validate_group("R007", {"planned_n": oracle["planned_n"]})
    planned_method_ok = any(function in planned_formula for function in ("ROUNDUP(", "CEILING(", "CEILING.MATH(", "CEILING.PRECISE("))
    checks["R007"] = 1.0 if planned_ok and planned_method_ok else 0.0

    chart_binding = chart_binding_details(workbook, engine, oracle, data_columns, data_rows)
    checks["R008"] = 1.0 if chart_binding else 0.0
    if not chart_binding:
        failures.append("CHART_NOT_BOUND_TO_ALL_PAIRED_DIFFERENCES")

    perturbations = {row["name"]: row for row in contract["perturbations"]}

    def perturbation_matches(case, require_decision=False):
        replay = FormulaEngine(workbook, case["overrides"])
        for address, target in {**case["expected"], **case["invariants"]}.items():
            sheet, cell = address.split("!", 1)
            observed = replay.value(sheet, cell)
            if target is None:
                if observed is not None:
                    return False
            elif isinstance(target, (int, float)):
                if not close(observed, target, 0.0001):
                    return False
            elif observed != target:
                return False
        if require_decision:
            if not decision_formula_ok(workbook, p_address, metrics.get("alpha")):
                return False
            if not chart_binding:
                return False
        return True

    try:
        observation_case = perturbations["paired_observation"]
        perturbed_group_2 = list(oracle_module.GROUP_2)
        perturb_address, perturb_value = next(iter(observation_case["overrides"].items()))
        perturb_row = int(re.search(r"(\d+)$", perturb_address).group(1))
        perturb_index = perturb_row - 4
        candidate_row = data_rows[perturb_index]
        candidate_perturb_address = f"Data!{data_columns.get('group_2', 'C')}{candidate_row}"
        candidate_overrides = {candidate_perturb_address: perturb_value}
        perturbed_group_2[perturb_index] = perturb_value
        perturbed_oracle = oracle_module.recompute(group_2=perturbed_group_2)
        dynamic_engine = FormulaEngine(workbook, candidate_overrides)
        expected_difference = perturbed_group_2[perturb_index] - oracle_module.GROUP_1[perturb_index]
        chart_row = chart_binding.get("row_map", {}).get(candidate_row) if chart_binding else None
        chart_cell = f"{chart_binding['column']}{chart_row}" if chart_binding and chart_row else None
        dynamic_ok = bool(
            p_address and chart_binding
            and close(dynamic_engine.value(*p_address.split("!", 1)), perturbed_oracle["p_value"], 0.0001)
            and close(dynamic_engine.value(chart_binding["sheet"], chart_cell), expected_difference, 0.0001)
        )
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        dynamic_ok = False
        failures.append(f"MEASUREMENT_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R009"] = 1.0 if dynamic_ok else 0.0

    decision_ok = decision_formula_ok(workbook, p_address, metrics.get("alpha"))
    checks["R010"] = 1.0 if decision_ok else 0.0
    if not decision_ok:
        failures.append("DECISION_RULE_NOT_LINKED_TO_TWO_SIDED_P_VALUE")

    baseline = openpyxl.load_workbook(TASK_ROOT / contract["baseline_workbook"], data_only=False, read_only=False)
    source_ok = all(
        workbook["Data"][f"{column}{row}"].value == baseline["Data"][f"{column}{row}"].value
        for row in range(4, 14) for column in ("B", "C")
    )
    checks["R011"] = 1.0 if source_ok else 0.0
    if not source_ok:
        failures.append("SOURCE_PAIRED_OBSERVATIONS_CHANGED")
    return checks, failures


def v3_token(value):
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).split())


def v3_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    if isinstance(value, str) and not value.startswith("="):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            text = text[:-1]
            scale = 0.01
        else:
            scale = 1.0
        try:
            return float(text) * scale
        except ValueError:
            return None
    return None


def v3_address(sheet, cell):
    return f"{sheet.title}!{cell.coordinate}"


def v3_cell(workbook, address):
    sheet, coordinate = address.split("!", 1)
    return workbook[sheet][coordinate]


def v3_value(workbook, engine, address, failures):
    raw = v3_cell(workbook, address).value
    if isinstance(raw, str) and raw.startswith("="):
        try:
            sheet, coordinate = address.split("!", 1)
            raw = engine.value(sheet, coordinate)
        except Exception as exc:
            failures.append(f"FORMULA_REPLAY_FAILED:{address}:{type(exc).__name__}")
            return None
    return v3_number(raw)


def v3_header_role(label):
    text = v3_token(label)
    if text in {"subject", "participant", "pair", "subject id", "participant id"}:
        return "subject"
    if "difference" in text and (
        "paired" in text
        or "group" in text
        or "treatment" in text
        or "g2" in text
        or "t2" in text
    ):
        return "difference"
    if re.search(r"(?:group|treatment|condition) 1\b", text):
        return "group_1"
    if re.search(r"(?:group|treatment|condition) 2\b", text):
        return "group_2"
    return None


def v3_find_data_table(workbook):
    candidates = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row_number in range(1, sheet.max_row + 1):
            roles = {}
            duplicates = set()
            for cell in sheet[row_number]:
                role = v3_header_role(cell.value)
                if not role:
                    continue
                if role in roles:
                    duplicates.add(role)
                roles[role] = col_name(cell.column)
            if duplicates or not {"subject", "group_1", "group_2", "difference"} <= set(roles):
                continue
            data_rows = []
            for data_row in range(row_number + 1, sheet.max_row + 1):
                subject = sheet[f"{roles['subject']}{data_row}"].value
                first = v3_number(sheet[f"{roles['group_1']}{data_row}"].value)
                second = v3_number(sheet[f"{roles['group_2']}{data_row}"].value)
                if subject not in (None, "") and first is not None and second is not None:
                    data_rows.append(data_row)
                elif data_rows:
                    break
            if data_rows:
                candidates.append((abs(len(data_rows) - 10), sheet_index, row_number, sheet.title, roles, data_rows))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[:3])
    return {
        "sheet": candidates[0][3],
        "columns": candidates[0][4],
        "rows": candidates[0][5],
    }


def v3_metric_match(role, label):
    text = v3_token(label)
    if role == "n":
        return (
            text in {"n", "paired n", "sample size n", "paired sample size", "complete pairs n"}
            or text.startswith("sample size n")
        ) and not any(word in text for word in ("required", "planned", "future"))
    if role == "mean":
        return "mean difference" in text or "mean paired difference" in text
    if role == "sd":
        return any(phrase in text for phrase in ("sd of difference", "sd of paired", "sd paired", "standard deviation", "std deviation", "std dev"))
    if role == "se":
        return any(phrase in text for phrase in ("standard error", "std error", "se of mean", "se of difference"))
    if role == "t":
        return "t statistic" in text or "test statistic t" in text
    if role == "df":
        return "degrees of freedom" in text or text == "df"
    if role == "p":
        return "p value" in text
    if role == "ci_lower":
        return text in {"lower bound", "lower limit", "ci lower bound", "95 ci lower"} or "ci lower" in text
    if role == "ci_upper":
        return text in {"upper bound", "upper limit", "ci upper bound", "95 ci upper"} or "ci upper" in text
    if role == "planned_n":
        return (
            "required sample size" in text
            or "required n" in text
            or "planned n" in text
            or "required complete paired subjects" in text
        )
    return False


def v3_paired_summary_column(sheet, label_cell):
    label = v3_token(label_cell.value)
    generic = (
        label in {"mean", "std dev s", "standard deviation", "se of mean", "sample size n"}
        or label.startswith("sample size n")
    )
    if not generic:
        return None
    for header_row in range(max(1, label_cell.row - 6), label_cell.row):
        for header in sheet[header_row]:
            if "paired difference" in v3_token(header.value):
                candidate = sheet.cell(label_cell.row, header.column)
                if candidate.value not in (None, ""):
                    return candidate
    return None


def v3_find_metric(workbook, role):
    candidates = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row in sheet.iter_rows():
            for label_cell in row:
                if not v3_metric_match(role, label_cell.value):
                    continue
                paired_cell = v3_paired_summary_column(sheet, label_cell)
                if paired_cell is not None:
                    candidates.append((0, 0, sheet_index, label_cell.row, paired_cell.column, v3_address(sheet, paired_cell)))
                label_text = v3_token(label_cell.value)
                specificity = 0 if any(word in label_text for word in ("difference", "paired", "required", "planned")) else 1
                for offset in (1, 2, 3):
                    column = label_cell.column + offset
                    if column > sheet.max_column:
                        continue
                    cell = sheet.cell(label_cell.row, column)
                    raw = cell.value
                    if v3_number(raw) is None and not (isinstance(raw, str) and raw.startswith("=")):
                        continue
                    candidates.append((specificity, offset, sheet_index, label_cell.row, column, v3_address(sheet, cell)))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][-1]


def v3_sheet_key(value):
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def v3_address_key(value):
    if "!" not in value:
        return v3_sheet_key(value)
    sheet, coordinate = value.split("!", 1)
    return f"{v3_sheet_key(sheet)}!{coordinate.replace('$', '').upper()}"


def v3_series_ranges(series):
    category = (
        getattr(getattr(getattr(series, "cat", None), "strRef", None), "f", "")
        or getattr(getattr(getattr(series, "cat", None), "numRef", None), "f", "")
        or getattr(getattr(getattr(series, "xVal", None), "numRef", None), "f", "")
    )
    values = (
        getattr(getattr(getattr(series, "val", None), "numRef", None), "f", "")
        or getattr(getattr(getattr(series, "yVal", None), "numRef", None), "f", "")
    )
    return parse_chart_range(category), parse_chart_range(values)


def v3_chart_binding(workbook, engine, data, expected_differences):
    data_sheet = data["sheet"]
    data_rows = data["rows"]
    columns = data["columns"]
    direct_keys = {
        v3_address_key(f"{data_sheet}!{columns['difference']}{row}")
        for row in data_rows
    }
    raw_pair_keys = [
        {
            v3_address_key(f"{data_sheet}!{columns['group_1']}{row}"),
            v3_address_key(f"{data_sheet}!{columns['group_2']}{row}"),
        }
        for row in data_rows
    ]
    for sheet in workbook.worksheets:
        for chart in getattr(sheet, "_charts", []):
            for series in chart.series:
                category, values = v3_series_ranges(series)
                if not category or not values:
                    continue
                value_sheet, value_column, value_start, value_end = values
                if value_end - value_start + 1 != len(data_rows):
                    continue
                if category[3] - category[2] + 1 != len(data_rows):
                    continue
                value_addresses = [f"{value_sheet}!{value_column}{row}" for row in range(value_start, value_end + 1)]
                if {
                    v3_address_key(item) for item in value_addresses
                } == direct_keys:
                    actual_sheet_name = next(
                        (name for name in workbook.sheetnames if v3_sheet_key(name) == v3_sheet_key(value_sheet)),
                        None,
                    )
                    if actual_sheet_name is not None:
                        value_addresses = [
                            f"{actual_sheet_name}!{item.split('!', 1)[1]}" for item in value_addresses
                        ]
                        return {"value_addresses": value_addresses, "mode": "direct"}
                helper_ok = True
                for index, helper_address in enumerate(value_addresses):
                    helper_sheet_name = next(
                        (name for name in workbook.sheetnames if v3_sheet_key(name) == v3_sheet_key(value_sheet)),
                        None,
                    )
                    if helper_sheet_name is None:
                        helper_ok = False
                        break
                    actual_address = f"{helper_sheet_name}!{helper_address.split('!', 1)[1]}"
                    if not formula_present(workbook, actual_address):
                        helper_ok = False
                        break
                    dependencies = {
                        v3_address_key(item)
                        for item in formula_dependency_closure(workbook, actual_address)
                    }
                    linked = v3_address_key(f"{data_sheet}!{columns['difference']}{data_rows[index]}") in dependencies
                    linked = linked or raw_pair_keys[index] <= dependencies
                    try:
                        observed = engine.value(helper_sheet_name, helper_address.split("!", 1)[1])
                    except Exception:
                        helper_ok = False
                        break
                    if not linked or not close(observed, expected_differences[index], 0.0001):
                        helper_ok = False
                        break
                    value_addresses[index] = actual_address
                if helper_ok:
                    return {"value_addresses": value_addresses, "mode": "helper"}
    return None


def v3_decision_ok(workbook, engine, reject_null):
    candidates = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                raw = cell.value
                rendered = raw
                if isinstance(raw, str) and raw.startswith("="):
                    try:
                        rendered = engine.value(sheet.title, cell.coordinate)
                    except Exception:
                        rendered = raw
                text = v3_token(rendered)
                raw_text = v3_token(raw)
                if any(token in text or token in raw_text for token in ("reject", "statistically significant", "significant difference")):
                    candidates.append(text)
                if text in {"yes", "no"}:
                    left = v3_token(sheet.cell(cell.row, max(1, cell.column - 1)).value)
                    if "significant" in left:
                        candidates.append("reject" if text == "yes" else "fail to reject")
    non_rejection_phrases = (
        "fail to reject", "do not reject", "not significant", "no significant"
    )
    rejection = [
        text for text in candidates
        if (
            "reject" in text
            or "statistically significant" in text
            or "significant difference" in text
        )
        and not any(phrase in text for phrase in non_rejection_phrases)
    ]
    non_rejection = [
        text for text in candidates
        if any(phrase in text for phrase in non_rejection_phrases)
    ]
    # A workbook that states both conclusions has not supplied one reviewable
    # decision, even if one of the two sentences happens to be correct.
    if rejection and non_rejection:
        return False
    return bool(rejection) if reject_null else bool(non_rejection)


def v3_task_checks(workbook, oracle, oracle_module, contract, data):
    checks, failures = {}, []
    engine = FormulaEngine(workbook)
    sheet = workbook[data["sheet"]]
    columns, rows = data["columns"], data["rows"]
    expected_differences = oracle.get("full_differences", oracle["differences"])

    row_count_ok = len(rows) == len(expected_differences)
    observed_differences = []
    formulas_linked = True
    for index, row in enumerate(rows):
        diff_address = f"{data['sheet']}!{columns['difference']}{row}"
        observed = v3_value(workbook, engine, diff_address, failures)
        observed_differences.append(observed)
        formulas_linked = formulas_linked and formula_present(workbook, diff_address)
        if index >= len(expected_differences) or not close(observed, expected_differences[index], 0.0001):
            row_count_ok = False
            failures.append(f"PAIRED_ROW_MISMATCH:{diff_address}")
    checks["R002"] = 1.0 if row_count_ok else 0.0

    metric_roles = ("n", "mean", "sd", "se", "t", "df", "p", "ci_lower", "ci_upper", "planned_n")
    metrics = {role: v3_find_metric(workbook, role) for role in metric_roles}

    def group_ok(criterion, expected):
        ok = True
        for role, target in expected.items():
            address = metrics.get(role)
            if not address:
                ok = False
                failures.append(f"METRIC_NOT_FOUND:{role}")
                continue
            observed = v3_value(workbook, engine, address, failures)
            tolerance = 0.0001 if role not in {"n", "df", "planned_n"} else 1e-9
            if not close(observed, target, tolerance):
                ok = False
                failures.append(f"INFERENTIAL_VALUE_MISMATCH:{role}:{address}")
        checks[criterion] = 1.0 if ok else 0.0
        return ok

    group_ok("R003", {"n": oracle["n"], "mean": oracle["mean_difference"]})
    group_ok("R004", {"sd": oracle["sd_difference"], "se": oracle["se"], "t": oracle["t"], "df": oracle["df"]})
    p_ok = group_ok("R005", {"p": oracle["p_value"]})
    all_text = " ".join(v3_token(cell.value) for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row)
    paired_method = "paired" in all_text and any(token in all_text for token in ("two sided", "two tailed"))
    if not paired_method:
        failures.append("PAIRED_TWO_SIDED_METHOD_NOT_DEMONSTRATED")
    checks["R005"] = 1.0 if p_ok and paired_method else 0.0
    group_ok("R006", {"ci_lower": oracle["ci_lower"], "ci_upper": oracle["ci_upper"]})
    group_ok("R007", {"planned_n": oracle["planned_n"]})

    chart_binding = v3_chart_binding(workbook, engine, data, expected_differences)
    checks["R008"] = 1.0 if chart_binding else 0.0
    if not chart_binding:
        failures.append("CHART_NOT_BOUND_TO_PAIRED_DIFFERENCES")

    dynamic_ok = False
    if chart_binding and rows:
        try:
            index = min(5, len(rows) - 1)
            row = rows[index]
            group_2_address = f"{data['sheet']}!{columns['group_2']}{row}"
            group_1_address = f"{data['sheet']}!{columns['group_1']}{row}"
            original_second = v3_number(v3_cell(workbook, group_2_address).value)
            original_first = v3_number(v3_cell(workbook, group_1_address).value)
            perturbed_second = original_second + 0.5
            replay = FormulaEngine(workbook, {group_2_address: perturbed_second})
            difference_address = f"{data['sheet']}!{columns['difference']}{row}"
            difference_sheet, difference_cell = difference_address.split("!", 1)
            chart_address = chart_binding["value_addresses"][index]
            chart_sheet, chart_cell = chart_address.split("!", 1)
            target = perturbed_second - original_first
            dynamic_ok = (
                formulas_linked
                and close(replay.value(difference_sheet, difference_cell), target, 0.0001)
                and close(replay.value(chart_sheet, chart_cell), target, 0.0001)
            )
        except Exception as exc:
            failures.append(f"PAIRED_CHART_RECALCULATION_FAILED:{type(exc).__name__}")
    checks["R009"] = 1.0 if dynamic_ok else 0.0

    checks["R010"] = 1.0 if v3_decision_ok(workbook, engine, oracle["p_value"] < 0.05) else 0.0
    if not checks["R010"]:
        failures.append("DECISION_NOT_CONSISTENT_WITH_P_VALUE")

    expected_subjects = contract.get("expected_subject_ids", [])

    def subject_key(value):
        if value in (None, "") or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            return str(int(value)) if float(value).is_integer() else str(float(value))
        return v3_token(value)

    observed_subjects = [
        subject_key(sheet[f"{columns['subject']}{row}"].value) for row in rows
    ]
    expected_subject_keys = [subject_key(value) for value in expected_subjects]
    subject_identity_ok = bool(expected_subject_keys) and (
        observed_subjects == expected_subject_keys
        and None not in observed_subjects
        and len(set(observed_subjects)) == len(observed_subjects)
    )
    source_ok = subject_identity_ok and len(rows) == len(oracle_module.GROUP_1) and all(
        close(v3_number(sheet[f"{columns['group_1']}{row}"].value), oracle_module.GROUP_1[index], 1e-9)
        and close(v3_number(sheet[f"{columns['group_2']}{row}"].value), oracle_module.GROUP_2[index], 1e-9)
        for index, row in enumerate(rows)
    )
    checks["R011"] = 1.0 if source_ok else 0.0
    if not source_ok:
        failures.append("SOURCE_PAIR_IDENTITIES_OR_OBSERVATIONS_CHANGED")
    return checks, failures


def evaluate(candidate):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"]
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    contract = load_contract()
    data = v3_find_data_table(workbook)
    if data is None:
        return criteria, ["PAIRED_DATA_TABLE_NOT_FOUND"]
    try:
        oracle, oracle_module = load_oracle(contract)
        checks, task_failures = v3_task_checks(workbook, oracle, oracle_module, contract, data)
        criteria["R001"] = 1.0
        criteria.update(checks)
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))



def main():
    candidate = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/output/answer.xlsx")
    criteria, failures = evaluate(candidate)
    payload = build_result(task=TASK, split=ACTIVE_SPLIT, candidate=str(candidate), criteria=criteria, failures=failures)
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True)
        (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\\n")
        if total is not None:
            (log_root / "reward.txt").write_text(str(total) + "\\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

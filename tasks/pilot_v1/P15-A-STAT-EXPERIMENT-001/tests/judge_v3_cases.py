#!/usr/bin/env python3
"""One focused V3 validation pass for the paired-analysis Judge."""
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
import openpyxl
from openpyxl.chart import BarChart, Reference

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tests"))
import evaluate  # noqa: E402
from judge_v2_support import build_result  # noqa: E402


def build_alternate(path):
    oracle, module = evaluate.load_oracle(evaluate.load_contract())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Paired Study"
    sheet.append(["Subject", "Treatment 1", "Treatment 2", "Paired difference (Treatment 2 - Treatment 1)", "", "Paired two-sided t-test", "Value"])
    for index, (first, second) in enumerate(zip(module.GROUP_1, module.GROUP_2), 2):
        sheet.append([index - 1, first, second, f"=C{index}-B{index}"])
    metrics = [
        (2, "Paired sample size", "=COUNT(D2:D11)"),
        (3, "Mean difference", "=AVERAGE(D2:D11)"),
        (4, "SD of difference", "=STDEV.S(D2:D11)"),
        (5, "SE of difference", "=G4/SQRT(G2)"),
        (6, "t statistic", "=G3/G5"),
        (7, "Degrees of freedom", "=G2-1"),
        (8, "Two-sided p-value", "=T.DIST.2T(ABS(G6),G7)"),
        (9, "95% CI lower", "=G3-T.INV.2T(0.05,G7)*G5"),
        (10, "95% CI upper", "=G3+T.INV.2T(0.05,G7)*G5"),
        (11, "Required sample size", oracle["planned_n"]),
        (12, "Decision (two-sided alpha 0.05)", '=IF(G8<0.05,"Reject null; statistically significant paired difference","Fail to reject null")'),
    ]
    for row, label, value in metrics:
        sheet[f"F{row}"] = label
        sheet[f"G{row}"] = value
    chart = BarChart()
    chart.title = "Paired differences by subject"
    chart.add_data(Reference(sheet, min_col=4, min_row=1, max_row=11), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=11))
    sheet.add_chart(chart, "I2")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def result(path):
    criteria, failures = evaluate.evaluate(path)
    payload = build_result(task=evaluate.TASK, split="dev", candidate=str(path), criteria=criteria, failures=failures)
    return {
        "score": payload["normalized_score"], "pass": payload["pass"],
        "evaluation_status": payload["evaluation_status"],
        "criterion_scores": payload["criterion_scores"], "failure_codes": payload["failure_codes"],
    }


def confirm_reference_result():
    process = subprocess.run(
        [
            sys.executable,
            str(ROOT / "tests/evaluate.py"),
            str(ROOT / "tests/confirm/reference.xlsx"),
            "--split",
            "confirm",
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    if process.returncode != 0:
        raise RuntimeError(process.stderr.strip() or "confirm Evaluator returned no diagnostic output")
    payload = json.loads(process.stdout)
    return {
        "score": payload["normalized_score"],
        "pass": payload["pass"],
        "evaluation_status": payload["evaluation_status"],
        "criterion_scores": payload["criterion_scores"],
        "failure_codes": payload["failure_codes"],
    }


def build_subject_identity_mutant(path):
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    data = evaluate.v3_find_data_table(workbook)
    sheet = workbook[data["sheet"]]
    subject_column = data["columns"]["subject"]
    first, second = data["rows"][:2]
    sheet[f"{subject_column}{first}"] = sheet[f"{subject_column}{second}"].value
    workbook.save(path)


def build_contradictory_decision_mutant(path):
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    sheet = workbook.create_sheet("Contradictory Decision")
    sheet["A1"] = "Fail to reject the null; the paired difference is not significant."
    workbook.save(path)


def main():
    alternate = ROOT / "fixtures/equivalent/candidate_v3_one_sheet_layout.xlsx"
    build_alternate(alternate)
    cases = {
        "standard": (ROOT / "solution/reference.xlsx", True, "Reference paired analysis must pass."),
        "equivalent": (ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", True, "Equivalent paired-analysis formulas must pass."),
        "alternate_layout": (alternate, True, "One clearly labeled sheet satisfies the analysis and chart requirements."),
        "noop": (ROOT / "fixtures/noop/candidate_noop.xlsx", False, "An uncompleted analysis workbook does not pass."),
        "malformed": (ROOT / "fixtures/malformed/candidate_malformed.xlsx", False, "A non-workbook cannot pass."),
    }
    for path in sorted((ROOT / "fixtures/mutants").glob("*.xlsx")):
        if path.stem == "independent_groups_error":
            cases[f"mutant:{path.stem}"] = (
                path, True,
                "This seed changes mean(differences) to mean(Group 2)-mean(Group 1), which is algebraically identical; the remaining SD, SE, and test stay paired, so it is an equivalent implementation rather than a true error.",
            )
        else:
            cases[f"mutant:{path.stem}"] = (path, False, "The seeded statistical-method defect must prevent a pass.")
    observed = {}
    with tempfile.TemporaryDirectory(prefix="p15-stat-v3-") as directory:
        subject_mutant = Path(directory) / "duplicate_subject_id.xlsx"
        decision_mutant = Path(directory) / "contradictory_decisions.xlsx"
        build_subject_identity_mutant(subject_mutant)
        build_contradictory_decision_mutant(decision_mutant)
        cases["mutant:duplicate_subject_identity"] = (
            subject_mutant, False,
            "Duplicate or changed subject IDs break the frozen pairing identity even when the numeric columns are unchanged.",
        )
        cases["mutant:contradictory_decisions"] = (
            decision_mutant, False,
            "Opposite inferential conclusions cannot jointly satisfy the decision summary.",
        )
        for name, (path, expected_pass, basis) in cases.items():
            row = result(path)
            row.update({"expected_pass": expected_pass, "expectation_basis": basis, "meets_expectation": row["pass"] is expected_pass})
            observed[name] = row
        confirm_row = confirm_reference_result()
        confirm_row.update(
            {
                "expected_pass": True,
                "expectation_basis": "The held-out batch-level reference uses valid baseline/intervention labels and must pass without renaming its columns.",
                "meets_expectation": confirm_row["pass"] is True and confirm_row["score"] == 1.0,
            }
        )
        observed["confirm_reference"] = confirm_row
    receipt = {
        "task_id": evaluate.TASK["task_id"], "judge_version": "v3-semantic",
        "generated_at": datetime.now(timezone.utc).isoformat(), "cases": observed,
        "all_expectations_met": all(row["meets_expectation"] for row in observed.values()),
    }
    destination = ROOT / "receipts/judge_v3_local_validation.json"
    destination.write_text(json.dumps(receipt, indent=2, sort_keys=True) + "\n")
    print(json.dumps(receipt, indent=2, sort_keys=True))
    if not receipt["all_expectations_met"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

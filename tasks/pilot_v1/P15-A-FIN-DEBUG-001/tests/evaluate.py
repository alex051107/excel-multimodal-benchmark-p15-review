#!/usr/bin/env python3
"""Deterministic semantic judge for P15-A-FIN-DEBUG-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-A-FIN-DEBUG-001",
  "pass_threshold": 0.7,
  "required_sheets": [
    "Assumptions",
    "Revenue_Build",
    "Working_Capital",
    "Debt_Schedule",
    "Summary",
    "Checks"
  ],
  "protected": {
    "Assumptions!B4": 12000,
    "Assumptions!C4": 12900,
    "Assumptions!D4": 13700,
    "Assumptions!B5": 48,
    "Assumptions!C5": 49,
    "Assumptions!D5": 50
  },
  "formula_cells": [
    "Revenue_Build!D6",
    "Working_Capital!D7",
    "Debt_Schedule!D5",
    "Summary!D8"
  ],
  "criteria": [
    {"id":"R001","description":"The workbook opens and retains every required model sheet.","weight":1,"type":"positive","dimension":"file_usability","method":"deterministic","method_params":{"implemented_check":"required_sheets"}},
    {"id":"R002","description":"The repaired 2027 revenue formula produces the independent root-cause result.","weight":4,"type":"positive","dimension":"root_cause_debugging","method":"deterministic","method_params":{"implemented_check":"root_revenue","oracle":"split_contract"}},
    {"id":"R003","description":"The 2027 accounts-receivable result reconciles to the independent operating-model replay.","weight":2,"type":"positive","dimension":"working_capital","method":"deterministic","method_params":{"implemented_check":"accounts_receivable","oracle":"split_contract"}},
    {"id":"R004","description":"The 2027 interest expense reconciles to the independent debt-schedule replay.","weight":2,"type":"positive","dimension":"debt_schedule","method":"deterministic","method_params":{"implemented_check":"interest_expense","oracle":"split_contract"}},
    {"id":"R005","description":"The 2027 free-cash-flow output reconciles to the independent operating-model replay.","weight":3,"type":"positive","dimension":"free_cash_flow","method":"deterministic","method_params":{"implemented_check":"free_cash_flow","oracle":"split_contract"}},
    {"id":"R006","description":"The repaired revenue and downstream cash-flow chain respond correctly to the declared 2027 price perturbation.","weight":2,"type":"positive","dimension":"native_recalculation","method":"deterministic","method_params":{"implemented_check":"price_perturbation","split_weights":{"dev":2,"confirm":3}}},
    {"id":"R007","description":"The summary free-cash-flow bridge ties exactly to contribution, working capital, and interest.","weight":2,"type":"positive","dimension":"fcf_bridge","method":"deterministic","method_params":{"implemented_check":"fcf_bridge"}},
    {"id":"R008","description":"The 2027 debt-service coverage ratio agrees with the independent replay.","weight":2,"type":"positive","dimension":"debt_service_coverage","method":"deterministic","method_params":{"implemented_check":"debt_service_coverage","oracle":"split_contract"}},
    {"id":"R009","description":"Only the declared root-cause formula cell differs from the starting workbook.","weight":3,"type":"positive","dimension":"repair_locality","method":"deterministic","method_params":{"implemented_check":"repair_locality"}},
    {"id":"R010","description":"Every declared assumption cell remains unchanged.","weight":2,"type":"positive","dimension":"assumption_protection","method":"deterministic","method_params":{"implemented_check":"assumption_protection"}},
    {"id":"R011","description":"Every declared downstream model output remains formula-linked.","weight":2,"type":"positive","dimension":"formula_integrity","method":"deterministic","method_params":{"implemented_check":"required_formulas"}},
    {"id":"R012","description":"The receivables and free-cash-flow chain responds correctly to the distinct 2027 DSO perturbation while revenue, contribution, and interest remain invariant.","weight":1,"type":"positive","dimension":"working_capital_recalculation","method":"deterministic","method_params":{"implemented_check":"dso_perturbation","applies_to":["dev"]}}
  ]
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R002", "R006", "R010"]
SHEET_ALIASES = {}


def cell_key(sheet, cell):
    return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column:
        value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


TASK_ROOT = Path(__file__).resolve().parents[1]


def requested_split():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    for index, argument in enumerate(sys.argv[1:]):
        if argument == "--split" and index + 2 <= len(sys.argv) - 1:
            split = sys.argv[index + 2].strip().lower()
        elif argument.startswith("--split="):
            split = argument.split("=", 1)[1].strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"unsupported split: {split}")
    return split


ACTIVE_SPLIT = requested_split()


def load_contract():
    relative = "tests/confirm/contract.json" if ACTIVE_SPLIT == "confirm" else "tests/private_contract.json"
    return json.loads((TASK_ROOT / relative).read_text())


def load_oracle(contract):
    oracle_path = TASK_ROOT / contract["oracle"]
    source_path = TASK_ROOT / "metadata/oracle_recompute.py"
    if ACTIVE_SPLIT == "dev" and source_path.exists() and source_path.read_bytes() != oracle_path.read_bytes():
        raise RuntimeError("DEV_ORACLE_COPY_OUT_OF_SYNC")
    spec = importlib.util.spec_from_file_location(f"task_oracle_{ACTIVE_SPLIT}", oracle_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute(), module


def _betacf(a, b, x):
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d, h = 1.0, 1.0 - qab * x / qap, 0.0
    if abs(d) < 3e-14: d = 3e-14
    d = 1.0 / d
    h = d
    for m in range(1, 201):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; delta = d * c; h *= delta
        if abs(delta - 1.0) < 3e-12: break
    return h


def _betai(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    factor = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x))
    return factor * _betacf(a, b, x) / a if x < (a + 1) / (a + b + 2) else 1 - factor * _betacf(b, a, 1 - x) / b


def student_t_two_tail(value, df):
    x = df / (df + value * value)
    return _betai(df / 2, 0.5, x)


def student_t_inv_two_tail(alpha, df):
    lo, hi = 0.0, 20.0
    for _ in range(90):
        mid = (lo + hi) / 2
        if student_t_two_tail(mid, df) > alpha: lo = mid
        else: hi = mid
    return (lo + hi) / 2


class UnsupportedFormulaError(RuntimeError):
    """Valid Excel syntax that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook
        self.overrides = overrides or {}
        self.memo = {}
        self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames:
            raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            if isinstance(raw, str) and raw.startswith("="):
                result = self.formula(raw[1:], sheet)
            else:
                result = raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        values = []
        for r in range(int(r1), int(r2) + 1):
            for c in range(col_index(c1), col_index(c2) + 1):
                values.append(self.value(sheet, f"{col_name(c)}{r}"))
        return values

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        expression = re.sub(r"\bSTDEV\.S\b", "STDEV_S", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.2T\b", "T_DIST_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.RT\b", "T_DIST_RT", expression, flags=re.I)
        expression = re.sub(r"\bT\.INV\.2T\b", "T_INV_2T", expression, flags=re.I)
        expression = re.sub(r"\bPI\(\)", "PI()", expression, flags=re.I)

        def replace_range(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))

        def sub_outside_strings(pattern, replacement, value):
            parts = re.split(r'("(?:[^"\\\\]|\\\\.)*")', value)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in parts)

        expression = sub_outside_strings(RANGE, replace_range, expression)

        def replace_ref(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))

        expression = sub_outside_strings(REF, replace_ref, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(
                f"UNSUPPORTED_FORMULA_SYNTAX:{expression}"
            ) from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name) and node.id.upper() in {"TRUE", "FALSE"}:
            return node.id.upper() == "TRUE"
        if isinstance(node, ast.List):
            return [self.safe_eval(v) for v in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand)
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operators = {ast.Add: lambda a,b:a+b, ast.Sub: lambda a,b:a-b, ast.Mult: lambda a,b:a*b, ast.Div: lambda a,b:a/b, ast.Pow: lambda a,b:a**b}
            for cls, fun in operators.items():
                if isinstance(node.op, cls): return fun(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for op, comp in zip(node.ops, node.comparators):
                right = self.safe_eval(comp)
                good = (left == right if isinstance(op, ast.Eq) else left != right if isinstance(op, ast.NotEq) else left >= right if isinstance(op, ast.GtE) else left <= right if isinstance(op, ast.LtE) else left > right if isinstance(op, ast.Gt) else left < right if isinstance(op, ast.Lt) else False)
                if not good: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                if self.safe_eval(node.args[0]):
                    return self.safe_eval(node.args[1])
                return self.safe_eval(node.args[2]) if len(node.args) == 3 else False
            args = [self.safe_eval(a) for a in node.args]
            values = [v for group in args for v in (group if isinstance(group, list) else [group]) if v is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            if name == "SUM": return sum(numeric_values)
            if name == "AVERAGE": return statistics.mean(numeric_values)
            if name == "COUNT": return len(numeric_values)
            if name == "STDEV_S": return statistics.stdev(numeric_values)
            if name == "SQRT": return math.sqrt(args[0])
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "PI": return math.pi
            if name == "AND": return all(args)
            if name == "OR": return any(args)
            if name == "NOT": return not args[0]
            if name == "POWER": return args[0] ** args[1]
            if name == "PRODUCT": return math.prod(numeric_values)
            if name == "MAX": return max(numeric_values)
            if name == "MIN": return min(numeric_values)
            if name == "NA": return None
            if name == "T_DIST_2T":
                return student_t_two_tail(args[0], args[1])
            if name == "T_DIST_RT":
                return student_t_two_tail(args[0], args[1]) / 2
            if name == "T_INV_2T":
                return student_t_inv_two_tail(args[0], args[1])
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def repair_locality(workbook, contract):
    baseline_path = TASK_ROOT / contract["baseline_workbook"]
    baseline = openpyxl.load_workbook(baseline_path, data_only=False, read_only=False)
    allowed = set(contract["allowed_changed_formula_cells"])
    differences = []
    for sheet_name in contract["required_sheets"]:
        candidate_sheet, baseline_sheet = workbook[sheet_name], baseline[sheet_name]
        max_row = max(candidate_sheet.max_row, baseline_sheet.max_row)
        max_col = max(candidate_sheet.max_column, baseline_sheet.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_col + 1):
                address = f"{sheet_name}!{candidate_sheet.cell(row, column).coordinate}"
                if address in allowed:
                    continue
                if candidate_sheet.cell(row, column).value != baseline_sheet.cell(row, column).value:
                    differences.append(address)
    return differences


def task_checks(workbook, engine, oracle, contract, oracle_module):
    checks, failures = {}, []
    required_formula_ok = True
    for address in contract["formula_cells"]:
        if not formula_present(workbook, address):
            required_formula_ok = False
            failures.append(f"FORMULA_REQUIRED:{address}")
    checks["R011"] = 1.0 if required_formula_ok else 0.0

    def check_value(criterion_id, address, target):
        sheet, cell = address.split("!", 1)
        try:
            observed = engine.value(sheet, cell)
            tolerance = 0.005 if criterion_id == "R008" else 0.5
            ok = formula_present(workbook, address) and close(observed, target, tolerance)
        except UnsupportedFormulaError:
            raise
        except Exception as exc:
            ok = False
            failures.append(f"MODEL_REPLAY_EVALUATION_FAILED:{address}:{type(exc).__name__}")
        if not ok:
            failures.append(f"MODEL_REPLAY_MISMATCH:{address}")
        checks[criterion_id] = 1.0 if ok else 0.0
        return ok

    root_ok = check_value("R002", "Revenue_Build!D6", oracle["revenue_2027"])
    ar_ok = check_value("R003", "Working_Capital!D5", oracle["ar_2027"])
    interest_ok = check_value("R004", "Debt_Schedule!D5", oracle["interest_2027"])
    fcf_ok = check_value("R005", "Summary!D8", oracle["fcf_2027"])
    dscr_ok = check_value("R008", "Summary!D9", oracle["dscr_2027"])

    perturbations = {row["name"]: row for row in contract["perturbations"]}

    def perturbation_matches(case):
        replay = FormulaEngine(workbook, case["overrides"])
        for address, target in {**case["expected"], **case["invariants"]}.items():
            sheet, cell = address.split("!", 1)
            observed = replay.value(sheet, cell)
            if target is None:
                if observed is not None:
                    return False
            elif isinstance(target, (int, float)):
                if not close(observed, target, 0.01):
                    return False
            elif observed != target:
                return False
        return True

    price_case = perturbations["price_2027"]
    try:
        if "expected" in price_case:
            dynamic_ok = perturbation_matches(price_case)
        else:
            price = price_case["overrides"]["Assumptions!D5"]
            dynamic = FormulaEngine(workbook, price_case["overrides"]).value("Revenue_Build", "D6")
            dynamic_ok = close(dynamic, oracle_module.ASSUMPTIONS["volume"][-1] * price, 0.01)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        dynamic_ok = False
        failures.append(f"PRICE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R006"] = 1.0 if dynamic_ok else 0.0

    dso_ok = True
    checks["R012"] = 0.0
    if ACTIVE_SPLIT == "dev":
        try:
            dso_ok = perturbation_matches(perturbations["dso_2027"])
        except UnsupportedFormulaError:
            raise
        except Exception as exc:
            dso_ok = False
            failures.append(f"DSO_PERTURBATION_FAILED:{type(exc).__name__}")
        checks["R012"] = 1.0 if dso_ok else 0.0

    try:
        fcf_bridge = engine.value("Summary", "D8") - (engine.value("Summary", "D5") + engine.value("Summary", "D6") - engine.value("Summary", "D7"))
        bridge_ok = close(fcf_bridge, 0.0, 0.01)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        bridge_ok = False
        failures.append(f"FCF_BRIDGE_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if bridge_ok else 0.0

    protected_ok = all(workbook[sheet][cell].value == value for key, value in contract["protected"].items() for sheet, cell in [key.split("!", 1)])
    checks["R010"] = 1.0 if protected_ok else 0.0
    collateral_changes = repair_locality(workbook, contract)
    locality_ok = not collateral_changes
    if collateral_changes:
        failures.append("COLLATERAL_EDIT:" + ",".join(collateral_changes[:12]))
    checks["R009"] = 1.0 if locality_ok else 0.0
    return checks, failures


def evaluate(candidate):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"]
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    contract = load_contract()
    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in contract["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, contract["required_sheets"], SHEET_ALIASES
    )
    criteria["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        return criteria, sorted(set(failures))
    try:
        oracle, oracle_module = load_oracle(contract)
        checks, task_failures = task_checks(workbook, FormulaEngine(workbook), oracle, contract, oracle_module)
        criteria.update(checks)
        failures.extend(task_failures)
    except UnsupportedFormulaError as exc:
        failures.append(f"UNSUPPORTED_FORMULA:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))



def main():
    candidate = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/output/answer.xlsx")
    criteria, failures = evaluate(candidate)
    payload = build_result(task=TASK, split=ACTIVE_SPLIT, candidate=str(candidate), criteria=criteria, failures=failures)
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True)
        (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\\n")
        if total is not None:
            (log_root / "reward.txt").write_text(str(total) + "\\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

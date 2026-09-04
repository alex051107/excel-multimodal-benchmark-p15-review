#!/usr/bin/env python3
"""Focused Judge V3 regression: equivalent layout plus real business errors."""
from __future__ import annotations

import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

import evaluate as judge

ROOT = Path(__file__).resolve().parents[1]


def run_case(path: Path, split: str = "dev") -> dict:
    criteria, failures = judge.evaluate(path, split)
    payload = judge.build_result(
        task=judge.TASK, split=split, candidate=str(path), criteria=criteria, failures=failures
    )
    return {
        "score": payload["normalized_score"], "pass": payload["pass"],
        "criteria": payload["criterion_scores"], "failure_codes": payload["failure_codes"],
    }


def alternative_layout(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Selected_Sources"].title = "Source Selection"
    workbook.save(path)


def registry_without_selection(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    sheet = workbook["Selected_Sources"]
    for cell in sheet[4]:
        cell.value = None
    workbook.save(path)


def selection_without_explanation(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Selected_Sources"]["E4"] = None
    workbook.save(path)


def main() -> None:
    expected_mutants = {"hardcoded_recognized_revenue": True}
    cases = {
        "reference": (ROOT / "solution/reference.xlsx", "dev", True),
        "confirm_reference": (ROOT / "tests/confirm/reference.xlsx", "confirm", True),
        "equivalent": (ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", "dev", True),
        "noop": (ROOT / "fixtures/noop/candidate_noop.xlsx", "dev", False),
        "malformed": (ROOT / "fixtures/malformed/candidate_malformed.xlsx", "dev", False),
    }
    cases.update({
        f"mutant:{path.stem}": (path, "dev", expected_mutants.get(path.stem, False))
        for path in sorted((ROOT / "fixtures/mutants").glob("*.xlsx"))
    })
    results = {}
    with tempfile.TemporaryDirectory() as directory:
        alternative = Path(directory) / "source_selection_with_spaces.xlsx"
        registry_only = Path(directory) / "registry_is_not_selection.xlsx"
        missing_reason = Path(directory) / "selection_without_explanation.xlsx"
        alternative_layout(alternative)
        registry_without_selection(registry_only)
        selection_without_explanation(missing_reason)
        cases["alternative_layout:source_selection_with_spaces"] = (alternative, "dev", True)
        cases["negative:registry_is_not_selection"] = (registry_only, "dev", False)
        cases["negative:selection_without_explanation"] = (missing_reason, "dev", False)
        for name, (path, split, expected_pass) in cases.items():
            result = run_case(path, split)
            assert result["pass"] is expected_pass, (name, result)
            results[name] = {**result, "expected_pass": expected_pass, "meets_expectation": True}
    receipt = {
        "task_id": judge.TASK["task_id"], "judge_version": "P15_JUDGE_V3",
        "generated_at": datetime.now(timezone.utc).isoformat(), "cases": results,
        "all_expectations_met": True,
    }
    (ROOT / "receipts/judge_v3_local_validation.json").write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n"
    )
    print(json.dumps(receipt, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

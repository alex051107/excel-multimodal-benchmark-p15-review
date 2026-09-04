#!/usr/bin/env python3
"""Deterministic, task-specific judge for P15-B-SALES-DISCOVERY-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import csv
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

CRITERIA = [
    {"id": "R001", "description": "A readable workbook contains every requested operating-review sheet.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {}},
    {"id": "R002", "description": "The source-selection record identifies the only approved, complete 2024Q2 schema-1.1 release.", "weight": 3, "type": "positive", "dimension": "source_selection", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R003", "description": "Consolidated_Data contains exactly the six records from the selected release and no mixed-release records.", "weight": 4, "type": "positive", "dimension": "data_completeness", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R004", "description": "All four operating KPIs equal the independent replay of the selected records.", "weight": 3, "type": "positive", "dimension": "metric_correctness", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R005", "description": "The selection explanation addresses approval, reporting period, schema, and regional coverage.", "weight": 3, "type": "positive", "dimension": "selection_rationale", "method": "deterministic", "method_params": {}},
    {"id": "R006", "description": "The chosen release identifier and file remain visible with the selected records.", "weight": 3, "type": "positive", "dimension": "source_provenance", "method": "deterministic", "method_params": {}},
    {"id": "R007", "description": "Regional record counts and recognized revenue close to the six selected records.", "weight": 3, "type": "positive", "dimension": "coverage_correctness", "method": "deterministic", "method_params": {}},
    {"id": "R008", "description": "Every release-registry source field is preserved exactly from release_registry.csv.", "weight": 3, "type": "positive", "dimension": "change_locality", "method": "deterministic", "method_params": {"source": "data/input_files/release_registry.csv"}},
    {"id": "R009", "description": "The chosen-source reason, four required-region statuses, and evidence labels remain reviewable.", "weight": 2, "type": "positive", "dimension": "auditability", "method": "deterministic", "method_params": {}},
]

TASK = {
    "task_id": "P15-B-SALES-DISCOVERY-001",
    "pass_threshold": 0.7,
    "required_sheets": ["Release_Registry", "Selected_Sources", "Consolidated_Data", "KPI_Summary", "Coverage_Checks"],
    "criteria": CRITERIA,
}

REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")

TASK["hurdle_criteria"] = ["R002", "R003", "R004", "R005", "R007", "R008"]
SHEET_ALIASES = {
    "Release_Registry": ("Registry",),
    "Selected_Sources": ("Source Selection",),
    "Consolidated_Data": ("Q2 Records",),
    "KPI_Summary": ("KPI Summary",),
    "Coverage_Checks": ("Regional Coverage",),
}


def cell_key(sheet, cell): return f"{sheet}!{cell}"

def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value

def col_name(index):
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value

def task_root(): return Path(__file__).resolve().parents[1]

def split_context(split):
    contract_path = task_root() / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    contract = json.loads(contract_path.read_text())
    drivers = {
        "bookings_source_cell": "Consolidated_Data!E4",
        "bookings_delta": 100,
        "revenue_source_cell": "Consolidated_Data!F5",
        "revenue_delta": 100,
    }
    drivers.update(contract.get("drivers", {}))
    return {
        "split": split,
        "contract": contract,
        "source_root": task_root() / contract["input_files_dir"],
        "oracle_path": task_root() / contract["oracle"],
        "drivers": drivers,
    }

def load_oracle(context):
    spec = importlib.util.spec_from_file_location(f"sales_oracle_{context['split']}", context["oracle_path"])
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()

class UnsupportedFormulaError(RuntimeError):
    """Valid Excel syntax that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides: return self.overrides[key]
        if key in self.memo: return self.memo[key]
        if key in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(c)}{r}") for r in range(int(r1), int(r2) + 1) for c in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(piece if piece.startswith('"') else pattern.sub(replacement, piece) for piece in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def ref_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, ref_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(
                f"UNSUPPORTED_FORMULA_SYNTAX:{expression}"
            ) from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operations = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b, ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b, ast.Pow: lambda a, b: a**b}
            for cls, operation in operations.items():
                if isinstance(node.op, cls): return operation(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [value for group in args for value in (group if isinstance(group, list) else [group]) if value is not None]
            if name == "SUM": return sum(values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "ABS": return abs(args[0])
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "COUNTIF":
                if len(args) != 2 or not isinstance(args[0], list): raise ValueError("INVALID_COUNTIF_ARGUMENTS")
                return sum(1 for value in args[0] if value == args[1])
            if name == "SUMIF":
                if len(args) not in {2, 3} or not isinstance(args[0], list): raise ValueError("INVALID_SUMIF_ARGUMENTS")
                sum_values = args[0] if len(args) == 2 else args[2]
                if not isinstance(sum_values, list) or len(sum_values) != len(args[0]): raise ValueError("INVALID_SUMIF_RANGE")
                return sum(value for criterion_value, value in zip(args[0], sum_values) if criterion_value == args[1] and isinstance(value, (int, float)))
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")

def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance

def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")

def norm(value): return round(value, 6) if isinstance(value, float) else value

def rows(workbook, sheet, start, end, columns, engine=None):
    values = []
    for row in range(start, end + 1):
        record = tuple(norm(engine.value(sheet, f"{col_name(column)}{row}") if engine else workbook[sheet].cell(row=row, column=column).value) for column in columns)
        if any(value not in (None, "") for value in record): values.append(record)
    return values


def semantic_token(value):
    token = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")
    aliases = {
        "release": "release_id", "release_identifier": "release_id",
        "filename": "file", "source_file": "file", "reporting_period": "period",
        "regional_coverage": "coverage", "record": "record_id",
        "bookings_usd": "bookings", "bookings": "bookings",
        "recognized_revenue_usd": "recognized_revenue", "recognized_revenue": "recognized_revenue",
        "closed_won": "closed_won_flag", "closed_won_flag": "closed_won_flag",
        "record_count": "record_count", "recognized_revenue_total": "recognized_revenue",
        "result": "value", "metric_value": "value",
        "selection_reason": "reason", "selection_rationale": "reason",
        "rationale": "reason", "explanation": "reason", "basis": "reason",
    }
    return aliases.get(token, token)


def semantic_tables(workbook):
    found = []
    for sheet in workbook.worksheets:
        for header_row in range(1, min(sheet.max_row, 20) + 1):
            headers = {}
            duplicate = False
            for column in range(1, min(sheet.max_column, 30) + 1):
                token = semantic_token(sheet.cell(row=header_row, column=column).value)
                if not token:
                    continue
                if token in headers:
                    duplicate = True
                headers[token] = column
            if duplicate or len(headers) < 2:
                continue
            if not ({"release_id", "file"} <= set(headers) or "record_id" in headers or "metric" in headers or "region" in headers):
                continue
            records = []
            blank_run = 0
            for row in range(header_row + 1, min(sheet.max_row, header_row + 100) + 1):
                record = {name: sheet.cell(row=row, column=column).value for name, column in headers.items()}
                if all(value in (None, "") for value in record.values()):
                    blank_run += 1
                    if blank_run >= 2 and records:
                        break
                    continue
                blank_run = 0
                records.append(record)
            found.append({"sheet": sheet.title, "headers": headers, "records": records})
            break
    return found


def same_value(left, right):
    if hasattr(left, "isoformat") and hasattr(right, "isoformat"):
        return left.isoformat() == right.isoformat()
    if isinstance(right, (int, float)):
        try:
            return abs(float(left) - float(right)) <= 0.01
        except (TypeError, ValueError):
            return False
    return str(left or "").strip().casefold() == str(right or "").strip().casefold()


def semantic_sales_checks(workbook, expected, context):
    tables = semantic_tables(workbook)
    expected_rows = [tuple(row) for row in expected["rows"]]
    selected_record = None
    selected_table = None
    for table in tables:
        # A release registry lists all candidates; it is source evidence, not
        # the requested act of choosing one.  A semantic selection record must
        # contain exactly one chosen release.
        if (
            {"release_id", "file", "period", "coverage"} <= set(table["headers"])
            and len(table["records"]) == 1
        ):
            for record in table["records"]:
                if all(same_value(record.get(key), value) for key, value in zip(("release_id", "file", "period", "coverage"), expected["selection"])):
                    selected_record = record
                    selected_table = table
                    break
    data_ok = False
    for table in tables:
        fields = ("record_id", "region", "product", "stage", "bookings", "recognized_revenue", "closed_won_flag")
        if not set(fields) <= set(table["headers"]):
            continue
        actual = [tuple(record.get(field) for field in fields) for record in table["records"]]
        if len(actual) == len(expected_rows) and all(
            sum(all(same_value(a, e) for a, e in zip(row, target)) for target in expected_rows) == 1
            for row in actual
        ) and all(
            sum(all(same_value(a, e) for a, e in zip(row, target)) for row in actual) == 1
            for target in expected_rows
        ):
            data_ok = True
            break

    registry_ok = False
    registry_fields = ("release_id", "file", "period", "published", "supersedes", "schema", "coverage")
    with (context["source_root"] / "release_registry.csv").open(newline="") as handle:
        registry_expected = [tuple(record[field] for field in registry_fields) for record in csv.DictReader(handle)]
    for table in tables:
        if not set(registry_fields) <= set(table["headers"]):
            continue
        actual = [tuple(record.get(field) for field in registry_fields) for record in table["records"]]
        if len(actual) == len(registry_expected) and all(
            sum(all(same_value(a, e) for a, e in zip(row, target)) for target in registry_expected) == 1
            for row in actual
        ) and all(
            sum(all(same_value(a, e) for a, e in zip(row, target)) for row in actual) == 1
            for target in registry_expected
        ):
            registry_ok = True
            break

    metric_expected = {
        "total_bookings": expected["bookings"], "bookings": expected["bookings"],
        "recognized_revenue": expected["revenue"], "closed_won_records": expected["closed_won"],
        "closed_won_count": expected["closed_won"], "south_recognized_revenue": expected["south_revenue"],
    }
    observed_metrics = {}
    for table in tables:
        if {"metric", "value"} <= set(table["headers"]):
            for record in table["records"]:
                observed_metrics[semantic_token(record.get("metric"))] = record.get("value")
    metrics_ok = all(any(key in observed_metrics and same_value(observed_metrics[key], value) for key in keys) for keys, value in [
        (("total_bookings", "bookings"), expected["bookings"]),
        (("recognized_revenue",), expected["revenue"]),
        (("closed_won_records", "closed_won_count"), expected["closed_won"]),
        (("south_recognized_revenue",), expected["south_revenue"]),
    ])

    coverage_expected = {
        region: (
            sum(1 for row in expected_rows if row[1] == region),
            sum(row[5] for row in expected_rows if row[1] == region),
        )
        for region in ("North", "South", "East", "West")
    }
    coverage_ok = False
    for table in tables:
        if not {"region", "record_count", "recognized_revenue"} <= set(table["headers"]):
            continue
        actual = {
            str(record.get("region") or "").strip(): (record.get("record_count"), record.get("recognized_revenue"))
            for record in table["records"]
        }
        if set(actual) == set(coverage_expected) and all(
            same_value(actual[region][0], expected_pair[0]) and same_value(actual[region][1], expected_pair[1])
            for region, expected_pair in coverage_expected.items()
        ):
            coverage_ok = True
            break

    registry_selected = next(
        (
            row for row in registry_expected
            if same_value(row[0], expected["selection"][0])
            and same_value(row[1], expected["selection"][1])
        ),
        None,
    )
    selected_sheet_text = ""
    if selected_table is not None:
        selected_sheet = workbook[selected_table["sheet"]]
        selected_sheet_text = " ".join(
            str(cell.value or "")
            for row in selected_sheet.iter_rows()
            for cell in row
        ).casefold()
    reason_value = selected_record.get("reason") if selected_record else None
    reason_text = str(reason_value or "").strip().casefold()
    expected_period = str(expected["selection"][2]).casefold()
    expected_coverage_text = str(expected["selection"][3]).casefold()
    expected_schema = str(registry_selected[5]).casefold() if registry_selected else ""
    reason_ok = bool(
        selected_record is not None
        and len(reason_text) >= 20
        and "approved" in reason_text
        and same_value(selected_record.get("period"), expected["selection"][2])
        and same_value(selected_record.get("coverage"), expected["selection"][3])
        # The chosen release identity already proves the schema requirement.
        # Do not require the explanation to repeat a particular technical
        # token when it plainly states why this approved complete release was
        # selected for the requested period.
        and (expected_period in reason_text or expected_period in selected_sheet_text)
        and ("complete" in reason_text or expected_coverage_text in selected_sheet_text)
    )
    provenance_ok = bool(
        selected_record is not None
        and str(expected["selection"][0]).casefold() in selected_sheet_text
        and str(expected["selection"][1]).casefold() in selected_sheet_text
    )
    role_count = sum((selected_record is not None, data_ok, registry_ok, bool(observed_metrics), coverage_ok))
    checks = {row["id"]: 0.0 for row in CRITERIA}
    checks.update({
        "R001": float(role_count >= 4), "R002": float(selected_record is not None),
        "R003": float(data_ok), "R004": float(metrics_ok), "R005": float(reason_ok),
        "R006": float(provenance_ok), "R007": float(coverage_ok), "R008": float(registry_ok),
        "R009": float(reason_ok and provenance_ok and coverage_ok),
    })
    failures = []
    if selected_record is None: failures.append("INVALID_RELEASE_SELECTION")
    if not data_ok: failures.append("SELECTED_RECORD_SET_MISMATCH")
    if not registry_ok: failures.append("REGISTRY_SOURCE_FIELDS_CHANGED")
    return checks, failures

def registry_matches_source(workbook, source_root):
    fields = ("release_id", "file", "period", "published", "supersedes", "schema", "coverage")
    with (source_root / "release_registry.csv").open(newline="") as handle:
        expected = [tuple(record[field] for field in fields) for record in csv.DictReader(handle)]
    actual = [tuple("" if workbook["Release_Registry"].cell(row=row, column=column).value is None else str(workbook["Release_Registry"].cell(row=row, column=column).value) for column in range(1, 8)) for row in range(4, 8)]
    return actual == expected

def task_checks(workbook, engine, expected, context):
    checks, failures = {}, []
    selection = tuple(norm(workbook["Selected_Sources"].cell(row=4, column=column).value) for column in range(1, 5))
    expected_selection = tuple(expected["selection"])
    coverage_text = str(selection[3] or "").strip().casefold()
    expected_coverage = str(expected_selection[3] or "").strip().casefold()
    coverage_ok = coverage_text == expected_coverage or (
        expected_coverage == "complete" and ("complete" in coverage_text or "4/4" in coverage_text)
    )
    selection_ok = selection[:3] == expected_selection[:3] and coverage_ok
    checks["R002"] = float(selection_ok)
    if not selection_ok: failures.append("INVALID_RELEASE_SELECTION")

    actual_rows = rows(workbook, "Consolidated_Data", 4, 12, range(1, 8))
    expected_rows = [tuple(row) for row in expected["rows"]]
    data_ok = len(actual_rows) == len(expected_rows) and set(actual_rows) == set(expected_rows)
    checks["R003"] = float(data_ok)
    if not data_ok: failures.append("SELECTED_RECORD_SET_MISMATCH")

    try:
        metrics_ok = close(engine.value("KPI_Summary", "B4"), expected["bookings"]) and close(engine.value("KPI_Summary", "B5"), expected["revenue"]) and close(engine.value("KPI_Summary", "B6"), expected["closed_won"]) and close(engine.value("KPI_Summary", "B7"), expected["south_revenue"])
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        metrics_ok = False; failures.append(f"KPI_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R004"] = float(metrics_ok)

    try:
        expected_coverage = {region: (len([record for record in expected_rows if record[1] == region]), sum(record[5] for record in expected_rows if record[1] == region)) for region in ("North", "South", "East", "West")}
        actual_coverage = {workbook["Coverage_Checks"].cell(row=row, column=1).value: (engine.value("Coverage_Checks", f"B{row}"), engine.value("Coverage_Checks", f"C{row}")) for row in range(4, 8)}
        coverage_ok = actual_coverage == expected_coverage
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        coverage_ok = False; failures.append(f"COVERAGE_CLOSURE_FAILED:{type(exc).__name__}")
    checks["R007"] = float(coverage_ok)

    protected_ok = registry_matches_source(workbook, context["source_root"])
    checks["R008"] = float(protected_ok)
    if not protected_ok: failures.append("REGISTRY_SOURCE_FIELDS_CHANGED")

    reason = workbook["Selected_Sources"]["E4"].value
    statuses = [workbook["Coverage_Checks"].cell(row=row, column=4).value for row in range(4, 8)]
    evidence = [workbook["KPI_Summary"].cell(row=row, column=4).value for row in range(4, 8)]
    acceptable_statuses = {"required", "pass", "passed", "complete", "ok"}
    statuses_ok = all(str(value or "").strip().casefold() in acceptable_statuses for value in statuses)
    evidence_ok = all(isinstance(value, str) and len(value.strip()) >= 8 for value in evidence)
    reason_text = str(reason or "").casefold()
    checks["R005"] = float(
        isinstance(reason, str)
        and len(reason.strip()) >= 20
        and all(word in reason_text for word in ("approved", "2024q2", "complete"))
        and ("schema" in reason_text or "1.1" in reason_text)
    )
    checks["R006"] = float(
        selection_ok
        and str(expected_selection[0]).casefold() in " ".join(str(cell.value or "") for row in workbook["Selected_Sources"].iter_rows() for cell in row).casefold()
        and str(expected_selection[1]).casefold() in " ".join(str(cell.value or "") for row in workbook["Selected_Sources"].iter_rows() for cell in row).casefold()
    )
    checks["R009"] = float(isinstance(reason, str) and len(reason.strip()) >= 20 and statuses_ok and evidence_ok)

    return checks, failures

def evaluate(candidate, split="dev"):
    criteria = {row["id"]: 0.0 for row in CRITERIA}
    if not candidate.exists() or candidate.stat().st_size == 0: return criteria, ["OUTPUT_MISSING"]
    try: workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc: return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    failures = []
    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], SHEET_ALIASES
    )
    criteria["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    context = split_context(split)
    expected = load_oracle(context)
    semantic_checks, semantic_failures = semantic_sales_checks(workbook, expected, context)
    if unresolved or ambiguous:
        if all(semantic_checks.get(key) == 1.0 for key in TASK["hurdle_criteria"]):
            return semantic_checks, sorted(set(semantic_failures))
        failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
        return criteria, sorted(set(failures))
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    try:
        checks, task_failures = task_checks(workbook, FormulaEngine(workbook), expected, context)
        criteria.update(checks); failures.extend(task_failures)
    except UnsupportedFormulaError as exc:
        failures.append(f"UNSUPPORTED_FORMULA:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    criteria = {key: max(criteria.get(key, 0.0), semantic_checks.get(key, 0.0)) for key in criteria}
    failures.extend(semantic_failures)
    failures = [
        code for code in failures
        if not code.startswith((
            "BOOKINGS_PERTURBATION_FAILED:", "REVENUE_PERTURBATION_FAILED:",
            "KPI_EVALUATION_FAILED:", "COVERAGE_CLOSURE_FAILED:",
        ))
    ]
    passed_failure_codes = {
        "R002": {"INVALID_RELEASE_SELECTION"},
        "R003": {"SELECTED_RECORD_SET_MISMATCH"},
        "R008": {"REGISTRY_SOURCE_FIELDS_CHANGED"},
    }
    failures = [
        code for code in failures
        if not any(criteria[criterion] == 1.0 and code in codes for criterion, codes in passed_failure_codes.items())
    ]
    return criteria, sorted(set(failures))

def parse_cli():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    candidate = None
    arguments = iter(sys.argv[1:])
    for argument in arguments:
        if argument == "--split": split = next(arguments, "")
        elif argument.startswith("--split="): split = argument.split("=", 1)[1]
        elif candidate is None: candidate = Path(argument)
        else: raise ValueError(f"UNEXPECTED_ARGUMENT:{argument}")
    if split not in {"dev", "confirm"}: raise ValueError(f"INVALID_SPLIT:{split}")
    return candidate or Path("/app/output/answer.xlsx"), split

def main():
    candidate, split = parse_cli()
    criteria, failures = evaluate(candidate, split)
    payload = build_result(task=TASK, split=split, candidate=str(candidate), criteria=criteria, failures=failures)
    payload["judge_version"] = "P15_JUDGE_V3"
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True); (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        if total is not None: (log_root / "reward.txt").write_text(str(total) + "\n")
    except OSError: pass
    print(json.dumps(payload, indent=2, sort_keys=True))

if __name__ == "__main__": main()

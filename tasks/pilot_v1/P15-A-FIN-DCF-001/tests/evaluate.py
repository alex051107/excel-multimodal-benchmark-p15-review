#!/usr/bin/env python3
"""Deterministic semantic judge for P15-A-FIN-DCF-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-A-FIN-DCF-001",
  "pass_threshold": 0.7,
  "required_sheets": [
    "Source_Data",
    "Assumptions",
    "Forecast",
    "Valuation",
    "Checks"
  ],
  "protected": {
    "Source_Data!D4": 391035,
    "Source_Data!D5": 123216,
    "Source_Data!D6": 11445,
    "Source_Data!D7": 9447,
    "Source_Data!D8": 29943
  },
  "formula_cells": [
    "Forecast!B5",
    "Forecast!F12",
    "Valuation!B7",
    "Valuation!B9",
    "Valuation!B13"
  ],
  "criteria": [
    {"id": "R001", "description": "The workbook opens and contains every required DCF sheet.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {"implemented_check": "required_sheets"}},
    {"id": "R002", "description": "Each forecast-period revenue formula agrees with the independent replay.", "weight": 3, "type": "positive", "dimension": "revenue_forecast", "method": "deterministic", "method_params": {"implemented_check": "forecast_revenue", "oracle": "split_contract"}},
    {"id": "R003", "description": "The EBIT-to-NOPAT bridge agrees with the independent replay in every forecast period.", "weight": 3, "type": "positive", "dimension": "operating_bridge", "method": "deterministic", "method_params": {"implemented_check": "forecast_operating_bridge", "oracle": "split_contract"}},
    {"id": "R004", "description": "The D&A, capex, NWC, and UFCF bridge agrees with the independent replay in every forecast period.", "weight": 3, "type": "positive", "dimension": "fcf_bridge", "method": "deterministic", "method_params": {"implemented_check": "forecast_fcf_bridge", "oracle": "split_contract"}},
    {"id": "R005", "description": "Forecast discount factors and present values agree with the independent replay.", "weight": 2, "type": "positive", "dimension": "discounting", "method": "deterministic", "method_params": {"implemented_check": "forecast_discount_pv", "oracle": "split_contract"}},
    {"id": "R006", "description": "The terminal FCF, terminal value, and discounted terminal value agree with the independent replay.", "weight": 3, "type": "positive", "dimension": "terminal_value", "method": "deterministic", "method_params": {"implemented_check": "terminal_value", "oracle": "split_contract"}},
    {"id": "R007", "description": "The forecast-PV, enterprise-value, net-debt, and equity-value bridge agrees with the independent replay.", "weight": 3, "type": "positive", "dimension": "valuation_bridge", "method": "deterministic", "method_params": {"implemented_check": "enterprise_equity_bridge", "oracle": "split_contract"}},
    {"id": "R008", "description": "Every declared WACC-by-terminal-growth sensitivity coordinate is formula-linked and correct.", "weight": 3, "type": "positive", "dimension": "sensitivity", "method": "deterministic", "method_params": {"implemented_check": "sensitivity_grid", "oracle": "split_contract"}},
    {"id": "R009", "description": "The terminal forecast revenue responds correctly to the declared growth perturbation.", "weight": 2, "type": "positive", "dimension": "growth_recalculation", "method": "deterministic", "method_params": {"implemented_check": "growth_perturbation"}},
    {"id": "R010", "description": "Enterprise value responds correctly to the declared WACC perturbation.", "weight": 2, "type": "positive", "dimension": "wacc_recalculation", "method": "deterministic", "method_params": {"implemented_check": "wacc_perturbation"}},
    {"id": "R011", "description": "The declared historical source cells remain unchanged.", "weight": 2, "type": "positive", "dimension": "change_locality", "method": "deterministic", "method_params": {"implemented_check": "historical_source_protection"}}
  ]
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R007", "R009", "R010", "R011"]
SHEET_ALIASES = {}


def cell_key(sheet, cell):
    return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column:
        value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


TASK_ROOT = Path(__file__).resolve().parents[1]


def requested_split():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    for index, argument in enumerate(sys.argv[1:]):
        if argument == "--split" and index + 2 <= len(sys.argv) - 1:
            split = sys.argv[index + 2].strip().lower()
        elif argument.startswith("--split="):
            split = argument.split("=", 1)[1].strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"unsupported split: {split}")
    return split


ACTIVE_SPLIT = requested_split()


def load_contract():
    relative = "tests/confirm/contract.json" if ACTIVE_SPLIT == "confirm" else "tests/private_contract.json"
    return json.loads((TASK_ROOT / relative).read_text())


def load_oracle(contract):
    oracle_path = TASK_ROOT / contract["oracle"]
    source_path = TASK_ROOT / "metadata/oracle_recompute.py"
    if ACTIVE_SPLIT == "dev" and source_path.exists() and source_path.read_bytes() != oracle_path.read_bytes():
        raise RuntimeError("DEV_ORACLE_COPY_OUT_OF_SYNC")
    spec = importlib.util.spec_from_file_location(f"task_oracle_{ACTIVE_SPLIT}", oracle_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute(), module


def _betacf(a, b, x):
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d, h = 1.0, 1.0 - qab * x / qap, 0.0
    if abs(d) < 3e-14: d = 3e-14
    d = 1.0 / d
    h = d
    for m in range(1, 201):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; delta = d * c; h *= delta
        if abs(delta - 1.0) < 3e-12: break
    return h


def _betai(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    factor = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x))
    return factor * _betacf(a, b, x) / a if x < (a + 1) / (a + b + 2) else 1 - factor * _betacf(b, a, 1 - x) / b


def student_t_two_tail(value, df):
    x = df / (df + value * value)
    return _betai(df / 2, 0.5, x)


def student_t_inv_two_tail(alpha, df):
    lo, hi = 0.0, 20.0
    for _ in range(90):
        mid = (lo + hi) / 2
        if student_t_two_tail(mid, df) > alpha: lo = mid
        else: hi = mid
    return (lo + hi) / 2


class UnsupportedFormulaError(ValueError):
    """The formula is valid Excel syntax outside this bounded replay engine."""


def excel_number(value):
    """Apply the narrow numeric-text coercion used by the observed DCF grids."""
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    stripped = value.strip().replace(",", "")
    percent = stripped.endswith("%")
    if percent:
        stripped = stripped[:-1].strip()
    try:
        parsed = float(stripped)
    except ValueError:
        return value
    return parsed / 100 if percent else parsed


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook
        self.overrides = overrides or {}
        self.memo = {}
        self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames:
            raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            if isinstance(raw, str) and raw.startswith("="):
                result = self.formula(raw[1:], sheet)
            else:
                result = raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        values = []
        for r in range(int(r1), int(r2) + 1):
            for c in range(col_index(c1), col_index(c2) + 1):
                values.append(self.value(sheet, f"{col_name(c)}{r}"))
        return values

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        expression = re.sub(r"\bSTDEV\.S\b", "STDEV_S", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.2T\b", "T_DIST_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.RT\b", "T_DIST_RT", expression, flags=re.I)
        expression = re.sub(r"\bT\.INV\.2T\b", "T_INV_2T", expression, flags=re.I)
        expression = re.sub(r"\bPI\(\)", "PI()", expression, flags=re.I)

        def replace_range(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))

        def sub_outside_strings(pattern, replacement, value):
            parts = re.split(r'("(?:[^"\\\\]|\\\\.)*")', value)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in parts)

        expression = sub_outside_strings(RANGE, replace_range, expression)

        def replace_ref(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))

        expression = sub_outside_strings(REF, replace_ref, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        tree = ast.parse(expression, mode="eval")
        return self.safe_eval(tree.body)

    def safe_eval(self, node, variables=None):
        variables = {} if variables is None else variables
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name):
            if node.id.upper() in {"TRUE", "FALSE"}:
                return node.id.upper() == "TRUE"
            variable = node.id.casefold()
            if variable in variables:
                return variables[variable]
        if isinstance(node, ast.List):
            return [self.safe_eval(v, variables) for v in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = excel_number(self.safe_eval(node.operand, variables))
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left = self.safe_eval(node.left, variables)
            right = self.safe_eval(node.right, variables)
            operators = {ast.Add: lambda a,b:a+b, ast.Sub: lambda a,b:a-b, ast.Mult: lambda a,b:a*b, ast.Div: lambda a,b:a/b, ast.Pow: lambda a,b:a**b}
            for cls, fun in operators.items():
                if isinstance(node.op, cls):
                    if isinstance(left, list) or isinstance(right, list):
                        left_values = left if isinstance(left, list) else [left] * len(right)
                        right_values = right if isinstance(right, list) else [right] * len(left)
                        if len(left_values) != len(right_values):
                            raise ValueError("ARRAY_LENGTH_MISMATCH")
                        return [fun(excel_number(a), excel_number(b)) for a, b in zip(left_values, right_values)]
                    return fun(excel_number(left), excel_number(right))
            raise ValueError("UNSUPPORTED_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left, variables)
            for op, comp in zip(node.ops, node.comparators):
                right = self.safe_eval(comp, variables)
                good = (left == right if isinstance(op, ast.Eq) else left != right if isinstance(op, ast.NotEq) else left >= right if isinstance(op, ast.GtE) else left <= right if isinstance(op, ast.LtE) else left > right if isinstance(op, ast.Gt) else left < right if isinstance(op, ast.Lt) else False)
                if not good: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "LET":
                if len(node.args) < 3 or len(node.args) % 2 == 0:
                    raise ValueError("LET_REQUIRES_NAME_VALUE_PAIRS_AND_RESULT")
                scope = dict(variables)
                for index in range(0, len(node.args) - 1, 2):
                    name_node = node.args[index]
                    if not isinstance(name_node, ast.Name):
                        raise ValueError("LET_NAME_MUST_BE_IDENTIFIER")
                    scope[name_node.id.casefold()] = self.safe_eval(node.args[index + 1], scope)
                return self.safe_eval(node.args[-1], scope)
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                if self.safe_eval(node.args[0], variables):
                    return self.safe_eval(node.args[1], variables)
                return self.safe_eval(node.args[2], variables) if len(node.args) == 3 else False
            args = [self.safe_eval(a, variables) for a in node.args]
            if name == "SUMPRODUCT":
                # Excel accepts SUMPRODUCT with one array and returns its sum.
                # This is a common, materially equivalent DCF formulation after
                # the element-wise discounting has already happened inside the
                # argument expression.
                if len(args) < 1 or any(not isinstance(arg, list) for arg in args):
                    raise ValueError("SUMPRODUCT_REQUIRES_ARRAYS")
                if len({len(arg) for arg in args}) != 1:
                    raise ValueError("SUMPRODUCT_ARRAY_LENGTH_MISMATCH")
                return sum(math.prod(items) for items in zip(*args))
            values = [v for group in args for v in (group if isinstance(group, list) else [group]) if v is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            if name == "SUM": return sum(numeric_values)
            if name == "AVERAGE": return statistics.mean(numeric_values)
            if name == "COUNT": return len(numeric_values)
            if name == "STDEV_S": return statistics.stdev(numeric_values)
            if name == "SQRT": return math.sqrt(args[0])
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "PI": return math.pi
            if name == "AND": return all(args)
            if name == "OR": return any(args)
            if name == "NOT": return not args[0]
            if name == "POWER": return args[0] ** args[1]
            if name == "PRODUCT": return math.prod(numeric_values)
            if name == "MAX": return max(numeric_values)
            if name == "MIN": return min(numeric_values)
            if name == "NA": return None
            if name == "VALUE":
                if len(args) != 1:
                    raise ValueError("VALUE_REQUIRES_ONE_ARGUMENT")
                converted = excel_number(args[0])
                if not isinstance(converted, (int, float)):
                    raise ValueError("VALUE_CANNOT_PARSE_TEXT")
                return converted
            if name == "SUBSTITUTE":
                if len(args) not in (3, 4):
                    raise ValueError("SUBSTITUTE_ARGUMENTS")
                source, old, new = map(str, args[:3])
                if len(args) == 3:
                    return source.replace(old, new)
                instance = int(args[3])
                if instance < 1:
                    raise ValueError("SUBSTITUTE_INSTANCE")
                matches = list(re.finditer(re.escape(old), source))
                if instance > len(matches):
                    return source
                match = matches[instance - 1]
                return source[:match.start()] + new + source[match.end():]
            if name == "T_DIST_2T":
                return student_t_two_tail(args[0], args[1])
            if name == "T_DIST_RT":
                return student_t_two_tail(args[0], args[1]) / 2
            if name == "T_INV_2T":
                return student_t_inv_two_tail(args[0], args[1])
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def task_checks(workbook, engine, oracle, contract, oracle_module):
    checks, failures = {}, []

    def validate_group(criterion_id, expected, formula_required=None):
        formula_required = set(expected) if formula_required is None else set(formula_required)
        ok = True
        for address, target in expected.items():
            if address in formula_required and not formula_present(workbook, address):
                ok = False
                failures.append(f"DCF_FORMULA_REQUIRED:{address}")
            sheet, cell = address.split("!", 1)
            try:
                observed = engine.value(sheet, cell)
            except UnsupportedFormulaError as exc:
                raise UnsupportedFormulaError(f"{address}:{exc}") from exc
            except Exception as exc:
                observed = None
                failures.append(f"DCF_FORMULA_EVALUATION_FAILED:{address}:{type(exc).__name__}")
            # Discount factors are unitless and displayed to much finer
            # precision than dollar outputs; a fifty-basis-point absolute
            # tolerance would hide a materially wrong discounting model.
            tolerance = 0.000001 if address.startswith("Forecast!") and address[10:11] and int(re.findall(r"\d+", address)[0]) == 14 else 0.5
            if not close(observed, target, tolerance):
                ok = False
                failures.append(f"DCF_VALUE_MISMATCH:{address}")
        checks[criterion_id] = 1.0 if ok else 0.0
        return ok

    def forecast_expected(rows):
        return {
            f"Forecast!{column}{row}": oracle[key][index]
            for row, key in rows.items()
            for index, column in enumerate("BCDEF")
        }

    revenue_ok = validate_group("R002", forecast_expected({5: "revenue"}))
    operating_ok = validate_group("R003", forecast_expected({6: "ebit", 7: "tax", 8: "nopat"}))
    fcf_ok = validate_group("R004", forecast_expected({9: "da", 10: "capex", 11: "nwc", 12: "fcf"}))
    discount_ok = validate_group("R005", forecast_expected({14: "discount", 15: "pv"}))
    terminal_ok = validate_group("R006", {
        "Valuation!B6": oracle["terminal_fcf"],
        "Valuation!B7": oracle["terminal_value"],
        "Valuation!B8": oracle["pv_terminal"],
    })
    bridge_ok = validate_group("R007", {
        "Valuation!B5": oracle["pv_forecast"],
        "Valuation!B9": oracle["enterprise_value"],
        "Valuation!B10": oracle["cash"],
        "Valuation!B11": oracle["debt"],
        "Valuation!B12": oracle["net_debt"],
        "Valuation!B13": oracle["equity_value"],
    }, {"Valuation!B5", "Valuation!B9", "Valuation!B10", "Valuation!B12", "Valuation!B13"})

    grid = contract["sensitivity_grid"]
    sensitivity_expected = {}
    for row_index, wacc in enumerate(grid["wacc_values"], start=grid["row_start"]):
        for column, terminal_growth in zip(grid["columns"], grid["terminal_growth_values"]):
            sensitivity_expected[f"Valuation!{column}{row_index}"] = oracle["sensitivity"][(wacc, terminal_growth)]
    sensitivity_ok = validate_group("R008", sensitivity_expected)

    perturbations = {row["name"]: row for row in contract["perturbations"]}
    growth_case = perturbations["growth"]
    growth_value = growth_case["overrides"]["Assumptions!B4"]
    growth = list(oracle_module.ASSUMPTIONS["growth"])
    growth[0] = growth_value
    try:
        growth_observed = FormulaEngine(workbook, growth_case["overrides"]).value("Forecast", "F5")
        growth_ok = close(growth_observed, oracle_module.recompute(growth=growth)["revenue"][-1], 0.5)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        growth_ok = False
        failures.append(f"DCF_GROWTH_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R009"] = 1.0 if growth_ok else 0.0

    wacc_case = perturbations["wacc"]
    wacc_value = wacc_case["overrides"]["Assumptions!B10"]
    try:
        wacc_observed = FormulaEngine(workbook, wacc_case["overrides"]).value("Valuation", "B9")
        wacc_ok = close(wacc_observed, oracle_module.recompute(wacc=wacc_value)["enterprise_value"], 0.5)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        wacc_ok = False
        failures.append(f"DCF_WACC_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R010"] = 1.0 if wacc_ok else 0.0

    protected_ok = all(
        workbook[sheet][cell].value == value
        for key, value in contract["protected"].items()
        for sheet, cell in [key.split("!", 1)]
    )
    checks["R011"] = 1.0 if protected_ok else 0.0
    if not protected_ok:
        failures.append("CRITICAL_PROTECTED_SOURCE_CHANGED")
    return checks, failures


def evaluate(candidate):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"]
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    contract = load_contract()
    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in contract["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, contract["required_sheets"], SHEET_ALIASES
    )
    criteria["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        return criteria, sorted(set(failures))
    try:
        oracle, oracle_module = load_oracle(contract)
        checks, task_failures = task_checks(workbook, FormulaEngine(workbook), oracle, contract, oracle_module)
        criteria.update(checks)
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))



def main():
    candidate = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/output/answer.xlsx")
    criteria, failures = evaluate(candidate)
    payload = build_result(
        task=TASK,
        split=ACTIVE_SPLIT,
        candidate=str(candidate),
        criteria=criteria,
        failures=failures,
    )
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True)
        (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\\n")
        if total is not None:
            (log_root / "reward.txt").write_text(str(total) + "\\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

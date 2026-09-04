#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-B-OPS-CLEAN-JOIN-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import csv
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

CRITERIA = [
    {"id": "R001", "description": "A readable workbook contains every requested operations sheet.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {}},
    {"id": "R002", "description": "Clean_Data classifies each raw order with the required normalization and duplicate policy.", "weight": 4, "type": "positive", "dimension": "data_cleaning", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R003", "description": "Joined_Data contains exactly the valid one-to-one joins with correct product, region, units, unit cost, and extended cost.", "weight": 4, "type": "positive", "dimension": "join_correctness", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R004", "description": "Each accepted order uses the correct units and product cost to produce its extended cost.", "weight": 3, "type": "positive", "dimension": "cost_arithmetic", "method": "deterministic", "method_params": {}},
    {"id": "R005", "description": "The accepted-order total is $3,940 and excludes duplicates and invalid joins.", "weight": 3, "type": "positive", "dimension": "total_correctness", "method": "deterministic", "method_params": {}},
    {"id": "R006", "description": "A reviewable weekly order-cost view is present for the accepted records.", "weight": 3, "type": "positive", "dimension": "weekly_view", "method": "deterministic", "method_params": {}},
    {"id": "R007", "description": "The four task-specific exceptions are complete and excluded from the accepted-order total.", "weight": 3, "type": "positive", "dimension": "exception_integrity", "method": "deterministic", "method_params": {}},
    {"id": "R008", "description": "Every raw-order, product-master, and location-master source field is preserved exactly.", "weight": 3, "type": "positive", "dimension": "change_locality", "method": "deterministic", "method_params": {}},
    {"id": "R009", "description": "Checks report the accepted count, exception count, and total cost in a reviewable form.", "weight": 2, "type": "positive", "dimension": "auditability", "method": "deterministic", "method_params": {}},
]

TASK = {
    "task_id": "P15-B-OPS-CLEAN-JOIN-001",
    "pass_threshold": 0.7,
    "required_sheets": ["Raw_Orders", "Product_Master", "Location_Master", "Clean_Data", "Joined_Data", "Exceptions", "Summary", "Checks"],
    "criteria": CRITERIA,
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R002", "R003", "R007", "R008"]
SHEET_ALIASES = {
    "Raw_Orders": ("Raw Orders",),
    "Product_Master": ("Product Master",),
    "Location_Master": ("Location Master",),
    "Clean_Data": ("Clean Records", "Normalized Orders"),
    "Joined_Data": ("Joined Valid Records", "Joined_Valid_Orders", "Valid Joined Orders"),
    "Exceptions": ("Exceptions Queue", "Exceptions_Queue"),
    "Checks": ("Data_Quality_Checks",),
}


def cell_key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def task_root(): return Path(__file__).resolve().parents[1]


def split_context(split):
    contract_path = task_root() / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    contract = json.loads(contract_path.read_text())
    drivers = {
        "units_order_id": "ORD-100",
        "raw_units_source_cell": "Raw_Orders!D4",
        "units_new_value": 11,
        "units_delta_cost": 150,
        "cost_order_id": "ORD-101",
        "product_cost_source_cell": "Product_Master!D5",
        "product_cost_new_value": 500,
        "product_cost_delta": 400,
        "summary_region": "Northeast",
    }
    drivers.update(contract.get("drivers", {}))
    return {"split": split, "contract": contract, "source_root": task_root() / contract["input_files_dir"], "oracle_path": task_root() / contract["oracle"], "drivers": drivers}


def load_oracle(context):
    global oracle_module
    spec = importlib.util.spec_from_file_location(f"task_oracle_{context['split']}", context["oracle_path"])
    oracle_module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(oracle_module)
    return oracle_module.recompute()


def excel_equal(left, right):
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, str) and isinstance(right, str):
        return left.casefold() == right.casefold()
    return left == right


def excel_criteria_match(value, criteria):
    operator = "="
    expected = criteria
    if isinstance(criteria, str):
        match = re.match(r"^(<=|>=|<>|!=|=|<|>)(.*)$", criteria)
        if match:
            operator, expected = match.groups()
        stripped = expected.strip()
        try:
            expected = float(stripped)
        except (TypeError, ValueError):
            expected = stripped
    if operator in {"=", "=="}:
        return excel_equal(value, expected)
    if operator in {"<>", "!="}:
        return not excel_equal(value, expected)
    try:
        if operator == "<": return value < expected
        if operator == "<=": return value <= expected
        if operator == ">": return value > expected
        if operator == ">=": return value >= expected
    except TypeError:
        return False
    raise ValueError(f"UNSUPPORTED_CRITERIA_OPERATOR:{operator}")


def as_range(value, function_name):
    if not isinstance(value, list):
        raise ValueError(f"{function_name}_REQUIRES_RANGE")
    return value


class UnsupportedFormulaError(RuntimeError):
    """Valid Excel syntax that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides: return self.overrides[key]
        if key in self.memo: return self.memo[key]
        if key in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
        finally:
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(c)}{r}") for r in range(int(r1), int(r2) + 1) for c in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(piece if piece.startswith('"') else pattern.sub(replacement, piece) for piece in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def ref_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, ref_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise UnsupportedFormulaError(
                f"UNSUPPORTED_FORMULA_SYNTAX:{expression}"
            ) from exc
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operations = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b, ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b, ast.Pow: lambda a, b: a ** b}
            for cls, operation in operations.items():
                if isinstance(node.op, cls): return operation(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [value for group in args for value in (group if isinstance(group, list) else [group]) if value is not None]
            if name == "SUM": return sum(values)
            if name == "SUMIFS":
                if len(args) < 3 or len(args) % 2 == 0:
                    raise ValueError("SUMIFS_ARGUMENTS")
                sum_range = as_range(args[0], name)
                criteria_pairs = [
                    (as_range(args[index], name), args[index + 1])
                    for index in range(1, len(args), 2)
                ]
                if any(len(criteria_range) != len(sum_range) for criteria_range, _ in criteria_pairs):
                    raise ValueError("SUMIFS_RANGE_LENGTH")
                return sum(
                    value
                    for row, value in enumerate(sum_range)
                    if isinstance(value, (int, float))
                    and all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                )
            if name == "COUNTA":
                return sum(value not in (None, "") for value in values)
            if name == "XLOOKUP":
                if len(args) != 3:
                    raise ValueError("XLOOKUP_ARGUMENTS")
                lookup_range = as_range(args[1], name)
                return_range = as_range(args[2], name)
                if len(lookup_range) != len(return_range):
                    raise ValueError("XLOOKUP_RANGE_LENGTH")
                for index, value in enumerate(lookup_range):
                    if excel_equal(value, args[0]):
                        return return_range[index]
                raise ValueError("XLOOKUP_NOT_FOUND")
            if name == "TRIM":
                if len(args) != 1:
                    raise ValueError("TRIM_ARGUMENTS")
                return re.sub(r" +", " ", str(args[0]).strip())
            if name == "UPPER":
                if len(args) != 1:
                    raise ValueError("UPPER_ARGUMENTS")
                return str(args[0]).upper()
            if name == "AVERAGE": return statistics.mean(values)
            if name == "ABS": return abs(args[0])
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def norm(value):
    if isinstance(value, float): return round(value, 6)
    return value


def category_token(value):
    if isinstance(value, bool):
        return "yes" if value else "no"
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().casefold()).strip("_")


def normalize_yes_no(value):
    token = category_token(value)
    return {
        "1": "yes", "true": "yes", "y": "yes", "yes": "yes",
        "0": "no", "false": "no", "n": "no", "no": "no",
    }.get(token, token)


def normalize_disposition(value):
    token = category_token(value)
    return {
        "valid": "join",
        "joined": "join",
        "include": "join",
        "included_in_joined_data": "join",
        "exception_duplicate_excluded": "exception_exact_duplicate",
        "exception_non_numeric_units": "exception_invalid_units",
    }.get(token, token)


def normalize_clean_record(record):
    values = list(record)
    for index in (4, 5, 6):
        values[index] = normalize_yes_no(values[index])
    values[7] = normalize_disposition(values[7])
    return tuple(values)


def rows(workbook, sheet, start, end, columns):
    values = []
    for row in range(start, end + 1):
        record = tuple(norm(workbook[sheet].cell(row=row, column=column).value) for column in columns)
        if any(value not in (None, "") for value in record): values.append(record)
    return values


def find_row(workbook, sheet, identifier, start=4, end=20):
    for row in range(start, end + 1):
        if workbook[sheet].cell(row=row, column=1).value == identifier:
            return row
    raise ValueError(f"MISSING_RECORD:{sheet}:{identifier}")


def source_tables_match(workbook, root):
    with (root / "raw_orders.csv").open(newline="") as handle:
        expected_raw = []
        for record in csv.DictReader(handle):
            units = int(record["units"]) if record["units"].isdigit() else record["units"]
            expected_raw.append((record["order_id"], record["product_code"], record["location_code"], units, record["order_date"]))
    with (root / "product_master.csv").open(newline="") as handle:
        expected_products = [(record["product_code"], record["product"], record["category"], int(record["unit_cost"])) for record in csv.DictReader(handle)]
    with (root / "location_master.csv").open(newline="") as handle:
        expected_locations = [(record["location_code"], record["region"], record["country"]) for record in csv.DictReader(handle)]
    actual_raw = rows(workbook, "Raw_Orders", 4, 10, range(1, 6))
    actual_products = rows(workbook, "Product_Master", 4, 6, range(1, 5))
    actual_locations = rows(workbook, "Location_Master", 4, 6, range(1, 4))
    return actual_raw == expected_raw and actual_products == expected_products and actual_locations == expected_locations


def bounded_tables(workbook):
    """Return small header-addressed tables without guessing workbook-wide roles."""
    tables = []
    for sheet in workbook.worksheets:
        max_column = min(sheet.max_column, 20)
        for header_row in range(1, min(sheet.max_row, 10) + 1):
            headers = {}
            for column in range(1, max_column + 1):
                header = sheet.cell(row=header_row, column=column).value
                token = "" if header in (None, "") else category_token(header)
                if token and token not in headers:
                    headers[token] = column
            if "order_id" not in headers and not {
                "product_code", "product", "category", "unit_cost"
            }.issubset(headers) and not {
                "location_code", "region", "country"
            }.issubset(headers) and not (
                "metric" in headers and ({"value", "count"} & set(headers))
            ) and not {
                "check", "status"
            }.issubset(headers):
                continue
            records = []
            for row in range(header_row + 1, min(sheet.max_row, header_row + 30) + 1):
                record = {
                    token: sheet.cell(row=row, column=column).value
                    for token, column in headers.items()
                }
                if any(value not in (None, "") for value in record.values()):
                    record["_row"] = row
                    records.append(record)
            tables.append({"sheet": sheet.title, "headers": headers, "records": records})
            break
    return tables


def comparable(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("$"):
            text = text[1:].replace(",", "")
        try:
            return round(float(text), 6)
        except ValueError:
            return text.casefold()
    if isinstance(value, float):
        return round(value, 6)
    return value


def record_equal(actual, expected):
    return all(comparable(actual.get(field)) == comparable(value) for field, value in expected.items())


def source_comparable(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return value
    return value


def source_record_equal(actual, expected):
    return all(
        source_comparable(actual.get(field)) == source_comparable(value)
        for field, value in expected.items()
    )


def find_exact_source_table(tables, expected_rows, required_headers):
    matches = []
    for table in tables:
        # A joined or normalized output may contain every master/source field;
        # that does not prove the original source table itself was preserved.
        if set(table["headers"]) != set(required_headers):
            continue
        actual = table["records"]
        if len(actual) != len(expected_rows):
            continue
        if all(source_record_equal(record, expected) for record, expected in zip(actual, expected_rows)):
            matches.append(table)
    return matches[0] if len(matches) == 1 else None


def expected_source_rows(root):
    with (root / "raw_orders.csv").open(newline="") as handle:
        raw = list(csv.DictReader(handle))
    with (root / "product_master.csv").open(newline="") as handle:
        products = list(csv.DictReader(handle))
    with (root / "location_master.csv").open(newline="") as handle:
        locations = list(csv.DictReader(handle))
    return raw, products, locations


def find_join_table(tables):
    required = {"order_id", "product", "region", "units", "unit_cost"}
    candidates = []
    for table in tables:
        headers = set(table["headers"])
        cost_header = next(
            (name for name in ("extended_cost", "order_cost", "total_cost") if name in headers),
            None,
        )
        if required.issubset(headers) and cost_header:
            candidates.append((table, cost_header))
    return candidates[0] if len(candidates) == 1 else (None, None)


def joined_records_match(table, cost_header, oracle, engine=None):
    if table is None:
        return False
    actual = {}
    duplicate_ids = set()
    for record in table["records"]:
        identifier = str(record.get("order_id") or "").strip()
        if not identifier:
            continue
        if identifier in actual:
            duplicate_ids.add(identifier)
        extended_cost = record.get(cost_header)
        if (
            engine is not None
            and isinstance(extended_cost, str)
            and extended_cost.startswith("=")
        ):
            extended_cost = engine.value(
                table["sheet"],
                f"{col_name(table['headers'][cost_header])}{record['_row']}",
            )
        actual[identifier] = {
            "product": record.get("product"),
            "region": record.get("region"),
            "units": record.get("units"),
            "unit_cost": record.get("unit_cost"),
            "extended_cost": extended_cost,
        }
    expected = {
        row[0]: {
            "product": row[1], "region": row[2], "units": row[3],
            "unit_cost": row[4], "extended_cost": row[5],
        }
        for row in oracle["joined"]
    }
    return not duplicate_ids and set(actual) == set(expected) and all(
        record_equal(actual[identifier], values)
        for identifier, values in expected.items()
    )


def find_clean_classification_table(tables, oracle):
    expected = {
        row[0]: {
            "product_code": row[1], "location_code": row[2], "units": row[3],
            "duplicate": normalize_yes_no(row[4]) == "yes",
            "valid": normalize_disposition(row[7]) == "join",
        }
        for row in oracle["clean"]
    }
    for table in tables:
        headers = set(table["headers"])
        validity_header = next(
            (name for name in ("valid", "disposition", "join_status") if name in headers),
            None,
        )
        duplicate_header = next(
            (name for name in ("is_duplicate", "duplicate", "duplicate_of") if name in headers),
            None,
        )
        if "order_id" not in headers or not validity_header or not duplicate_header:
            continue
        product_header = next(
            (name for name in ("product_code_normalized", "product_normalized", "product_code") if name in headers),
            None,
        )
        location_header = next(
            (name for name in ("location_code_normalized", "location_normalized", "location_code") if name in headers),
            None,
        )
        units_header = next(
            (name for name in ("units_numeric", "units_normalized", "units") if name in headers),
            None,
        )
        if not {product_header, location_header, units_header}.issubset(headers):
            continue
        actual = {}
        duplicate_ids = set()
        for record in table["records"]:
            identifier = str(record.get("order_id") or "").strip()
            if identifier:
                if identifier in actual:
                    duplicate_ids.add(identifier)
                actual[identifier] = {
                    "product_code": record.get(product_header),
                    "location_code": record.get(location_header),
                    "units": (
                        record.get(units_header)
                        if isinstance(record.get(units_header), (int, float))
                        else None
                    ),
                    "duplicate": (
                        normalize_yes_no(record.get(duplicate_header)) == "yes"
                        if duplicate_header != "duplicate_of"
                        else record.get(duplicate_header) not in (None, "")
                    ),
                    "valid": normalize_disposition(record.get(validity_header)) == "join",
                }
        if not duplicate_ids and len(table["records"]) == len(expected) and set(actual) == set(expected) and all(
            record_equal(actual[identifier], values)
            for identifier, values in expected.items()
        ):
            return table
    return None


def exception_kind(record):
    text = " ".join(
        str(record.get(field) or "")
        for field in ("exception_type", "exception_reason", "issue", "reason", "field", "original_value", "value")
    ).casefold()
    if "duplicate" in text:
        return "DUPLICATE"
    if "product" in text and any(token in text for token in ("not found", "not in", "invalid", "missing", "unmatched")):
        return "UNMATCHED_PRODUCT"
    if "location" in text and any(token in text for token in ("not found", "not in", "invalid", "missing", "unmatched")):
        return "UNMATCHED_LOCATION"
    if "unit" in text and any(token in text for token in ("non-numeric", "nonnumeric", "not numeric", "invalid", "numeric value")):
        return "INVALID_UNITS"
    return None


def business_exceptions_match(tables, oracle):
    expected = {(row[0], row[1]) for row in oracle["exceptions"]}
    candidates = []
    for table in tables:
        headers = set(table["headers"])
        if "order_id" not in headers or not headers.intersection(
            {"exception_type", "exception_reason", "issue", "reason", "field"}
        ):
            continue
        classified = set()
        for record in table["records"]:
            identifier = str(record.get("order_id") or "").strip()
            kind = exception_kind(record)
            if identifier and kind:
                classified.add((identifier, kind))
        candidates.append(classified)
    return any(classified == expected for classified in candidates)


def semantic_layout_checks(workbook, oracle, context):
    """Score business facts by headers and values, independent of sheet/column names."""
    tables = bounded_tables(workbook)
    if not tables:
        return None
    raw_rows, product_rows, location_rows = expected_source_rows(context["source_root"])
    raw = find_exact_source_table(
        tables, raw_rows, ("order_id", "product_code", "location_code", "units", "order_date")
    )
    products = find_exact_source_table(
        tables, product_rows, ("product_code", "product", "category", "unit_cost")
    )
    locations = find_exact_source_table(
        tables, location_rows, ("location_code", "region", "country")
    )
    joined, cost_header = find_join_table(tables)
    clean = find_clean_classification_table(tables, oracle)
    exceptions_ok = business_exceptions_match(tables, oracle)

    summary_present = any(
        "metric" in table["headers"]
        and ({"value", "count"} & set(table["headers"]))
        for table in tables
    )
    checks_present = any({"check", "status"}.issubset(table["headers"]) for table in tables)
    clean_present = clean is not None or any(
        "clean" in category_token(table["sheet"]) or "normaliz" in category_token(table["sheet"])
        for table in tables
    )
    exception_present = any("exception" in category_token(table["sheet"]) for table in tables)
    all_roles_present = all(
        (raw, products, locations, clean_present, joined, exception_present, summary_present, checks_present)
    )

    checks = {row["id"]: 0.0 for row in TASK["criteria"]}
    checks["R001"] = 1.0 if all_roles_present else 0.0
    checks["R002"] = 1.0 if clean is not None else 0.0
    joined_ok = joined_records_match(joined, cost_header, oracle, FormulaEngine(workbook))
    checks["R003"] = 1.0 if joined_ok else 0.0
    checks["R004"] = 1.0 if joined_ok else 0.0
    checks["R005"] = 1.0 if joined_ok else 0.0
    workbook_text = " ".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ).casefold()
    weekly_view = "week" in workbook_text and any(
        phrase in workbook_text for phrase in ("extended cost", "order cost", "total cost")
    )
    checks["R006"] = 1.0 if weekly_view else 0.0
    checks["R007"] = 1.0 if exceptions_ok and joined_ok else 0.0
    checks["R008"] = 1.0 if raw and products and locations else 0.0
    checks["R009"] = 1.0 if checks_present and joined_ok and exceptions_ok else 0.0
    failures = []
    if checks["R001"] == 0.0:
        failures.append("REQUESTED_OPERATIONS_SECTION_MISSING")
    if checks["R002"] == 0.0:
        failures.append("CLEAN_CLASSIFICATION_NOT_PROVEN")
    if checks["R003"] == 0.0:
        failures.append("JOINED_RECORDS_MISMATCH")
    if not exceptions_ok:
        failures.append("BUSINESS_EXCEPTIONS_MISMATCH")
    if checks["R008"] == 0.0:
        failures.append("PROTECTED_SOURCE_TABLE_NOT_PROVEN")
    return checks, failures


def task_checks(workbook, engine, oracle, context):
    checks, failures = {}, []
    expected = oracle
    clean_actual = []
    try:
        for row in range(4, 14):
            if workbook["Clean_Data"].cell(row=row, column=1).value in (None, ""):
                continue
            clean_actual.append(tuple(norm(engine.value("Clean_Data", f"{col_name(column)}{row}")) for column in range(1, 9)))
        clean_ok = {
            record[0]: normalize_clean_record(record)[1:]
            for record in clean_actual
        } == {
            record[0]: normalize_clean_record(record)[1:]
            for record in expected["clean"]
        }
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        clean_ok = False
        failures.append(f"CLEAN_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R002"] = 1.0 if clean_ok else 0.0
    if not clean_ok: failures.append("CLEAN_RECORDS_MISMATCH")

    joined_actual = []
    try:
        for row in range(4, 14):
            if workbook["Joined_Data"].cell(row=row, column=1).value in (None, ""): continue
            joined_actual.append(tuple(norm(engine.value("Joined_Data", f"{col_name(column)}{row}")) for column in range(1, 7)))
        joined_ok = {record[0]: record[1:] for record in joined_actual} == {record[0]: tuple(record[1:]) for record in expected["joined"]}
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        joined_ok = False; failures.append(f"JOIN_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R003"] = 1.0 if joined_ok else 0.0

    # The instruction requires correct order-cost arithmetic and a weekly view;
    # it does not prescribe formulas or hidden source-cell perturbations.
    checks["R004"] = 1.0 if joined_ok else 0.0

    exceptions_actual = {(row[0], category_token(row[1])) for row in rows(workbook, "Exceptions", 4, 12, range(1, 3))}
    exceptions_ok = exceptions_actual == {(row[0], category_token(row[1])) for row in expected["exceptions"]}
    try:
        summary_region_total = expected.get("summary_region_total", expected.get("northeast"))
        summary_ok = close(engine.value("Summary", "B4"), len(expected["joined"])) and close(engine.value("Summary", "B5"), expected["total"]) and close(engine.value("Summary", "B6"), len(expected["exceptions"])) and close(engine.value("Summary", "B7"), summary_region_total)
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        summary_ok = False; failures.append(f"SUMMARY_CLOSURE_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if exceptions_ok and summary_ok else 0.0
    checks["R005"] = 1.0 if summary_ok else 0.0
    checks["R006"] = 1.0 if summary_ok else 0.0

    protected_ok = source_tables_match(workbook, context["source_root"])
    checks["R008"] = 1.0 if protected_ok else 0.0
    if not protected_ok: failures.append("PROTECTED_SOURCE_TABLE_CHANGED")

    try:
        checks_ok = close(engine.value("Checks", "B4"), len(expected["joined"])) and close(engine.value("Checks", "B5"), len(expected["exceptions"])) and close(engine.value("Checks", "B6"), expected["total"])
    except UnsupportedFormulaError:
        raise
    except Exception as exc:
        checks_ok = False; failures.append(f"CHECK_LINKAGE_FAILED:{type(exc).__name__}")
    checks["R009"] = 1.0 if checks_ok else 0.0

    return checks, failures


def evaluate(candidate, split="dev"):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0: return criteria, ["OUTPUT_MISSING"]
    try: workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc: return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], SHEET_ALIASES
    )
    criteria["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    context = split_context(split)
    oracle = load_oracle(context)
    try:
        alternative = semantic_layout_checks(workbook, oracle, context)
    except Exception as exc:
        alternative = None
        failures.append(f"SEMANTIC_LAYOUT_EVALUATION_FAILED:{type(exc).__name__}:{exc}")
    if unresolved or ambiguous:
        if alternative is not None and all(alternative[0].get(key) == 1.0 for key in TASK["hurdle_criteria"]):
            checks, task_failures = alternative
            criteria.update(checks)
            failures.extend(task_failures)
            return criteria, sorted(set(failures))
        failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
        return criteria, sorted(set(failures))
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    try:
        checks, task_failures = task_checks(workbook, FormulaEngine(workbook), oracle, context)
        criteria.update(checks); failures.extend(task_failures)
    except UnsupportedFormulaError as exc:
        failures.append(f"UNSUPPORTED_FORMULA:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    if alternative is not None:
        semantic_checks, semantic_failures = alternative
        criteria = {key: max(criteria.get(key, 0.0), semantic_checks.get(key, 0.0)) for key in criteria}
        failures.extend(semantic_failures)
        failures = [
            code for code in failures
            if not code.startswith((
                "RAW_UNITS_PERTURBATION_FAILED:",
                "MASTER_COST_PERTURBATION_FAILED:",
                "CHECK_LINKAGE_FAILED:",
            ))
        ]
    passed_failure_codes = {
        "R001": {"REQUESTED_OPERATIONS_SECTION_MISSING"},
        "R002": {"CLEAN_CLASSIFICATION_NOT_PROVEN", "CLEAN_RECORDS_MISMATCH"},
        "R003": {"JOINED_RECORDS_MISMATCH"},
        "R007": {"BUSINESS_EXCEPTIONS_MISMATCH"},
        "R008": {"PROTECTED_SOURCE_TABLE_NOT_PROVEN", "PROTECTED_SOURCE_TABLE_CHANGED"},
    }
    failures = [
        code for code in failures
        if not any(criteria[criterion] == 1.0 and code in codes for criterion, codes in passed_failure_codes.items())
    ]
    return criteria, sorted(set(failures))



def parse_cli():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    candidate = None
    arguments = iter(sys.argv[1:])
    for argument in arguments:
        if argument == "--split": split = next(arguments, "")
        elif argument.startswith("--split="): split = argument.split("=", 1)[1]
        elif candidate is None: candidate = Path(argument)
        else: raise ValueError(f"UNEXPECTED_ARGUMENT:{argument}")
    if split not in {"dev", "confirm"}: raise ValueError(f"INVALID_SPLIT:{split}")
    return candidate or Path("/app/output/answer.xlsx"), split


def main():
    candidate, split = parse_cli()
    criteria, failures = evaluate(candidate, split)
    payload = build_result(task=TASK, split=split, candidate=str(candidate), criteria=criteria, failures=failures)
    payload["judge_version"] = "P15_JUDGE_V3"
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True); (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        if total is not None: (log_root / "reward.txt").write_text(str(total) + "\n")
    except OSError: pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__": main()

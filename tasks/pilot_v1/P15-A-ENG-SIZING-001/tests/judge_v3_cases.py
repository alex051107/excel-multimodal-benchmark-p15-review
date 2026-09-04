#!/usr/bin/env python3
"""One focused V3 validation pass for the engineering-sizing Judge."""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tests"))
import evaluate  # noqa: E402
from judge_v2_support import build_result  # noqa: E402


def build_alternate(path):
    workbook = Workbook()
    design = workbook.active
    design.title = "Design"
    rows = [
        ("Design flow", 42, "L/s"), ("Static head", 18, "m"),
        ("Pipe length", 140, "m"), ("Pipe internal diameter", 150, "mm"),
        ("Darcy friction factor", 0.02, ""), ("Pump efficiency", 0.72, ""),
        ("Safety factor", 0.15, ""), ("Fluid density", 998, "kg/m3"),
        ("Gravity", 9.81, "m/s2"),
    ]
    for index, row in enumerate(rows, 2):
        for column, value in enumerate(row, 1):
            design.cell(index, column, value)
    sizing = workbook.create_sheet("Sizing")
    sizing_rows = [
        ("Flow rate", "=Design!B2/1000", "m3/s"),
        ("Pipe diameter", "=Design!B5/1000", "m"),
        ("Cross-sectional area", "=PI()*B2^2/4", "m2"),
        ("Flow velocity", "=B1/B3", "m/s"),
        ("Velocity head", "=B4^2/(2*Design!B10)", "m"),
        ("Friction head loss", "=Design!B6*(Design!B4/B2)*B5", "m"),
        ("Total dynamic head", "=(Design!B3+B6)*(1+Design!B8)", "m"),
        ("Hydraulic power", "=Design!B9*Design!B10*B1*B7/1000", "kW"),
        ("Shaft power", "=B8/Design!B7", "kW"),
        ("Minimum pump flow", "=Design!B2*(1+Design!B8)", "L/s"),
        ("Required motor rating", "=B9*(1+Design!B8)", "kW"),
    ]
    for row in sizing_rows:
        sizing.append(row)
    catalog = workbook.create_sheet("Catalog")
    catalog.append(["Pump ID", "Max flow (L/s)", "Max head (m)", "Motor (kW)", "Eligibility"])
    for pump, flow, head, motor in (("P-120", 45, 25, 11), ("P-180", 58, 36, 18.5), ("P-250", 80, 48, 30)):
        catalog.append([pump, flow, head, motor, f'=IF(AND(B{catalog.max_row+1}>=Sizing!B10,C{catalog.max_row+1}>=Sizing!B7,D{catalog.max_row+1}>=Sizing!B11),"Eligible","Ineligible")'])
    catalog["A6"] = "Selected pump"
    catalog["B6"] = '=IF(E2="Eligible",A2,IF(E3="Eligible",A3,IF(E4="Eligible",A4,"No eligible pump")))'
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_constant_formula_mutant(path):
    workbook = load_workbook(ROOT / "solution/reference.xlsx")
    oracle, _ = evaluate.load_oracle(evaluate.load_contract())
    workbook["Calculations"]["B10"] = f"={oracle['minimum_flow']}"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def result(path):
    criteria, failures = evaluate.evaluate(path)
    payload = build_result(task=evaluate.TASK, split="dev", candidate=str(path), criteria=criteria, failures=failures)
    return {
        "score": payload["normalized_score"], "pass": payload["pass"],
        "evaluation_status": payload["evaluation_status"],
        "criterion_scores": payload["criterion_scores"], "failure_codes": payload["failure_codes"],
    }


def main():
    alternate = ROOT / "fixtures/equivalent/candidate_v3_three_sheet_layout.xlsx"
    constant_formula = ROOT / "fixtures/mutants/constant_formula_flow.xlsx"
    build_alternate(alternate)
    build_constant_formula_mutant(constant_formula)
    cases = {
        "standard": (ROOT / "solution/reference.xlsx", True, "Reference implementation must pass."),
        "equivalent": (ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", True, "Algebraically equivalent formulas must pass."),
        "alternate_layout": (alternate, True, "Three semantically labeled sheets satisfy the task without the reference sheet names."),
        "noop": (ROOT / "fixtures/noop/candidate_noop.xlsx", False, "An unchanged blank calculation block does not complete the task."),
        "malformed": (ROOT / "fixtures/malformed/candidate_malformed.xlsx", False, "A non-workbook cannot pass."),
    }
    for path in sorted((ROOT / "fixtures/mutants").glob("*.xlsx")):
        cases[f"mutant:{path.stem}"] = (path, False, "The seeded engineering defect must prevent a pass.")
    observed = {}
    for name, (path, expected_pass, basis) in cases.items():
        row = result(path)
        row.update({"expected_pass": expected_pass, "expectation_basis": basis, "meets_expectation": row["pass"] is expected_pass})
        observed[name] = row
    receipt = {
        "task_id": evaluate.TASK["task_id"], "judge_version": "v3-semantic",
        "generated_at": datetime.now(timezone.utc).isoformat(), "cases": observed,
        "all_expectations_met": all(row["meets_expectation"] for row in observed.values()),
    }
    destination = ROOT / "receipts/judge_v3_local_validation.json"
    destination.write_text(json.dumps(receipt, indent=2, sort_keys=True) + "\n")
    print(json.dumps(receipt, indent=2, sort_keys=True))
    if not receipt["all_expectations_met"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

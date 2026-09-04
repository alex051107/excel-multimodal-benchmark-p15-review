#!/usr/bin/env python3
"""Deterministic semantic judge for P15-A-ENG-SIZING-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-A-ENG-SIZING-001",
  "pass_threshold": 0.7,
  "canonical_reference_sheets": [
    "Inputs",
    "Unit_Conversions",
    "Calculations",
    "Equipment_Selection",
    "Checks"
  ],
  "protected": {
    "Inputs!B4": 42,
    "Inputs!B7": 150,
    "Equipment_Selection!A4": "P-120",
    "Equipment_Selection!B4": 45,
    "Equipment_Selection!C4": 25,
    "Equipment_Selection!D4": 11,
    "Equipment_Selection!A5": "P-180",
    "Equipment_Selection!B5": 58,
    "Equipment_Selection!C5": 36,
    "Equipment_Selection!D5": 18.5,
    "Equipment_Selection!A6": "P-250",
    "Equipment_Selection!B6": 80,
    "Equipment_Selection!C6": 48,
    "Equipment_Selection!D6": 30
  },
  "formula_cells": [
    "Unit_Conversions!C4",
    "Calculations!B7",
    "Calculations!B11",
    "Equipment_Selection!E5",
    "Equipment_Selection!B7"
  ],
  "criteria": [
    {"id":"R001","description":"The workbook opens and contains the required sizing content in a deterministically identifiable layout.","weight":1,"type":"positive","dimension":"file_usability","method":"deterministic","method_params":{"implemented_check":"required_sizing_content"}},
    {"id":"R002","description":"Flow, diameter, and cross-sectional-area conversions agree with the independent SI replay.","weight":3,"type":"positive","dimension":"unit_conversion","method":"deterministic","method_params":{"implemented_check":"si_conversions","oracle":"split_contract"}},
    {"id":"R003","description":"Velocity, velocity head, friction head, and total dynamic head agree with the hydraulic replay.","weight":3,"type":"positive","dimension":"hydraulic_head","method":"deterministic","method_params":{"implemented_check":"hydraulic_head_chain","oracle":"split_contract"}},
    {"id":"R004","description":"Hydraulic and shaft power agree with the independent replay.","weight":2,"type":"positive","dimension":"power_sizing","method":"deterministic","method_params":{"implemented_check":"power_chain","oracle":"split_contract"}},
    {"id":"R005","description":"Safety-adjusted minimum flow and motor rating agree with the independent replay.","weight":2,"type":"positive","dimension":"safety_sizing","method":"deterministic","method_params":{"implemented_check":"safety_adjusted_requirements","oracle":"split_contract"}},
    {"id":"R006","description":"Each catalog row's eligibility status matches all three sizing constraints.","weight":3,"type":"positive","dimension":"catalog_eligibility","method":"deterministic","method_params":{"implemented_check":"catalog_eligibility","oracle":"split_contract"}},
    {"id":"R007","description":"The selected pump is the first eligible catalog option.","weight":3,"type":"positive","dimension":"constraint_selection","method":"deterministic","method_params":{"implemented_check":"selected_pump","oracle":"split_contract"}},
    {"id":"R008","description":"Hydraulic sizing and pump selection respond correctly to the declared design-flow perturbation.","weight":2,"type":"positive","dimension":"native_recalculation","method":"deterministic","method_params":{"implemented_check":"flow_perturbation","split_weights":{"dev":2,"confirm":3}}},
    {"id":"R009","description":"Every declared engineering output carries the required unit label.","weight":1,"type":"positive","dimension":"unit_labels","method":"deterministic","method_params":{"implemented_check":"unit_labels"}},
    {"id":"R010","description":"Any check rows the workbook declares evaluate to PASS; a separate check sheet is optional.","weight":1,"type":"positive","dimension":"model_checks","method":"deterministic","method_params":{"implemented_check":"optional_check_rows"}},
    {"id":"R011","description":"The declared design inputs and catalog capacity values remain unchanged.","weight":2,"type":"positive","dimension":"change_locality","method":"deterministic","method_params":{"implemented_check":"input_catalog_protection"}},
    {"id":"R012","description":"The hydraulic head, motor requirement, eligibility, and selection respond correctly to a distinct internal-diameter perturbation while design flow and catalog values remain invariant.","weight":1,"type":"positive","dimension":"diameter_recalculation","method":"deterministic","method_params":{"implemented_check":"diameter_perturbation","applies_to":["dev"]}}
  ]
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R007", "R008", "R011"]
SHEET_ALIASES = {}


def cell_key(sheet, cell):
    return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column:
        value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


TASK_ROOT = Path(__file__).resolve().parents[1]


def requested_split():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    for index, argument in enumerate(sys.argv[1:]):
        if argument == "--split" and index + 2 <= len(sys.argv) - 1:
            split = sys.argv[index + 2].strip().lower()
        elif argument.startswith("--split="):
            split = argument.split("=", 1)[1].strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"unsupported split: {split}")
    return split


ACTIVE_SPLIT = requested_split()


def load_contract():
    relative = "tests/confirm/contract.json" if ACTIVE_SPLIT == "confirm" else "tests/private_contract.json"
    return json.loads((TASK_ROOT / relative).read_text())


def load_oracle(contract):
    oracle_path = TASK_ROOT / contract["oracle"]
    source_path = TASK_ROOT / "metadata/oracle_recompute.py"
    if ACTIVE_SPLIT == "dev" and source_path.exists() and source_path.read_bytes() != oracle_path.read_bytes():
        raise RuntimeError("DEV_ORACLE_COPY_OUT_OF_SYNC")
    spec = importlib.util.spec_from_file_location(f"task_oracle_{ACTIVE_SPLIT}", oracle_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute(), module


def _betacf(a, b, x):
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d, h = 1.0, 1.0 - qab * x / qap, 0.0
    if abs(d) < 3e-14: d = 3e-14
    d = 1.0 / d
    h = d
    for m in range(1, 201):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < 3e-14: d = 3e-14
        c = 1.0 + aa / c
        if abs(c) < 3e-14: c = 3e-14
        d = 1.0 / d; delta = d * c; h *= delta
        if abs(delta - 1.0) < 3e-12: break
    return h


def _betai(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    factor = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x))
    return factor * _betacf(a, b, x) / a if x < (a + 1) / (a + b + 2) else 1 - factor * _betacf(b, a, 1 - x) / b


def student_t_two_tail(value, df):
    x = df / (df + value * value)
    return _betai(df / 2, 0.5, x)


def student_t_inv_two_tail(alpha, df):
    lo, hi = 0.0, 20.0
    for _ in range(90):
        mid = (lo + hi) / 2
        if student_t_two_tail(mid, df) > alpha: lo = mid
        else: hi = mid
    return (lo + hi) / 2


class ExcelFormulaError(ValueError):
    """An Excel value error that IFERROR may legitimately handle."""


def excel_equal(left, right):
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-12)
    if isinstance(left, str) or isinstance(right, str):
        return " ".join(str(left or "").strip().casefold().split()) == " ".join(str(right or "").strip().casefold().split())
    return left == right


def excel_criteria_match(value, criterion):
    if not isinstance(criterion, str):
        return excel_equal(value, criterion)
    match = re.fullmatch(r"\s*(<=|>=|<>|=|<|>)?\s*(.*?)\s*", criterion)
    operator, operand_text = match.groups() if match else (None, criterion)
    try:
        candidate = float(value)
        operand = float(operand_text)
    except (TypeError, ValueError):
        candidate = " ".join(str(value or "").strip().casefold().split())
        operand = " ".join(operand_text.strip().casefold().split())
    if operator in (None, "="):
        return candidate == operand
    if operator == "<>":
        return candidate != operand
    if operator == "<":
        return candidate < operand
    if operator == "<=":
        return candidate <= operand
    if operator == ">":
        return candidate > operand
    if operator == ">=":
        return candidate >= operand
    return False


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook
        self.overrides = overrides or {}
        self.memo = {}
        self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames:
            raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            if isinstance(raw, str) and raw.startswith("="):
                result = self.formula(raw[1:], sheet)
            else:
                result = raw
        finally:
            # A formula that raises must not leave its cell on the visiting
            # stack.  The engine is shared across criteria, so a stale entry
            # turns every later read of that cell -- and of anything that
            # depends on it -- into a fake CIRCULAR_REFERENCE, which zeroes
            # criteria that have nothing to do with the original fault.
            self.stack.discard(key)
        self.memo[key] = result
        return result

    def range_values(self, sheet, c1, r1, c2, r2):
        values = []
        for r in range(int(r1), int(r2) + 1):
            for c in range(col_index(c1), col_index(c2) + 1):
                values.append(self.value(sheet, f"{col_name(c)}{r}"))
        return values

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        expression = re.sub(r"\bSTDEV\.S\b", "STDEV_S", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.2T\b", "T_DIST_2T", expression, flags=re.I)
        expression = re.sub(r"\bT\.DIST\.RT\b", "T_DIST_RT", expression, flags=re.I)
        expression = re.sub(r"\bT\.INV\.2T\b", "T_INV_2T", expression, flags=re.I)
        expression = re.sub(r"\bPI\(\)", "PI()", expression, flags=re.I)

        def replace_range(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))

        def sub_outside_strings(pattern, replacement, value):
            parts = re.split(r'("(?:[^"\\\\]|\\\\.)*")', value)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in parts)

        expression = sub_outside_strings(RANGE, replace_range, expression)

        def replace_ref(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))

        expression = sub_outside_strings(REF, replace_ref, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        tree = ast.parse(expression, mode="eval")
        return self.safe_eval(tree.body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name) and node.id.upper() in {"TRUE", "FALSE"}:
            return node.id.upper() == "TRUE"
        if isinstance(node, ast.List):
            return [self.safe_eval(v) for v in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand)
            return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operators = {ast.Add: lambda a,b:a+b, ast.Sub: lambda a,b:a-b, ast.Mult: lambda a,b:a*b, ast.Div: lambda a,b:a/b, ast.Pow: lambda a,b:a**b}
            for cls, fun in operators.items():
                if isinstance(node.op, cls): return fun(left, right)
            raise ValueError("UNSUPPORTED_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for op, comp in zip(node.ops, node.comparators):
                right = self.safe_eval(comp)
                good = (left == right if isinstance(op, ast.Eq) else left != right if isinstance(op, ast.NotEq) else left >= right if isinstance(op, ast.GtE) else left <= right if isinstance(op, ast.LtE) else left > right if isinstance(op, ast.Gt) else left < right if isinstance(op, ast.Lt) else False)
                if not good: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "IF":
                if len(node.args) not in (2, 3):
                    raise ValueError("IF_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                if self.safe_eval(node.args[0]):
                    return self.safe_eval(node.args[1])
                return self.safe_eval(node.args[2]) if len(node.args) == 3 else False
            if name == "IFERROR":
                if len(node.args) != 2:
                    raise ValueError("IFERROR_REQUIRES_TWO_ARGUMENTS")
                try:
                    return self.safe_eval(node.args[0])
                except (ExcelFormulaError, ArithmeticError):
                    return self.safe_eval(node.args[1])
            args = [self.safe_eval(a) for a in node.args]
            values = [v for group in args for v in (group if isinstance(group, list) else [group]) if v is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]
            if name == "SUM": return sum(numeric_values)
            if name == "AVERAGE": return statistics.mean(numeric_values)
            if name == "COUNT": return len(numeric_values)
            if name == "STDEV_S": return statistics.stdev(numeric_values)
            if name == "SQRT": return math.sqrt(args[0])
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "PI": return math.pi
            if name == "AND": return all(args)
            if name == "OR": return any(args)
            if name == "NOT": return not args[0]
            if name == "POWER": return args[0] ** args[1]
            if name == "PRODUCT": return math.prod(numeric_values)
            if name == "MAX": return max(numeric_values)
            if name == "MIN": return min(numeric_values)
            if name == "MATCH":
                if len(args) not in {2, 3}:
                    raise ValueError("MATCH_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                candidates = args[1] if isinstance(args[1], list) else [args[1]]
                match_type = args[2] if len(args) == 3 else 1
                if match_type != 0:
                    raise ValueError("ONLY_EXACT_MATCH_MODE_IS_SUPPORTED")
                for index, candidate in enumerate(candidates, start=1):
                    if excel_equal(candidate, args[0]):
                        return index
                raise ExcelFormulaError("MATCH_NOT_FOUND")
            if name == "INDEX":
                if len(args) not in {2, 3}:
                    raise ValueError("INDEX_REQUIRES_TWO_OR_THREE_ARGUMENTS")
                candidates = args[0] if isinstance(args[0], list) else [args[0]]
                if len(args) == 3 and int(args[2]) != 1:
                    raise ValueError("ONLY_SINGLE_COLUMN_INDEX_IS_SUPPORTED")
                row = int(args[1])
                if row < 1 or row > len(candidates):
                    raise ExcelFormulaError("INDEX_OUT_OF_RANGE")
                return candidates[row - 1]
            if name == "COUNTIF":
                if len(args) != 2:
                    raise ValueError("COUNTIF_REQUIRES_TWO_ARGUMENTS")
                candidates = args[0] if isinstance(args[0], list) else [args[0]]
                return sum(1 for candidate in candidates if excel_criteria_match(candidate, args[1]))
            if name == "NA": return None
            if name == "T_DIST_2T":
                return student_t_two_tail(args[0], args[1])
            if name == "T_DIST_RT":
                return student_t_two_tail(args[0], args[1]) / 2
            if name == "T_INV_2T":
                return student_t_inv_two_tail(args[0], args[1])
        raise ValueError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def normalized_text(value):
    return " ".join(str(value or "").strip().casefold().split())


def eligibility_status(value):
    aliases = {
        "eligible": "eligible",
        "meets criteria": "eligible",
        "meets all constraints": "eligible",
        "pass": "eligible",
        "ineligible": "ineligible",
        "not eligible": "ineligible",
        "does not meet criteria": "ineligible",
        "does not meet all constraints": "ineligible",
        "fail": "ineligible",
    }
    return aliases.get(normalized_text(value))


def normalized_unit(value):
    return normalized_text(value).replace("\u00b2", "2").replace("\u00b3", "3").replace("^", "").replace(" ", "")


class UnsupportedSingleSheetLayout(ValueError):
    """The semantic content cannot be located without guessing."""
    pass


def semantic_label(value):
    rendered = normalized_text(value).replace("_", " ").replace("-", " ")
    rendered = rendered.replace("\u00b2", "2").replace("\u00b3", "3")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", rendered).split())


INPUT_LABELS = {
    "design_flow": {"design flow", "q design"},
    "static_head": {"static head", "h static"},
    "pipe_length": {"pipe length", "l pipe"},
    "diameter": {"pipe id", "pipe inside diameter", "pipe internal diameter", "d pipe"},
    "friction_factor": {"darcy friction factor", "f darcy"},
    "efficiency": {"pump efficiency", "eta pump"},
    "safety_factor": {"safety factor", "sf"},
    "density": {"fluid density", "rho"},
    "gravity": {"gravity", "g"},
}

CANONICAL_INPUT_CELLS = {
    "design_flow": ("flow_lps", "Inputs!B4"),
    "static_head": ("static_head_m", "Inputs!B5"),
    "pipe_length": ("pipe_length_m", "Inputs!B6"),
    "diameter": ("diameter_mm", "Inputs!B7"),
    "friction_factor": ("friction", "Inputs!B8"),
    "efficiency": ("efficiency", "Inputs!B9"),
    "safety_factor": ("safety", "Inputs!B10"),
    "density": ("density", "Inputs!B11"),
    "gravity": ("gravity", "Inputs!B12"),
}


METRIC_SPECS = {
    "flow_m3s": ([r"(?:flow rate|flow)(?: q)?", r"q"], {"m3/s": 1.0}),
    "diameter_m": ([r"pipe (?:inside )?diameter(?: d)?", r"d"], {"m": 1.0}),
    "area_m2": ([r"(?:pipe )?(?:cross sectional|cross section) area(?: a)?", r"area(?: a)?", r"a"], {"m2": 1.0}),
    "velocity": ([r"(?:flow )?velocity(?: v)?", r"v"], {"m/s": 1.0}),
    "velocity_head": ([r"velocity head.*", r"h v"], {"m": 1.0}),
    "friction_head": ([r"friction head(?: loss)?.*", r"h f"], {"m": 1.0}),
    "tdh": ([r"total dynamic head(?: tdh)?", r"tdh"], {"m": 1.0}),
    "hydraulic_power": ([r"hydraulic power.*", r"p hydraulic", r"p hyd"], {"kw": 1.0, "w": 0.001}),
    "shaft_power": ([r"(?:required )?shaft power.*", r"p shaft"], {"kw": 1.0, "w": 0.001}),
    "minimum_flow": ([r"min(?:imum)? (?:pump )?flow.*", r"q min"], {"l/s": 1.0}),
    "minimum_motor": ([r"min(?:imum)? motor.*", r"required motor rating.*", r"motor min", r"motor rating (?:with|including) safety.*"], {"kw": 1.0, "w": 0.001}),
}


def formula_cell(sheet, address):
    return isinstance(sheet[address].value, str) and sheet[address].value.startswith("=")


def address(sheet, cell):
    return f"{sheet.title}!{cell}"


def split_address(value):
    return value.split("!", 1)


def find_semantic_input(workbook, role):
    candidates = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row in sheet.iter_rows():
            for label_cell in row:
                if semantic_label(label_cell.value) not in INPUT_LABELS[role]:
                    continue
                for offset in (1, 2):
                    column = label_cell.column + offset
                    if column > sheet.max_column:
                        continue
                    value_cell = sheet.cell(label_cell.row, column)
                    value = value_cell.value
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        candidates.append((offset, sheet_index, label_cell.row, column, address(sheet, value_cell.coordinate)))
    if len(candidates) != 1:
        rendered = ",".join(item[-1] for item in candidates) or "missing"
        raise UnsupportedSingleSheetLayout(f"INPUT_{role}:{rendered}")
    return candidates[0][-1]


def find_semantic_metric(workbook, role):
    patterns, accepted_units = METRIC_SPECS[role]
    candidates = []
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row in sheet.iter_rows():
            for label_cell in row:
                label = semantic_label(label_cell.value)
                rank = next((index for index, pattern in enumerate(patterns) if re.fullmatch(pattern, label)), None)
                if rank is None:
                    continue
                for value_offset in (1, 2, 3):
                    value_column = label_cell.column + value_offset
                    if value_column > sheet.max_column:
                        continue
                    value_cell = sheet.cell(label_cell.row, value_column)
                    raw = value_cell.value
                    if not (
                        isinstance(raw, (int, float)) and not isinstance(raw, bool)
                        or isinstance(raw, str) and raw.startswith("=")
                    ):
                        continue
                    unit_candidates = []
                    for unit_offset in (1, 2):
                        unit_column = value_column + unit_offset
                        if unit_column <= sheet.max_column:
                            unit_cell = sheet.cell(label_cell.row, unit_column)
                            unit_candidates.append((unit_offset, unit_cell))
                    accepted = [
                        (distance, cell, normalized_unit(cell.value))
                        for distance, cell in unit_candidates
                        if normalized_unit(cell.value) in accepted_units
                    ]
                    if accepted:
                        unit_distance, unit_cell, unit = accepted[0]
                    else:
                        unit_distance, unit_cell = unit_candidates[0] if unit_candidates else (9, value_cell)
                        unit = normalized_unit(unit_cell.value)
                    candidates.append(
                        (
                            rank,
                            0 if unit in accepted_units else 1,
                            unit_distance,
                            0 if isinstance(raw, str) and raw.startswith("=") else 1,
                            value_offset,
                            sheet_index,
                            label_cell.row,
                            value_column,
                            address(sheet, value_cell.coordinate),
                            address(sheet, unit_cell.coordinate),
                            accepted_units.get(unit, 1.0),
                            unit in accepted_units,
                        )
                    )
    if not candidates:
        raise UnsupportedSingleSheetLayout(f"METRIC_{role}:missing")
    candidates.sort()
    best = candidates[0]
    if len(candidates) > 1 and candidates[1][:5] == best[:5]:
        raise UnsupportedSingleSheetLayout(f"METRIC_{role}:ambiguous")
    return {
        "address": best[8],
        "unit_address": best[9],
        "scale": best[10],
        "unit_valid": best[11],
    }


def header_above(sheet, row, column):
    for offset in range(1, 11):
        if row - offset < 1:
            break
        raw = sheet.cell(row=row - offset, column=column).value
        if isinstance(raw, str) and raw.startswith("="):
            continue
        label = semantic_label(raw)
        if (
            "eligible" in label
            or "eligibility" in label
            or label == "status"
            or ("flow" in label and "ok" in label)
            or ("head" in label and "ok" in label)
            or ("motor" in label and "ok" in label)
        ):
            return label
    return ""


def eligibility_bool(value):
    if isinstance(value, bool):
        return value
    status = normalized_text(value)
    if status in {"eligible", "yes", "true", "pass", "meets criteria", "meets all constraints"}:
        return True
    if status in {"ineligible", "not eligible", "no", "false", "fail", "does not meet criteria", "does not meet all constraints"}:
        return False
    raise ValueError(f"UNRECOGNIZED_ELIGIBILITY:{value}")


def single_sheet_split_truth(contract, oracle_module):
    """Bind the semantic adapter to the selected DEV/CONFIRM truth."""
    oracle_inputs = getattr(oracle_module, "INPUT", None)
    oracle_catalog = getattr(oracle_module, "CATALOG", None)
    if not isinstance(oracle_inputs, dict) or not isinstance(oracle_catalog, (list, tuple)):
        raise RuntimeError("SPLIT_ORACLE_MISSING_INPUT_OR_CATALOG")

    protected = contract.get("protected", {})
    inputs = {}
    for role, (oracle_key, contract_address) in CANONICAL_INPUT_CELLS.items():
        if oracle_key not in oracle_inputs or contract_address not in protected:
            raise RuntimeError(f"SPLIT_TRUTH_MISSING:{role}")
        if not excel_equal(oracle_inputs[oracle_key], protected[contract_address]):
            raise RuntimeError(f"SPLIT_TRUTH_MISMATCH:{role}")
        inputs[role] = protected[contract_address]

    catalog = []
    for offset, oracle_row in enumerate(oracle_catalog, start=4):
        if not isinstance(oracle_row, (list, tuple)) or len(oracle_row) != 4:
            raise RuntimeError("SPLIT_ORACLE_INVALID_CATALOG")
        contract_row = tuple(
            protected.get(f"Equipment_Selection!{column}{offset}")
            for column in ("A", "B", "C", "D")
        )
        if any(value is None for value in contract_row):
            raise RuntimeError(f"SPLIT_TRUTH_MISSING:catalog_row_{offset}")
        if not all(excel_equal(left, right) for left, right in zip(oracle_row, contract_row)):
            raise RuntimeError(f"SPLIT_TRUTH_MISMATCH:catalog_row_{offset}")
        catalog.append(contract_row)

    pump_ids = tuple(str(row[0]) for row in catalog)
    if len(set(pump_ids)) != len(pump_ids):
        raise RuntimeError("SPLIT_ORACLE_DUPLICATE_PUMP_ID")
    return {"inputs": inputs, "catalog": tuple(catalog), "pump_ids": pump_ids}


def find_catalog_rows(workbook, pump_ids):
    rows = {}
    wanted = set(pump_ids)
    for sheet in workbook.worksheets:
        for row in range(1, sheet.max_row + 1):
            for column in range(1, max(1, sheet.max_column - 2)):
                pump = str(sheet.cell(row, column).value or "").strip()
                values = [sheet.cell(row=row, column=column + offset).value for offset in (1, 2, 3)]
                if pump in wanted and all(
                    isinstance(value, (int, float)) and not isinstance(value, bool) for value in values
                ):
                    if pump in rows:
                        raise UnsupportedSingleSheetLayout(f"CATALOG_{pump}:ambiguous")
                    rows[pump] = {
                        "sheet": sheet.title,
                        "row": row,
                        "cells": tuple(address(sheet, sheet.cell(row, column + offset).coordinate) for offset in (1, 2, 3)),
                    }
    if set(rows) != wanted:
        raise UnsupportedSingleSheetLayout("CATALOG_ROWS:missing")
    return rows


def find_eligibility_refs(workbook, engine, pump_ids):
    refs = {}
    for pump in pump_ids:
        composite, parts = [], []
        for sheet_index, sheet in enumerate(workbook.worksheets):
            row_candidates = []
            for row in range(1, sheet.max_row + 1):
                for column in range(1, sheet.max_column + 1):
                    raw = sheet.cell(row, column).value
                    try:
                        observed = engine.value(sheet.title, sheet.cell(row, column).coordinate) if isinstance(raw, str) and raw.startswith("=") else raw
                    except Exception:
                        observed = raw
                    if str(observed or "").strip() == pump:
                        row_candidates.append(row)
                        break
            for row in row_candidates:
                row_parts = {}
                for column in range(1, min(sheet.max_column, 10) + 1):
                    cell = sheet.cell(row, column)
                    if cell.value in (None, ""):
                        continue
                    header = header_above(sheet, row, column)
                    ref = address(sheet, cell.coordinate)
                    if "eligible" in header or "eligibility" in header or header == "status":
                        composite.append((0 if "eligible" in header else 1, sheet_index, row, ref))
                    elif "flow" in header and "ok" in header:
                        row_parts["flow"] = ref
                    elif "head" in header and "ok" in header:
                        row_parts["head"] = ref
                    elif "motor" in header and "ok" in header:
                        row_parts["motor"] = ref
                if set(row_parts) == {"flow", "head", "motor"}:
                    parts.append((sheet_index, row, row_parts))
        if composite:
            composite.sort()
            if len(composite) > 1 and composite[1][:3] == composite[0][:3]:
                raise UnsupportedSingleSheetLayout(f"ELIGIBILITY_{pump}:ambiguous")
            refs[pump] = {"composite": composite[0][3]}
        elif len(parts) == 1:
            refs[pump] = {"parts": parts[0][2]}
        else:
            raise UnsupportedSingleSheetLayout(f"ELIGIBILITY_{pump}:missing_or_ambiguous")
    return refs


def find_selected_pump_cell(workbook, engine, pump_ids):
    candidates = []
    valid_values = {*pump_ids, "No eligible pump", "NONE ELIGIBLE"}
    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row in range(1, sheet.max_row + 1):
            for label_cell in sheet[row]:
                label = semantic_label(label_cell.value)
                if not any(token in label for token in ("selected pump", "first eligible pump", "recommended pump")):
                    continue
                for target_row in range(row, min(sheet.max_row, row + 2) + 1):
                    for column in range(1, min(sheet.max_column, label_cell.column + 4) + 1):
                        cell = sheet.cell(target_row, column)
                        if cell.coordinate == label_cell.coordinate or cell.value in (None, ""):
                            continue
                        try:
                            observed = engine.value(sheet.title, cell.coordinate)
                        except Exception:
                            observed = cell.value
                        if observed in valid_values:
                            candidates.append((target_row - row, abs(column - label_cell.column), sheet_index, target_row, column, address(sheet, cell.coordinate)))
    if not candidates:
        raise UnsupportedSingleSheetLayout("SELECTED_PUMP:missing")
    candidates.sort()
    return candidates[0][-1]


def build_semantic_layout(workbook, pump_ids):
    inputs = {role: find_semantic_input(workbook, role) for role in INPUT_LABELS}
    metrics = {role: find_semantic_metric(workbook, role) for role in METRIC_SPECS}
    engine = FormulaEngine(workbook)
    catalog_rows = find_catalog_rows(workbook, pump_ids)
    eligibility = find_eligibility_refs(workbook, engine, pump_ids)
    selected = find_selected_pump_cell(workbook, engine, pump_ids)
    return {
        "inputs": inputs,
        "metrics": metrics,
        "catalog_rows": catalog_rows,
        "pump_ids": pump_ids,
        "eligibility": eligibility,
        "selected": selected,
    }


def single_metric_value(engine, layout, role):
    metric = layout["metrics"][role]
    sheet, cell = split_address(metric["address"])
    return engine.value(sheet, cell) * metric["scale"]


def single_eligibility(engine, layout, pump):
    ref = layout["eligibility"][pump]
    if "composite" in ref:
        sheet, cell = split_address(ref["composite"])
        return eligibility_bool(engine.value(sheet, cell))
    values = []
    for value in ref["parts"].values():
        sheet, cell = split_address(value)
        values.append(eligibility_bool(engine.value(sheet, cell)))
    return all(values)


def task_checks_single_sheet(workbook, oracle, contract, oracle_module, layout, split_truth):
    checks, failures = {}, []
    engine = FormulaEngine(workbook)

    def metric_group(criterion, expected):
        ok = True
        for role, target in expected.items():
            try:
                observed = single_metric_value(engine, layout, role)
            except Exception as exc:
                observed = None
                failures.append(f"SIZING_EVALUATION_FAILED:{role}:{type(exc).__name__}")
            tolerance = max(0.000001, abs(target) * 0.0001)
            if not close(observed, target, tolerance):
                ok = False
                failures.append(f"SIZING_VALUE_MISMATCH:{role}")
        checks[criterion] = 1.0 if ok else 0.0

    metric_group("R002", {"flow_m3s": oracle["flow_m3s"], "diameter_m": oracle["diameter_m"], "area_m2": oracle["area_m2"]})
    metric_group("R003", {"velocity": oracle["velocity"], "velocity_head": oracle["velocity_head"], "friction_head": oracle["friction_head"], "tdh": oracle["tdh"]})
    metric_group("R004", {"hydraulic_power": oracle["hydraulic_power"], "shaft_power": oracle["shaft_power"]})
    metric_group("R005", {"minimum_flow": oracle["minimum_flow"], "minimum_motor": oracle["minimum_motor"]})

    pumps = split_truth["pump_ids"]
    try:
        observed_eligibility = [single_eligibility(engine, layout, pump) for pump in pumps]
        expected_eligibility = [eligibility_status(value) == "eligible" for value in oracle["eligibility"]]
        checks["R006"] = 1.0 if observed_eligibility == expected_eligibility else 0.0
    except Exception as exc:
        checks["R006"] = 0.0
        failures.append(f"ELIGIBILITY_EVALUATION_FAILED:{type(exc).__name__}")
    try:
        selected_sheet, selected_cell = split_address(layout["selected"])
        selected = engine.value(selected_sheet, selected_cell)
        checks["R007"] = 1.0 if selected == oracle["selected_pump"] else 0.0
    except Exception as exc:
        checks["R007"] = 0.0
        failures.append(f"SELECTION_EVALUATION_FAILED:{type(exc).__name__}")

    def replay_ok(overrides, expected, expected_eligibility=None):
        replay = FormulaEngine(workbook, {layout["inputs"][role]: value for role, value in overrides.items()})
        numeric_expected = {role: target for role, target in expected.items() if role != "selected_pump"}
        if not all(
            close(single_metric_value(replay, layout, role), target, max(0.000001, abs(target) * 0.0001))
            for role, target in numeric_expected.items()
        ):
            return False
        selected_sheet, selected_cell = split_address(layout["selected"])
        selected = replay.value(selected_sheet, selected_cell)
        selected_target = expected.get("selected_pump")
        if selected_target is not None and selected != selected_target:
            return False
        if expected_eligibility is not None:
            observed = [single_eligibility(replay, layout, pump) for pump in pumps]
            wanted = [eligibility_status(value) == "eligible" for value in expected_eligibility]
            if observed != wanted:
                return False
        return True

    try:
        flow_value = contract["perturbations"][0]["overrides"]["Inputs!B4"]
        flow_oracle = oracle_module.recompute(flow_lps=flow_value)
        flow_expected = {role: flow_oracle[role] for role in ("tdh", "shaft_power", "minimum_flow", "minimum_motor")}
        flow_expected["selected_pump"] = flow_oracle["selected_pump"]
        flow_ok = replay_ok(
            {"design_flow": flow_value},
            flow_expected,
            flow_oracle["eligibility"],
        )
    except Exception as exc:
        flow_ok = False
        failures.append(f"FLOW_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R008"] = 1.0 if flow_ok else 0.0

    checks["R009"] = 1.0 if all(
        metric["unit_valid"] for metric in layout["metrics"].values()
    ) else 0.0
    declared_checks = []
    for sheet in workbook.worksheets:
        status_columns = {
            cell.column
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20))
            for cell in row
            if normalized_text(cell.value) in {"status", "result", "check result"}
        }
        for row in sheet.iter_rows():
            for cell in row:
                formula_cell = isinstance(cell.value, str) and cell.value.startswith("=")
                try:
                    result = normalized_text(
                        engine.value(sheet.title, cell.coordinate)
                        if formula_cell
                        else cell.value
                    )
                except Exception:
                    continue
                if (
                    (formula_cell or cell.column in status_columns)
                    and result in {"pass", "fail", "check", "all checks pass", "failures present"}
                ):
                    declared_checks.append(result)
    checks["R010"] = 1.0 if not declared_checks or all(value in {"pass", "all checks pass"} for value in declared_checks) else 0.0
    expected_inputs = split_truth["inputs"]
    input_ok = all(workbook[split_address(layout["inputs"][role])[0]][split_address(layout["inputs"][role])[1]].value == value for role, value in expected_inputs.items())
    catalog_expected = {row[0]: tuple(row[1:]) for row in split_truth["catalog"]}
    catalog_ok = list(layout["catalog_rows"]) == list(pumps) and all(
        tuple(workbook[split_address(cell)[0]][split_address(cell)[1]].value for cell in layout["catalog_rows"][pump]["cells"]) == values
        for pump, values in catalog_expected.items()
    )
    checks["R011"] = 1.0 if input_ok and catalog_ok else 0.0

    checks["R012"] = 0.0
    if ACTIVE_SPLIT == "dev":
        try:
            diameter_value = contract["perturbations"][1]["overrides"]["Inputs!B7"]
            diameter_oracle = oracle_module.recompute(diameter_mm=diameter_value)
            diameter_expected = {role: diameter_oracle[role] for role in ("diameter_m", "area_m2", "velocity", "tdh", "shaft_power", "minimum_motor")}
            diameter_expected["selected_pump"] = diameter_oracle["selected_pump"]
            diameter_ok = replay_ok({"diameter": diameter_value}, diameter_expected, diameter_oracle["eligibility"])
        except Exception as exc:
            diameter_ok = False
            failures.append(f"DIAMETER_PERTURBATION_FAILED:{type(exc).__name__}")
        checks["R012"] = 1.0 if diameter_ok else 0.0
    return checks, failures


def task_checks(workbook, engine, oracle, contract, oracle_module):
    checks, failures = {}, []
    def validate_group(criterion_id, expected):
        ok = True
        for address, target in expected.items():
            if not formula_present(workbook, address):
                ok = False
                failures.append(f"FORMULA_REQUIRED:{address}")
            sheet, cell = address.split("!", 1)
            try:
                observed = engine.value(sheet, cell)
            except Exception as exc:
                observed = None
                failures.append(f"SIZING_EVALUATION_FAILED:{address}:{type(exc).__name__}")
            if not close(observed, target, 0.01):
                ok = False
                failures.append(f"SIZING_VALUE_MISMATCH:{address}")
        checks[criterion_id] = 1.0 if ok else 0.0
        return ok

    conversion_ok = validate_group("R002", {
        "Unit_Conversions!C4": oracle["flow_m3s"],
        "Unit_Conversions!C5": oracle["diameter_m"],
        "Unit_Conversions!C6": oracle["area_m2"],
    })
    head_ok = validate_group("R003", {
        "Calculations!B4": oracle["velocity"], "Calculations!B5": oracle["velocity_head"],
        "Calculations!B6": oracle["friction_head"], "Calculations!B7": oracle["tdh"],
    })
    power_ok = validate_group("R004", {
        "Calculations!B8": oracle["hydraulic_power"], "Calculations!B9": oracle["shaft_power"],
    })
    safety_ok = validate_group("R005", {
        "Calculations!B10": oracle["minimum_flow"], "Calculations!B11": oracle["minimum_motor"],
    })

    eligibility = [engine.value("Equipment_Selection", f"E{row}") for row in range(4, 7)]
    eligibility_formula_ok = all(formula_present(workbook, f"Equipment_Selection!E{row}") for row in range(4, 7))
    eligibility_ok = eligibility_formula_ok and [eligibility_status(value) for value in eligibility] == [eligibility_status(value) for value in oracle["eligibility"]]
    checks["R006"] = 1.0 if eligibility_ok else 0.0
    selected = engine.value("Equipment_Selection", "B7")
    selection_ok = formula_present(workbook, "Equipment_Selection!B7") and selected == oracle["selected_pump"]
    checks["R007"] = 1.0 if selection_ok else 0.0

    perturbations = {row["name"]: row for row in contract["perturbations"]}

    def perturbation_matches(case):
        replay = FormulaEngine(workbook, case["overrides"])
        for address, target in {**case["expected"], **case["invariants"]}.items():
            sheet, cell = address.split("!", 1)
            observed = replay.value(sheet, cell)
            if isinstance(target, (int, float)):
                if not close(observed, target, 0.01):
                    return False
            elif observed != target:
                return False
        if "expected_eligibility" in case:
            observed = [replay.value("Equipment_Selection", f"E{row}") for row in range(4, 7)]
            if [eligibility_status(value) for value in observed] != [eligibility_status(value) for value in case["expected_eligibility"]]:
                return False
        return True

    flow_case = perturbations["design_flow"]
    try:
        if "expected" in flow_case:
            dynamic_ok = perturbation_matches(flow_case)
        else:
            flow_value = flow_case["overrides"]["Inputs!B4"]
            dynamic_engine = FormulaEngine(workbook, flow_case["overrides"])
            expected_dynamic = oracle_module.recompute(flow_lps=flow_value)
            dynamic_ok = close(dynamic_engine.value("Calculations", "B10"), expected_dynamic["minimum_flow"], 0.01) and dynamic_engine.value("Equipment_Selection", "B7") == expected_dynamic["selected_pump"]
    except Exception as exc:
        dynamic_ok = False
        failures.append(f"FLOW_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R008"] = 1.0 if dynamic_ok else 0.0

    diameter_ok = True
    checks["R012"] = 0.0
    if ACTIVE_SPLIT == "dev":
        try:
            diameter_ok = perturbation_matches(perturbations["internal_diameter"])
        except Exception as exc:
            diameter_ok = False
            failures.append(f"DIAMETER_PERTURBATION_FAILED:{type(exc).__name__}")
        checks["R012"] = 1.0 if diameter_ok else 0.0
    expected_units = {
        "Unit_Conversions!D4": "m³/s", "Unit_Conversions!D5": "m", "Unit_Conversions!D6": "m²",
        "Calculations!C4": "m/s", "Calculations!C5": "m", "Calculations!C6": "m",
        "Calculations!C7": "m", "Calculations!C8": "kW", "Calculations!C9": "kW",
        "Calculations!C10": "L/s", "Calculations!C11": "kW",
    }
    units_ok = all(normalized_unit(workbook[sheet][cell].value) == normalized_unit(value) for key, value in expected_units.items() for sheet, cell in [key.split("!")])
    checks["R009"] = 1.0 if units_ok else 0.0
    try:
        check_ok = all(normalized_text(engine.value("Checks", f"D{row}")) == "pass" for row in (4, 5, 6))
    except Exception as exc:
        check_ok = False
        failures.append(f"CHECK_ROW_EVALUATION_FAILED:{type(exc).__name__}")
    checks["R010"] = 1.0 if check_ok else 0.0
    protected_ok = all(workbook[sheet][cell].value == value for key, value in contract["protected"].items() for sheet, cell in [key.split("!", 1)])
    checks["R011"] = 1.0 if protected_ok else 0.0
    return checks, failures


def evaluate(candidate):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return criteria, ["OUTPUT_MISSING"]
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    contract = load_contract()
    canonical_sheets = contract.get("canonical_reference_sheets", contract.get("required_sheets", []))
    role_workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, canonical_sheets, SHEET_ALIASES
    )
    try:
        oracle, oracle_module = load_oracle(contract)
        split_truth = single_sheet_split_truth(contract, oracle_module)
        layout = build_semantic_layout(workbook, split_truth["pump_ids"])
        checks, task_failures = task_checks_single_sheet(
            workbook, oracle, contract, oracle_module, layout, split_truth
        )
        criteria["R001"] = 1.0
        criteria.update(checks)
        return criteria, sorted(set(task_failures))
    except UnsupportedSingleSheetLayout as semantic_exc:
        if unresolved or ambiguous:
            return criteria, [f"UNSUPPORTED_LAYOUT:A_ENG_SEMANTIC:{semantic_exc}"]
        criteria["R001"] = 1.0
        failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
        try:
            oracle, oracle_module = load_oracle(contract)
            checks, task_failures = task_checks(role_workbook, FormulaEngine(role_workbook), oracle, contract, oracle_module)
            criteria.update(checks)
            failures.extend(task_failures)
        except Exception as exc:
            failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    except Exception as exc:
        failures.append(f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return criteria, sorted(set(failures))



def main():
    candidate = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/output/answer.xlsx")
    criteria, failures = evaluate(candidate)
    payload = build_result(
        task=TASK,
        split=ACTIVE_SPLIT,
        candidate=str(candidate),
        criteria=criteria,
        failures=failures,
    )
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True)
        (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\\n")
        if total is not None:
            (log_root / "reward.txt").write_text(str(total) + "\\n")
    except OSError:
        pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

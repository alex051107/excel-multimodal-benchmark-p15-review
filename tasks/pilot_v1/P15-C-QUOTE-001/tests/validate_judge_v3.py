#!/usr/bin/env python3
"""Focused Judge V3 regression: business totals independent of hidden formulas/rows."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


TASK_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TASK_DIR / "tests"))
SPEC = importlib.util.spec_from_file_location("quote_evaluator", TASK_DIR / "tests" / "evaluate.py")
EVALUATOR = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(EVALUATOR)


def find_value_cell(sheet, label: str):
    expected = EVALUATOR.norm(label)
    for row in sheet.iter_rows():
        for cell in row[:-1]:
            if EVALUATOR.norm(cell.value).startswith(expected):
                return sheet.cell(cell.row, cell.column + 1)
    raise AssertionError(f"label not found: {label}")


def build_static_shifted_layout(path: Path) -> None:
    oracle = EVALUATOR.load_oracle("dev")
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    header = workbook["Quote_Header"]
    for label, value in (
        ("Base scope subtotal", oracle["base"]),
        ("Discount", -oracle["discount"]),
        ("Tax", oracle["tax"]),
        ("Base-scope total", oracle["total"]),
        ("Optional alternate", oracle["alternate"]),
    ):
        find_value_cell(header, label).value = value
    for sheet in workbook.worksheets:
        sheet.insert_rows(2, amount=2)
    workbook.save(path)


def build_renamed_sheets(path: Path) -> None:
    workbook = load_workbook(TASK_DIR / "solution/reference.xlsx", data_only=False)
    names = ("Overview", "Scope rollup", "Quoted lines", "Buyer options", "Source register", "Review notes")
    for sheet, name in zip(workbook.worksheets, names):
        sheet.title = name
    workbook.save(path)


def result_record(path: Path, expected: str, predicate, display_path: str | None = None) -> dict:
    result = EVALUATOR.score(path, "dev")
    return {
        "path": display_path or str(path.relative_to(TASK_DIR)),
        "score": result["normalized_score"],
        "pass": result["pass"],
        "criterion_scores": result["criterion_scores"],
        "failure_codes": result["failure_codes"],
        "expected": expected,
        "expected_pass": expected.startswith(("full pass", "pass with")),
        "expectation_met": bool(predicate(result)),
    }


def main() -> None:
    renamed_path = TASK_DIR / "fixtures/equivalent/candidate_renamed_sheets.xlsx"
    build_renamed_sheets(renamed_path)
    fixtures = {
        "reference": (TASK_DIR / "solution/reference.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "equivalent": (TASK_DIR / "fixtures/equivalent/candidate_equivalent.xlsx", "full pass", lambda r: r["pass"] and r["normalized_score"] == 1.0),
        "noop": (TASK_DIR / "fixtures/noop/candidate_noop.xlsx", "fail", lambda r: not r["pass"]),
        "malformed": (TASK_DIR / "fixtures/malformed/candidate_malformed.xlsx", "fail with zero score", lambda r: not r["pass"] and r["normalized_score"] == 0.0),
    }
    for path in sorted((TASK_DIR / "fixtures/mutants").glob("*.xlsx")):
        fixtures[f"mutant:{path.stem}"] = (path, "fail", lambda r: not r["pass"])
    fixture_results = {
        name: result_record(path, expected, predicate)
        for name, (path, expected, predicate) in fixtures.items()
    }

    with tempfile.TemporaryDirectory(prefix="quote-judge-v3-") as directory:
        shifted = Path(directory) / "static_shifted.xlsx"
        build_static_shifted_layout(shifted)
        shifted_result = result_record(
            shifted,
            "full pass: correct static totals, negative discount, and shifted table rows",
            lambda r: r["pass"] and r["normalized_score"] == 1.0,
            "generated:static_shifted_layout",
        )
        broken = Path(directory) / "wrong_total.xlsx"
        workbook = load_workbook(shifted)
        find_value_cell(workbook["Quote_Header"], "Base-scope total").value = 999.0
        workbook.save(broken)
        broken_result = result_record(
            broken,
            "fail R006/R007: base-scope total is arithmetically wrong",
            lambda r: not r["pass"] and r["criterion_scores"]["R006"] == 0.0 and r["criterion_scores"]["R007"] == 0.0,
            "generated:wrong_total",
        )

    renamed_result = result_record(
        renamed_path,
        "full pass: every sheet is renamed while visible headers preserve its business role",
        lambda r: r["pass"] and r["normalized_score"] == 1.0,
    )
    new_cases = {"renamed_sheets": renamed_result, "static_shifted_layout": shifted_result, "wrong_total": broken_result}
    all_met = all(item["expectation_met"] for item in (*fixture_results.values(), *new_cases.values()))
    receipt = {
        "task_id": "P15-C-QUOTE-001",
        "judge_version": "p15_v3.0",
        "validation_scope": "reference, equivalent, noop, malformed, all mutants, shifted static layout, genuine total error",
        "expectation_basis": {
            "static_shifted_layout": "The instruction asks for correct scope grouping and totals; it does not require formulas, fixed rows, or a positive discount convention.",
            "renamed_sheets": "The instruction does not prescribe worksheet names; visible headers and labels establish each business role.",
            "wrong_total": "The quoted base-scope total is a required business result and must reconcile arithmetically.",
        },
        "fixtures": fixture_results,
        "new_cases": new_cases,
        "all_expectations_met": all_met,
    }
    receipt_path = TASK_DIR / "receipts/judge_v3_local_validation.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"task_id": receipt["task_id"], "all_expectations_met": all_met, "scores": {key: [value["score"], value["pass"]] for key, value in {**fixture_results, **new_cases}.items()}}, ensure_ascii=False))
    if not all_met:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

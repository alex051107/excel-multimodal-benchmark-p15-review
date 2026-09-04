#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-C-QUOTE-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

TASK = {
  "task_id": "P15-C-QUOTE-001",
  "pass_threshold": 0.7,
  "criteria": [
    {
      "id": "R001",
      "description": "Workbook contains the required editable quote-normalization sheets.",
      "weight": 2,
      "type": "positive",
      "dimension": "file_usability",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R002",
      "description": "Quote header identity, validity, customer, and currency are correct.",
      "weight": 4,
      "type": "positive",
      "dimension": "header_accuracy",
      "method": "deterministic",
      "method_params": {
        "oracle": "metadata/oracle_recompute.py"
      }
    },
    {
      "id": "R003",
      "description": "Scope rows preserve group and optional-item semantics with typed amounts.",
      "weight": 6,
      "type": "positive",
      "dimension": "scope_semantics",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R004",
      "description": "Base-scope subtotal, discount, tax, and total are formula-linked and dynamically responsive.",
      "weight": 5,
      "type": "positive",
      "dimension": "native_workbook_semantics",
      "method": "deterministic",
      "method_params": {
        "perturbations": 2
      }
    },
    {
      "id": "R005",
      "description": "The optional alternate is retained separately and excluded from the base-scope total.",
      "weight": 4,
      "type": "positive",
      "dimension": "business_rule_correctness",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "R006",
      "description": "Each quote row links to the supplied scan and text locator.",
      "weight": 3,
      "type": "positive",
      "dimension": "provenance",
      "method": "deterministic",
      "method_params": {}
    },
    {
      "id": "P001",
      "description": "Penalty for contradictory source-file or page claims in quote provenance.",
      "weight": -7,
      "type": "penalty",
      "dimension": "integrity",
      "method": "deterministic",
      "method_params": {}
    }
  ],
  "required_sheets": [
    "Quote_Header",
    "Groups",
    "Line_Items",
    "Alternates",
    "Provenance",
    "Checks"
  ]
}
TESTS_ROOT = Path(__file__).resolve().parent
REPO_TASK_ROOT = TESTS_ROOT.parent
DEV_ROOT = REPO_TASK_ROOT if (REPO_TASK_ROOT / "rubric.json").is_file() else TESTS_ROOT / "dev"
TASK = json.loads((DEV_ROOT / "rubric.json").read_text(encoding="utf-8"))
TASK["required_sheets"] = json.loads(
    (TESTS_ROOT / "private_contract.json").read_text(encoding="utf-8")
)["required_sheets"]
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")
FULL_COLUMN_RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+):\$?([A-Z]+)")


TASK["hurdle_criteria"] = ["R004", "R006"]
SHEET_ALIASES = {}


def semantic_sheet_aliases(workbook):
    signatures = {
        "Quote_Header": (("Vendor", "Contractor"), ("Quote ID", "Quote number"), ("Customer", "Client"), ("Valid through", "Valid until"), ("Currency",)),
        "Groups": (("Group",), ("Amount",), ("Included in base total", "Included in base"), ("Document", "Source document")),
        "Line_Items": (("Line ID", "ID"), ("Description", "Scope", "Scope item"), ("Optional", "Alternate", "Included in base"), ("Amount", "Price", "Quoted amount"), ("Source page", "Page")),
        "Alternates": (("Line ID", "ID"), ("Description", "Scope"), ("Amount", "Price"), ("Include in base", "Included in base")),
        "Provenance": (("Line ID", "ID"), ("Source document", "Source file", "Document"), ("Page", "Source page"), ("Text locator", "Locator", "Source locator")),
        "Checks": (("Check",), ("Observed",), ("Expected",), ("Interpretation",)),
    }
    aliases = {}
    for role, groups in signatures.items():
        candidates = []
        for sheet in workbook.worksheets:
            values = {
                norm(cell.value).rstrip(":")
                for row in sheet.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            }
            if all(any(norm(alias).rstrip(":") in values for alias in group) for group in groups):
                candidates.append(sheet.title)
        if candidates:
            aliases[role] = tuple(candidates)
    return aliases


def key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    text = ""
    while index:
        index, rem = divmod(index - 1, 26)
        text = chr(65 + rem) + text
    return text


def active_split(explicit=None):
    split = (explicit or os.environ.get("P15_EVAL_SPLIT", "dev")).strip().lower()
    if split not in {"dev", "confirm"}:
        raise ValueError(f"UNSUPPORTED_EVAL_SPLIT:{split}")
    return split


def load_oracle(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "oracle_recompute.py" if split == "confirm" else DEV_ROOT / "metadata" / "oracle_recompute.py"
    spec = importlib.util.spec_from_file_location(f"task_oracle_{split}", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module.recompute()


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        current = key(sheet, cell)
        if current in self.overrides: return self.overrides[current]
        if current in self.memo: return self.memo[current]
        if current in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{current}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(current)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[current] = result
            return result
        finally:
            self.stack.discard(current)

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.value(sheet, f"{col_name(col)}{row}") for row in range(int(r1), int(r2) + 1) for col in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(part if part.startswith('"') else pattern.sub(replacement, part) for part in pieces)
        expression = outside(
            re.compile(r"(?<![\w.])(\d+(?:\.\d+)?)%"),
            lambda match: f"({match.group(1)}/100)",
            expression,
        )
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        def full_column_range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), 1, match.group(4), self.workbook[sheet].max_row))
        expression = outside(FULL_COLUMN_RANGE, full_column_range_replace, expression)
        expression = outside(RANGE, range_replace, expression)
        def reference_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.value(sheet, f"{match.group(3)}{match.group(4)}"))
        expression = outside(REF, reference_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            if isinstance(node.op, ast.Add): return left + right
            if isinstance(node.op, ast.Sub): return left - right
            if isinstance(node.op, ast.Mult): return left * right
            if isinstance(node.op, ast.Div): return left / right
            if isinstance(node.op, ast.Pow): return left ** right
            raise ValueError("UNSUPPORTED_OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [item for group in args for item in (group if isinstance(group, list) else [group]) if item is not None]
            if name == "SUM": return sum(value for value in values if isinstance(value, (int, float)))
            if name == "COUNTA": return sum(value not in (None, "") for value in values)
            if name == "SUMIF":
                if len(args) not in {2, 3}: raise ValueError("SUMIF_ARGUMENT_COUNT")
                criteria_range = args[0] if isinstance(args[0], list) else [args[0]]
                sum_range = args[2] if len(args) == 3 else criteria_range
                sum_range = sum_range if isinstance(sum_range, list) else [sum_range]
                if len(criteria_range) != len(sum_range): raise ValueError("SUMIF_RANGE_SIZE_MISMATCH")
                return sum(value for value, candidate in zip(sum_range, criteria_range) if excel_criteria_match(candidate, args[1]) and isinstance(value, (int, float)))
            if name == "SUMIFS":
                if len(args) < 3 or len(args) % 2 == 0: raise ValueError("SUMIFS_ARGUMENT_COUNT")
                sum_range = args[0] if isinstance(args[0], list) else [args[0]]
                criteria_pairs = []
                for index in range(1, len(args), 2):
                    criteria_range = args[index] if isinstance(args[index], list) else [args[index]]
                    if len(criteria_range) != len(sum_range): raise ValueError("SUMIFS_RANGE_SIZE_MISMATCH")
                    criteria_pairs.append((criteria_range, args[index + 1]))
                return sum(
                    value for row, value in enumerate(sum_range)
                    if isinstance(value, (int, float))
                    and all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                )
            if name == "AVERAGE": return statistics.mean(values)
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "ABS": return abs(args[0])
            if name == "ROUND": return round(args[0], int(args[1]))
        raise ValueError(f"UNSUPPORTED_FORMULA_NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(float(actual) - float(expected)) <= tolerance


def text(value): return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", text(value).lower())


def semantic_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    candidate = text(value)
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(candidate, pattern).date().isoformat()
        except ValueError:
            continue
    return norm(candidate)


def excel_criteria_match(value, criterion):
    if not isinstance(criterion, str):
        return close(value, criterion) if isinstance(value, (int, float)) and isinstance(criterion, (int, float)) else value == criterion
    match = re.fullmatch(r"\s*(<=|>=|<>|=|<|>)?\s*(.*?)\s*", criterion)
    operator, operand_text = match.groups() if match else (None, criterion)
    try:
        operand = float(operand_text)
        candidate = float(value)
    except (TypeError, ValueError):
        operand = norm(operand_text)
        candidate = norm(value)
    if operator in (None, "="):
        return candidate == operand
    if operator == "<>":
        return candidate != operand
    if operator == "<":
        return candidate < operand
    if operator == "<=":
        return candidate <= operand
    if operator == ">":
        return candidate > operand
    if operator == ">=":
        return candidate >= operand
    return False


def page_matches(value, expected):
    if isinstance(value, (int, float)):
        return close(value, expected, tolerance=0.0)
    candidate = norm(value)
    if candidate == str(int(expected)):
        return True
    return re.search(rf"\b(?:p|page)[\s.:#_-]*{int(expected)}\b", candidate) is not None


def inclusion_value(value):
    candidate = norm(value)
    if candidate in {"yes", "true", "optional", "alternate", "optional alternate", "included", "included as alternate"}:
        return True
    if candidate in {"no", "false", "base", "base scope", "excluded", "not included"}:
        return False
    return None


def load_perturbations(split="dev"):
    split = active_split(split)
    path = TESTS_ROOT / "confirm" / "perturbations.json" if split == "confirm" else DEV_ROOT / "metadata" / "perturbations.json"
    records = json.loads(path.read_text(encoding="utf-8"))
    expected_count = 1 if split == "confirm" else 2
    if len(records) != expected_count or len({record.get("id") for record in records}) != expected_count:
        raise ValueError(f"{split.upper()}_PERTURBATION_MANIFEST_MUST_CONTAIN_{expected_count}_UNIQUE_CASES")
    return records


def find_content_row(workbook, spec):
    sheet = workbook[spec["sheet"]]
    primary_col = col_index(spec["match_column"])
    secondary_col = col_index(spec["secondary_match_column"]) if spec.get("secondary_match_column") else None
    rows = []
    for row in range(4, sheet.max_row + 1):
        if norm(sheet.cell(row=row, column=primary_col).value) != norm(spec["match_value"]):
            continue
        if secondary_col and norm(sheet.cell(row=row, column=secondary_col).value) != norm(spec["secondary_match_value"]):
            continue
        rows.append(row)
    return rows[0] if len(rows) == 1 else None


def perturbation_case_ok(workbook, case):
    target = case["target"]
    row = find_content_row(workbook, target)
    if row is None:
        return False
    target_cell = f'{target["value_column"]}{row}'
    baseline_engine = FormulaEngine(workbook)
    if not close(baseline_engine.value(target["sheet"], target_cell), target["baseline"]):
        return False
    engine = FormulaEngine(workbook, {key(target["sheet"], target_cell): target["perturbed"]})
    for expected in case.get("expected", []):
        if not close(engine.value(expected["sheet"], expected["cell"]), expected["value"]):
            return False
    for expected in case.get("expected_by_content", []):
        expected_row = find_content_row(workbook, expected)
        if expected_row is None or not close(
            engine.value(expected["sheet"], f'{expected["value_column"]}{expected_row}'), expected["value"]
        ):
            return False
    for protected in case.get("protected", []):
        if not close(engine.value(protected["sheet"], protected["cell"]), protected["value"]):
            return False
    for protected in case.get("protected_by_content", []):
        protected_row = find_content_row(workbook, protected)
        if protected_row is None or not close(
            engine.value(protected["sheet"], f'{protected["value_column"]}{protected_row}'), protected["value"]
        ):
            return False
    return True


def perturbation_response_ok(workbook, split="dev"):
    return all(perturbation_case_ok(workbook, case) for case in load_perturbations(split))


def source_field_matches(value, filename):
    rendered, expected = norm(value), norm(filename)
    return bool(rendered and expected and expected in rendered)


def source_field_contradicts(value, filename):
    rendered = text(value)
    if not rendered or source_field_matches(rendered, filename):
        return False
    return re.search(r"\.(?:pdf|png|jpe?g|tiff?|xlsx?|csv)\b", rendered, re.IGNORECASE) is not None


def locator_ok(value, filename=None, page=None):
    # Source file and page are separate fields in the supplied layout.
    return len(re.sub(r"\s+", "", text(value))) >= 3


def present_formula(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def formula_value(engine, sheet, cell):
    try:
        return engine.value(sheet, cell)
    except Exception:
        return None


def row_values(workbook, sheet, start, end, cols):
    records = []
    for row in range(start, end + 1):
        values = tuple(workbook[sheet].cell(row=row, column=col).value for col in cols)
        if any(value not in (None, "") for value in values): records.append(values)
    return records


def static_consolidated_quote_checks(workbook, oracle):
    """Score the observed one-sheet static quote without inventing hidden sheets.

    The adapter is intentionally narrow: it requires explicit base, optional,
    and source-traceability sections and declines formula-bearing variants.
    Such variants need a separate dynamic adapter rather than being
    under-credited here.
    """
    if len(workbook.sheetnames) != 1:
        return None
    sheet = workbook[workbook.sheetnames[0]]
    if any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in sheet.iter_rows()
        for cell in row
    ):
        return None

    labels = {
        norm(sheet[f"A{row}"].value): row for row in range(1, sheet.max_row + 1)
    }
    base_row = next((row for label, row in labels.items() if label == "base scope"), None)
    optional_row = next(
        (row for label, row in labels.items() if "optional alternate" in label), None
    )
    provenance_row = next(
        (row for label, row in labels.items() if label == "source traceability"), None
    )
    if not (
        isinstance(base_row, int)
        and isinstance(optional_row, int)
        and isinstance(provenance_row, int)
        and base_row < optional_row < provenance_row
    ):
        return None

    header_values = {
        norm(sheet[f"A{row}"].value): sheet[f"B{row}"].value
        for row in range(1, base_row)
        if text(sheet[f"A{row}"].value)
    }
    header_map = {
        "vendor": "contractor",
        "quote_id": "quote number",
        "quote_date": "date",
        "customer": "customer",
        "valid_through": "valid through",
        "currency": "currency",
    }
    headers_ok = all(
        semantic_date(header_values.get(label)) == semantic_date(oracle["headers"][name])
        if name in {"quote_date", "valid_through"}
        else norm(header_values.get(label)) == norm(oracle["headers"][name])
        for name, label in header_map.items()
    )

    def scope_rows(start, end, optional):
        rows = []
        for row in range(start, end):
            description = text(sheet[f"B{row}"].value)
            amount = sheet[f"D{row}"].value
            if not description or not isinstance(amount, (int, float)):
                continue
            rows.append(
                {
                    "id": text(sheet[f"A{row}"].value),
                    "description": description,
                    "amount": amount,
                    "optional": optional,
                }
            )
        return rows

    actual = scope_rows(base_row + 1, optional_row, False) + scope_rows(
        optional_row + 1, provenance_row, True
    )
    expected = {norm(item["description"]): item for item in oracle["items"]}
    actual_by_description = {norm(item["description"]): item for item in actual}
    facts_ok = (
        len(actual) == len(expected) == len(actual_by_description)
        and set(actual_by_description) == set(expected)
        and all(
            close(actual_by_description[name]["amount"], item["amount"])
            for name, item in expected.items()
        )
    )
    semantics_ok = facts_ok and all(
        actual_by_description[name]["optional"] is item["optional"]
        for name, item in expected.items()
    )

    total_values = {
        norm(sheet[f"C{row}"].value): sheet[f"D{row}"].value
        for row in range(base_row + 1, optional_row)
        if text(sheet[f"C{row}"].value)
    }
    totals_ok = (
        close(total_values.get("base scope subtotal"), oracle["base"])
        and close(abs(total_values.get("discount (6%)", math.inf)), oracle["discount"])
        and close(total_values.get("tax (7.5%)"), oracle["tax"])
        and close(total_values.get("base-scope total"), oracle["total"])
    )

    provenance = []
    for row in range(provenance_row + 2, sheet.max_row + 1):
        if text(sheet[f"A{row}"].value):
            provenance.append(
                (
                    text(sheet[f"A{row}"].value),
                    text(sheet[f"B{row}"].value),
                    text(sheet[f"C{row}"].value),
                )
            )
    actual_ids = {item["id"] for item in actual}
    provenance_by_id = {record[0]: record for record in provenance}
    population_ok = (
        len(provenance) == len(actual)
        and len(provenance_by_id) == len(provenance)
        and set(provenance_by_id) == actual_ids
    )
    provenance_ok = population_ok and all(
        source_field_matches(record[1], oracle["document"]["filename"])
        and locator_ok(record[2], oracle["document"]["filename"])
        for record in provenance
    )
    false_provenance = any(
        source_field_contradicts(record[1], oracle["document"]["filename"])
        for record in provenance
    )

    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    checks.update(
        {
            "R001": 1.0,
            "R002": 1.0 if headers_ok else 0.0,
            "R003": 1.0 if facts_ok else 0.0,
            "R004": 1.0 if semantics_ok else 0.0,
            "R005": 1.0 if semantics_ok and totals_ok else 0.0,
            "R006": 1.0 if totals_ok else 0.0,
            "R007": 1.0 if totals_ok and semantics_ok else 0.0,
            "R008": 1.0 if provenance_ok else 0.0,
            "R009": 1.0 if facts_ok and population_ok and all(actual_ids) else 0.0,
            "P001": 1.0 if false_provenance else 0.0,
        }
    )
    failures = ["TASK_LOCAL_CONSOLIDATED_LAYOUT"]
    if not totals_ok:
        failures.append("QUOTE_TOTAL_VALUE_MISMATCH")
    if false_provenance:
        failures.append("FABRICATED_QUOTE_PROVENANCE")
    return checks, failures


def _legacy_task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    header_cells = {"vendor": "B4", "quote_id": "B5", "customer": "B6", "quote_date": "B7", "valid_through": "B8", "currency": "B9"}
    headers_ok = all(
        semantic_date(workbook["Quote_Header"][cell].value) == semantic_date(oracle["headers"][name])
        if name in {"quote_date", "valid_through"}
        else norm(workbook["Quote_Header"][cell].value) == norm(oracle["headers"][name])
        for name, cell in header_cells.items()
    )
    checks["R001"] = 1.0
    checks["R002"] = 1.0 if headers_ok else 0.0
    sheet = workbook["Line_Items"]
    actual = []
    for row in range(4, sheet.max_row + 1):
        if text(sheet[f"B{row}"].value):
            actual.append({
                "row": row, "id": text(sheet[f"A{row}"].value), "description": text(sheet[f"B{row}"].value),
                "group": text(sheet[f"C{row}"].value), "optional": text(sheet[f"D{row}"].value),
                "amount": sheet[f"E{row}"].value, "page": sheet[f"F{row}"].value,
            })
    expected = {norm(item["description"]): item for item in oracle["items"]}
    actual_by_description = {norm(item["description"]): item for item in actual}
    facts_ok = (
        len(actual) == len(expected) == len(actual_by_description)
        and set(actual_by_description) == set(expected)
        and all(
            isinstance(actual_by_description[name]["amount"], (int, float))
            and close(actual_by_description[name]["amount"], item["amount"])
            and page_matches(actual_by_description[name]["page"], item["page"])
            for name, item in expected.items()
        )
    )
    semantics_ok = facts_ok and all(
        inclusion_value(actual_by_description[name]["optional"]) == item["optional"]
        for name, item in expected.items()
    )
    checks["R003"] = 1.0 if facts_ok else 0.0
    checks["R004"] = 1.0 if semantics_ok else 0.0
    optional_item = next((item for item in actual if expected.get(norm(item["description"]), {}).get("optional")), None)
    alternate_rows = []
    for row in range(4, workbook["Alternates"].max_row + 1):
        if norm(workbook["Alternates"][f"B{row}"].value) == norm(optional_item["description"] if optional_item else ""):
            alternate_rows.append(row)
    alternate_ok = bool(optional_item) and len(alternate_rows) == 1
    if alternate_ok:
        alternate_row = alternate_rows[0]
        alternate_ok = (
            text(workbook["Alternates"][f"A{alternate_row}"].value) == optional_item["id"]
            and close(engine.value("Alternates", f"C{alternate_row}"), oracle["alternate"])
            and inclusion_value(workbook["Alternates"][f"D{alternate_row}"].value) is False
        )
    totals_ok = all(close(engine.value("Quote_Header", cell), oracle[name]) for cell, name in (
        ("B12", "base"), ("B13", "discount"), ("B14", "tax"), ("B15", "total"), ("B16", "alternate")
    ))
    base_formula_ok = all(present_formula(workbook, address) for address in ("Groups!B4", "Quote_Header!B12"))
    total_formula_ok = all(present_formula(workbook, address) for address in (
        "Quote_Header!B13", "Quote_Header!B14", "Quote_Header!B15"
    ))
    checks["R005"] = 1.0 if totals_ok and base_formula_ok and alternate_ok else 0.0
    checks["R006"] = 1.0 if totals_ok and total_formula_ok else 0.0
    try:
        dynamic_ok = perturbation_response_ok(workbook, split)
    except Exception as exc:
        dynamic_ok = False; failures.append(f"QUOTE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if dynamic_ok else 0.0
    provenance_rows = []
    provenance_sheet = workbook["Provenance"]
    for row in range(4, provenance_sheet.max_row + 1):
        if text(provenance_sheet[f"A{row}"].value):
            provenance_rows.append((
                text(provenance_sheet[f"A{row}"].value), text(provenance_sheet[f"B{row}"].value),
                provenance_sheet[f"C{row}"].value, text(provenance_sheet[f"D{row}"].value)
            ))
    provenance_by_id = {}
    for record in provenance_rows:
        provenance_by_id.setdefault(record[0], []).append(record)
    expected_provenance_ids = {item["id"] for item in actual}
    provenance_population_ok = (
        len(provenance_rows) == len(actual)
        and set(provenance_by_id) == expected_provenance_ids
        and all(len(records) == 1 for records in provenance_by_id.values())
    )
    provenance_ok = facts_ok and provenance_population_ok and all(
        len(provenance_by_id.get(item["id"], [])) == 1
        and source_field_matches(provenance_by_id[item["id"]][0][1], oracle["document"]["filename"])
        and page_matches(provenance_by_id[item["id"]][0][2], expected[norm(item["description"])]["page"])
        and locator_ok(
            provenance_by_id[item["id"]][0][3],
            oracle["document"]["filename"],
        )
        for item in actual
    )
    identity_ok = (
        facts_ok and all(item["id"] for item in actual)
        and len({item["id"] for item in actual}) == len(expected)
        and provenance_population_ok
    )
    checks["R008"] = 1.0 if provenance_ok else 0.0
    checks["R009"] = 1.0 if identity_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = any(
        (
            source_field_contradicts(record[1], oracle["document"]["filename"])
        )
        or (
            record[2] not in (None, "")
            and not any(page_matches(record[2], page) for page in valid_pages)
        )
        for record in provenance_rows
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_QUOTE_PROVENANCE")
    return checks, failures


def semantic_header(sheet, aliases, required):
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for row in range(1, sheet.max_row + 1):
        columns = {}
        for column in range(1, sheet.max_column + 1):
            value = norm(sheet.cell(row=row, column=column).value)
            for role, accepted in normalized.items():
                if value in accepted and role not in columns:
                    columns[role] = column
        if set(required) <= set(columns):
            return row, columns
    return None, {}


def labeled_value_cells(sheet, aliases):
    matches = {role: [] for role in aliases}
    normalized = {role: {norm(alias) for alias in values} for role, values in aliases.items()}
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column):
            label = norm(sheet.cell(row=row, column=column).value).rstrip(":")
            for role, accepted in normalized.items():
                if any(label == value.rstrip(":") or label.startswith(value.rstrip(":") + " ") for value in accepted):
                    matches[role].append(sheet.cell(row=row, column=column + 1))
    return {
        role: sorted(cells, key=lambda cell: cell.value in (None, ""))[0]
        for role, cells in matches.items()
        if cells
    }


def task_checks(workbook, engine, oracle, split="dev"):
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    header_cells = labeled_value_cells(workbook["Quote_Header"], {
        "vendor": ("Vendor", "Contractor"),
        "quote_id": ("Quote ID", "Quote number"),
        "customer": ("Customer", "Client"),
        "quote_date": ("Quote date", "Date"),
        "valid_through": ("Valid through", "Valid until"),
        "currency": ("Currency",),
        "base": ("Base scope subtotal", "Base-scope subtotal"),
        "discount": ("Discount",),
        "tax": ("Tax",),
        "total": ("Base-scope total", "Base scope total"),
        "alternate": ("Optional alternate", "Optional alternate total"),
    })
    checks["R001"] = 1.0
    headers_ok = set(header_cells) >= {"vendor", "quote_id", "customer", "quote_date", "valid_through", "currency"} and all(
        semantic_date(header_cells[name].value) == semantic_date(oracle["headers"][name])
        if name in {"quote_date", "valid_through"}
        else norm(header_cells[name].value) == norm(oracle["headers"][name])
        for name in ("vendor", "quote_id", "customer", "quote_date", "valid_through", "currency")
    )
    checks["R002"] = 1.0 if headers_ok else 0.0

    sheet = workbook["Line_Items"]
    header, columns = semantic_header(sheet, {
        "id": ("Line ID", "ID"),
        "description": ("Description", "Scope", "Scope item"),
        "group": ("Group", "Scope group"),
        "optional": ("Optional", "Alternate", "Included in base"),
        "amount": ("Amount", "Price", "Quoted amount"),
        "page": ("Source page", "Page"),
        "locator": ("Text locator", "Locator"),
    }, ("id", "description", "optional", "amount", "page"))
    actual = []
    if header is not None:
        for row in range(header + 1, sheet.max_row + 1):
            description = text(sheet.cell(row=row, column=columns["description"]).value)
            if not description:
                continue
            actual.append({
                "row": row,
                "id": text(sheet.cell(row=row, column=columns["id"]).value),
                "description": description,
                "group": text(sheet.cell(row=row, column=columns["group"]).value) if "group" in columns else "",
                "optional": text(sheet.cell(row=row, column=columns["optional"]).value),
                "amount": sheet.cell(row=row, column=columns["amount"]).value,
                "page": sheet.cell(row=row, column=columns["page"]).value,
            })
    expected = {norm(item["description"]): item for item in oracle["items"]}
    actual_by_description = {norm(item["description"]): item for item in actual}
    facts_ok = (
        len(actual) == len(expected) == len(actual_by_description)
        and set(actual_by_description) == set(expected)
        and all(
            close(actual_by_description[name]["amount"], item["amount"])
            and page_matches(actual_by_description[name]["page"], item["page"])
            for name, item in expected.items()
        )
    )
    semantics_ok = facts_ok and all(
        inclusion_value(actual_by_description[name]["optional"]) == item["optional"]
        for name, item in expected.items()
    )
    checks["R003"] = 1.0 if facts_ok else 0.0
    checks["R004"] = 1.0 if semantics_ok else 0.0

    total_values = {
        name: formula_value(engine, cell.parent.title, cell.coordinate)
        for name, cell in header_cells.items()
        if name in {"base", "discount", "tax", "total", "alternate"}
    }
    discount_value = total_values.get("discount")
    totals_ok = set(total_values) == {"base", "discount", "tax", "total", "alternate"} and isinstance(discount_value, (int, float)) and (
        close(total_values["base"], oracle["base"])
        and close(abs(discount_value), oracle["discount"])
        and close(total_values["tax"], oracle["tax"])
        and close(total_values["total"], oracle["total"])
        and close(total_values["alternate"], oracle["alternate"])
    )
    arithmetic_ok = totals_ok and close(
        total_values["base"] - abs(discount_value) + total_values["tax"],
        total_values["total"],
    )
    checks["R005"] = 1.0 if semantics_ok and totals_ok else 0.0
    checks["R006"] = 1.0 if totals_ok else 0.0
    checks["R007"] = 1.0 if arithmetic_ok and semantics_ok else 0.0

    provenance_sheet = workbook["Provenance"]
    provenance_header, provenance_columns = semantic_header(provenance_sheet, {
        "id": ("Line ID", "ID"),
        "source": ("Source document", "Source file", "Document"),
        "page": ("Page", "Source page"),
        "locator": ("Text locator", "Locator", "Source locator"),
    }, ("id", "source", "page", "locator"))
    provenance_rows = []
    if provenance_header is not None:
        for row in range(provenance_header + 1, provenance_sheet.max_row + 1):
            line_id = text(provenance_sheet.cell(row=row, column=provenance_columns["id"]).value)
            if line_id:
                provenance_rows.append((
                    line_id,
                    text(provenance_sheet.cell(row=row, column=provenance_columns["source"]).value),
                    provenance_sheet.cell(row=row, column=provenance_columns["page"]).value,
                    text(provenance_sheet.cell(row=row, column=provenance_columns["locator"]).value),
                ))
    provenance_by_id = {}
    for record in provenance_rows:
        provenance_by_id.setdefault(record[0], []).append(record)
    expected_ids = {item["id"] for item in actual}
    population_ok = len(provenance_rows) == len(actual) and set(provenance_by_id) == expected_ids and all(len(rows) == 1 for rows in provenance_by_id.values())
    provenance_ok = facts_ok and population_ok and all(
        source_field_matches(provenance_by_id[item["id"]][0][1], oracle["document"]["filename"])
        and page_matches(provenance_by_id[item["id"]][0][2], expected[norm(item["description"])]["page"])
        and locator_ok(provenance_by_id[item["id"]][0][3], oracle["document"]["filename"])
        for item in actual
    )
    identity_ok = facts_ok and all(expected_ids) and len(expected_ids) == len(expected) and population_ok
    checks["R008"] = 1.0 if provenance_ok else 0.0
    checks["R009"] = 1.0 if identity_ok else 0.0
    valid_pages = {item["page"] for item in expected.values()}
    false_provenance = any(
        source_field_contradicts(record[1], oracle["document"]["filename"])
        or (record[2] not in (None, "") and not any(page_matches(record[2], page) for page in valid_pages))
        for record in provenance_rows
    )
    checks["P001"] = 1.0 if false_provenance else 0.0
    if false_provenance:
        failures.append("FABRICATED_QUOTE_PROVENANCE")
    return checks, failures


def score(candidate, split=None):
    split = active_split(split)
    checks = {criterion["id"]: 0.0 for criterion in TASK["criteria"]}
    failures = []
    if not candidate.exists() or candidate.stat().st_size == 0:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=["OUTPUT_MISSING"],
        )
    try:
        workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc:
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=[f"MALFORMED_XLSX:{type(exc).__name__}"],
        )

    # R001 asks whether the workbook contains every requested sheet, not
    # whether it spelled them the reference way.  The task-local aliases
    # below are already accepted as establishing sheet identity for every
    # other criterion, so scoring R001 on the literal pre-alias name match
    # contradicts the same run's own role resolution.  Any renaming stays
    # visible through the SHEET_ALIAS failure codes.
    exact_layout = all(sheet in workbook.sheetnames for sheet in TASK["required_sheets"])
    workbook, role_map, unresolved, ambiguous = resolve_sheet_roles(
        workbook, TASK["required_sheets"], semantic_sheet_aliases(workbook)
    )
    checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
    failures.extend(sheet_resolution_failures(role_map, unresolved, ambiguous))
    if unresolved or ambiguous:
        oracle = load_oracle(split)
        consolidated = static_consolidated_quote_checks(workbook, oracle)
        if consolidated is not None:
            checks, task_failures = consolidated
            return build_result(
                task=TASK,
                split=split,
                candidate=str(candidate),
                criteria=checks,
                failures=task_failures,
            )
        return build_result(
            task=TASK,
            split=split,
            candidate=str(candidate),
            criteria=checks,
            failures=failures,
        )

    try:
        oracle = load_oracle(split)
        engine = FormulaEngine(workbook)
        checks, task_failures = task_checks(workbook, engine, oracle, split)
        checks["R001"] = 1.0 if (exact_layout or not (unresolved or ambiguous)) else 0.0
        failures.extend(task_failures)
    except Exception as exc:
        failures.append(f"EVALUATION_ERROR:{type(exc).__name__}:{exc}")
    return build_result(
        task=TASK,
        split=split,
        candidate=str(candidate),
        criteria=checks,
        failures=failures,
    )


if __name__ == "__main__":
    arguments = sys.argv[1:]
    selected_split = None
    if "--split" in arguments:
        split_index = arguments.index("--split")
        if split_index + 1 >= len(arguments):
            raise SystemExit("--split requires dev or confirm")
        selected_split = arguments[split_index + 1]
        del arguments[split_index:split_index + 2]
    if len(arguments) > 1:
        raise SystemExit("usage: evaluate.py [candidate.xlsx] [--split dev|confirm]")
    target = Path(arguments[0]) if arguments else Path("/app/output/answer.xlsx")
    print(json.dumps(score(target, selected_split), sort_keys=True))

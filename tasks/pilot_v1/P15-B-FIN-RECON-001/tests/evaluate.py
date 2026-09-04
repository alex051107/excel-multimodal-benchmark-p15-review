#!/usr/bin/env python3
"""Deterministic task-specific judge for P15-B-FIN-RECON-001."""
# Reviewer note: keep this implementation aligned with ../rubric.json["review_notes"].
from __future__ import annotations

import ast
import csv
import importlib.util
import json
import math
import os
import re
import statistics
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

from judge_v2_support import build_result, resolve_sheet_roles, sheet_resolution_failures

CRITERIA = [
    {"id": "R001", "description": "A readable workbook contains every requested reconciliation sheet.", "weight": 1, "type": "positive", "dimension": "file_usability", "method": "deterministic", "method_params": {}},
    {"id": "R002", "description": "All ten ledger/subledger records use the correct transaction-date FX rate and USD normalization.", "weight": 4, "type": "positive", "dimension": "fx_normalization", "method": "deterministic", "method_params": {"oracle": "metadata/oracle_recompute.py"}},
    {"id": "R003", "description": "Only in-period invoice pairs are matched, with correct source amounts, approved adjustment, differences, and the $1 tolerance status.", "weight": 4, "type": "positive", "dimension": "reconciliation_correctness", "method": "deterministic", "method_params": {"tolerance_usd": 1.0}},
    {"id": "R004", "description": "The variance bridge and final decision close to the matched and unmatched records.", "weight": 3, "type": "positive", "dimension": "reconciliation_closure", "method": "deterministic", "method_params": {}},
    {"id": "R005", "description": "The workbook shows the transaction-date FX evidence used for normalization.", "weight": 3, "type": "positive", "dimension": "fx_evidence", "method": "deterministic", "method_params": {}},
    {"id": "R006", "description": "The approved ADJ-105 evidence is retained and applied to INV-105 before tolerance is assessed.", "weight": 3, "type": "positive", "dimension": "adjustment_evidence", "method": "deterministic", "method_params": {}},
    {"id": "R007", "description": "Ledger-only and subledger-only exceptions remain distinct with correct identities and amounts.", "weight": 3, "type": "positive", "dimension": "exception_integrity", "method": "deterministic", "method_params": {}},
    {"id": "R008", "description": "Every ledger, subledger, FX-rate, and adjustment-evidence source field is preserved exactly.", "weight": 3, "type": "positive", "dimension": "change_locality", "method": "deterministic", "method_params": {}},
    {"id": "R009", "description": "Checks expose the matched difference, exception count, and final closure residual in reviewable form.", "weight": 2, "type": "positive", "dimension": "auditability", "method": "deterministic", "method_params": {}},
]

TASK = {
    "task_id": "P15-B-FIN-RECON-001",
    "pass_threshold": 0.7,
    "required_sheets": ["Ledger", "Subledger", "FX_Rates", "Adjustment_Evidence", "Normalized_Records", "Matched_Items", "Unmatched_Items", "Variance_Bridge", "Final_Reconciliation", "Checks"],
    "criteria": CRITERIA,
}
REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+)")
RANGE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")


TASK["hurdle_criteria"] = ["R002", "R003", "R004", "R007", "R008"]
SHEET_ALIASES = {
    "Matched_Items": ("Matched Items", "Matched"),
    "Variance_Bridge": ("Variance Bridge",),
    "Adjustment_Evidence": ("Adjustment Evidence",),
    "Final_Reconciliation": ("Decision_Sheet", "Decision Summary", "Final_Decision"),
    "Checks": ("Integrity_Checks", "Integrity Checks"),
}


def cell_key(sheet, cell): return f"{sheet}!{cell}"


def col_index(column):
    value = 0
    for char in column: value = value * 26 + ord(char) - 64
    return value


def col_name(index):
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def task_root(): return Path(__file__).resolve().parents[1]


def split_context(split):
    contract_path = task_root() / "tests" / ("confirm/contract.json" if split == "confirm" else "private_contract.json")
    contract = json.loads(contract_path.read_text())
    drivers = {
        "fx_source_cell": "FX_Rates!C6",
        "fx_invoice": "INV-101",
        "fx_new_value": 1.08,
        "adjustment_source_cell": "Adjustment_Evidence!C4",
        "adjustment_invoice": "INV-105",
        "adjustment_new_value": -40,
        "adjusted_subledger": 260,
        "adjustment_difference": -10,
    }
    drivers.update(contract.get("drivers", {}))
    return {"split": split, "contract": contract, "source_root": task_root() / contract["input_files_dir"], "oracle_path": task_root() / contract["oracle"], "drivers": drivers}


def load_oracle(context):
    global oracle_module
    spec = importlib.util.spec_from_file_location(f"task_oracle_{context['split']}", context["oracle_path"])
    oracle_module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(oracle_module)
    return oracle_module.recompute()


def excel_equal(left, right):
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, str) and isinstance(right, str):
        return left.casefold() == right.casefold()
    return left == right


def excel_criteria_match(value, criteria):
    operator = "="
    expected = criteria
    if isinstance(criteria, str):
        match = re.match(r"^(<=|>=|<>|!=|=|<|>)(.*)$", criteria)
        if match:
            operator, expected = match.groups()
        stripped = expected.strip()
        try:
            expected = float(stripped)
        except (TypeError, ValueError):
            expected = stripped
    if operator in {"=", "=="}:
        return excel_equal(value, expected)
    if operator in {"<>", "!="}:
        return not excel_equal(value, expected)
    try:
        if operator == "<": return value < expected
        if operator == "<=": return value <= expected
        if operator == ">": return value > expected
        if operator == ">=": return value >= expected
    except TypeError:
        return False
    raise ValueError(f"UNSUPPORTED_CRITERIA_OPERATOR:{operator}")


def as_range(value, function_name):
    if not isinstance(value, list):
        raise ValueError(f"{function_name}_REQUIRES_RANGE")
    return value


class UnsupportedFormulaError(ValueError):
    """A valid Excel formula that this bounded replay engine cannot evaluate."""


class FormulaEngine:
    def __init__(self, workbook, overrides=None):
        self.workbook = workbook; self.overrides = overrides or {}; self.memo = {}; self.stack = set()

    def value(self, sheet, cell):
        key = cell_key(sheet, cell)
        if key in self.overrides: return self.overrides[key]
        if key in self.memo: return self.memo[key]
        if key in self.stack: raise ValueError(f"CIRCULAR_REFERENCE:{key}")
        if sheet not in self.workbook.sheetnames: raise ValueError(f"MISSING_SHEET:{sheet}")
        self.stack.add(key)
        try:
            raw = self.workbook[sheet][cell].value
            result = self.formula(raw[1:], sheet) if isinstance(raw, str) and raw.startswith("=") else raw
            self.memo[key] = result
            return result
        finally:
            self.stack.discard(key)

    @staticmethod
    def formula_scalar(value):
        if isinstance(value, datetime): return value.date().isoformat()
        if isinstance(value, date): return value.isoformat()
        return value

    def range_values(self, sheet, c1, r1, c2, r2):
        return [self.formula_scalar(self.value(sheet, f"{col_name(c)}{r}")) for r in range(int(r1), int(r2) + 1) for c in range(col_index(c1), col_index(c2) + 1)]

    def formula(self, expression, current_sheet):
        expression = expression.replace("^", "**").replace("<>", "!=")
        def outside(pattern, replacement, text):
            pieces = re.split(r'("(?:[^"\\]|\\.)*")', text)
            return "".join(piece if piece.startswith('"') else pattern.sub(replacement, piece) for piece in pieces)
        def range_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.range_values(sheet, match.group(3), match.group(4), match.group(5), match.group(6)))
        expression = outside(RANGE, range_replace, expression)
        def ref_replace(match):
            sheet = match.group(1) or match.group(2) or current_sheet
            return repr(self.formula_scalar(self.value(sheet, f"{match.group(3)}{match.group(4)}")))
        expression = outside(REF, ref_replace, expression)
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)
        return self.safe_eval(ast.parse(expression, mode="eval").body)

    def safe_eval(self, node):
        if isinstance(node, ast.Constant): return node.value
        if isinstance(node, ast.List): return [self.safe_eval(item) for item in node.elts]
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.USub, ast.UAdd)):
            value = self.safe_eval(node.operand); return -value if isinstance(node.op, ast.USub) else value
        if isinstance(node, ast.BinOp):
            left, right = self.safe_eval(node.left), self.safe_eval(node.right)
            operations = {ast.Add: lambda a, b: a + b, ast.Sub: lambda a, b: a - b, ast.Mult: lambda a, b: a * b, ast.Div: lambda a, b: a / b, ast.Pow: lambda a, b: a ** b}
            for cls, operation in operations.items():
                if isinstance(node.op, cls): return operation(left, right)
            raise UnsupportedFormulaError("UNSUPPORTED_FORMULA:OPERATOR")
        if isinstance(node, ast.Compare):
            left = self.safe_eval(node.left)
            for operator, comparator in zip(node.ops, node.comparators):
                right = self.safe_eval(comparator)
                ok = left == right if isinstance(operator, ast.Eq) else left != right if isinstance(operator, ast.NotEq) else left > right if isinstance(operator, ast.Gt) else left >= right if isinstance(operator, ast.GtE) else left < right if isinstance(operator, ast.Lt) else left <= right if isinstance(operator, ast.LtE) else False
                if not ok: return False
                left = right
            return True
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            args = [self.safe_eval(item) for item in node.args]; name = node.func.id.upper()
            values = [value for group in args for value in (group if isinstance(group, list) else [group]) if value is not None]
            if name == "SUM": return sum(value for value in values if isinstance(value, (int, float)))
            if name == "SUMIF":
                if len(args) not in {2, 3}:
                    raise ValueError("SUMIF_ARGUMENTS")
                criteria_range = as_range(args[0], name)
                sum_range = as_range(args[2], name) if len(args) == 3 else criteria_range
                if len(criteria_range) != len(sum_range):
                    raise ValueError("SUMIF_RANGE_LENGTH")
                return sum(
                    value
                    for value, candidate in zip(sum_range, criteria_range)
                    if isinstance(value, (int, float)) and excel_criteria_match(candidate, args[1])
                )
            if name == "SUMIFS":
                if len(args) < 3 or len(args) % 2 == 0:
                    raise ValueError("SUMIFS_ARGUMENTS")
                sum_range = as_range(args[0], name)
                criteria_pairs = [
                    (as_range(args[index], name), args[index + 1])
                    for index in range(1, len(args), 2)
                ]
                if any(len(criteria_range) != len(sum_range) for criteria_range, _ in criteria_pairs):
                    raise ValueError("SUMIFS_RANGE_LENGTH")
                return sum(
                    value
                    for row, value in enumerate(sum_range)
                    if isinstance(value, (int, float))
                    and all(excel_criteria_match(criteria_range[row], criterion) for criteria_range, criterion in criteria_pairs)
                )
            if name == "COUNTIF":
                if len(args) != 2:
                    raise ValueError("COUNTIF_ARGUMENTS")
                criteria_range = as_range(args[0], name)
                return sum(excel_criteria_match(value, args[1]) for value in criteria_range)
            if name == "COUNTA": return sum(value not in (None, "") for value in values)
            if name == "AVERAGE": return statistics.mean(values)
            if name == "MAX": return max(values)
            if name == "MIN": return min(values)
            if name == "ABS":
                return [abs(value) for value in args[0]] if isinstance(args[0], list) else abs(args[0])
            if name == "IF": return args[1] if args[0] else args[2]
            if name == "AND": return all(args)
            if name == "ROUND": return round(args[0], int(args[1]))
            if name == "LEFT": return str(args[0])[: int(args[1])]
            if name == "RIGHT": return str(args[0])[-int(args[1]) :]
            if name == "MID": return str(args[0])[int(args[1]) - 1 : int(args[1]) - 1 + int(args[2])]
            if name == "DATE": return f"{int(args[0]):04d}-{int(args[1]):02d}-{int(args[2]):02d}"
            if name == "TEXT":
                if len(args) != 2:
                    raise ValueError("TEXT_ARGUMENTS")
                return str(args[0])
        raise UnsupportedFormulaError(f"UNSUPPORTED_FORMULA:NODE:{ast.dump(node)}")


def close(actual, expected, tolerance=0.01):
    return isinstance(actual, (int, float)) and math.isfinite(actual) and abs(actual - expected) <= tolerance


def formula_present(workbook, address):
    sheet, cell = address.split("!", 1)
    return sheet in workbook.sheetnames and isinstance(workbook[sheet][cell].value, str) and workbook[sheet][cell].value.startswith("=")


def norm(value):
    if isinstance(value, float): return round(value, 6)
    return value


def semantic_token(value):
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def comparable(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return norm(value)


def table_layout(workbook, sheet, aliases):
    worksheet = workbook[sheet]
    normalized = {
        field: {semantic_token(alias) for alias in names}
        for field, names in aliases.items()
    }
    for row in range(1, min(worksheet.max_row, 20) + 1):
        columns = {}
        for column in range(1, worksheet.max_column + 1):
            token = semantic_token(worksheet.cell(row=row, column=column).value)
            for field, accepted in normalized.items():
                if token in accepted and field not in columns:
                    columns[field] = column
        if len(columns) == len(aliases):
            return row, columns
    raise ValueError(f"REQUIRED_HEADERS_NOT_FOUND:{sheet}")


def table_records(workbook, engine, sheet, aliases, identifier):
    header_row, columns = table_layout(workbook, sheet, aliases)
    worksheet = workbook[sheet]
    records = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=columns[identifier]).value in (None, ""):
            continue
        values = {
            field: comparable(engine.value(sheet, f"{col_name(column)}{row}"))
            for field, column in columns.items()
        }
        records.append((row, values))
    return columns, records


def metric_layout(workbook, sheet):
    return table_layout(
        workbook,
        sheet,
        {
            "label": ("metric", "bridge item", "check", "control"),
            "value": ("value", "amount usd", "calculated value", "calculated usd"),
        },
    )


def metric_cell(workbook, sheet, aliases):
    header_row, columns = metric_layout(workbook, sheet)
    accepted = {semantic_token(alias) for alias in aliases}
    worksheet = workbook[sheet]
    for row in range(header_row + 1, worksheet.max_row + 1):
        if semantic_token(worksheet.cell(row=row, column=columns["label"]).value) in accepted:
            return row, columns["value"]
    raise ValueError(f"REQUIRED_METRIC_NOT_FOUND:{sheet}:{'/'.join(sorted(accepted))}")


def metric_value(workbook, engine, sheet, aliases):
    row, column = metric_cell(workbook, sheet, aliases)
    return engine.value(sheet, f"{col_name(column)}{row}"), row, column


def formula_cell(workbook, sheet, row, column):
    value = workbook[sheet].cell(row=row, column=column).value
    return isinstance(value, str) and value.startswith("=")


def normalize_reconciliation_status(value):
    if not isinstance(value, str):
        return value
    text = unicodedata.normalize("NFKC", value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    if text in {"matched", "reconciled", "within tolerance", "within 1 tolerance"}:
        return "MATCHED"
    if text in {
        "out of tolerance",
        "outside tolerance",
        "out of 1 tolerance",
        "outside 1 tolerance",
        "variance exceeds tolerance",
    }:
        return "OUT_OF_TOLERANCE"
    return value


def rows(workbook, sheet, start, end, columns):
    values = []
    for row in range(start, end + 1):
        record = tuple(norm(workbook[sheet].cell(row=row, column=column).value) for column in columns)
        if any(value not in (None, "") for value in record): values.append(record)
    return values


def find_row(workbook, sheet, identifier, id_column=1, start=4, end=30):
    for row in range(start, end + 1):
        if workbook[sheet].cell(row=row, column=id_column).value == identifier:
            return row
    raise ValueError(f"MISSING_RECORD:{sheet}:{identifier}")


def source_tables_match(workbook, root):
    def records(filename):
        with (root / filename).open(newline="") as handle:
            return list(csv.DictReader(handle))
    expected_ledger = [(r["record_id"], r["date"], r["invoice"], r["currency"], float(r["original_amount"])) for r in records("ledger.csv")]
    expected_subledger = [(r["record_id"], r["date"], r["invoice"], r["currency"], float(r["original_amount"])) for r in records("subledger.csv")]
    expected_fx = [(r["date"], r["currency"], float(r["usd_per_unit"])) for r in records("fx_rates.csv")]
    expected_adjustments = [(r["adjustment_id"], r["invoice"], float(r["amount_usd"]), r["reason"]) for r in records("adjustments.csv")]
    actual_ledger = rows(workbook, "Ledger", 4, 8, range(1, 6))
    actual_subledger = rows(workbook, "Subledger", 4, 8, range(1, 6))
    actual_fx = rows(workbook, "FX_Rates", 4, 6, range(1, 4))
    actual_adjustments = rows(workbook, "Adjustment_Evidence", 4, 10, range(1, 5))
    return actual_ledger == expected_ledger and actual_subledger == expected_subledger and actual_fx == expected_fx and actual_adjustments == expected_adjustments


def v3_header(value):
    token = semantic_token(value)
    aliases = {
        "transaction date": "date", "posting date": "date",
        "journal id": "record id", "invoice id": "invoice",
        "source system": "source", "source amount": "original amount",
        "amount": "original amount", "original": "original amount", "exchange rate": "fx rate", "usd per unit": "fx rate",
        "normalized usd": "usd normalized", "usd amount": "usd normalized", "amount usd": "usd normalized",
        "ledger amount usd": "ledger usd", "subledger amount usd": "subledger usd",
        "approved adjustment": "adjustment usd", "adjustment": "adjustment usd",
        "adjusted amount": "adjusted subledger usd", "adjusted subledger": "adjusted subledger usd",
        "variance": "difference", "difference usd": "difference",
        "bridge item": "metric", "bridge step": "metric", "control": "metric", "check": "metric",
        "calculated value": "value", "calculated usd": "value", "observed value": "value",
    }
    return aliases.get(token, token)


def v3_tables(workbook, engine):
    tables = []
    for sheet in workbook.worksheets:
        for header_row in range(1, min(sheet.max_row, 25) + 1):
            headers = {}
            duplicate = False
            for column in range(1, min(sheet.max_column, 30) + 1):
                key = v3_header(sheet.cell(row=header_row, column=column).value)
                if not key:
                    continue
                if key in headers:
                    duplicate = True
                headers[key] = column
            if duplicate or len(headers) < 2:
                continue
            # In reconciliation/controls tables, ``Amount USD`` is the value
            # column.  In normalized/source tables it remains ``usd normalized``.
            if "metric" in headers and "value" not in headers and "usd normalized" in headers:
                headers["value"] = headers["usd normalized"]
            if not (
                {"record id", "invoice"} <= set(headers)
                or {"invoice", "ledger usd", "subledger usd"} <= set(headers)
                or {"invoice", "source", "usd normalized"} <= set(headers)
                or "metric" in headers
                or {"date", "currency"} <= set(headers)
                or "adjustment id" in headers
            ):
                continue
            records = []
            blank_run = 0
            for row in range(header_row + 1, min(sheet.max_row, header_row + 100) + 1):
                record = {}
                for key, column in headers.items():
                    cell = f"{col_name(column)}{row}"
                    try:
                        record[key] = engine.value(sheet.title, cell)
                    except Exception:
                        record[key] = sheet[cell].value
                if all(value in (None, "") for value in record.values()):
                    blank_run += 1
                    if blank_run >= 2 and records:
                        break
                    continue
                blank_run = 0
                records.append(record)
            tables.append({"sheet": sheet.title, "headers": headers, "records": records})
            break
    return tables


def v3_comparable(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 6)
    text = str(value or "").strip()
    if text.startswith("$"):
        text = text[1:].replace(",", "")
    try:
        return round(float(text), 6)
    except ValueError:
        return re.sub(r"\s+", " ", text).casefold()


def v3_equal(left, right):
    return v3_comparable(left) == v3_comparable(right)


def v3_source_preserved(tables, source_root):
    specifications = [
        ("ledger.csv", ("record id", "date", "invoice", "currency", "original amount")),
        ("subledger.csv", ("record id", "date", "invoice", "currency", "original amount")),
        ("fx_rates.csv", ("date", "currency", "fx rate")),
        ("adjustments.csv", ("adjustment id", "invoice", "usd normalized", "reason")),
    ]
    source_fields = {
        "record id": "record_id", "date": "date", "invoice": "invoice", "currency": "currency",
        "original amount": "original_amount", "fx rate": "usd_per_unit", "adjustment id": "adjustment_id",
        "adjustment usd": "amount_usd", "usd normalized": "amount_usd", "reason": "reason",
    }
    for filename, fields in specifications:
        with (source_root / filename).open(newline="") as handle:
            expected = [tuple(row[source_fields[field]] for field in fields) for row in csv.DictReader(handle)]
        matched = False
        for table in tables:
            if not set(fields) <= set(table["headers"]):
                continue
            actual = [tuple(row.get(field) for field in fields) for row in table["records"]]
            if len(actual) == len(expected) and {
                tuple(v3_comparable(value) for value in row) for row in actual
            } == {
                tuple(v3_comparable(value) for value in row) for row in expected
            }:
                matched = True
                break
        if not matched:
            return False
    return True


def close_period_truth(oracle):
    """Derive the requested close period from the dated source rows.

    Both splits contain eight rows in the close month and a later two-row pair
    that must be retained for audit but excluded from matching and closure.
    """
    month_counts = {}
    for row in oracle["normalized"]:
        match = re.match(r"^(\d{4}-\d{2})", str(row[3]))
        if match:
            month_counts[match.group(1)] = month_counts.get(match.group(1), 0) + 1
    if not month_counts:
        raise ValueError("CLOSE_PERIOD_NOT_DERIVABLE")
    close_month = max(month_counts, key=lambda key: (month_counts[key], key))
    scoped = [row for row in oracle["normalized"] if str(row[3]).startswith(close_month)]
    by_source = {"Ledger": {}, "Subledger": {}}
    for row in scoped:
        by_source[row[0]][row[2]] = float(row[7])
    adjustment_by_invoice = {row[0]: float(row[3]) for row in oracle["matched"]}
    matched = []
    unmatched = []
    for invoice in sorted(set(by_source["Ledger"]) | set(by_source["Subledger"])):
        ledger = by_source["Ledger"].get(invoice)
        subledger = by_source["Subledger"].get(invoice)
        if ledger is None:
            unmatched.append((invoice, "Subledger", subledger))
        elif subledger is None:
            unmatched.append((invoice, "Ledger", ledger))
        else:
            adjustment = adjustment_by_invoice.get(invoice, 0.0)
            adjusted = subledger + adjustment
            matched.append((invoice, ledger, subledger, adjustment, adjusted, ledger - adjusted))
    ledger_total = sum(by_source["Ledger"].values())
    subledger_total = sum(by_source["Subledger"].values())
    adjustment_total = sum(row[3] for row in matched)
    adjusted_subledger_total = subledger_total + adjustment_total
    ledger_only = sum(row[2] for row in unmatched if row[1] == "Ledger")
    subledger_only = sum(row[2] for row in unmatched if row[1] == "Subledger")
    return {
        "close_month": close_month,
        "normalized": oracle["normalized"],
        "matched": matched,
        "unmatched": unmatched,
        "ledger_total": ledger_total,
        "subledger_total": subledger_total,
        "adjustment_total": adjustment_total,
        "adjusted_subledger_total": adjusted_subledger_total,
        "matched_total": sum(row[4] for row in matched),
        "ledger_only": ledger_only,
        "subledger_only": subledger_only,
        "investigation": subledger_only - ledger_only,
        "out_of_period_count": len(oracle["normalized"]) - len(scoped),
    }


def normalized_status(value):
    token = semantic_token(value)
    if "matched" in token and "out" not in token and "outside" not in token:
        return "MATCHED"
    return normalize_reconciliation_status(value)


def metric_record_map(table):
    result = {}
    if not {"metric", "value"} <= set(table["headers"]):
        return result
    for record in table["records"]:
        label = semantic_token(record.get("metric"))
        if label:
            result.setdefault(label, []).append(record)
    return result


def metric_match(metric_map, aliases, expected, tolerance=0.01):
    accepted = {semantic_token(alias) for alias in aliases}
    records = [record for label, rows in metric_map.items() if label in accepted for record in rows]
    return bool(records) and all(close(v3_comparable(record.get("value")), expected, tolerance) for record in records)


def metric_match_any(metric_map, aliases, expected_values, tolerance=0.01):
    accepted = {semantic_token(alias) for alias in aliases}
    records = [record for label, rows in metric_map.items() if label in accepted for record in rows]
    return bool(records) and all(
        any(close(v3_comparable(record.get("value")), expected, tolerance) for expected in expected_values)
        for record in records
    )


def semantic_finance_checks(workbook, oracle, context):
    engine = FormulaEngine(workbook)
    tables = v3_tables(workbook, engine)
    expected = close_period_truth(oracle)
    normalized_records = []
    for table in tables:
        headers = set(table["headers"])
        required = {"record id", "invoice", "currency", "original amount", "fx rate", "usd normalized"}
        if not required <= headers:
            continue
        inferred_source = ""
        title = semantic_token(table["sheet"])
        if "subledger" in title:
            inferred_source = "Subledger"
        elif "ledger" in title:
            inferred_source = "Ledger"
        for record in table["records"]:
            source = record.get("source") or inferred_source
            normalized_records.append((
                source, record.get("record id"), record.get("invoice"), record.get("date"),
                record.get("currency"), record.get("original amount"), record.get("fx rate"),
                record.get("usd normalized"),
            ))
    expected_normalized = [tuple(row) for row in expected["normalized"]]
    normalized_ok = len(normalized_records) == len(expected_normalized) and {
        tuple(v3_comparable(value) for value in row) for row in normalized_records
    } == {
        tuple(v3_comparable(value) for value in row) for row in expected_normalized
    }

    actual_matches = []
    matched_status_ok = True
    for table in tables:
        order = ("invoice", "ledger usd", "subledger usd", "adjustment usd", "adjusted subledger usd", "difference")
        if set(order) <= set(table["headers"]) and "status" in table["headers"]:
            for record in table["records"]:
                actual_matches.append(tuple(v3_comparable(record.get(field)) for field in order))
                matched_status_ok = matched_status_ok and normalized_status(record.get("status")) == "MATCHED"
    expected_matches = [tuple(v3_comparable(value) for value in row) for row in expected["matched"]]
    actual_match_ids = [row[0] for row in actual_matches]
    matched_ok = (
        len(actual_matches) == len(expected_matches)
        and len(set(actual_match_ids)) == len(actual_match_ids)
        and sorted(actual_matches) == sorted(expected_matches)
        and matched_status_ok
    )

    unmatched_actual = []
    for table in tables:
        if (
            {"invoice", "source", "usd normalized"} <= set(table["headers"])
            and (
                "unmatched" in semantic_token(table["sheet"])
                or {"reason", "disposition"} & set(table["headers"])
            )
        ):
            for record in table["records"]:
                source = str(record.get("source") or "").casefold()
                source = "Subledger" if "subledger" in source else "Ledger" if "ledger" in source else source
                unmatched_actual.append((v3_comparable(record.get("invoice")), v3_comparable(source), v3_comparable(record.get("usd normalized"))))
        elif {"invoice", "usd normalized"} <= set(table["headers"]) and "unmatched" in semantic_token(table["sheet"]):
            title = semantic_token(table["sheet"])
            source = "Subledger" if "subledger" in title else "Ledger" if "ledger" in title else ""
            for record in table["records"]:
                unmatched_actual.append((v3_comparable(record.get("invoice")), v3_comparable(source), v3_comparable(record.get("usd normalized"))))
    unmatched_expected = [tuple(v3_comparable(value) for value in row) for row in expected["unmatched"]]
    unmatched_keys = [(row[0], row[1]) for row in unmatched_actual]
    unmatched_ok = (
        len(unmatched_actual) == len(unmatched_expected)
        and len(set(unmatched_keys)) == len(unmatched_keys)
        and sorted(unmatched_actual) == sorted(unmatched_expected)
    )

    protected_ok = v3_source_preserved(tables, context["source_root"])

    bridge_ok = False
    final_ok = False
    checks_ok = False
    for table in tables:
        metric_map = metric_record_map(table)
        if not metric_map:
            continue
        bridge_ok = bridge_ok or all((
            metric_match(metric_map, ("close period ledger total", "june ledger total", "september ledger total", "ledger total"), expected["ledger_total"]),
            metric_match(metric_map, ("close period adjusted subledger total", "adjusted june subledger total", "adjusted september subledger total", "adjusted subledger total"), expected["adjusted_subledger_total"]),
            metric_match_any(
                metric_map,
                ("ledger only exception", "ledger only exception amount", "less ledger only exception"),
                (expected["ledger_only"], -expected["ledger_only"]),
            ),
            metric_match(
                metric_map,
                ("subledger only exception", "subledger only exception amount", "add subledger only exception"),
                expected["subledger_only"],
            ),
            metric_match(metric_map, ("bridge residual", "variance bridge residual", "exception bridge closure"), 0.0),
        ))

        table_text = " ".join(str(value or "") for record in table["records"] for value in record.values()).casefold()
        final_metrics_ok = all((
            metric_match(metric_map, ("matched invoices", "matched invoice count"), len(expected["matched"])),
            metric_match_any(metric_map, ("ledger only exceptions", "ledger only exception"), (1, expected["ledger_only"])),
            metric_match_any(metric_map, ("subledger only exceptions", "subledger only exception"), (1, expected["subledger_only"])),
            metric_match(metric_map, ("investigation amount", "adjusted net variance", "net investigation difference"), expected["investigation"]),
        ))
        decision_records = [
            record for label, records in metric_map.items()
            if label in {"final decision", "reconciliation decision", "decision"}
            for record in records
        ]
        decision_ok = bool(decision_records) and all(
            "review ready" in " ".join(str(value or "") for value in record.values()).casefold()
            and ("open exception" in " ".join(str(value or "") for value in record.values()).casefold()
                 or "resolve" in " ".join(str(value or "") for value in record.values()).casefold())
            for record in decision_records
        )
        final_ok = final_ok or (final_metrics_ok and decision_ok)

        checks_ok = checks_ok or all((
            metric_match(metric_map, ("matched item differences", "maximum matched absolute difference", "matched difference", "matched totals agree"), 0.0),
            metric_match(metric_map, ("unmatched exception count", "exception count", "exceptions"), len(expected["unmatched"])),
            metric_match(metric_map, ("final closure residual", "exception bridge closure", "variance bridge residual", "bridge residual"), 0.0),
        ))

    closure_ok = matched_ok and unmatched_ok and bridge_ok and final_ok
    fx_evidence_ok = normalized_ok
    adjustment_ok = any(
        row[3] != 0 and row in expected_matches for row in actual_matches
    )
    requested_roles = sum((normalized_ok, matched_ok, unmatched_ok, bridge_ok, final_ok, checks_ok))
    checks = {row["id"]: 0.0 for row in CRITERIA}
    checks.update({
        "R001": float(requested_roles == 6), "R002": float(normalized_ok), "R003": float(matched_ok),
        "R004": float(closure_ok), "R005": float(fx_evidence_ok), "R006": float(adjustment_ok),
        "R007": float(unmatched_ok), "R008": float(protected_ok), "R009": float(checks_ok),
    })
    failures = []
    if not normalized_ok: failures.append("NORMALIZED_RECORDS_MISMATCH")
    if not matched_ok: failures.append("MATCHED_RECORDS_MISMATCH")
    if not unmatched_ok: failures.append("UNMATCHED_RECORDS_MISMATCH")
    if not protected_ok: failures.append("PROTECTED_FINANCE_SOURCE_CHANGED")
    if not bridge_ok: failures.append("VARIANCE_BRIDGE_VALUES_MISMATCH")
    if not final_ok: failures.append("FINAL_RECONCILIATION_MISMATCH")
    if not checks_ok: failures.append("RECONCILIATION_CHECKS_MISMATCH")
    return checks, failures


def task_checks(workbook, engine, oracle, context):
    checks, failures = {}, []
    expected = oracle
    try:
        normalized_aliases = {
            "source": ("source", "source system"),
            "record_id": ("record id", "journal id"),
            "invoice": ("invoice", "invoice id"),
            "date": ("date", "transaction date"),
            "currency": ("currency",),
            "original": ("original", "original amount", "source amount"),
            "fx_rate": ("fx rate", "exchange rate", "usd per unit"),
            "usd": ("usd normalized", "normalized usd", "usd amount", "amount usd"),
        }
        normalized_columns, normalized_rows = table_records(
            workbook, engine, "Normalized_Records", normalized_aliases, "record_id"
        )
        normalized_actual = {
            (values["source"], values["record_id"]): tuple(
                values[field]
                for field in ("invoice", "date", "currency", "original", "fx_rate", "usd")
            )
            for _, values in normalized_rows
        }
        normalized_expected = {
            (record[0], record[1]): tuple(comparable(value) for value in record[2:])
            for record in expected["normalized"]
        }
        normalized_ok = (
            len(normalized_rows) == len(normalized_expected)
            and len(normalized_actual) == len(normalized_rows)
            and normalized_actual == normalized_expected
        )
        normalized_formulas = all(
            formula_cell(workbook, "Normalized_Records", row, normalized_columns["usd"])
            and (
                values["currency"] == "USD"
                or formula_cell(workbook, "Normalized_Records", row, normalized_columns["fx_rate"])
            )
            for row, values in normalized_rows
        )
    except Exception as exc:
        normalized_ok = normalized_formulas = False
        failures.append(f"FX_REPLAY_FAILED:{type(exc).__name__}")
    checks["R002"] = 1.0 if normalized_ok else 0.0

    try:
        matched_aliases = {
            "invoice": ("invoice", "invoice id"),
            "ledger": ("ledger usd", "ledger amount usd"),
            "subledger": ("subledger usd", "subledger amount usd"),
            "adjustment": ("adjustment usd", "approved adjustment", "adjustment"),
            "adjusted": ("adjusted subledger", "adjusted subledger usd"),
            "difference": ("difference", "variance", "difference usd"),
            "status": ("status", "match status"),
        }
        matched_columns, matched_rows = table_records(
            workbook, engine, "Matched_Items", matched_aliases, "invoice"
        )
        tolerance_ok = True
        matched_actual = {}
        for row, values in matched_rows:
            matched_actual[values["invoice"]] = tuple(
                values[field]
                for field in ("ledger", "subledger", "adjustment", "adjusted", "difference")
            )
            tolerance_ok = (
                tolerance_ok
                and close(values["difference"], 0.0, 1.0)
                and normalize_reconciliation_status(values["status"]) == "MATCHED"
            )
        matched_expected = {
            record[0]: tuple(comparable(value) for value in record[1:])
            for record in expected["matched"]
        }
        matched_ok = (
            len(matched_rows) == len(matched_expected)
            and len(matched_actual) == len(matched_rows)
            and matched_actual == matched_expected
        )
        matched_formulas = all(
            all(
                formula_cell(workbook, "Matched_Items", row, matched_columns[field])
                for field in ("ledger", "subledger", "adjustment", "adjusted", "difference", "status")
            )
            for row, _ in matched_rows
        )
    except Exception as exc:
        matched_ok = tolerance_ok = matched_formulas = False
        failures.append(f"MATCH_REPLAY_FAILED:{type(exc).__name__}")
    checks["R003"] = 1.0 if matched_ok and tolerance_ok else 0.0

    try:
        bridge_specs = (
            (("ledger total",), expected["ledger_total"]),
            (("adjusted matched subledger", "matched total"), expected["matched_total"]),
            (("ledger only exception", "ledger only exceptions"), expected["ledger_only"]),
            (("subledger only exception", "subledger only exceptions"), expected["subledger_only"]),
            (("net investigation difference", "investigation amount"), expected["investigation"]),
        )
        bridge_ok = True
        bridge_formulas = True
        for aliases, expected_value in bridge_specs:
            value, row, column = metric_value(workbook, engine, "Variance_Bridge", aliases)
            bridge_ok = bridge_ok and close(value, expected_value)
            bridge_formulas = bridge_formulas and formula_cell(
                workbook, "Variance_Bridge", row, column
            )

        final_specs = (
            (("matched invoices", "matched invoice count"), (len(expected["matched"]),)),
            (
                ("ledger only exception", "ledger only exceptions"),
                (sum(row[1] == "Ledger" for row in expected["unmatched"]), expected["ledger_only"]),
            ),
            (
                ("subledger only exception", "subledger only exceptions"),
                (sum(row[1] == "Subledger" for row in expected["unmatched"]), expected["subledger_only"]),
            ),
            (("investigation amount", "net investigation difference"), (expected["investigation"],)),
        )
        final_ok = True
        final_formulas = True
        for aliases, accepted_values in final_specs:
            value, row, column = metric_value(workbook, engine, "Final_Reconciliation", aliases)
            final_ok = final_ok and any(close(value, item) for item in accepted_values)
            final_formulas = final_formulas and formula_cell(
                workbook, "Final_Reconciliation", row, column
            )
    except Exception as exc:
        bridge_ok = final_ok = bridge_formulas = final_formulas = False
        failures.append(f"RECONCILIATION_CLOSURE_FAILED:{type(exc).__name__}")
    checks["R004"] = 1.0 if bridge_ok and final_ok else 0.0

    try:
        invoice = context["drivers"]["fx_invoice"]
        normalized_columns, normalized_rows = table_records(
            workbook, engine, "Normalized_Records", normalized_aliases, "record_id"
        )
        ledger_row = next(row for row, values in normalized_rows if values["source"] == "Ledger" and values["invoice"] == invoice)
        subledger_row = next(row for row, values in normalized_rows if values["source"] == "Subledger" and values["invoice"] == invoice)
        matched_columns, matched_rows = table_records(
            workbook, engine, "Matched_Items", matched_aliases, "invoice"
        )
        matched_row = next(row for row, values in matched_rows if values["invoice"] == invoice)
        source = context["drivers"]["fx_source_cell"]
        source_sheet, source_cell = source.split("!", 1)
        baseline_rate = workbook[source_sheet][source_cell].value
        new_rate = context["drivers"]["fx_new_value"]
        original = engine.value("Normalized_Records", f"{col_name(normalized_columns['original'])}{ledger_row}")
        expected_usd = original * new_rate
        delta = original * (new_rate - baseline_rate)
        perturb = FormulaEngine(workbook, {source: new_rate})
        ledger_total_row, ledger_total_column = metric_cell(workbook, "Variance_Bridge", ("ledger total",))
        matched_total_row, matched_total_column = metric_cell(workbook, "Variance_Bridge", ("adjusted matched subledger", "matched total"))
        fx_dynamic = (
            close(perturb.value("Normalized_Records", f"{col_name(normalized_columns['usd'])}{ledger_row}"), expected_usd)
            and close(perturb.value("Normalized_Records", f"{col_name(normalized_columns['usd'])}{subledger_row}"), expected_usd)
            and close(perturb.value("Matched_Items", f"{col_name(matched_columns['difference'])}{matched_row}"), 0)
            and close(perturb.value("Variance_Bridge", f"{col_name(ledger_total_column)}{ledger_total_row}"), expected["ledger_total"] + delta)
            and close(perturb.value("Variance_Bridge", f"{col_name(matched_total_column)}{matched_total_row}"), expected["matched_total"] + delta)
        )
    except Exception as exc:
        fx_dynamic = False; failures.append(f"FX_SOURCE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R005"] = 1.0 if normalized_ok else 0.0

    try:
        matched_columns, matched_rows = table_records(
            workbook, engine, "Matched_Items", matched_aliases, "invoice"
        )
        row = next(row for row, values in matched_rows if values["invoice"] == context["drivers"]["adjustment_invoice"])
        new_adjustment = context["drivers"]["adjustment_new_value"]
        perturb = FormulaEngine(workbook, {context["drivers"]["adjustment_source_cell"]: new_adjustment})
        adjustment_dynamic = (
            close(perturb.value("Matched_Items", f"{col_name(matched_columns['adjustment'])}{row}"), new_adjustment)
            and close(perturb.value("Matched_Items", f"{col_name(matched_columns['adjusted'])}{row}"), context["drivers"]["adjusted_subledger"])
            and close(perturb.value("Matched_Items", f"{col_name(matched_columns['difference'])}{row}"), context["drivers"]["adjustment_difference"])
            and normalize_reconciliation_status(perturb.value("Matched_Items", f"{col_name(matched_columns['status'])}{row}")) == "OUT_OF_TOLERANCE"
        )
    except Exception as exc:
        adjustment_dynamic = False; failures.append(f"ADJUSTMENT_SOURCE_PERTURBATION_FAILED:{type(exc).__name__}")
    checks["R006"] = 1.0 if matched_ok else 0.0

    try:
        unmatched_aliases = {
            "invoice": ("invoice", "invoice id"),
            "source": ("source", "source system"),
            "amount": ("usd amount", "amount usd", "normalized amount"),
        }
        _, unmatched_rows = table_records(
            workbook, engine, "Unmatched_Items", unmatched_aliases, "invoice"
        )
        unmatched = {
            (values["invoice"], values["source"], values["amount"])
            for _, values in unmatched_rows
        }
        expected_unmatched = {
            (row[0], row[1], comparable(row[2])) for row in expected["unmatched"]
        }
        exceptions_ok = (
            len(unmatched_rows) == len(expected_unmatched)
            and len(unmatched) == len(unmatched_rows)
            and unmatched == expected_unmatched
        )
        final_exception_formulas = all(
            formula_cell(workbook, "Final_Reconciliation", *metric_cell(workbook, "Final_Reconciliation", aliases))
            for aliases in (
                ("matched invoices", "matched invoice count"),
                ("ledger only exception", "ledger only exceptions"),
                ("subledger only exception", "subledger only exceptions"),
                ("investigation amount", "net investigation difference"),
            )
        )
    except Exception as exc:
        exceptions_ok = final_exception_formulas = False
        failures.append(f"EXCEPTION_CLOSURE_FAILED:{type(exc).__name__}")
    checks["R007"] = 1.0 if exceptions_ok and final_ok else 0.0

    protected_ok = source_tables_match(workbook, context["source_root"])
    checks["R008"] = 1.0 if protected_ok else 0.0
    if not protected_ok: failures.append("PROTECTED_FINANCE_SOURCE_CHANGED")

    try:
        check_specs = (
            (("matched item differences", "maximum matched absolute difference", "matched difference"), 0),
            (("unmatched exception count", "exception count", "exceptions"), len(expected["unmatched"])),
            (("final closure residual", "exception bridge closure", "fx date sensitivity", "ledger normalization tie out", "subledger normalization tie out"), 0),
        )
        check_values_ok = True
        check_formulas = True
        for aliases, expected_value in check_specs:
            value, row, column = metric_value(workbook, engine, "Checks", aliases)
            check_values_ok = check_values_ok and close(value, expected_value)
            check_formulas = check_formulas and formula_cell(workbook, "Checks", row, column)
    except Exception as exc:
        check_values_ok = check_formulas = False
        failures.append(f"CHECK_LINKAGE_FAILED:{type(exc).__name__}")
    checks["R009"] = 1.0 if check_values_ok else 0.0

    return checks, failures


def evaluate(candidate, split="dev"):
    criteria = {row["id"]: 0.0 for row in TASK["criteria"]}
    if not candidate.exists() or candidate.stat().st_size == 0: return criteria, ["OUTPUT_MISSING"]
    try: workbook = openpyxl.load_workbook(candidate, data_only=False, read_only=False)
    except Exception as exc: return criteria, [f"MALFORMED_XLSX:{type(exc).__name__}"]
    context = split_context(split)
    oracle = load_oracle(context)
    try:
        # Judge V3 has one semantic source of truth.  The fixed-layout replay
        # remains below as migration/reference code, but is deliberately not
        # merged criterion-by-criterion: two inconsistent representations
        # must never combine into a passing reconciliation.
        return semantic_finance_checks(workbook, oracle, context)
    except Exception as exc:
        return criteria, [f"SEMANTIC_EVALUATION_ERROR:{type(exc).__name__}:{exc}"]



def parse_cli():
    split = os.environ.get("P15_EVAL_SPLIT", "dev").strip().lower()
    candidate = None
    arguments = iter(sys.argv[1:])
    for argument in arguments:
        if argument == "--split": split = next(arguments, "")
        elif argument.startswith("--split="): split = argument.split("=", 1)[1]
        elif candidate is None: candidate = Path(argument)
        else: raise ValueError(f"UNEXPECTED_ARGUMENT:{argument}")
    if split not in {"dev", "confirm"}: raise ValueError(f"INVALID_SPLIT:{split}")
    return candidate or Path("/app/output/answer.xlsx"), split


def main():
    candidate, split = parse_cli()
    criteria, failures = evaluate(candidate, split)
    payload = build_result(task=TASK, split=split, candidate=str(candidate), criteria=criteria, failures=failures)
    payload["judge_version"] = "P15_JUDGE_V3"
    total = payload["normalized_score"]
    log_root = Path(os.environ.get("P15_VERIFIER_LOG_DIR", "/logs/verifier"))
    try:
        log_root.mkdir(parents=True, exist_ok=True); (log_root / "report.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
        if total is not None: (log_root / "reward.txt").write_text(str(total) + "\n")
    except OSError: pass
    print(json.dumps(payload, indent=2, sort_keys=True))


if __name__ == "__main__": main()

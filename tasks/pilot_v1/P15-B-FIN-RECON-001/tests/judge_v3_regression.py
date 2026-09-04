#!/usr/bin/env python3
"""Focused Judge V3 regression: equivalent layout plus real business errors."""
from __future__ import annotations

import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

import evaluate as judge

ROOT = Path(__file__).resolve().parents[1]


def run_case(path: Path, split: str = "dev") -> dict:
    criteria, failures = judge.evaluate(path, split)
    payload = judge.build_result(
        task=judge.TASK, split=split, candidate=str(path), criteria=criteria, failures=failures
    )
    return {
        "score": payload["normalized_score"], "pass": payload["pass"],
        "criteria": payload["criterion_scores"], "failure_codes": payload["failure_codes"],
    }


def alternative_layout(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Checks"].title = "Integrity Checks"
    workbook.save(path)


def add_out_of_period_match(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Matched_Items"].append((
        "INV-102", 500, 500, 0, 500, 0, "MATCHED", "July pair incorrectly included in June",
    ))
    workbook.save(path)


def duplicate_match(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    sheet = workbook["Matched_Items"]
    sheet.append(tuple(cell.value for cell in sheet[4]))
    workbook.save(path)


def duplicate_unmatched(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    sheet = workbook["Unmatched_Items"]
    sheet.append(tuple(cell.value for cell in sheet[4]))
    workbook.save(path)


def wrong_bridge_with_keywords(path: Path) -> None:
    workbook = openpyxl.load_workbook(ROOT / "solution/reference.xlsx")
    workbook["Variance_Bridge"]["B4"] = 999999
    workbook.save(path)


def main() -> None:
    cases = {
        "reference": (ROOT / "solution/reference.xlsx", "dev", True),
        "confirm_reference": (ROOT / "tests/confirm/reference.xlsx", "confirm", True),
        "equivalent": (ROOT / "fixtures/equivalent/candidate_equivalent.xlsx", "dev", True),
        "noop": (ROOT / "fixtures/noop/candidate_noop.xlsx", "dev", False),
        "malformed": (ROOT / "fixtures/malformed/candidate_malformed.xlsx", "dev", False),
    }
    cases.update({
        f"mutant:{path.stem}": (path, "dev", False)
        for path in sorted((ROOT / "fixtures/mutants").glob("*.xlsx"))
    })
    results = {}
    with tempfile.TemporaryDirectory() as directory:
        alternative = Path(directory) / "renamed_checks.xlsx"
        out_of_period = Path(directory) / "july_match.xlsx"
        duplicated_match = Path(directory) / "duplicate_match.xlsx"
        duplicated_unmatched = Path(directory) / "duplicate_unmatched.xlsx"
        wrong_bridge = Path(directory) / "wrong_bridge_with_keywords.xlsx"
        alternative_layout(alternative)
        add_out_of_period_match(out_of_period)
        duplicate_match(duplicated_match)
        duplicate_unmatched(duplicated_unmatched)
        wrong_bridge_with_keywords(wrong_bridge)
        cases["alternative_layout:renamed_checks"] = (alternative, "dev", True)
        cases["negative:out_of_period_match"] = (out_of_period, "dev", False)
        cases["negative:duplicate_match"] = (duplicated_match, "dev", False)
        cases["negative:duplicate_unmatched"] = (duplicated_unmatched, "dev", False)
        cases["negative:wrong_bridge_keywords_do_not_close"] = (wrong_bridge, "dev", False)
        for name, (path, split, expected_pass) in cases.items():
            result = run_case(path, split)
            assert result["pass"] is expected_pass, (name, result)
            results[name] = {**result, "expected_pass": expected_pass, "meets_expectation": True}
    receipt = {
        "task_id": judge.TASK["task_id"], "judge_version": "P15_JUDGE_V3",
        "generated_at": datetime.now(timezone.utc).isoformat(), "cases": results,
        "all_expectations_met": True,
    }
    (ROOT / "receipts/judge_v3_local_validation.json").write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n"
    )
    print(json.dumps(receipt, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

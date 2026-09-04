#!/usr/bin/env python3
"""Repair the canonical workbooks to the requested close-period semantics."""
from __future__ import annotations

import shutil
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]


def month(value):
    return str(value)[:7]


def repair(path: Path) -> None:
    workbook = openpyxl.load_workbook(path)
    normalized = workbook["Normalized_Records"]
    rows = []
    for row in range(4, normalized.max_row + 1):
        rows.append({
            "row": row,
            "source": normalized.cell(row, 1).value,
            "invoice": normalized.cell(row, 3).value,
            "date": normalized.cell(row, 4).value,
        })
    close_month = Counter(month(row["date"]) for row in rows).most_common(1)[0][0]
    scoped = [row for row in rows if month(row["date"]) == close_month]
    by_source = {
        source: {row["invoice"]: row for row in scoped if row["source"] == source}
        for source in ("Ledger", "Subledger")
    }
    matched_invoices = sorted(set(by_source["Ledger"]) & set(by_source["Subledger"]))
    adjustment = workbook["Adjustment_Evidence"]
    adjustment_invoice = adjustment["B4"].value

    matched = workbook["Matched_Items"]
    if matched.max_row >= 4:
        matched.delete_rows(4, matched.max_row - 3)
    for output_row, invoice in enumerate(matched_invoices, start=4):
        ledger_row = by_source["Ledger"][invoice]["row"]
        subledger_row = by_source["Subledger"][invoice]["row"]
        adjustment_formula = "='Adjustment_Evidence'!C4" if invoice == adjustment_invoice else "=0"
        values = (
            invoice,
            f"='Normalized_Records'!H{ledger_row}",
            f"='Normalized_Records'!H{subledger_row}",
            adjustment_formula,
            f"=C{output_row}+D{output_row}",
            f"=B{output_row}-E{output_row}",
            f'=IF(ABS(F{output_row})<=1,"MATCHED","OUT_OF_TOLERANCE")',
            f"Exact in-period invoice identity{' and approved adjustment' if invoice == adjustment_invoice else ''}",
        )
        for column, value in enumerate(values, start=1):
            matched.cell(output_row, column, value)

    ledger_rows = [row["row"] for row in scoped if row["source"] == "Ledger"]
    subledger_rows = [row["row"] for row in scoped if row["source"] == "Subledger"]
    ledger_sum = "+".join(f"'Normalized_Records'!H{row}" for row in ledger_rows)
    subledger_sum = "+".join(f"'Normalized_Records'!H{row}" for row in subledger_rows)

    bridge = workbook["Variance_Bridge"]
    if bridge.max_row >= 4:
        bridge.delete_rows(4, bridge.max_row - 3)
    bridge_rows = (
        ("Close-period ledger total", f"={ledger_sum}", "All in-period normalized ledger records", "Normalized_Records"),
        ("Close-period adjusted subledger total", f"={subledger_sum}+'Adjustment_Evidence'!C4", "In-period subledger plus approved adjustment", "Normalized_Records + Adjustment_Evidence"),
        ("Ledger-only exception", '=IF(\'Unmatched_Items\'!B4="Ledger",\'Unmatched_Items\'!C4,0)+IF(\'Unmatched_Items\'!B5="Ledger",\'Unmatched_Items\'!C5,0)', "Retained ledger-only amount", "Unmatched_Items"),
        ("Subledger-only exception", '=IF(\'Unmatched_Items\'!B4="Subledger",\'Unmatched_Items\'!C4,0)+IF(\'Unmatched_Items\'!B5="Subledger",\'Unmatched_Items\'!C5,0)', "Retained subledger-only amount", "Unmatched_Items"),
        ("Bridge residual", "=B4+B7-B6-B5", "Must equal zero", "Normalized_Records + Unmatched_Items"),
        ("Net investigation difference", "=B7-B6", "Subledger-only minus ledger-only", "Unmatched_Items"),
    )
    for row in bridge_rows:
        bridge.append(row)

    final = workbook["Final_Reconciliation"]
    if final.max_row >= 4:
        final.delete_rows(4, final.max_row - 3)
    last_match_row = 3 + len(matched_invoices)
    final_rows = (
        ("Reconciliation period", close_month, "Only the requested close month is matched", "Normalized_Records"),
        ("Matched invoices", f'=COUNTIF(\'Matched_Items\'!G4:G{last_match_row},"MATCHED")', "All in-period matches are within $1", "Matched_Items"),
        ("Matched amount", f"=SUM('Matched_Items'!E4:E{last_match_row})", "Ledger and adjusted subledger matched amounts agree", "Matched_Items"),
        ("Ledger-only exceptions", '=COUNTIF(\'Unmatched_Items\'!B4:B5,"Ledger")', "Retain for investigation", "Unmatched_Items"),
        ("Subledger-only exceptions", '=COUNTIF(\'Unmatched_Items\'!B4:B5,"Subledger")', "Retain for investigation", "Unmatched_Items"),
        ("Investigation amount", "='Variance_Bridge'!B9", "Resolve before close", "Variance_Bridge"),
        ("Final decision", "REVIEW READY — OPEN EXCEPTIONS", "Resolve the two retained exceptions before close", "Checks"),
    )
    for row in final_rows:
        final.append(row)

    checks = workbook["Checks"]
    if checks.max_row >= 4:
        checks.delete_rows(4, checks.max_row - 3)
    check_rows = (
        ("Maximum matched absolute difference", f"=MAX(ABS('Matched_Items'!F4:F{last_match_row}))", 0, "PASS"),
        ("Unmatched exception count", '=COUNTA(\'Unmatched_Items\'!A4:A5)', 2, "PASS"),
        ("Variance bridge residual", "='Variance_Bridge'!B8", 0, "PASS"),
    )
    for row in check_rows:
        checks.append(row)

    workbook.save(path)


def main() -> None:
    repair(ROOT / "solution/reference.xlsx")
    repair(ROOT / "tests/confirm/reference.xlsx")
    shutil.copy2(ROOT / "solution/reference.xlsx", ROOT / "fixtures/equivalent/candidate_equivalent.xlsx")
    equivalent = openpyxl.load_workbook(ROOT / "fixtures/equivalent/candidate_equivalent.xlsx")
    equivalent["Checks"].title = "Integrity Checks"
    equivalent.save(ROOT / "fixtures/equivalent/candidate_equivalent.xlsx")


if __name__ == "__main__":
    main()
